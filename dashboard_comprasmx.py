"""
Dashboard de Integridad en Contrataciones Públicas - ComprasMX 2026
Álvaro Quintero Casillas | División de Monitoreo de la Integridad Institucional | IMSS

Instrucciones:
1. Coloca este archivo en la misma carpeta que tu CSV (contratos_comprasmx_2026.csv)
2. Abre tu terminal en esa carpeta y ejecuta:
   pip install streamlit plotly pandas openpyxl python-dateutil
   streamlit run dashboard_comprasmx.py
"""

import json
import re as _re
from datetime import date as _date, datetime as _datetime
from pathlib import Path

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# ─────────────────────────────────────────────
# PALETA CROMÁTICA INSTITUCIONAL IMSS
# ─────────────────────────────────────────────
IMSS_VERDE       = "#0B5445"   # PANTONE IMSS 561 C
IMSS_VERDE_OSC   = "#0F2621"   # PANTONE 627 C
IMSS_ROJO        = "#8C1033"   # PANTONE 7420 C
IMSS_ROJO_OSC    = "#541024"   # PANTONE 7421 C
IMSS_NEGRO       = "#171B19"   # PANTONE Neutral Black C
IMSS_GRIS        = "#868688"   # PANTONE Cool Gray C
IMSS_ORO_CLARO   = "#E8D188"   # PANTONE 7402 C
IMSS_ORO         = "#9D7119"   # PANTONE 1255 C

COLORES_TIPO = {
    "Licitación Pública":              IMSS_VERDE,
    "Invitación a 3 personas":         IMSS_ORO,
    "Adjudicación Directa":            IMSS_ROJO,
    "Adjudicación Directa — Fr. I":    "#C05078",  # Rosa-rojo: AD estructural por patente/exclusividad
    "Adjudicación Directa — Patentes": "#C05078",  # Alias de display para la misma categoría
    "Entre Entes Públicos":            IMSS_ORO_CLARO,
    "Sin clasificar":                  IMSS_GRIS,
    "Otro":                            IMSS_GRIS,
}

# ─────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Dashboard de Integridad – ComprasMX 2026",
    page_icon="🔍",
    layout="wide",
    menu_items={}
)

# CSS: tipografía NotoSans + estilos institucionales
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans:wght@300;400;600;700&display=swap');

html, body, [class*="css"], .stApp, .stMarkdown, .stMetric,
.stSelectbox, .stNumberInput, h1, h2, h3, p, div {
    font-family: 'Noto Sans', sans-serif !important;
}

/* Fondo general */
.stApp {
    background-color: #F5F5F5;
}

/* Sidebar */
section[data-testid="stSidebar"] {
    background-color: #0F2621 !important;
}
section[data-testid="stSidebar"] * {
    color: #E8D188 !important;
    /* font-family NO se hereda aquí para no sobreescribir la fuente
       de iconos Material Symbols que usa el botón de colapso */
}

/* Título principal */
h1 { color: #0B5445 !important; font-weight: 700 !important; }
h2, h3 { color: #0F2621 !important; font-weight: 600 !important; }

/* Tarjetas de métricas */
[data-testid="stMetric"] {
    background-color: #ffffff;
    border-left: 5px solid #0B5445;
    border-radius: 6px;
    padding: 12px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
}
[data-testid="stMetricLabel"] {
    color: #0F2621 !important;
    font-weight: 600 !important;
    white-space: normal !important;
    word-break: break-word !important;
    font-size: 0.82rem !important;
    line-height: 1.3 !important;
}
[data-testid="stMetricValue"] { color: #0B5445 !important; font-weight: 700 !important; font-size: 1.35rem !important; }
[data-testid="stMetricDelta"]  { font-weight: 600 !important; }

/* Divisor */
hr { border-color: #0B5445; opacity: 0.3; }

/* Tabs */
[data-testid="stTab"] {
    font-family: 'Noto Sans', sans-serif !important;
    font-weight: 600 !important;
}

/* Ocultar branding de Streamlit */
#MainMenu                        { visibility: hidden !important; }
footer                           { visibility: hidden !important; }
[data-testid="stToolbar"]        { visibility: hidden !important; }
[data-testid="stDecoration"]     { display: none !important; }
[data-testid="stStatusWidget"]   { display: none !important; }

/* Restaurar controles de colapso/expansión del sidebar.
   Todos estos botones viven DENTRO de stToolbar y heredan su
   visibility:hidden. Se restauran explícitamente para que las
   flechas << >> funcionen en escritorio y móvil.
   - stExpandSidebarButton : botón >> para ABRIR el sidebar (Streamlit ≥1.44)
   - stSidebarCollapseButton: botón << para CERRAR el sidebar (dentro del sidebar)
   - collapsedControl       : nombre alternativo en versiones anteriores */
[data-testid="stExpandSidebarButton"],
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapseButton"] {
    visibility: visible !important;
    pointer-events: auto !important;
}
</style>
""", unsafe_allow_html=True)

st.title("🔍 Dashboard de Integridad en Contrataciones Públicas")

# ─────────────────────────────────────────────
# CARGA DE DATOS
# ─────────────────────────────────────────────
@st.cache_data
def cargar_datos(filename):
    df = pd.read_csv(filename, encoding="utf-8", low_memory=False)
    df["Importe DRC"] = pd.to_numeric(df["Importe DRC"], errors="coerce")
    # Normalizar Institución a MAYÚSCULAS — formato antiguo usa Title Case,
    # nuevo usa MAYÚSCULAS. Así el filtro del sidebar funciona con todos los años.
    if "Institución" in df.columns:
        df["Institución"] = df["Institución"].str.strip().str.upper()
    if "Partida específica" not in df.columns:
        df["Partida específica"] = ""
    df["Partida específica"] = df["Partida específica"].astype(str).str.strip().str.zfill(5)
    # "00000" proviene de celdas vacías → dejar vacío
    df.loc[df["Partida específica"] == "00000", "Partida específica"] = ""
    def clasificar(tipo):
        if pd.isna(tipo):
            return "Sin clasificar"
        t = str(tipo).upper()
        if "LICITACIÓN PÚBLICA" in t:
            return "Licitación Pública"
        elif "INVITACIÓN" in t:
            return "Invitación a 3 personas"
        elif "ADJUDICACIÓN" in t:
            return "Adjudicación Directa"
        elif "ENTRE ENTES" in t:
            return "Entre Entes Públicos"
        return "Otro"
    df["Tipo Simplificado"] = df["Tipo Procedimiento"].apply(clasificar)
    # Separar ADs por patente/exclusividad (Art. 54 Fr. I LAASSP vigente
    # o Art. 41 Fr. I de la ley anterior) del resto de adjudicaciones directas
    _col_exc = "Artículo de excepción"
    if _col_exc in df.columns:
        _exc_upper = df[_col_exc].astype(str).str.upper().str.strip()
        _mask_fri = (
            (df["Tipo Simplificado"] == "Adjudicación Directa") &
            _exc_upper.isin({"ART. 54 FR. I", "ART. 41 FR. I"})
        )
        df.loc[_mask_fri, "Tipo Simplificado"] = "Adjudicación Directa — Fr. I"
    return df

@st.cache_data
def cargar_cucop():
    df_c = pd.read_excel("cucop_20260301.xlsx", sheet_name="cucop", dtype=str)
    df_c.columns = df_c.columns.str.strip()
    for col in ["PARTIDA ESPECÍFICA", "DESC. PARTIDA ESPECÍFICA",
                "PARTIDA GENÉRICA", "DESC. PARTIDA GENÉRICA",
                "CONCEPTO", "DESC. CONCEPTO", "CAPÍTULO", "DESC. CAPÍTULO"]:
        if col in df_c.columns:
            df_c[col] = df_c[col].str.strip()
    df_c["PARTIDA ESPECÍFICA"] = df_c["PARTIDA ESPECÍFICA"].str.zfill(5)
    return df_c

df_cucop = cargar_cucop()

@st.cache_data
def cargar_directorio_uc():
    try:
        df_uc = pd.read_excel("Base_UC_2025_V2.xlsx", dtype=str)
        df_uc.columns = df_uc.columns.str.strip()
        df_uc = df_uc.dropna(subset=["Clave_UC"])
        df_uc["Clave_UC"]       = df_uc["Clave_UC"].str.strip()
        df_uc["Nombre_editado"] = df_uc["Nombre_editado"].str.strip()
        # Normalizar Tipo_UC: quitar espacios y estandarizar "NIVEL CENTRAL" → "Nivel Central"
        # (OOAD y UMAE quedan en mayúsculas tal como están en el archivo)
        df_uc["Tipo_UC"] = (
            df_uc["Tipo_UC"]
            .str.strip()
            .replace("NIVEL CENTRAL", "Nivel Central")
            .fillna("Sin clasificar")
        )
        # Crear columna Adscripción con tres niveles de fallback:
        #  ① columna OOAD válida             → usar ese valor (estado de la delegación)
        #  ② OOAD/UMAE sin valor en col OOAD → usar Entidad federativa
        #     (cubre: Centro Vacacional Oaxtepec y las 35 UMAEs sin OOAD territorial)
        #  ③ Nivel Central                   → literal "Nivel Central"
        _ooad    = df_uc["OOAD"].str.strip()
        _entidad = df_uc["Entidad federativa"].str.strip()
        _ooad_valido = _ooad.notna() & (_ooad.str.upper() != "NA") & (_ooad != "")
        df_uc["Adscripción"] = (
            _ooad.where(_ooad_valido,                                # ① OOAD válida
            _entidad.where(df_uc["Tipo_UC"] != "Nivel Central",     # ② OOAD/UMAE → Entidad fed.
            "Nivel Central"))                                         # ③ Nivel Central
            .fillna("Nivel Central")                                  # seguridad ante NaN residuales
        )
        # Renombrar Tipo_UC → Tipo UC para compatibilidad con el resto del código
        df_uc = df_uc.rename(columns={"Tipo_UC": "Tipo UC"})
        return df_uc
    except FileNotFoundError:
        return pd.DataFrame(columns=["Clave_UC", "Nombre_editado", "Tipo UC", "Adscripción"])

df_dir_uc = cargar_directorio_uc()

@st.cache_data
def cargar_precios_unitarios():
    """Carga la hoja 'Insumo principal' de AnaliticaPreciosUnitarios.xlsx."""
    df_pu = pd.read_excel(
        "AnaliticaPreciosUnitarios.xlsx",
        sheet_name="Insumo principal",
        dtype=str
    )
    df_pu.columns = df_pu.columns.str.strip()

    # Convertir columnas numéricas
    _num_cols = [
        "Precio unitario", "Monto partida", "Cantidad",
        "Mediana (P)", "Límite inferior (P)", "Límite superior (P)",
        "Precio estandarizado", "Cantidad estandarizada",
        "Índice de alto riesgo", "Montos analizados",
        "Muestra significativa", "Reciente creación",
    ]
    for _c in _num_cols:
        if _c in df_pu.columns:
            df_pu[_c] = pd.to_numeric(df_pu[_c], errors="coerce")

    # Normalizar columnas de texto
    _str_cols = [
        "Precio atípico", "Cantidad atípica", "Consolidada",
        "Tipo de proveedor por historial", "Caso de atención crítico",
        "Fuente Compras MX", "Dentro de la fecha de análisis",
        "Clave UC", "RFC del proveedor adjudicado", "UC", "Proveedor",
    ]
    for _c in _str_cols:
        if _c in df_pu.columns:
            df_pu[_c] = df_pu[_c].str.strip()

    return df_pu


# ─────────────────────────────────────────────
# FUNCIONES DE DATOS — INDICADORES DE RIESGO
# ─────────────────────────────────────────────
@st.cache_data
def cargar_sancionados():
    import openpyxl
    from datetime import datetime, date
    from dateutil.relativedelta import relativedelta

    wb = openpyxl.load_workbook(
        "BD_01_2026__Me_trica_empresas_inhabilitadas__DOF.xlsm",
        read_only=True, keep_vba=True
    )
    ws = wb['Empresas_inhabilitadas']
    rows = list(ws.iter_rows(values_only=True))
    df_s = pd.DataFrame(rows[1:], columns=rows[0])
    df_s = df_s[df_s['RFC de la empresa o persona fisica'].notna()].copy()

    hoy = date.today()

    def calcular_nivel(row):
        try:
            inicio = row['Fecha de inicio de inhabilitación ']
            if pd.isna(inicio):
                return "Sin datos"
            if hasattr(inicio, 'date'):
                inicio = inicio.date()

            meses = int(row['Periodo de inhabilitación (meses)']) if pd.notna(row['Periodo de inhabilitación (meses)']) else 0
            dias  = int(row['Periodo de inhabilitación (días)'])  if pd.notna(row['Periodo de inhabilitación (días)'])  else 0

            from dateutil.relativedelta import relativedelta
            fecha_termino = inicio + relativedelta(months=meses) + pd.Timedelta(days=dias)
            if hasattr(fecha_termino, 'date'):
                fecha_termino = fecha_termino.date()

            circular_term = row.get('Circular de término de inhabilitación')
            terminado = (
                (circular_term and str(circular_term).strip() not in ['', 'nan', 'None'])
                or fecha_termino <= hoy
            )

            fecha_susp    = row.get('Fecha de suspensión')
            circular_susp = row.get('Circular de suspensión de inhabilitación')
            suspendido = (
                (pd.notna(fecha_susp) and str(fecha_susp).strip() not in ['', 'NaT'])
                or (circular_susp and str(circular_susp).strip() not in ['', 'nan', 'None'])
            )

            if terminado:
                return "🟡 Riesgo medio — Historial de inhabilitación"
            elif suspendido:
                return "🟠 Riesgo alto — Inhabilitación suspendida judicialmente"
            else:
                return "🔴 Riesgo crítico — Inhabilitación vigente"

        except Exception:
            return "Sin datos"

    df_s['Nivel de Riesgo'] = df_s.apply(calcular_nivel, axis=1)
    df_s = df_s.rename(columns={
        'Nombre de la empresa o persona fisica': 'Empresa',
        'RFC de la empresa o persona fisica':    'RFC',
        'Fecha de inicio de inhabilitación ':    'Inicio inhabilitación',
        'Periodo de inhabilitación (meses)':     'Meses',
    })
    return df_s[['Empresa', 'RFC', 'Inicio inhabilitación', 'Meses', 'Nivel de Riesgo']]

@st.cache_data
def cargar_efos():
    df_efos = pd.read_csv(
        "Listado_completo_69-B.csv",
        encoding="utf-8",   # versión slim generada por etl_slim.py (sin filas de encabezado SAT)
        dtype=str
    )
    df_efos.columns = df_efos.columns.str.strip()
    df_efos["RFC"] = df_efos["RFC"].str.strip().str.upper()
    df_efos["Situación del contribuyente"] = df_efos["Situación del contribuyente"].str.strip()
    return df_efos

@st.cache_data
def cargar_umbrales_pef():
    """Carga umbrales máximos PEF para AD e I3P (LAASSP y LOPSRM) desde UmbralPEF.xlsx."""
    try:
        df_u = pd.read_excel("UmbralPEF.xlsx", sheet_name="Hoja1", dtype=str)
        df_u.columns = df_u.columns.str.strip()
        result = {}
        def _pm(val):
            return float(str(val).replace(",", "").replace(" ", "").strip())
        for _, row in df_u.iterrows():
            try:
                año = int(str(row.get("Año", "")).strip())
                result[año] = {
                    "ad_laassp":       _pm(row["Adjudicación Directa - LAASSP"]),
                    "i3p_laassp":      _pm(row["Invitación cuando menos a tres personas - LAASSP"]),
                    "ad_obra_lopsrm":  _pm(row["Adjudicación Directa - Obra Pública LOPSRM"]),
                    "ad_serv_lopsrm":  _pm(row["Adjudicación Directa - Servicios LOPSRM"]),
                    "i3p_obra_lopsrm": _pm(row["Invitación cuando menos a tres personas - Obra Pública LOPSRM"]),
                    "i3p_serv_lopsrm": _pm(row["Invitación cuando menos a tres personas - Servicios LOPSRM"]),
                }
            except Exception:
                continue
        return result
    except FileNotFoundError:
        return {}
    except Exception:
        return {}


def nivel_efos(sit):
    s = str(sit).strip()
    if s == "Definitivo":
        return "🔴 EFOS definitivo — Operaciones simuladas confirmadas"
    elif s == "Presunto":
        return "🟡 EFOS presunto — Proceso en curso"
    elif s == "Desvirtuado":
        return "🟢 Desvirtuado — Sin riesgo"
    elif "Sentencia" in s:
        return "🟢 Sentencia favorable — Sin riesgo"
    return "⚪ Sin clasificar"


# Helper para aplicar fuente en figuras Plotly
def plotly_font():
    return dict(family="Noto Sans, sans-serif", size=13, color=IMSS_NEGRO)

# ─────────────────────────────────────────────
# BARRA LATERAL – FILTROS
# ─────────────────────────────────────────────
st.sidebar.header("⚙️ Filtros")

_ARCHIVOS_ANIO = {
    "2026": "contratos_comprasmx_2026.csv",
    "2025": "contratos_comprasmx_2025.csv",
    "2024": "contratos_compranet_2024.csv",
    "2023": "contratos_compranet_2023.csv",
}
_anios_disponibles = list(_ARCHIVOS_ANIO.keys())
anios_sel = st.sidebar.multiselect(
    "📅 Año(s) de datos",
    options=_anios_disponibles,
    default=["2026"],
    help="Selecciona uno o más años para analizar en conjunto. Al combinar años se habilita el análisis comparativo."
)
# Garantizar que siempre haya al menos un año seleccionado
if not anios_sel:
    anios_sel = ["2026"]
    st.sidebar.warning("Selecciona al menos un año.")

# Cargar y concatenar los DataFrames de los años seleccionados
_dfs_anio = []
for _a in anios_sel:
    _archivo = _ARCHIVOS_ANIO.get(_a)
    if _archivo:
        try:
            _df_a = cargar_datos(_archivo)
            _df_a["Año"] = _a  # columna para distinguir origen
            _dfs_anio.append(_df_a)
        except Exception:
            st.sidebar.warning(f"No se pudo cargar el archivo de {_a}.")

df = pd.concat(_dfs_anio, ignore_index=True) if _dfs_anio else cargar_datos(_ARCHIVOS_ANIO["2026"])

# Aplicar nombres editados de UC desde Base_UC_2025_V2.xlsx a todo el dashboard
if len(df_dir_uc) > 0 and "Clave_UC" in df_dir_uc.columns:
    _uc_nombre_map = (
        df_dir_uc.dropna(subset=["Nombre_editado"])
        .set_index("Clave_UC")["Nombre_editado"]
        .to_dict()
    )
    df["Nombre de la UC"] = (
        df["Clave de la UC"].map(_uc_nombre_map).fillna(df["Nombre de la UC"])
    )

# Caption dinámico
_anios_label = ", ".join(anios_sel)
st.caption(f"Fuente: ComprasMX {_anios_label} | División de Monitoreo de la Integridad Institucional – IMSS")

instituciones = ["Todas"] + sorted(df["Institución"].dropna().unique().tolist())
default_imss = instituciones.index("INSTITUTO MEXICANO DEL SEGURO SOCIAL") if "INSTITUTO MEXICANO DEL SEGURO SOCIAL" in instituciones else 0
inst_sel = st.sidebar.selectbox("Institución", instituciones, index=default_imss)

tipos = ["Todos"] + sorted(df["Tipo Simplificado"].unique().tolist())
tipo_sel = st.sidebar.selectbox("Tipo de procedimiento", tipos)

contrataciones = ["Todas"] + sorted(df["Tipo de contratación"].dropna().unique().tolist())
contratacion_sel = st.sidebar.selectbox("Tipo de contratación", contrataciones)

excl_consolidadas = st.sidebar.checkbox("Excluir compras consolidadas", value=False)

# ── Aplicar filtros base ──
dff = df.copy()
if inst_sel != "Todas":
    dff = dff[dff["Institución"] == inst_sel]
if tipo_sel != "Todos":
    dff = dff[dff["Tipo Simplificado"] == tipo_sel]
if contratacion_sel != "Todas":
    dff = dff[dff["Tipo de contratación"] == contratacion_sel]
if excl_consolidadas:
    dff = dff[dff["Compra consolidada"].str.upper().str.strip() != "SI"]

# ── Filtro de Unidad Compradora (sidebar) ──
_ucs_sidebar = ["Todas"] + sorted(dff["Nombre de la UC"].dropna().unique().tolist())
uc_sel = st.sidebar.selectbox("🏢 Unidad Compradora", _ucs_sidebar)
if uc_sel != "Todas":
    dff = dff[dff["Nombre de la UC"] == uc_sel]

# ── Última actualización de datos (sidebar) ──
_META_PATH = Path(__file__).parent / "metadata.json"
_meta_labels = {
    "contratos_comprasmx_2026.csv": "Contratos 2026",
    "contratos_comprasmx_2025.csv": "Contratos 2025",
    "contratos_compranet_2024.csv": "Contratos 2024",
    "Listado_completo_69-B.csv":    "EFOS 69-B (SAT)",
    "AnaliticaPreciosUnitarios.xlsx": "Analítica precios",
}
try:
    with open(_META_PATH, encoding="utf-8") as _mf:
        _meta = json.load(_mf)
    # Fecha de actualización más reciente entre los contratos
    _fechas_contratos = [
        _meta[k]["actualizado"]
        for k in ("contratos_comprasmx_2026.csv",
                  "contratos_comprasmx_2025.csv",
                  "contratos_compranet_2024.csv")
        if k in _meta
    ]
    _ultima = max(_fechas_contratos) if _fechas_contratos else None
    st.sidebar.divider()
    if _ultima:
        st.sidebar.caption(f"🗓 **Datos actualizados:** {_ultima}")
    with st.sidebar.expander("📋 Detalle por fuente"):
        for _fname, _label in _meta_labels.items():
            if _fname in _meta:
                _m = _meta[_fname]
                _filas = f"{_m['filas']:,} registros" if "filas" in _m else ""
                st.caption(
                    f"**{_label}**  \n"
                    f"🕐 {_m.get('actualizado', '—')}"
                    + (f"  \n📊 {_filas}" if _filas else "")
                )
except FileNotFoundError:
    pass  # metadata.json no existe aún — no mostrar nada
except Exception:
    pass  # cualquier error de lectura — no interrumpir el dashboard

# ─────────────────────────────────────────────
# KPIs PRINCIPALES (siempre visibles, fuera de tabs)
# ─────────────────────────────────────────────
st.subheader("📊 Indicadores Generales")

total_contratos = len(dff)
monto_total     = dff["Importe DRC"].sum()
n_adj_directas  = dff["Tipo Simplificado"].isin(["Adjudicación Directa", "Adjudicación Directa — Fr. I"]).sum()
pct_adj         = (n_adj_directas / total_contratos * 100) if total_contratos > 0 else 0
n_proveedores   = dff["Proveedor o contratista"].nunique()

n_ucs_activas = dff["Nombre de la UC"].nunique()

col1, col2, col3, col4 = st.columns(4)
col1.metric("📄 Total de contratos", f"{total_contratos:,}")
col2.metric(
    "💰 Monto total",
    f"${monto_total/1e9:,.2f} miles de millones MXN" if monto_total >= 1e9 else f"${monto_total/1e6:,.1f} M MXN"
)
col3.metric("🏥 Unidades Compradoras activas", f"{n_ucs_activas:,}")
col4.metric("🏢 Proveedores únicos", f"{n_proveedores:,}")

st.divider()


# ───────────────────────────────────────────────────────────────
# PÁGINA 1: PAGINA_DESCRIPCION
# ───────────────────────────────────────────────────────────────
def pagina_descripcion():

    # ── SECCIÓN 1: DISTRIBUCIÓN POR TIPO DE PROCEDIMIENTO ──
    st.subheader("1️⃣ Distribución por Tipo de Procedimiento")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            La **Ley de Adquisiciones, Arrendamientos y Servicios del Sector Público (LAASSP)**
            establece tres modalidades de contratación:

            - 🟢 **Licitación pública** — procedimiento abierto y competitivo; garantiza las mejores
              condiciones para el Estado (Art. 26 Fr. I LAASSP).
            - 🟡 **Invitación a cuando menos tres personas** — proceso restringido a un mínimo de tres
              participantes; procede en supuestos específicos de menor cuantía.
            - 🔴 **Adjudicación directa** — asignación sin concurso; permitida solo en los casos de
              excepción del Art. 41 LAASSP (urgencia, exclusividad, etc.).
            - 🔴 **Adjudicación directa — Fr. I** — casos estructurales de excepción por **patente,
              licencia exclusiva o derechos de autor** (Art. 54 Fr. I LAASSP vigente o Art. 41 Fr. I
              de la ley anterior). Se separan del resto de adjudicaciones directas porque su causa
              jurídica no es discrecional sino vinculada al titular del derecho.

            El **Decreto presidencial del 18 de noviembre de 2024** establece que al menos el **65 %**
            del monto total de adquisiciones debe canalizarse por licitación pública. La DMII monitorea
            trimestralmente el cumplimiento de esta meta.
            """
        )

    col_a, col_b = st.columns(2)

    _TIPO_RENAME = {"Adjudicación Directa — Fr. I": "Adjudicación Directa — Patentes"}

    with col_a:
        dist_num = dff["Tipo Simplificado"].replace(_TIPO_RENAME).value_counts().reset_index()
        dist_num.columns = ["Tipo", "Contratos"]
        fig1 = px.pie(dist_num, names="Tipo", values="Contratos",
                      color="Tipo", color_discrete_map=COLORES_TIPO,
                      title="Por número de contratos",
                      hole=0.35)
        fig1.update_traces(
            textinfo="percent",
            textposition="inside",
            insidetextorientation="horizontal",
            textfont=dict(family="Noto Sans, sans-serif", size=13),
            hovertemplate="<b>%{label}</b><br>Contratos: %{value:,}<br>%{percent}<extra></extra>"
        )
        fig1.update_layout(
            font=plotly_font(),
            title_font_color=IMSS_VERDE_OSC,
            legend=dict(
                orientation="v",
                font=dict(family="Noto Sans, sans-serif", size=11),
                x=1.01, y=0.5, xanchor="left"
            ),
            margin=dict(r=160)
        )
        st.plotly_chart(fig1, use_container_width=True)

    with col_b:
        dist_monto = dff.copy()
        dist_monto["Tipo Simplificado"] = dist_monto["Tipo Simplificado"].replace(_TIPO_RENAME)
        dist_monto = dist_monto.groupby("Tipo Simplificado")["Importe DRC"].sum().reset_index()
        dist_monto.columns = ["Tipo", "Monto"]
        fig2 = px.pie(dist_monto, names="Tipo", values="Monto",
                      color="Tipo", color_discrete_map=COLORES_TIPO,
                      title="Por monto total",
                      hole=0.35)
        fig2.update_traces(
            textinfo="percent",
            textposition="inside",
            insidetextorientation="horizontal",
            textfont=dict(family="Noto Sans, sans-serif", size=13),
            hovertemplate="<b>%{label}</b><br>Monto: $%{value:,.0f}<br>%{percent}<extra></extra>"
        )
        fig2.update_layout(
            font=plotly_font(),
            title_font_color=IMSS_VERDE_OSC,
            legend=dict(
                orientation="v",
                font=dict(family="Noto Sans, sans-serif", size=11),
                x=1.01, y=0.5, xanchor="left"
            ),
            margin=dict(r=160)
        )
        st.plotly_chart(fig2, use_container_width=True)

    _monto_ad_s1  = dff.loc[dff["Tipo Simplificado"].isin(["Adjudicación Directa", "Adjudicación Directa — Fr. I"]), "Importe DRC"].sum()
    _pct_ad_monto = (_monto_ad_s1 / monto_total * 100) if monto_total > 0 else 0
    if _pct_ad_monto > 35:
        st.warning(
            f"⚠️ **Alerta:** El **{_pct_ad_monto:.1f}%** del monto contratado corresponde a "
            f"adjudicaciones directas. La licitación pública es la regla general (Art. 35 LAASSP 2025); "
            f"un porcentaje elevado de AD puede indicar uso excesivo de procedimientos de excepción."
        )

    st.divider()

    # ── SECCIÓN 2: CONCENTRACIÓN POR PROVEEDOR Y UC ──
    st.subheader("2️⃣ Concentración por Proveedor y Unidad Compradora")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            Identifica **concentración de contratos** en un número reducido de proveedores o
            Unidades Compradoras, lo que puede señalar dependencia excesiva, riesgos de corrupción
            o captura del proceso de contratación.

            - **TreeMap de proveedores** — los proveedores que concentran el mayor valor contratado.
              El tamaño de cada celda es proporcional al monto. Una alta concentración puede indicar
              falta de competencia real.
            - **TreeMap de Unidades Compradoras** — las UCs con mayor monto contratado en el período.
              Permite identificar qué dependencias ejercen el mayor presupuesto.
            """
        )

    col_c, col_d = st.columns(2)

    with col_c:
        st.markdown("**Proveedores por monto contratado**")
        top_prov = (dff.groupby("Proveedor o contratista")["Importe DRC"]
                    .sum().sort_values(ascending=False).head(20).reset_index())
        top_prov.columns = ["Proveedor", "Monto"]
        top_prov["Monto_fmt"] = top_prov["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
        fig3 = px.treemap(
            top_prov,
            path=["Proveedor"],
            values="Monto",
            color="Monto",
            color_continuous_scale=[[0, IMSS_VERDE_OSC], [1, IMSS_VERDE]],
            custom_data=["Monto_fmt"],
        )
        fig3.update_traces(
            texttemplate="<b>%{label}</b><br>%{customdata[0]}",
            hovertemplate="<b>%{label}</b><br>Monto: %{customdata[0]}<extra></extra>",
            textfont=dict(family="Noto Sans, sans-serif", size=12),
        )
        fig3.update_layout(
            font=plotly_font(),
            paper_bgcolor="#ffffff",
            margin=dict(t=10, l=0, r=0, b=0),
            height=380,
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig3, use_container_width=True)

    with col_d:
        st.markdown("**Unidades Compradoras por monto contratado**")
        _uc_monto = (dff.groupby("Nombre de la UC")["Importe DRC"]
                     .sum().sort_values(ascending=False).head(20).reset_index())
        _uc_monto.columns = ["UC", "Monto"]
        _uc_monto["Monto_fmt"] = _uc_monto["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
        fig4 = px.treemap(
            _uc_monto,
            path=["UC"],
            values="Monto",
            color="Monto",
            color_continuous_scale=[[0, IMSS_ROJO_OSC], [1, IMSS_ROJO]],
            custom_data=["Monto_fmt"],
        )
        fig4.update_traces(
            texttemplate="<b>%{label}</b><br>%{customdata[0]}",
            hovertemplate="<b>%{label}</b><br>Monto: %{customdata[0]}<extra></extra>",
            textfont=dict(family="Noto Sans, sans-serif", size=12),
        )
        fig4.update_layout(
            font=plotly_font(),
            paper_bgcolor="#ffffff",
            margin=dict(t=10, l=0, r=0, b=0),
            height=380,
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig4, use_container_width=True)

    if len(top_prov) >= 3 and monto_total > 0:
        pct_top3 = top_prov.head(3)["Monto"].sum() / monto_total * 100
        if pct_top3 > 50:
            st.warning(f"⚠️ **Concentración elevada:** Los 3 principales proveedores concentran "
                       f"el {pct_top3:.1f}% del monto total.")

    st.divider()

    # ── SECCIÓN 2b: DISTRIBUCIÓN POR TIPO DE UC ──
    st.subheader("🗺️  Distribución del Gasto por Tipo de UC")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown("""
            Desglosa el gasto total por tipo de Unidad Compradora: **OOAD** (Delegaciones Regionales),
            **UMAE** (Unidades de Alta Especialidad) y **Nivel Central**. Para el Nivel Central,
            se distingue entre el gasto correspondiente a **compras consolidadas** (adquisiciones
            coordinadas a nivel APF) y el gasto **no consolidado** de esa Unidad Compradora.
        """)
    _tj_df = dff[["Clave de la UC", "Importe DRC", "Compra consolidada"]].merge(
        df_dir_uc[["Clave_UC", "Tipo UC"]],
        left_on="Clave de la UC", right_on="Clave_UC", how="left"
    )
    _tj_df["Tipo UC"] = _tj_df["Tipo UC"].fillna("Sin clasificar")
    def _cat_uc_tj(row):
        t = row["Tipo UC"]
        if t == "Nivel Central":
            return ("Nivel Central — Consolidada"
                    if str(row["Compra consolidada"]).strip().upper() == "SI"
                    else "Nivel Central — No consolidada")
        return t
    _tj_df["Cat_uc"] = _tj_df.apply(_cat_uc_tj, axis=1)
    _agg_tj = _tj_df.groupby("Cat_uc")["Importe DRC"].sum().reset_index()
    _total_tj = _agg_tj["Importe DRC"].sum()
    _agg_tj["Pct"] = (_agg_tj["Importe DRC"] / _total_tj * 100).fillna(0) if _total_tj > 0 else 0
    _agg_tj["Monto_fmt"] = _agg_tj["Importe DRC"].apply(lambda x: f"${x/1e6:,.1f} M")
    _agg_tj["Pct_fmt"]   = _agg_tj["Pct"].apply(lambda x: f"{x:.1f}%")
    _colores_cat_uc = {
        "OOAD":                             IMSS_VERDE,
        "UMAE":                             IMSS_ROJO,
        "Nivel Central — No consolidada":    IMSS_ORO,
        "Nivel Central — Consolidada":       IMSS_ORO_CLARO,
        "Sin clasificar":                   IMSS_GRIS,
    }
    fig_tj = px.treemap(
        _agg_tj, path=["Cat_uc"], values="Importe DRC",
        color="Cat_uc", color_discrete_map=_colores_cat_uc,
        custom_data=["Monto_fmt", "Pct_fmt"],
    )
    fig_tj.update_traces(
        texttemplate="<b>%{label}</b><br>%{customdata[1]}<br>%{customdata[0]}",
        hovertemplate=(
            "<b>%{label}</b><br>Monto: %{customdata[0]}<br>"
            "% del total: %{customdata[1]}<extra></extra>"
        ),
        textfont=dict(family="Noto Sans, sans-serif", size=13),
    )
    fig_tj.update_layout(
        font=plotly_font(), height=400,
        margin=dict(l=10, r=10, t=20, b=10),
        showlegend=False,
    )
    st.plotly_chart(fig_tj, use_container_width=True)
    st.divider()

    # ── SECCIÓN 3: ANÁLISIS POR PARTIDA PRESUPUESTARIA (CUCOP) ──
    st.subheader("3️⃣ Gasto por Partida Presupuestaria (CUCoP)")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            Distribuye el gasto conforme al **Catálogo de Unidades de Compra y Objeto del Gasto (CUCoP)**,
            que es la clasificación presupuestaria oficial para el gasto en adquisiciones de la APF.

            - El catálogo tiene cuatro niveles de agregación: **Capítulo → Concepto → Partida genérica →
              Partida específica**. Puedes cambiar el nivel de análisis con el selector inferior.
            - Cuando un contrato tiene **múltiples partidas**, el monto se distribuye de forma proporcional
              entre ellas para evitar doble conteo.
            - El análisis permite identificar en qué bienes y servicios se concentra el gasto y detectar
              posibles desvíos respecto a los objetivos institucionales.
            """
        )

    # Explotar contratos con múltiples partidas presupuestarias (ej: "15401, 27101, 27201, 27301")
    # El monto se divide en partes iguales entre todas las partidas del contrato
    _dff_exp = dff.copy()
    # Filtrar contratos sin partida presupuestaria
    _dff_exp = _dff_exp[_dff_exp["Partida específica"].str.strip() != ""]
    _dff_exp["_lista"] = _dff_exp["Partida específica"].str.split(",")
    _dff_exp["_n"] = _dff_exp["_lista"].apply(len)
    _dff_exp["Importe DRC"] = _dff_exp["Importe DRC"] / _dff_exp["_n"]
    _dff_exp = _dff_exp.explode("_lista")
    _dff_exp["Partida específica"] = _dff_exp["_lista"].str.strip().str.zfill(5)
    _dff_exp = _dff_exp.drop(columns=["_lista", "_n"])

    # Unir contratos con descripciones CUCOP
    dff_cucop = _dff_exp.merge(
        df_cucop[["PARTIDA ESPECÍFICA", "DESC. PARTIDA ESPECÍFICA",
                  "DESC. PARTIDA GENÉRICA", "DESC. CAPÍTULO"]].drop_duplicates("PARTIDA ESPECÍFICA"),
        left_on="Partida específica",
        right_on="PARTIDA ESPECÍFICA",
        how="left"
    )
    dff_cucop["Etiqueta partida"] = (
        dff_cucop["Partida específica"] + " — " +
        dff_cucop["DESC. PARTIDA ESPECÍFICA"].fillna("Sin descripción")
    )

    # Sub-filtro de Unidad Compradora (inline, específico para esta sección)
    ucs_disponibles = ["Todas las UCs"] + sorted(dff_cucop["Nombre de la UC"].dropna().unique().tolist())
    col_filtro1, col_filtro2, col_filtro3 = st.columns([2, 1, 1])
    with col_filtro1:
        uc_sel = st.selectbox("🏥 Ver por Unidad Compradora", ucs_disponibles, key="uc_partida")
    with col_filtro2:
        top_n = st.selectbox("Mostrar top", [10, 20, 30, 50], index=1, key="top_n_partida")
    with col_filtro3:
        nivel_agrup = st.selectbox(
            "Agrupar por",
            ["Partida específica", "Partida genérica", "Capítulo"],
            key="nivel_partida"
        )

    dff_part = dff_cucop.copy()
    if uc_sel != "Todas las UCs":
        dff_part = dff_part[dff_part["Nombre de la UC"] == uc_sel]

    # Definir columna de agrupación según selección
    if nivel_agrup == "Partida específica":
        col_grupo  = "Etiqueta partida"
        titulo_eje = "Partida específica"
    elif nivel_agrup == "Partida genérica":
        dff_part["Etiqueta genérica"] = (
            dff_part["Partida específica"].str[:4] + " — " +
            dff_part["DESC. PARTIDA GENÉRICA"].fillna("Sin descripción")
        )
        col_grupo  = "Etiqueta genérica"
        titulo_eje = "Partida genérica"
    else:
        dff_part["Etiqueta capítulo"] = (
            dff_part["Partida específica"].str[:1] + "000 — " +
            dff_part["DESC. CAPÍTULO"].fillna("Sin descripción")
        )
        col_grupo  = "Etiqueta capítulo"
        titulo_eje = "Capítulo"

    gasto_partida = (
        dff_part.groupby(col_grupo)["Importe DRC"]
        .sum()
        .sort_values(ascending=False)
        .head(top_n)
        .reset_index()
    )
    gasto_partida.columns = [titulo_eje, "Monto"]
    gasto_partida["Monto_fmt"] = gasto_partida["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")

    # Truncar etiquetas largas para legibilidad
    MAX_LABEL = 55
    gasto_partida["Etiqueta"] = gasto_partida[titulo_eje].apply(
        lambda s: s if len(str(s)) <= MAX_LABEL else str(s)[:MAX_LABEL] + "…"
    )
    # Tooltip con texto completo
    gasto_partida["Tooltip"] = gasto_partida[titulo_eje]

    if len(gasto_partida) > 0:
        _titulo_part = (
            f"Top {top_n} por {titulo_eje.lower()}"
            + (f" · UC: {uc_sel}" if uc_sel != "Todas las UCs" else "")
        )
        fig_part = px.treemap(
            gasto_partida,
            path=["Etiqueta"],
            values="Monto",
            color="Monto",
            color_continuous_scale=[[0, IMSS_VERDE_OSC], [1, IMSS_VERDE]],
            custom_data=["Tooltip", "Monto_fmt"],
            title=_titulo_part,
        )
        fig_part.update_traces(
            texttemplate="<b>%{label}</b><br>%{customdata[1]}",
            hovertemplate="<b>%{customdata[0]}</b><br>Monto: %{customdata[1]}<extra></extra>",
            textfont=dict(family="Noto Sans, sans-serif", size=11),
        )
        fig_part.update_layout(
            font=plotly_font(),
            paper_bgcolor="#ffffff",
            title_font_color=IMSS_VERDE_OSC,
            margin=dict(t=40, l=0, r=0, b=0),
            height=max(440, top_n * 18),
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig_part, use_container_width=True)

        # Tabla resumen
        with st.expander("📋 Ver tabla de datos"):
            tabla_part = gasto_partida.copy()
            tabla_part["Monto"] = tabla_part["Monto"].apply(lambda x: f"${x:,.0f}")
            tabla_part.index += 1
            st.dataframe(tabla_part, use_container_width=True)
    else:
        st.info("ℹ️ No hay datos de partidas para los filtros actuales.")

    st.divider()

    # ── SECCIÓN 4 (PARTE I): PATRONES DE CONTRATACIÓN ──
    st.subheader("4️⃣ Patrones de Contratación")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            Analiza la **estructura del gasto** por tamaño de proveedor y detecta señales de alerta
            en adjudicaciones de alto valor.

            - **Estratificación por tamaño** — clasifica a los proveedores según el registro en el
              sistema (micro, pequeña, mediana, grande empresa). Permite verificar si el gasto favorece
              a sectores específicos o si hay concentración en grandes empresas a costa de las MIPYMES.
            - **Distribución por monto** — complementa el conteo al mostrar el valor económico que
              cada segmento representa, que puede diferir significativamente del número de contratos.
            - **Cuotas Mipymes y cooperativas** — el nuevo Reglamento de la LAASSP establece cuotas
              específicas para la contratación con micro, pequeñas y medianas empresas (Mipymes), así
              como con sociedades cooperativas de producción. El análisis por estratificación permite
              verificar el cumplimiento de estos porcentajes mínimos de gasto con dichos sectores.
            """
        )

    # Paleta consistente por nombre de categoría — misma estratificación = mismo color en ambas gráficas
    _estrat_palette = [IMSS_VERDE, IMSS_ORO, IMSS_ROJO,
                       IMSS_GRIS, IMSS_ORO_CLARO, IMSS_VERDE_OSC, IMSS_NEGRO]
    _estrat_cats = sorted(dff["Estratificación"].dropna().unique().tolist())
    _estrat_color_map = {
        cat: _estrat_palette[i % len(_estrat_palette)]
        for i, cat in enumerate(_estrat_cats)
    }

    col_e, col_f = st.columns(2)

    with col_e:
        st.markdown("**Distribución por estratificación — Número de contratos**")
        estrat = dff["Estratificación"].value_counts().reset_index()
        estrat.columns = ["Estratificación", "Contratos"]
        estrat["Pct"] = (estrat["Contratos"] / estrat["Contratos"].sum() * 100).round(1)
        fig5 = px.pie(
            estrat, names="Estratificación", values="Contratos",
            color="Estratificación", color_discrete_map=_estrat_color_map,
            hole=0.42,
        )
        fig5.update_traces(
            textinfo="percent+label",
            textposition="inside",
            insidetextorientation="horizontal",
            textfont=dict(family="Noto Sans, sans-serif", size=12),
            hovertemplate="<b>%{label}</b><br>Contratos: %{value:,}<br>%{percent}<extra></extra>"
        )
        fig5.update_layout(
            font=plotly_font(),
            showlegend=True,
            legend=dict(orientation="h", font=dict(size=10), x=0.5, xanchor="center", y=-0.15),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            margin=dict(t=30, b=40)
        )
        st.plotly_chart(fig5, use_container_width=True)

    with col_f:
        st.markdown("**Distribución por estratificación — Monto contratado**")
        _estrat_monto = (
            dff.groupby("Estratificación")["Importe DRC"]
            .sum().reset_index()
            .sort_values("Importe DRC", ascending=False)
        )
        _estrat_monto.columns = ["Estratificación", "Monto"]
        _estrat_monto["Monto_fmt"] = _estrat_monto["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
        fig6 = px.pie(
            _estrat_monto, names="Estratificación", values="Monto",
            color="Estratificación", color_discrete_map=_estrat_color_map,
            hole=0.42,
        )
        fig6.update_traces(
            textinfo="percent+label",
            textposition="inside",
            insidetextorientation="horizontal",
            textfont=dict(family="Noto Sans, sans-serif", size=12),
            hovertemplate="<b>%{label}</b><br>Monto: $%{value:,.0f}<br>%{percent}<extra></extra>"
        )
        fig6.update_layout(
            font=plotly_font(),
            showlegend=True,
            legend=dict(orientation="h", font=dict(size=10), x=0.5, xanchor="center", y=-0.15),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            margin=dict(t=30, b=40)
        )
        st.plotly_chart(fig6, use_container_width=True)

    # ── Mipyme / Cooperativa — Reglamento Art. 113 ──────────────────
    st.markdown("**Cumplimiento del cupo Mipyme y Cooperativa (Reglamento LAASSP Art. 113)**")
    st.caption(
        "El Art. 113 del Reglamento LAASSP establece que en adjudicaciones directas (Art. 55 LAASSP), "
        "**≥ 50%** del monto debe contratarse con **Mipymes**, y de ese monto Mipyme, "
        "**≥ 25%** debe ir a **cooperativas** (i.e., ≥ 12.5% del total AD). "
        "El análisis se realiza sobre todos los contratos del año en vista con la selección de filtros actual."
    )

    # Adjudicaciones directas (todas las causales) del conjunto filtrado
    _dff_ad_m = dff[dff["Tipo Simplificado"].str.contains("Adjudicación Directa", na=False)].copy()
    _dff_ad_m["_estrat_norm"] = _dff_ad_m["Estratificación"].fillna("").str.upper().str.strip()

    # Mipymes: Micro, Pequeña, Mediana
    _mipyme_vals = {"MICRO", "PEQUEÑA", "MEDIANA", "PEQUEÃ'A", "PEQUEA"}
    _dff_ad_mipyme = _dff_ad_m[_dff_ad_m["_estrat_norm"].isin(_mipyme_vals)]

    # Cooperativas: keyword en nombre del proveedor o estratificación con COOP
    _prov_up = _dff_ad_m["Proveedor o contratista"].fillna("").str.upper()
    _dff_ad_coop = _dff_ad_m[
        _prov_up.str.contains(r"COOPERATIVA|S\.C\.L\.|S\.C\b", na=False, regex=True)
    ]

    monto_ad_m      = _dff_ad_m["Importe DRC"].sum()
    monto_mipyme_ad = _dff_ad_mipyme["Importe DRC"].sum()
    monto_coop_ad   = _dff_ad_coop["Importe DRC"].sum()

    pct_mipyme_ad = monto_mipyme_ad / monto_ad_m * 100 if monto_ad_m > 0 else 0
    pct_coop_of_mipyme = monto_coop_ad / monto_mipyme_ad * 100 if monto_mipyme_ad > 0 else 0
    pct_coop_ad   = monto_coop_ad / monto_ad_m * 100 if monto_ad_m > 0 else 0

    _m1, _m2, _m3, _m4 = st.columns(4)
    _m1.metric("💰 Monto AD total", f"${monto_ad_m/1e6:,.1f} M MXN")
    _m2.metric(
        "🏭 Mipyme — % del monto AD",
        f"{pct_mipyme_ad:.1f}%",
        delta=f"{pct_mipyme_ad - 50:.1f} pp vs objetivo 50%",
        delta_color="normal",   # verde si ≥0 (cumple), rojo si <0 (incumple)
    )
    _m3.metric(
        "🤝 Cooperativas — % del monto AD",
        f"{pct_coop_ad:.1f}%",
        delta=f"{pct_coop_of_mipyme - 25:.1f} pp (coop/mipyme vs 25%)",
        delta_color="normal",   # verde si ≥0 (cumple), rojo si <0 (incumple)
    )
    _m4.metric("📋 Contratos a Mipymes", f"{len(_dff_ad_mipyme):,}")

    # Banner de cumplimiento
    if monto_ad_m > 0:
        _alerts_mipyme = []
        if pct_mipyme_ad < 50:
            _alerts_mipyme.append(f"Mipyme: {pct_mipyme_ad:.1f}% < 50% requerido")
        if pct_coop_of_mipyme < 25 and monto_mipyme_ad > 0:
            _alerts_mipyme.append(f"Cooperativas: {pct_coop_of_mipyme:.1f}% del monto Mipyme < 25% requerido")
        if _alerts_mipyme:
            st.warning(
                "⚠️ **Posible incumplimiento del cupo Mipyme/Cooperativa (Reglamento Art. 113):** "
                + " | ".join(_alerts_mipyme)
                + ". Nota: el cálculo se basa en la columna 'Estratificación' del sistema."
            )
        else:
            st.success(
                f"✅ Se cumple el cupo Mipyme ({pct_mipyme_ad:.1f}% ≥ 50%) y "
                f"cooperativas ({pct_coop_of_mipyme:.1f}% del monto Mipyme ≥ 25%) sobre el monto de AD."
            )

    # Gráfica de barras por UC — distribución Mipyme vs Grande en AD
    with st.expander("📊 Distribución Mipyme/Grande por Unidad Compradora (AD)"):
        if monto_ad_m == 0:
            st.info("No hay adjudicaciones directas con los filtros actuales.")
        else:
            _by_uc_m = (
                _dff_ad_m.groupby(["Nombre de la UC", "_estrat_norm"])["Importe DRC"]
                .sum().reset_index()
            )
            _by_uc_m["Categoría"] = _by_uc_m["_estrat_norm"].apply(
                lambda x: "Mipyme" if x in _mipyme_vals else ("Cooperativa" if "COOP" in x else "Grande / Otro")
            )
            _by_uc_grp = _by_uc_m.groupby(["Nombre de la UC", "Categoría"])["Importe DRC"].sum().reset_index()
            _totals_uc_m = _by_uc_grp.groupby("Nombre de la UC")["Importe DRC"].sum().rename("Total")
            _by_uc_grp = _by_uc_grp.merge(_totals_uc_m, on="Nombre de la UC")
            _by_uc_grp["Pct"] = _by_uc_grp["Importe DRC"] / _by_uc_grp["Total"] * 100

            _top_uc_m = (
                _totals_uc_m.nlargest(20).index.tolist()
            )
            _by_uc_plot = _by_uc_grp[_by_uc_grp["Nombre de la UC"].isin(_top_uc_m)].copy()
            _by_uc_plot["UC_corta"] = _by_uc_plot["Nombre de la UC"].apply(
                lambda s: s[:45] + "…" if len(s) > 45 else s
            )
            _ord_uc_m = (
                _by_uc_plot[_by_uc_plot["Categoría"] == "Mipyme"]
                .sort_values("Pct")["UC_corta"].tolist()
            )
            fig_mipyme = px.bar(
                _by_uc_plot, x="Pct", y="UC_corta",
                color="Categoría",
                color_discrete_map={
                    "Mipyme": IMSS_VERDE, "Cooperativa": IMSS_ORO, "Grande / Otro": IMSS_ROJO
                },
                orientation="h",
                barmode="stack",
                category_orders={"UC_corta": _ord_uc_m},
                title="% del monto AD por categoría Mipyme — top 20 UCs",
                text=_by_uc_plot["Pct"].apply(lambda x: f"{x:.0f}%" if x >= 8 else ""),
            )
            fig_mipyme.add_vline(
                x=50, line_dash="dash", line_color=IMSS_ORO_CLARO,
                annotation_text="50% objetivo Mipyme",
                annotation_position="top right",
                annotation_font=dict(family="Noto Sans, sans-serif", size=11)
            )
            fig_mipyme.update_layout(
                font=plotly_font(), xaxis_title="% del monto AD",
                yaxis_title="", plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis=dict(range=[0, 105]),
                legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0)
            )
            fig_mipyme.update_traces(
                textfont=dict(family="Noto Sans, sans-serif"),
                textposition="inside"
            )
            st.plotly_chart(fig_mipyme, use_container_width=True)

    # ── Timeline mensual de contratos ────────────────────────────
    st.markdown("**Monto contratado por mes y tipo de procedimiento**")
    _dff_tl = dff.copy()
    _dff_tl["Tipo Display"] = _dff_tl["Tipo Simplificado"].replace(
        {"Adjudicación Directa — Fr. I": "Adjudicación Directa — Patentes"}
    )
    _MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
                 7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
    _fecha_tl = pd.to_datetime(
        _dff_tl["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
    )
    _dff_tl["_mes_sort"]  = (
        _fecha_tl.dt.year.astype("Int64").astype(str)
        + _fecha_tl.dt.month.astype("Int64").astype(str).str.zfill(2)
    )
    _dff_tl["_mes_label"] = (
        _fecha_tl.dt.month.map(_MESES_ES)
        + " " + _fecha_tl.dt.year.astype("Int64").astype(str)
    )
    _tl_data = (
        _dff_tl[_dff_tl["_mes_sort"].notna() & (_dff_tl["_mes_sort"] != "NANaN")]
        .groupby(["_mes_sort", "_mes_label", "Tipo Display"])
        .agg(Contratos=("Importe DRC", "count"), Monto=("Importe DRC", "sum"))
        .reset_index()
        .sort_values("_mes_sort")
    )
    # Orden categórico de los meses para el eje X
    _orden_meses = (
        _tl_data[["_mes_sort", "_mes_label"]]
        .drop_duplicates()
        .sort_values("_mes_sort")["_mes_label"]
        .tolist()
    )
    _tl_data["Monto_M"] = (_tl_data["Monto"] / 1e6).round(2)
    _tl_data["Monto_fmt"] = _tl_data["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")

    if len(_tl_data) > 0:
        _colores_tl = {k.replace("Adjudicación Directa — Fr. I", "Adjudicación Directa — Patentes"): v
                       for k, v in COLORES_TIPO.items()}
        fig_tl = px.bar(
            _tl_data,
            x="_mes_label", y="Monto_M",
            color="Tipo Display",
            color_discrete_map=_colores_tl,
            barmode="stack",
            category_orders={"_mes_label": _orden_meses},
            custom_data=["Tipo Display", "Contratos", "Monto_fmt"],
        )
        fig_tl.update_traces(
            hovertemplate="<b>%{x}</b> · %{customdata[0]}<br>"
                          "Monto: %{customdata[2]}<br>"
                          "Contratos: %{customdata[1]:,}<extra></extra>",
        )
        fig_tl.update_layout(
            font=plotly_font(),
            xaxis_title="",
            yaxis_title="Monto (M MXN)",
            xaxis=dict(tickangle=-35, tickfont=dict(size=11)),
            legend_title="Tipo",
            legend=dict(orientation="h", y=-0.28, font=dict(size=10)),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            height=350,
            margin=dict(t=10, b=90),
        )
        st.plotly_chart(fig_tl, use_container_width=True)



    st.divider()



# ───────────────────────────────────────────────────────────────
# PÁGINA 2: PAGINA_RIESGO
# ───────────────────────────────────────────────────────────────
def pagina_riesgo():

    # ── Filtro de Unidad Compradora (aplica a todos los indicadores de riesgo) ──
    _ucs_t2 = ["Todas"] + sorted(dff["Nombre de la UC"].dropna().unique().tolist())
    _uc_sel_t2 = st.selectbox(
        "🏢 Filtrar por Unidad Compradora",
        _ucs_t2,
        key="uc_filtro_tab2",
        help="Selecciona una UC para concentrar todos los indicadores de riesgo en esa unidad compradora."
    )
    _dff2 = dff[dff["Nombre de la UC"] == _uc_sel_t2].copy() if _uc_sel_t2 != "Todas" else dff.copy()
    _monto_total2 = _dff2["Importe DRC"].sum()

    st.divider()

    # ── SECCIÓN 4 (PARTE II): PROVEEDORES SANCIONADOS ──
    st.subheader("4️⃣ Cruce con Proveedores Sancionados (SABG)")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            Cruza los contratos con la **base de empresas inhabilitadas** publicada por la
            Secretaría Anticorrupción y Buen Gobierno (SABG) en el Diario Oficial de la Federación.

            **Criterio jurídico (Art. 46 LAASSP):** La fecha de referencia para determinar si una
            contratación viola la ley es la **fecha de fallo**, no la firma del contrato. Con la
            notificación del fallo las obligaciones son exigibles; si el fallo se emitió durante
            un período de inhabilitación vigente, se configura la violación.

            **Niveles de riesgo:**
            - 🔴 **Riesgo crítico** — inhabilitación vigente sin suspensión judicial al momento del fallo.
            - 🟠 **Riesgo alto** — inhabilitación suspendida judicialmente (controversia en curso).
            - 🟡 **Riesgo medio** — inhabilitación ya concluida (historial de sanción).
            - ⚫ **Fallo anterior** — el fallo fue previo al inicio de la inhabilitación; sin violación activa.
            """
        )

    try:
        df_san = cargar_sancionados()

        rfcs_san = df_san['RFC'].str.strip().unique()
        cruce = _dff2[_dff2['rfc'].str.strip().isin(rfcs_san)].copy()
        cruce = cruce.merge(
            df_san[['RFC', 'Empresa', 'Nivel de Riesgo', 'Inicio inhabilitación']],
            left_on='rfc', right_on='RFC', how='left'
        )

        # Fecha de referencia: Fecha de fallo (Art. 46 LAASSP)
        cruce['Fecha_fallo_d'] = pd.to_datetime(cruce['Fecha de fallo'], errors='coerce').dt.date
        cruce['Inicio_d']      = pd.to_datetime(cruce['Inicio inhabilitación'], errors='coerce').dt.date

        def ajustar_nivel(row):
            nivel = row['Nivel de Riesgo']
            if nivel != "🔴 Riesgo crítico — Inhabilitación vigente":
                return nivel
            ff  = row['Fecha_fallo_d']
            ini = row['Inicio_d']
            if pd.isna(ff) or pd.isna(ini):
                return "⚪ Sin fecha de fallo (verificar manualmente)"
            if ff >= ini:
                return "🔴 Riesgo crítico — Inhabilitación vigente"
            else:
                return "⚫ Fallo anterior a inhabilitación (sin violación)"

        cruce['Nivel de Riesgo'] = cruce.apply(ajustar_nivel, axis=1)

        n_critico     = (cruce['Nivel de Riesgo'] == "🔴 Riesgo crítico — Inhabilitación vigente").sum()
        n_alto        = (cruce['Nivel de Riesgo'] == "🟠 Riesgo alto — Inhabilitación suspendida judicialmente").sum()
        n_medio       = (cruce['Nivel de Riesgo'] == "🟡 Riesgo medio — Historial de inhabilitación").sum()
        monto_critico = cruce[cruce['Nivel de Riesgo'] == "🔴 Riesgo crítico — Inhabilitación vigente"]["Importe DRC"].sum()

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("🔴 Inhabilitación vigente",      f"{n_critico:,} contratos")
        k2.metric("🟠 Suspendida judicialmente",    f"{n_alto:,} contratos")
        k3.metric("🟡 Historial de inhabilitación", f"{n_medio:,} contratos")
        k4.metric("💰 Monto en riesgo crítico",
                  f"${monto_critico/1e6:,.1f} M MXN" if monto_critico > 0 else "N/D")

        if n_critico > 0:
            st.error(f"🚨 **ALERTA CRÍTICA:** Se detectaron {n_critico:,} contratos con proveedores cuya inhabilitación "
                     f"está **vigente**. Esto constituye una posible violación a la LAASSP.")
        if n_alto > 0:
            st.warning(f"⚠️ **Atención:** {n_alto:,} contratos con proveedores cuya inhabilitación está "
                       f"**suspendida judicialmente**. El proceso sancionador sigue abierto.")

        if len(cruce) > 0:
            # Orden fijo de izquierda a derecha en la gráfica; solo niveles de alerta
            _orden_graf = [
                "🔴 Riesgo crítico — Inhabilitación vigente",
                "🟠 Riesgo alto — Inhabilitación suspendida judicialmente",
                "🟡 Riesgo medio — Historial de inhabilitación",
            ]
            color_map_sabg = {
                "🔴 Riesgo crítico — Inhabilitación vigente":              IMSS_ROJO,
                "🟠 Riesgo alto — Inhabilitación suspendida judicialmente": "#E07B00",
                "🟡 Riesgo medio — Historial de inhabilitación":            IMSS_ORO,
            }
            dist_riesgo = cruce['Nivel de Riesgo'].value_counts().reset_index()
            dist_riesgo.columns = ['Nivel', 'Contratos']
            # Filtrar solo los niveles que se muestran y ordenar
            dist_riesgo = dist_riesgo[dist_riesgo['Nivel'].isin(_orden_graf)].copy()
            dist_riesgo['_ord'] = dist_riesgo['Nivel'].map(
                {n: i for i, n in enumerate(_orden_graf)}
            )
            dist_riesgo = dist_riesgo.sort_values('_ord').drop(columns=['_ord'])
            # Monto por nivel de riesgo
            dist_monto_sabg = (
                cruce[cruce['Nivel de Riesgo'].isin(_orden_graf)]
                .groupby('Nivel de Riesgo')['Importe DRC'].sum()
                .reset_index()
            )
            dist_monto_sabg.columns = ['Nivel', 'Monto']
            dist_monto_sabg['_ord'] = dist_monto_sabg['Nivel'].map(
                {n: i for i, n in enumerate(_orden_graf)}
            )
            dist_monto_sabg = dist_monto_sabg.sort_values('_ord').drop(columns=['_ord'])
            dist_monto_sabg['Monto_fmt'] = dist_monto_sabg['Monto'].apply(
                lambda x: f"${x/1e6:,.1f} M"
            )

            _col_sabg1, _col_sabg2 = st.columns(2)

            with _col_sabg1:
                fig7 = px.bar(
                    dist_riesgo, x='Nivel', y='Contratos',
                    color='Nivel', color_discrete_map=color_map_sabg,
                    title="Contratos por nivel de riesgo",
                    category_orders={"Nivel": _orden_graf}
                )
                fig7.update_layout(font=plotly_font(), showlegend=False,
                                   plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                                   title_font_color=IMSS_VERDE_OSC,
                                   xaxis_title="", yaxis_title="Número de contratos")
                st.plotly_chart(fig7, use_container_width=True)

            with _col_sabg2:
                fig7b = px.bar(
                    dist_monto_sabg, x='Nivel', y='Monto',
                    color='Nivel', color_discrete_map=color_map_sabg,
                    text='Monto_fmt',
                    title="Monto contratado por nivel de riesgo",
                    category_orders={"Nivel": _orden_graf}
                )
                fig7b.update_layout(
                    font=plotly_font(), showlegend=False,
                    plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                    title_font_color=IMSS_VERDE_OSC,
                    xaxis_title="", yaxis_title="Monto (MXN)"
                )
                fig7b.update_traces(
                    textfont=dict(family="Noto Sans, sans-serif"),
                    textposition="outside", cliponaxis=False
                )
                st.plotly_chart(fig7b, use_container_width=True)

        st.markdown("**📋 Detalle de contratos con proveedores sancionados**")
        if len(cruce) > 0:
            # Orden de riesgo: crítico primero; solo niveles de alerta
            _orden_riesgo = {
                "🔴 Riesgo crítico — Inhabilitación vigente": 0,
                "🟠 Riesgo alto — Inhabilitación suspendida judicialmente": 1,
                "🟡 Riesgo medio — Historial de inhabilitación": 2,
            }
            # Filtro por nivel de riesgo (solo niveles de alerta)
            niveles_sabg = ["Todos"] + [n for n in _orden_riesgo if n in cruce['Nivel de Riesgo'].values]
            nivel_filtro_sabg = st.selectbox(
                "🔍 Filtrar por nivel de riesgo:", niveles_sabg, key="filtro_riesgo_sabg"
            )
            cols_san = [c for c in [
                'Nivel de Riesgo', 'Institución', 'Nombre de la UC',
                'Proveedor o contratista', 'rfc', 'Tipo Procedimiento',
                'Importe DRC', 'Descripción del contrato',
                'Dirección del anuncio'
            ] if c in cruce.columns]
            tabla_cruce = cruce[cols_san].copy()
            # Excluir niveles no accionables de la tabla de detalle
            tabla_cruce = tabla_cruce[
                ~tabla_cruce["Nivel de Riesgo"].isin([
                    "⚪ Sin fecha de fallo (verificar manualmente)",
                    "⚫ Fallo anterior a inhabilitación (sin violación)",
                ])
            ].copy()
            tabla_cruce["_orden"] = tabla_cruce["Nivel de Riesgo"].map(_orden_riesgo).fillna(99)
            tabla_cruce = tabla_cruce.sort_values("_orden").drop(columns=["_orden"]).reset_index(drop=True)
            if nivel_filtro_sabg != "Todos":
                tabla_cruce = tabla_cruce[
                    tabla_cruce["Nivel de Riesgo"] == nivel_filtro_sabg
                ].reset_index(drop=True)
            tabla_cruce["Importe DRC"] = tabla_cruce["Importe DRC"].apply(
                lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
            )
            tabla_cruce = tabla_cruce.rename(columns={"Importe DRC": "Importe"})
            tabla_cruce.index += 1
            st.dataframe(
                tabla_cruce,
                column_config={
                    "Dirección del anuncio": st.column_config.LinkColumn("🔗 ComprasMX", display_text="Ver contrato")
                },
                use_container_width=True
            )
        else:
            st.success("✅ No se encontraron contratos con proveedores en el listado de sancionados.")

    except FileNotFoundError:
        st.info("ℹ️ Para activar esta sección, coloca el archivo "
                "`BD_01_2026__Me_trica_empresas_inhabilitadas__DOF.xlsm` "
                "en la misma carpeta que el dashboard.")

    st.divider()

    # ── SECCIÓN 5: CRUCE CON EFOS (ART. 69-B CFF — SAT) ──
    st.subheader("5️⃣ Cruce con EFOS Art. 69-B CFF (SAT)")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            Cruza los contratos con el **Listado de Empresas que Facturan Operaciones Simuladas
            (EFOS)** publicado por el SAT conforme al **Art. 69-B del Código Fiscal de la
            Federación (CFF)**.

            Contratar con una EFOS puede implicar que los bienes o servicios pagados nunca se
            entregaron, y que los comprobantes fiscales emitidos por esas empresas son inválidos
            para efectos fiscales y de comprobación del gasto público.

            **Clasificación de riesgo:**
            - 🔴 **EFOS definitivo** — el SAT confirmó que la empresa facturó operaciones simuladas
              (resolución definitiva publicada en el DOF). Riesgo máximo.
            - 🟡 **EFOS presunto** — el SAT notificó presunción; la empresa está en proceso de
              desvirtuar. Riesgo alto mientras no concluya el proceso.
            - 🟢 **Desvirtuado / Sentencia favorable** — la empresa acreditó la materialidad de sus
              operaciones o ganó un juicio; se incluye como referencia informativa sin alerta activa.
            """
        )




    try:
        df_efos = cargar_efos()

        rfcs_alerta = df_efos[
            df_efos["Situación del contribuyente"].isin(["Definitivo", "Presunto"])
        ]["RFC"].unique()

        cruce_alerta = _dff2[_dff2["rfc"].str.strip().str.upper().isin(rfcs_alerta)].copy()
        if len(cruce_alerta) > 0:
            cruce_alerta = cruce_alerta.merge(
                df_efos[["RFC", "Nombre del Contribuyente", "Situación del contribuyente"]],
                left_on=cruce_alerta["rfc"].str.strip().str.upper(),
                right_on="RFC",
                how="left"
            ).drop(columns=["key_0"], errors="ignore")
            cruce_alerta["Nivel 69-B"] = cruce_alerta["Situación del contribuyente"].apply(nivel_efos)

        rfcs_todos  = df_efos["RFC"].unique()
        cruce_todos = _dff2[_dff2["rfc"].str.strip().str.upper().isin(rfcs_todos)].copy()
        if len(cruce_todos) > 0:
            cruce_todos = cruce_todos.merge(
                df_efos[["RFC", "Nombre del Contribuyente", "Situación del contribuyente"]],
                left_on=cruce_todos["rfc"].str.strip().str.upper(),
                right_on="RFC",
                how="left"
            ).drop(columns=["key_0"], errors="ignore")
            cruce_todos["Nivel 69-B"] = cruce_todos["Situación del contribuyente"].apply(nivel_efos)

        n_definitivo      = (cruce_alerta["Nivel 69-B"].str.startswith("🔴")).sum() if len(cruce_alerta) > 0 else 0
        n_presunto        = (cruce_alerta["Nivel 69-B"].str.startswith("🟡")).sum() if len(cruce_alerta) > 0 else 0
        monto_riesgo      = cruce_alerta["Importe DRC"].sum() if len(cruce_alerta) > 0 else 0
        n_proveedores_69b = cruce_alerta["rfc"].nunique() if len(cruce_alerta) > 0 else 0

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("🔴 EFOS definitivo",  f"{n_definitivo:,} contratos")
        k2.metric("🟡 EFOS presunto",    f"{n_presunto:,} contratos")
        k3.metric("💰 Monto en riesgo",  f"${monto_riesgo/1e6:,.1f} M MXN" if monto_riesgo > 0 else "N/D")
        k4.metric("🏢 Proveedores EFOS", f"{n_proveedores_69b:,}")

        if n_definitivo > 0:
            st.error(
                f"🚨 **ALERTA ROJA:** {n_definitivo:,} contrato(s) con empresa(s) en el listado **definitivo** "
                f"Art. 69-B CFF. Las operaciones facturadas se consideran inexistentes para efectos fiscales."
            )
        if n_presunto > 0:
            st.warning(
                f"⚠️ **ALERTA AMARILLA:** {n_presunto:,} contrato(s) con empresa(s) en el listado **presunto** "
                f"Art. 69-B CFF. El proceso de aclaración ante el SAT sigue en curso."
            )
        if n_definitivo == 0 and n_presunto == 0:
            st.success("✅ No se encontraron contratos con proveedores en el listado de EFOS (definitivos o presuntos).")

        if len(cruce_todos) > 0:
            dist_efos = cruce_todos["Nivel 69-B"].value_counts().reset_index()
            dist_efos.columns = ["Nivel", "Contratos"]
            color_map_efos = {
                "🔴 EFOS definitivo — Operaciones simuladas confirmadas": IMSS_ROJO,
                "🟡 EFOS presunto — Proceso en curso":                    IMSS_ORO,
                "🟢 Desvirtuado — Sin riesgo":                            IMSS_VERDE,
                "🟢 Sentencia favorable — Sin riesgo":                    IMSS_VERDE,
            }
            fig8 = px.bar(
                dist_efos, x="Nivel", y="Contratos",
                color="Nivel", color_discrete_map=color_map_efos,
                title="Contratos con proveedores en lista 69-B por situación"
            )
            fig8.update_layout(
                font=plotly_font(), showlegend=False,
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis_title="", yaxis_title="Número de contratos"
            )
            st.plotly_chart(fig8, use_container_width=True)

        st.markdown("**📋 Detalle de contratos en alerta (definitivos y presuntos)**")
        if len(cruce_alerta) > 0:
            cols_tabla = [c for c in [
                "Nivel 69-B", "Institución", "Nombre de la UC",
                "Proveedor o contratista", "rfc", "Nombre del Contribuyente",
                "Tipo Procedimiento", "Importe DRC", "Descripción del contrato",
                "Dirección del anuncio"
            ] if c in cruce_alerta.columns]
            tabla_69b = cruce_alerta[cols_tabla].sort_values("Nivel 69-B").reset_index(drop=True)
            tabla_69b["Importe DRC"] = tabla_69b["Importe DRC"].apply(
                lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
            )
            tabla_69b.index += 1
            st.dataframe(
                tabla_69b,
                column_config={
                    "Dirección del anuncio": st.column_config.LinkColumn("🔗 ComprasMX", display_text="Ver contrato")
                },
                use_container_width=True
            )
        else:
            st.info("ℹ️ No hay contratos con alertas activas (definitivos o presuntos) con los filtros actuales.")

    except FileNotFoundError:
        st.info("ℹ️ Para activar esta sección, coloca el archivo "
                "`Listado_completo_69-B.csv` en la misma carpeta que el dashboard.")

    st.divider()

    # ── SECCIÓN 6: EMPRESAS DE RECIENTE CREACIÓN ──
    st.subheader("6️⃣ Empresas de Reciente Creación")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            Identifica contratos adjudicados a empresas con **menos de un año de constitución**
            al inicio del contrato. Este patrón puede indicar empresas creadas *ad hoc* para
            obtener contratos públicos sin trayectoria comprobable.

            **Metodología:**
            - La fecha de constitución se extrae directamente del **RFC de persona moral**
              (12 caracteres: 3 letras + 6 dígitos de fecha YYMMDD + 3 alfanuméricos).
            - Regla de siglo: `YY ≤ 30 → año 2000+YY`; `YY > 30 → año 1900+YY`.
            - Solo aplica a personas morales; las personas físicas (RFC de 13 caracteres) se excluyen.
            - La fecha de referencia es la **Fecha de inicio del contrato** (momento en que el
              contrato entra en vigor). Umbral de alerta: antigüedad < 365 días.

            Un proveedor de reciente creación que recibe contratos de alto valor sin antecedentes
            puede ser señal de simulación o empresa fachada.
            """
        )

    def parse_fecha_rfc(rfc_str):
        """Extrae fecha de constitución de RFC de persona moral (LLL-YYMMDD-HHH, 12 chars)."""
        try:
            s = str(rfc_str).strip().upper()
            if not _re.match(r'^[A-ZÑ&]{3}[0-9]{6}[A-Z0-9]{3}$', s):
                return None
            yy, mm, dd = int(s[3:5]), int(s[5:7]), int(s[7:9])
            if mm < 1 or mm > 12 or dd < 1 or dd > 31:
                return None
            yr = 2000 + yy if yy <= 30 else 1900 + yy
            return _date(yr, mm, dd)
        except Exception:
            return None

    dff_rc = _dff2.copy()
    dff_rc["Fecha_inicio_d"] = pd.to_datetime(
        dff_rc["Fecha de inicio del contrato"], format="%d/%m/%Y", errors="coerce"
    ).dt.date
    dff_rc["Fecha_RFC"] = dff_rc["rfc"].apply(parse_fecha_rfc)
    dff_rc = dff_rc[dff_rc["Fecha_RFC"].notna() & dff_rc["Fecha_inicio_d"].notna()].copy()
    dff_rc["Edad_dias"] = dff_rc.apply(
        lambda r: (r["Fecha_inicio_d"] - r["Fecha_RFC"]).days, axis=1
    )

    recientes = dff_rc[(dff_rc["Edad_dias"] >= 0) & (dff_rc["Edad_dias"] < 365)].copy()
    recientes["Edad_meses"] = (recientes["Edad_dias"] // 30).clip(0, 11)

    n_rec      = len(recientes)
    n_prov_rec = recientes["rfc"].nunique()
    monto_rec  = recientes["Importe DRC"].sum()
    edad_min   = int(recientes["Edad_dias"].min()) if n_rec > 0 else None

    r1, r2, r3, r4 = st.columns(4)
    r1.metric("📋 Contratos",        f"{n_rec:,}")
    r2.metric("🏢 Proveedores",      f"{n_prov_rec:,}")
    r3.metric("💰 Monto total",      f"${monto_rec/1e6:,.1f} M MXN" if monto_rec > 0 else "N/D")
    r4.metric("⚡ Menor antigüedad", f"{edad_min} días" if edad_min is not None else "N/D")

    if n_rec > 0:
        st.warning(
            f"⚠️ **{n_rec:,} contratos** adjudicados a **{n_prov_rec:,} proveedores** "
            f"con menos de un año de constitución al inicio del contrato."
        )

        dist_meses = (
            recientes["Edad_meses"].value_counts()
            .reindex(range(12), fill_value=0)
            .reset_index()
        )
        dist_meses.columns = ["Meses", "Contratos"]
        dist_meses["Antigüedad"] = dist_meses["Meses"].apply(
            lambda m: f"{m} mes" if m == 1 else f"{m} meses"
        )
        fig_rec = px.bar(
            dist_meses, x="Antigüedad", y="Contratos",
            color_discrete_sequence=[IMSS_ROJO],
            title="Contratos por antigüedad de la empresa al inicio del contrato"
        )
        fig_rec.update_layout(
            font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            xaxis_title="Antigüedad al inicio del contrato",
            yaxis_title="Número de contratos",
            title_font_color=IMSS_VERDE_OSC
        )
        st.plotly_chart(fig_rec, use_container_width=True)

        st.markdown("**Top proveedores por monto contratado (reciente creación)**")
        top_rec = (
            recientes.groupby(["Proveedor o contratista", "rfc"])
            .agg(Contratos=("rfc", "count"),
                 Monto=("Importe DRC", "sum"),
                 Edad_min=("Edad_dias", "min"))
            .sort_values("Monto", ascending=False)
            .head(10)
            .reset_index()
        )
        # UC con mayor monto contratado para cada proveedor de reciente creación
        _uc_rec = (
            recientes.groupby(["rfc", "Nombre de la UC"])["Importe DRC"]
            .sum().reset_index()
            .sort_values("Importe DRC", ascending=False)
            .groupby("rfc", as_index=False).first()[["rfc", "Nombre de la UC"]]
            .rename(columns={"Nombre de la UC": "UC_principal"})
        )
        top_rec = top_rec.merge(_uc_rec, on="rfc", how="left")
        top_rec["Monto"]    = top_rec["Monto"].apply(lambda x: f"${x/1e6:.2f} M")
        top_rec["Edad_min"] = top_rec["Edad_min"].apply(lambda x: f"{int(x)} días")
        top_rec.columns     = ["Proveedor", "RFC", "Contratos", "Monto total", "Menor antigüedad", "UC principal"]
        top_rec.index += 1
        st.dataframe(top_rec, use_container_width=True)

        st.markdown("**📋 Detalle de contratos con empresas de reciente creación**")
        cols_rec = [c for c in [
            "Proveedor o contratista", "rfc", "Fecha_RFC", "Edad_dias",
            "Nombre de la UC", "Tipo Procedimiento",
            "Importe DRC", "Descripción del contrato",
            "Dirección del anuncio"
        ] if c in recientes.columns]
        tabla_rec = recientes[cols_rec].copy().rename(columns={
            "Fecha_RFC": "Fecha constitución (RFC)",
            "Edad_dias": "Días de antigüedad",
            "Importe DRC": "Importe",
        }).sort_values("Días de antigüedad").reset_index(drop=True)
        tabla_rec["Importe"] = tabla_rec["Importe"].apply(
            lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
        )
        tabla_rec.index += 1
        st.dataframe(
            tabla_rec,
            column_config={
                "Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                )
            },
            use_container_width=True
        )
    else:
        st.success(
            "✅ No se encontraron contratos con proveedores de reciente creación "
            "(menos de 1 año de constitución) con los filtros actuales."
        )

    st.divider()

    # ── SECCIÓN 7: CONTRATOS POR PROCESOS DE EXCEPCIÓN ──
    st.subheader("7️⃣ Contratos por Procesos de Excepción")
    with st.expander("ℹ️ Metodología y contexto", expanded=False):
        st.markdown(
            """
            Identifica contratos adjudicados bajo las **figuras de excepción** a la licitación pública
            y verifica si las Unidades Compradoras superan los **límites legales** establecidos:

            - **Art. 55 / 42 LAASSP** — excepción a la licitación pública por adquisición de bienes o
              servicios. Límite: el monto bajo excepción no debe rebasar el **30 %** del presupuesto
              autorizado de adquisiciones de la dependencia.
            - **Art. 43 LOPSRM** — excepción en obra pública y servicios relacionados. Límite: **20 %**
              del presupuesto autorizado.

            El análisis se realiza **por Unidad Compradora**, comparando el monto ejercido bajo excepción
            respecto al total de cada UC. Superar el umbral es una señal de alerta que puede indicar
            evasión sistemática del concurso público.
            """
        )

    # Filtros por artículo de excepción
    _mask_art55_s7  = _dff2["Artículo de excepción"].str.upper().str.startswith("ART. 55", na=False)
    _mask_art42_s7  = _dff2["Artículo de excepción"].str.upper().str.startswith("ART. 42", na=False)
    _mask_art20_s7  = _dff2["Artículo de excepción"].str.upper().str.startswith("ART. 43", na=False)
    _mask_laassp_s7 = _mask_art55_s7 | _mask_art42_s7   # umbral 30%
    _mask_lopsrm_s7 = _mask_art20_s7                     # umbral 20%
    _mask_exc_s7    = _mask_laassp_s7 | _mask_lopsrm_s7

    art_excepcion = _dff2[_mask_exc_s7].copy()
    art_laassp    = _dff2[_mask_laassp_s7].copy()
    art_lopsrm    = _dff2[_mask_lopsrm_s7].copy()

    n_excepcion      = len(art_excepcion)
    monto_excepcion  = art_excepcion["Importe DRC"].sum()
    pct_exc_gl       = (monto_excepcion / _monto_total2 * 100) if _monto_total2 > 0 else 0
    n_ucs_excepcion  = art_excepcion["Nombre de la UC"].nunique()

    # KPIs globales
    a1, a2, a3, a4 = st.columns(4)
    a1.metric("📋 Contratos por excepción",  f"{n_excepcion:,}")
    a2.metric("💰 Monto total excepción",     f"${monto_excepcion/1e6:,.1f} M MXN")
    a3.metric("📊 % del gasto total",         f"{pct_exc_gl:.1f}%")
    a4.metric("🏥 UCs involucradas",          f"{n_ucs_excepcion:,}")

    if n_excepcion == 0:
        st.info("ℹ️ No se encontraron contratos por procesos de excepción con los filtros actuales.")
    else:
        # ── Análisis por UC: LAASSP (30%) y LOPSRM (20%) ──
        _uc_total_s7 = (
            _dff2.groupby("Nombre de la UC")["Importe DRC"]
            .sum().rename("Monto_Total").reset_index()
        )
        _uc_laassp_s7 = (
            art_laassp.groupby("Nombre de la UC")["Importe DRC"]
            .sum().rename("Monto_LAASSP").reset_index()
        )
        _uc_lopsrm_s7 = (
            art_lopsrm.groupby("Nombre de la UC")["Importe DRC"]
            .sum().rename("Monto_LOPSRM").reset_index()
        )
        _uc_contratos_s7 = (
            art_excepcion.groupby("Nombre de la UC").size()
            .rename("Contratos").reset_index()
        )
        resumen_uc = (
            _uc_contratos_s7
            .merge(_uc_total_s7,   on="Nombre de la UC", how="left")
            .merge(_uc_laassp_s7,  on="Nombre de la UC", how="left")
            .merge(_uc_lopsrm_s7,  on="Nombre de la UC", how="left")
        )
        resumen_uc["Monto_LAASSP"]   = resumen_uc["Monto_LAASSP"].fillna(0)
        resumen_uc["Monto_LOPSRM"]   = resumen_uc["Monto_LOPSRM"].fillna(0)
        resumen_uc["Monto_Exc_Total"] = resumen_uc["Monto_LAASSP"] + resumen_uc["Monto_LOPSRM"]

        resumen_uc["Pct_LAASSP"] = (
            resumen_uc["Monto_LAASSP"] / resumen_uc["Monto_Total"].replace(0, pd.NA) * 100
        ).fillna(0).round(1)
        resumen_uc["Pct_LOPSRM"] = (
            resumen_uc["Monto_LOPSRM"] / resumen_uc["Monto_Total"].replace(0, pd.NA) * 100
        ).fillna(0).round(1)
        resumen_uc["Pct_Total_Exc"] = (
            resumen_uc["Monto_Exc_Total"] / resumen_uc["Monto_Total"].replace(0, pd.NA) * 100
        ).fillna(0).round(1)

        resumen_uc["Riesgo"] = resumen_uc.apply(
            lambda r: "🔴 Supera límite" if (r["Pct_LAASSP"] > 30 or r["Pct_LOPSRM"] > 20)
            else "🟢 Dentro del límite",
            axis=1
        )
        n_ucs_riesgo = (
            (resumen_uc["Pct_LAASSP"] > 30) | (resumen_uc["Pct_LOPSRM"] > 20)
        ).sum()

        if n_ucs_riesgo > 0:
            st.error(
                f"🚨 **{n_ucs_riesgo} Unidad(es) Compradora(s)** superan el umbral legal "
                f"(30% LAASSP o 20% LOPSRM) respecto a su gasto total registrado."
            )
        else:
            st.success("✅ Ninguna Unidad Compradora supera los umbrales legales de excepción.")

        color_map_art = {"🔴 Supera límite": IMSS_ROJO, "🟢 Dentro del límite": IMSS_VERDE}
        col_art1, col_art2 = st.columns(2)

        with col_art1:
            st.markdown("**Monto bajo excepción**")
            top20_monto = resumen_uc.sort_values("Monto_Exc_Total", ascending=False).head(20).copy()
            top20_monto["Monto_fmt"] = top20_monto["Monto_Exc_Total"].apply(
                lambda x: f"${x/1e6:,.1f} M"
            )
            top20_monto["UC_corta"] = top20_monto["Nombre de la UC"].apply(
                lambda s: s if len(s) <= 30 else s[:30] + "…"
            )
            fig_art1 = px.bar(
                top20_monto.sort_values("Monto_Exc_Total"),
                x="Monto_Exc_Total", y="UC_corta",
                orientation="h", text="Monto_fmt",
                color="Riesgo", color_discrete_map=color_map_art,
                custom_data=["Nombre de la UC", "Pct_Total_Exc"],
            )
            fig_art1.update_traces(
                textfont=dict(family="Noto Sans, sans-serif"),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Monto excepción: %{text}<br>"
                    "% del total UC: %{customdata[1]:.1f}%<extra></extra>"
                )
            )
            fig_art1.update_layout(
                font=plotly_font(), xaxis_title="Monto (MXN)", yaxis_title="",
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                height=max(420, 20 * 26),
                legend_title_text="",
                legend=dict(font=dict(family="Noto Sans, sans-serif", size=11)),
                yaxis=dict(tickfont=dict(size=10, family="Noto Sans, sans-serif"))
            )
            st.plotly_chart(fig_art1, use_container_width=True)

        with col_art2:
            st.markdown("**% de gasto por excepción respecto al gasto total**")
            pct_chart = resumen_uc.sort_values("Pct_Total_Exc", ascending=False).head(20).copy()
            pct_chart["UC_corta"] = pct_chart["Nombre de la UC"].apply(
                lambda s: s if len(s) <= 30 else s[:30] + "…"
            )
            pct_chart["Pct_fmt"] = pct_chart["Pct_Total_Exc"].apply(lambda x: f"{x:.1f}%")
            pct_chart_sorted = pct_chart.sort_values("Pct_Total_Exc").reset_index(drop=True)
            fig_art2 = px.bar(
                pct_chart_sorted,
                x="Pct_Total_Exc", y="UC_corta",
                orientation="h",
                color="Riesgo", color_discrete_map=color_map_art,
                text="Pct_fmt",
                custom_data=["Nombre de la UC", "Contratos"],
            )
            # Línea de umbral de referencia (30% LAASSP)
            fig_art2.add_vline(
                x=30, line_dash="dash", line_color=IMSS_ORO, line_width=2,
                annotation_text="Límite 30%",
                annotation_font=dict(family="Noto Sans, sans-serif", color=IMSS_ORO, size=10),
                annotation_position="top right"
            )
            fig_art2.update_traces(
                textfont=dict(family="Noto Sans, sans-serif"),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "% excepción total: %{x:.1f}%<br>"
                    "Contratos: %{customdata[1]}<extra></extra>"
                )
            )
            fig_art2.update_layout(
                font=plotly_font(), xaxis_title="% del gasto total de la UC",
                yaxis_title="", plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                height=max(420, 20 * 26),
                legend_title_text="",
                legend=dict(font=dict(family="Noto Sans, sans-serif", size=11)),
                yaxis=dict(tickfont=dict(size=10, family="Noto Sans, sans-serif"))
            )
            st.plotly_chart(fig_art2, use_container_width=True)

        # Tabla resumen por UC
        with st.expander("📋 Ver tabla completa por Unidad Compradora"):
            tabla_exc_uc = resumen_uc.sort_values("Pct_Total_Exc", ascending=False).copy()
            tabla_exc_uc["Monto_LAASSP"]   = tabla_exc_uc["Monto_LAASSP"].apply(lambda x: f"${x:,.0f}")
            tabla_exc_uc["Monto_LOPSRM"]   = tabla_exc_uc["Monto_LOPSRM"].apply(lambda x: f"${x:,.0f}")
            tabla_exc_uc["Monto_Exc_Total"] = tabla_exc_uc["Monto_Exc_Total"].apply(lambda x: f"${x:,.0f}")
            tabla_exc_uc["Monto_Total"]     = tabla_exc_uc["Monto_Total"].apply(lambda x: f"${x:,.0f}")
            tabla_exc_uc = tabla_exc_uc.rename(columns={
                "Nombre de la UC": "Unidad Compradora",
                "Contratos":       "Contratos excepción",
                "Monto_LAASSP":    "Monto LAASSP (Art.55/42)",
                "Monto_LOPSRM":    "Monto LOPSRM (Art.43)",
                "Monto_Exc_Total": "Monto total excepción",
                "Monto_Total":     "Monto Total UC",
                "Pct_LAASSP":      "% LAASSP / Total",
                "Pct_LOPSRM":      "% LOPSRM / Total",
                "Pct_Total_Exc":   "% Excepción / Total",
                "Riesgo":          "Nivel de riesgo",
            })
            tabla_exc_uc.index = range(1, len(tabla_exc_uc) + 1)
            st.dataframe(tabla_exc_uc, use_container_width=True)

        # Detalle de contratos por excepción
        st.markdown("**📋 Detalle de contratos por procesos de excepción**")
        cols_exc = [c for c in [
            "Artículo de excepción", "Descripción excepción",
            "Nombre de la UC", "Proveedor o contratista",
            "Tipo Procedimiento", "Importe DRC",
            "Descripción del contrato", "Dirección del anuncio"
        ] if c in art_excepcion.columns]
        tabla_det_exc = (
            art_excepcion[cols_exc]
            .sort_values("Importe DRC", ascending=False)
            .reset_index(drop=True)
            .rename(columns={"Importe DRC": "Importe"})
        )
        tabla_det_exc["Importe"] = tabla_det_exc["Importe"].apply(
            lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
        )
        tabla_det_exc.index += 1
        st.dataframe(
            tabla_det_exc,
            column_config={
                "Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                )
            },
            use_container_width=True
        )

    st.divider()

    # ── SECCIÓN 8: TIPO DE PROCEDIMIENTO — DISTRIBUCIÓN POR TIPO ──
    st.subheader("8️⃣ Tipo de Procedimiento — Distribución por Tipo de Procedimiento")
    st.caption(
        "La licitación pública es la **regla general** del sistema de contrataciones (Art. 35 LAASSP 2025). "
        "Las adjudicaciones directas por **Fracción I** (patentes, licencias exclusivas u oferente único — "
        "Art. 54 Fr. I) se separan por su naturaleza estructural, al no depender de discrecionalidad."
    )

    # Categorizar contratos en 4 grupos de análisis
    _mask_lp    = _dff2["Tipo Simplificado"] == "Licitación Pública"
    _mask_i3p   = _dff2["Tipo Simplificado"] == "Invitación a 3 personas"
    _mask_ep    = _dff2["Tipo Simplificado"] == "Entre Entes Públicos"
    _mask_frac1    = _dff2["Tipo Simplificado"] == "Adjudicación Directa — Fr. I"
    _mask_ad_otras = _dff2["Tipo Simplificado"] == "Adjudicación Directa"

    monto_lp      = _dff2.loc[_mask_lp,       "Importe DRC"].sum()
    monto_frac1   = _dff2.loc[_mask_frac1,    "Importe DRC"].sum()
    monto_ad_ot   = _dff2.loc[_mask_ad_otras, "Importe DRC"].sum()
    monto_i3p     = _dff2.loc[_mask_i3p,      "Importe DRC"].sum()
    monto_ep      = _dff2.loc[_mask_ep,       "Importe DRC"].sum()

    n_lp      = _mask_lp.sum()
    n_frac1   = _mask_frac1.sum()
    n_ad_ot   = _mask_ad_otras.sum()

    pct_lp_m    = monto_lp    / _monto_total2 * 100 if _monto_total2 > 0 else 0
    pct_frac1_m = monto_frac1 / _monto_total2 * 100 if _monto_total2 > 0 else 0
    pct_ad_ot_m = monto_ad_ot / _monto_total2 * 100 if _monto_total2 > 0 else 0

    # KPIs
    s1, s2, s3, s4 = st.columns(4)
    s1.metric(
        "🟢 % LP del monto total",
        f"{pct_lp_m:.1f}%"
    )
    s2.metric("📋 Contratos LP",              f"{n_lp:,}")
    s3.metric(
        "🔴 AD otras causales (% monto)",
        f"{pct_ad_ot_m:.1f}%  ({n_ad_ot:,} contratos)"
    )
    s4.metric(
        "🟡 AD Fracc. I — Patente (% monto)",
        f"{pct_frac1_m:.1f}%  ({n_frac1:,} contratos)"
    )

    # ── Gráfica de composición del gasto por tipo ──
    _resumen_tipo = pd.DataFrame({
        "Tipo": [
            "🟢 Licitación pública",
            "🔵 Invitación a 3 personas",
            "🔵 Entre Entes Públicos",
            "🟡 AD — Fracción I (Patente)",
            "🔴 AD — Otras causales",
        ],
        "Monto": [monto_lp, monto_i3p, monto_ep, monto_frac1, monto_ad_ot],
    })
    _resumen_tipo = _resumen_tipo[_resumen_tipo["Monto"] > 0]
    _resumen_tipo["Monto_fmt"] = _resumen_tipo["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
    _resumen_tipo["Pct"] = _resumen_tipo["Monto"] / _monto_total2 * 100

    _color_tipo8 = {
        "🟢 Licitación pública":         IMSS_VERDE,
        "🔵 Invitación a 3 personas":    IMSS_ORO,
        "🔵 Entre Entes Públicos":       IMSS_ORO_CLARO,
        "🟡 AD — Fracción I (Patente)":  "#C89F30",
        "🔴 AD — Otras causales":        IMSS_ROJO,
    }
    fig_tipo8 = px.pie(
        _resumen_tipo, names="Tipo", values="Monto",
        color="Tipo", color_discrete_map=_color_tipo8,
        title="Composición del gasto contratado por tipo de procedimiento",
        hole=0.4
    )
    fig_tipo8.update_traces(
        texttemplate="<b>%{label}</b><br>%{percent:.1%}",
        textposition="outside",
        hovertemplate="<b>%{label}</b><br>Monto: %{customdata}<br>%{percent:.1%}<extra></extra>",
        customdata=_resumen_tipo["Monto_fmt"],
    )
    fig_tipo8.update_layout(
        font=plotly_font(), showlegend=True,
        legend=dict(orientation="v", x=1.02, y=0.5),
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff"
    )
    st.plotly_chart(fig_tipo8, use_container_width=True)

    # ── Análisis por Unidad Compradora ──
    st.markdown("**% de Licitación Pública por Unidad Compradora (top 30 por monto total)**")

    _dff_tipo8 = _dff2.copy()
    _dff_tipo8["Cat8"] = "AD — Otras causales"
    _dff_tipo8.loc[_mask_lp,    "Cat8"] = "Licitación pública"
    _dff_tipo8.loc[_mask_i3p,   "Cat8"] = "Invitación a 3 personas"
    _dff_tipo8.loc[_mask_ep,    "Cat8"] = "Entre Entes Públicos"
    _dff_tipo8.loc[_mask_frac1, "Cat8"] = "AD — Fracción I (Patente)"

    _by_uc8 = (
        _dff_tipo8.groupby(["Nombre de la UC", "Cat8"])["Importe DRC"]
        .sum().unstack(fill_value=0).reset_index()
    )
    # Asegurar que todas las columnas existan
    for _col in ["Licitación pública", "Invitación a 3 personas", "Entre Entes Públicos",
                 "AD — Fracción I (Patente)", "AD — Otras causales"]:
        if _col not in _by_uc8.columns:
            _by_uc8[_col] = 0.0

    _by_uc8["Monto_total"] = _by_uc8[
        ["Licitación pública", "Invitación a 3 personas", "Entre Entes Públicos",
         "AD — Fracción I (Patente)", "AD — Otras causales"]
    ].sum(axis=1)
    _by_uc8["Pct_LP"] = (
        _by_uc8["Licitación pública"] / _by_uc8["Monto_total"].replace(0, pd.NA) * 100
    ).fillna(0)
    # Top 30 UCs por monto total
    _top30 = _by_uc8.nlargest(30, "Monto_total").sort_values("Pct_LP")
    _top30["UC_corta"] = _top30["Nombre de la UC"].apply(
        lambda s: s[:35] + "…" if len(s) > 35 else s
    )

    fig_uc8 = px.bar(
        _top30, x="Pct_LP", y="UC_corta",
        orientation="h",
        text=_top30["Pct_LP"].apply(lambda x: f"{x:.1f}%"),
        title="% del gasto por licitación pública — top 30 UCs por monto",
        custom_data=["Nombre de la UC", "Monto_total"]
    )
    fig_uc8.update_traces(marker_color=IMSS_VERDE)
    fig_uc8.update_layout(
        font=plotly_font(), xaxis_title="% del monto contratado por LP",
        yaxis_title="", showlegend=False,
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        xaxis=dict(range=[0, 115]),
        height=max(650, len(_top30) * 24 + 120),
        margin=dict(l=270, r=80, t=60, b=40),
    )
    fig_uc8.update_traces(
        textfont=dict(family="Noto Sans, sans-serif"),
        textposition="outside", cliponaxis=False,
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "% LP: %{x:.1f}%<br>"
            "Monto total: $%{customdata[1]:,.0f}<extra></extra>"
        )
    )
    st.plotly_chart(fig_uc8, use_container_width=True)

    with st.expander("📋 Ver tabla detallada por UC"):
        _tbl8 = _by_uc8.sort_values("Pct_LP").copy()
        _tbl8["Licitación pública"]      = _tbl8["Licitación pública"].apply(lambda x: f"${x:,.0f}")
        _tbl8["AD — Otras causales"]     = _tbl8["AD — Otras causales"].apply(lambda x: f"${x:,.0f}")
        _tbl8["AD — Fracción I (Patente)"] = _tbl8["AD — Fracción I (Patente)"].apply(lambda x: f"${x:,.0f}")
        _tbl8["Invitación a 3 personas"] = _tbl8["Invitación a 3 personas"].apply(lambda x: f"${x:,.0f}")
        _tbl8["Monto_total"]             = _tbl8["Monto_total"].apply(lambda x: f"${x:,.0f}")
        _tbl8["Pct_LP"]                  = _tbl8["Pct_LP"].apply(lambda x: f"{x:.1f}%")
        _tbl8 = _tbl8.drop(columns=["UC_corta", "Entre Entes Públicos"], errors="ignore")
        _tbl8 = _tbl8.rename(columns={"Nombre de la UC": "Unidad Compradora",
                                       "Monto_total": "Monto total"})
        _tbl8 = _tbl8.reset_index(drop=True)
        _tbl8.index += 1
        st.dataframe(_tbl8, use_container_width=True)

    st.divider()

    # ── SECCIÓN 9: CONTRATOS POR CASO FORTUITO / URGENCIA ──
    st.subheader("9️⃣ Contratos por Caso Fortuito (Fr. II) y Tiempo de Urgencia (Fr. V) — Art. 54 LAASSP")
    st.caption(
        "**Fr. II — Caso fortuito o fuerza mayor:** eventos imprevisibles que impiden licitación. "
        "**Fr. V — Tiempo de urgencia:** circunstancias urgentes *no atribuibles a falta de planeación* "
        "que no pueden atenderse por licitación. La ley explícitamente excluye la falta de planeación como "
        "justificación para Fr. V — un patrón recurrente en la misma UC puede indicar uso indebido de esta excepción."
    )

    # Fr. II — Caso fortuito (filtro por Artículo de excepción, con fallback a descripción)
    _art_exc9 = _dff2["Artículo de excepción"].fillna("").str.upper().str.strip()
    if _art_exc9.str.contains("ART. 54 FR. II", na=False).any():
        caso_f9 = _dff2[_art_exc9 == "ART. 54 FR. II"].copy()
    else:
        caso_f9 = _dff2[
            _dff2["Descripción excepción"].str.upper().str.contains("CASO FORTUITO", na=False)
        ].copy()

    # Fr. V — Tiempo de urgencia (filtro por Artículo de excepción)
    frac5_f9 = _dff2[_art_exc9 == "ART. 54 FR. V"].copy()

    n_cf9      = len(caso_f9)
    monto_cf9  = caso_f9["Importe DRC"].sum()
    pct_cf9    = monto_cf9 / _monto_total2 * 100 if _monto_total2 > 0 else 0
    n_uc_cf9   = caso_f9["Nombre de la UC"].nunique()

    n_frv      = len(frac5_f9)
    monto_frv  = frac5_f9["Importe DRC"].sum()
    pct_frv    = monto_frv / _monto_total2 * 100 if _monto_total2 > 0 else 0
    n_uc_frv   = frac5_f9["Nombre de la UC"].nunique()

    # KPIs — Fr. II
    st.markdown("##### Art. 54 Fr. II — Caso fortuito o fuerza mayor")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("⚡ Contratos caso fortuito",   f"{n_cf9:,}")
    c2.metric("🏥 UCs con caso fortuito",    f"{n_uc_cf9:,}")
    c3.metric("💰 Monto total",
              f"${monto_cf9/1e9:,.2f} miles de millones MXN" if monto_cf9 >= 1e9
              else f"${monto_cf9/1e6:,.1f} M MXN")
    c4.metric("📊 % del gasto total",        f"{pct_cf9:.1f}%")

    # Monto total por UC (usado por Fr. II y Fr. V)
    _monto_uc_total = _dff2.groupby("Nombre de la UC")["Importe DRC"].sum().rename("Monto_uc")

    if n_cf9 == 0:
        st.success("✅ No se detectaron contratos por caso fortuito con los filtros actuales.")
    else:
        st.warning(
            f"⚠️ **{n_cf9:,} contratos** por caso fortuito suman "
            f"**${monto_cf9/1e6:,.1f} M MXN** ({pct_cf9:.1f}% del gasto total). "
            f"Verificar que correspondan a situaciones realmente imprevisibles."
        )

        # ── % de caso fortuito por UC (top 20 por monto) ──
        _monto_uc_cf    = caso_f9.groupby("Nombre de la UC")["Importe DRC"].sum().rename("Monto_cf")

        _cf_uc = pd.concat([_monto_uc_cf, _monto_uc_total], axis=1).fillna(0).reset_index()
        _cf_uc.columns = ["Nombre de la UC", "Monto_cf", "Monto_uc"]
        _cf_uc["Pct_CF"] = (_cf_uc["Monto_cf"] / _cf_uc["Monto_uc"].replace(0, pd.NA) * 100).fillna(0)
        _cf_uc = _cf_uc[_cf_uc["Monto_cf"] > 0]

        # Top 20 por monto absoluto de caso fortuito
        _top20_cf = _cf_uc.nlargest(20, "Monto_cf").sort_values("Pct_CF")
        _top20_cf["Pct_fmt"] = _top20_cf["Pct_CF"].apply(lambda x: f"{x:.1f}%")
        _top20_cf["Color"]   = _top20_cf["Pct_CF"].apply(
            lambda x: IMSS_ROJO if x >= 30 else (IMSS_ORO if x >= 15 else IMSS_VERDE)
        )
        # Orden y etiquetas: nombre completo como clave (sin duplicados), truncado solo para display
        _cf9_order     = _top20_cf["Nombre de la UC"].tolist()
        _cf9_ticktext  = [s[:50] + "…" if len(str(s)) > 50 else s for s in _cf9_order]

        fig_cf9 = px.bar(
            _top20_cf, x="Pct_CF", y="Nombre de la UC",
            orientation="h", text="Pct_fmt",
            color="Color",
            color_discrete_map={IMSS_ROJO: IMSS_ROJO, IMSS_ORO: IMSS_ORO, IMSS_VERDE: IMSS_VERDE},
            title="% del gasto por UC que corresponde a caso fortuito (top 20 por monto absoluto)",
            category_orders={"Nombre de la UC": _cf9_order},
            custom_data=["Nombre de la UC", "Monto_cf", "Monto_uc"]
        )
        fig_cf9.update_layout(
            font=plotly_font(), xaxis_title="% del gasto de la UC",
            yaxis=dict(
                title="",
                tickvals=_cf9_order,
                ticktext=_cf9_ticktext,
                tickfont=dict(size=10),
            ),
            showlegend=False,
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff"
        )
        fig_cf9.update_traces(
            textfont=dict(family="Noto Sans, sans-serif"),
            textposition="outside", cliponaxis=False,
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Caso fortuito: $%{customdata[1]:,.0f}<br>"
                "Total UC: $%{customdata[2]:,.0f}<br>"
                "% caso fortuito: %{x:.1f}%<extra></extra>"
            )
        )
        st.plotly_chart(fig_cf9, use_container_width=True)

        # ── Tabla de contratos más grandes de caso fortuito ──
        st.markdown("**📋 Contratos de caso fortuito de mayor monto**")
        _cols_cf9 = [c for c in [
            "Nombre de la UC", "Proveedor o contratista",
            "Importe DRC", "Descripción del contrato",
            "Fecha de fallo", "Dirección del anuncio"
        ] if c in caso_f9.columns]
        _tabla_cf9 = (
            caso_f9[_cols_cf9]
            .sort_values("Importe DRC", ascending=False)
            .head(50)
            .reset_index(drop=True)
        )
        _tabla_cf9["Importe DRC"] = _tabla_cf9["Importe DRC"].apply(
            lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
        )
        _tabla_cf9.index += 1
        st.dataframe(
            _tabla_cf9,
            column_config={
                "Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                )
            },
            use_container_width=True
        )

    # ── Fr. V — Tiempo de urgencia ──
    st.markdown("---")
    st.markdown("##### Art. 54 Fr. V — Tiempo de urgencia")
    st.caption(
        "La Fr. V requiere que la urgencia **no sea resultado de falta de planeación** (Art. 54 Fr. V LAASSP 2025). "
        "Contratos recurrentes bajo esta fracción en la misma Unidad Compradora son una señal de alerta."
    )

    # KPIs Fr. V
    v1, v2, v3, v4 = st.columns(4)
    v1.metric("⏱️ Contratos Fr. V (urgencia)",  f"{n_frv:,}")
    v2.metric("🏥 UCs con Fr. V",               f"{n_uc_frv:,}")
    v3.metric("💰 Monto total",
              f"${monto_frv/1e9:,.2f} miles de millones MXN" if monto_frv >= 1e9
              else f"${monto_frv/1e6:,.1f} M MXN")
    v4.metric("📊 % del gasto total",           f"{pct_frv:.1f}%")

    if n_frv == 0:
        st.success("✅ No se detectaron contratos por Fr. V (urgencia) con los filtros actuales.")
    else:
        st.warning(
            f"⚠️ **{n_frv:,} contratos** por tiempo de urgencia (Fr. V) suman "
            f"**${monto_frv/1e6:,.1f} M MXN** ({pct_frv:.1f}% del gasto total). "
            f"Verificar que ninguno sea atribuible a falta de planeación."
        )

        # Tabla Fr. V con link ComprasMX
        st.markdown("**📋 Contratos bajo Art. 54 Fr. V — Verificar falta de planeación**")
        _cols_frv = [c for c in [
            "Nombre de la UC", "Proveedor o contratista",
            "Importe DRC", "Descripción del contrato",
            "Descripción excepción", "Artículo de excepción",
            "Fecha de fallo", "Dirección del anuncio"
        ] if c in frac5_f9.columns]
        _tabla_frv = (
            frac5_f9[_cols_frv]
            .sort_values("Importe DRC", ascending=False)
            .head(100)
            .reset_index(drop=True)
        )
        _tabla_frv["Importe DRC"] = _tabla_frv["Importe DRC"].apply(
            lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
        )
        _tabla_frv.index += 1
        st.dataframe(
            _tabla_frv,
            column_config={
                "Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                )
            },
            use_container_width=True
        )

        # % de Fr. V por UC (top 15 por monto)
        with st.expander("📊 % de gasto Fr. V por Unidad Compradora"):
            _monto_uc_frv = frac5_f9.groupby("Nombre de la UC")["Importe DRC"].sum().rename("Monto_frv")
            _cf_uc_frv = pd.concat([_monto_uc_frv, _monto_uc_total], axis=1).fillna(0).reset_index()
            _cf_uc_frv.columns = ["Nombre de la UC", "Monto_frv", "Monto_uc"]
            _cf_uc_frv["Pct_FrV"] = (
                _cf_uc_frv["Monto_frv"] / _cf_uc_frv["Monto_uc"].replace(0, pd.NA) * 100
            ).fillna(0)
            _cf_uc_frv = _cf_uc_frv[_cf_uc_frv["Monto_frv"] > 0]
            _top15_frv = _cf_uc_frv.nlargest(15, "Monto_frv").sort_values("Pct_FrV")
            _top15_frv["UC_corta"] = _top15_frv["Nombre de la UC"].apply(
                lambda s: s[:50] + "…" if len(s) > 50 else s
            )
            fig_frv = px.bar(
                _top15_frv, x="Pct_FrV", y="UC_corta",
                orientation="h",
                text=_top15_frv["Pct_FrV"].apply(lambda x: f"{x:.1f}%"),
                title="% del gasto por UC bajo Fr. V — urgencia (top 15 por monto)",
                custom_data=["Nombre de la UC", "Monto_frv"]
            )
            fig_frv.update_traces(
                marker_color=IMSS_ORO,
                textfont=dict(family="Noto Sans, sans-serif"),
                textposition="outside", cliponaxis=False,
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Fr. V: $%{customdata[1]:,.0f}<br>"
                    "% urgencia: %{x:.1f}%<extra></extra>"
                )
            )
            fig_frv.update_layout(
                font=plotly_font(), xaxis_title="% del gasto de la UC",
                yaxis_title="", showlegend=False,
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff"
            )
            st.plotly_chart(fig_frv, use_container_width=True)

    st.divider()

    # ── SECCIÓN 10: CONCENTRACIÓN DE PROVEEDORES ──────────────────────
    st.subheader("🔟 Concentración de Proveedores por Unidad Compradora")
    st.caption(
        "Mide qué tanto depende cada Unidad Compradora de un proveedor individual. "
        "Un proveedor que concentra más del **50 %** del gasto de una UC puede indicar "
        "falta de competencia o adjudicación sistemática. El **Índice HHI** "
        "(Herfindahl-Hirschman) cuantifica la concentración total: "
        "> 2,500 = alta · 1,500-2,500 = moderada · < 1,500 = competitiva."
    )

    # ── Calcular concentración por (UC, Proveedor) ────────────────────
    _conc_grp = (
        _dff2.groupby(["Nombre de la UC", "Proveedor o contratista", "rfc"])["Importe DRC"]
        .sum()
        .reset_index()
        .rename(columns={
            "Nombre de la UC":         "UC",
            "Proveedor o contratista": "Proveedor",
            "rfc":                     "RFC",
            "Importe DRC":             "Monto_Prov"
        })
    )
    _conc_uc_total = _conc_grp.groupby("UC")["Monto_Prov"].sum().rename("Monto_UC")
    _conc_grp = _conc_grp.merge(_conc_uc_total, on="UC")
    _conc_grp["Share"] = (
        _conc_grp["Monto_Prov"] / _conc_grp["Monto_UC"].replace(0, pd.NA)
    ).fillna(0)

    # HHI por UC  (suma de cuadrados de market shares × 10 000)
    _hhi_uc = (
        _conc_grp.groupby("UC")["Share"]
        .apply(lambda x: round((x ** 2).sum() * 10_000, 0))
        .rename("HHI")
        .reset_index()
    )

    # Proveedor principal por UC (mayor share)
    _top_prov_uc = (
        _conc_grp.sort_values("Share", ascending=False)
        .groupby("UC", as_index=False)
        .first()
        [["UC", "Proveedor", "RFC", "Monto_Prov", "Monto_UC", "Share"]]
        .rename(columns={
            "Proveedor":  "Top_Proveedor",
            "RFC":        "Top_RFC",
            "Monto_Prov": "Monto_Top",
            "Share":      "Share_Top"
        })
    )

    # Número de proveedores y contratos por UC
    _n_prov_uc  = _conc_grp.groupby("UC")["Proveedor"].nunique().rename("N_Proveedores").reset_index()
    _n_contr_uc = (
        _dff2.groupby("Nombre de la UC").size().rename("N_Contratos")
        .reset_index().rename(columns={"Nombre de la UC": "UC"})
    )

    _conc_uc = (
        _top_prov_uc
        .merge(_hhi_uc,     on="UC")
        .merge(_n_prov_uc,  on="UC")
        .merge(_n_contr_uc, on="UC", how="left")
    )

    def _nivel_conc(share):
        if share >= 0.80: return "🔴 Alta — ≥ 80 %"
        if share >= 0.50: return "🟠 Media — 50-79 %"
        return "🟢 Baja — < 50 %"

    def _nivel_hhi(hhi):
        if hhi > 2500:  return "🔴 Alto (> 2,500)"
        if hhi >= 1500: return "🟠 Moderado (1,500-2,500)"
        return "🟢 Bajo (< 1,500)"

    _conc_uc["Nivel_Conc"] = _conc_uc["Share_Top"].apply(_nivel_conc)
    _conc_uc["Nivel_HHI"]  = _conc_uc["HHI"].apply(_nivel_hhi)

    _color_conc = {
        "🔴 Alta — ≥ 80 %":   IMSS_ROJO,
        "🟠 Media — 50-79 %": "#E07B00",
        "🟢 Baja — < 50 %":   IMSS_VERDE,
    }

    # ── KPIs ──────────────────────────────────────────────────────────
    _n_conc_alta  = (_conc_uc["Share_Top"] >= 0.80).sum()
    _n_conc_media = ((_conc_uc["Share_Top"] >= 0.50) & (_conc_uc["Share_Top"] < 0.80)).sum()
    _n_conc_hhi   = (_conc_uc["HHI"] > 2500).sum()
    _n_conc_total = len(_conc_uc)

    _ck1, _ck2, _ck3, _ck4 = st.columns(4)
    _ck1.metric("🔴 UCs con proveedor dominante (≥ 80 %)",    f"{_n_conc_alta:,}")
    _ck2.metric("🟠 UCs con proveedor mayoritario (50-79 %)", f"{_n_conc_media:,}")
    _ck3.metric("📊 UCs con HHI alto (> 2,500)",              f"{_n_conc_hhi:,}")
    _ck4.metric("📋 UCs analizadas",                           f"{_n_conc_total:,}")

    if _n_conc_alta > 0:
        st.error(
            f"🚨 **{_n_conc_alta:,} Unidades Compradoras** tienen un proveedor que concentra "
            f"el **80 % o más** de su gasto total. Esto puede indicar falta de competencia "
            f"o adjudicación sistemática."
        )
    if _n_conc_media > 0:
        st.warning(
            f"⚠️ **{_n_conc_media:,} Unidades Compradoras** tienen un proveedor con entre el "
            f"50 % y 79 % de su gasto total."
        )

    # ── Gráfica 1: Scatter — tamaño de UC vs. dominancia ─────────────
    st.markdown("**Panorama general: gasto de la UC vs. concentración del proveedor principal**")
    _sc = _conc_uc.copy()
    _sc["Share_Pct"]     = (_sc["Share_Top"] * 100).round(2)
    _sc["UC_c"]          = _sc["UC"].apply(lambda s: str(s)[:50] + "…" if len(str(s)) > 50 else str(s))
    _sc["Monto_UC_fmt"]  = _sc["Monto_UC"].apply(lambda x: f"${x/1e6:,.1f} M")
    _sc["Monto_Top_fmt"] = _sc["Monto_Top"].apply(lambda x: f"${x/1e6:,.1f} M")

    fig_scatter_conc = px.scatter(
        _sc,
        x="Monto_UC", y="Share_Pct",
        color="Nivel_Conc",
        color_discrete_map=_color_conc,
        size="Monto_Top", size_max=45,
        hover_name="UC_c",
        custom_data=["UC", "Top_Proveedor", "Share_Pct",
                     "Monto_Top_fmt", "Monto_UC_fmt", "N_Contratos", "HHI", "N_Proveedores"],
        labels={
            "Monto_UC":   "Monto total contratado por la UC (MXN)",
            "Share_Pct":  "% del gasto concentrado en el proveedor principal",
            "Nivel_Conc": "Concentración"
        },
        title="Gasto total de la UC vs. % concentrado en el proveedor principal"
    )
    for _yval, _lcolor, _label in [
        (80, IMSS_ROJO, "80 % — Riesgo alto"),
        (50, "#E07B00", "50 % — Riesgo medio"),
    ]:
        fig_scatter_conc.add_hline(
            y=_yval, line_dash="dash", line_color=_lcolor,
            annotation_text=_label, annotation_position="top right",
            annotation_font=dict(family="Noto Sans, sans-serif", color=_lcolor, size=11)
        )
    fig_scatter_conc.update_traces(
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "Proveedor principal: <b>%{customdata[1]}</b><br>"
            "% del gasto de la UC: %{customdata[2]:.1f} %<br>"
            "Monto proveedor: %{customdata[3]}<br>"
            "Monto total UC: %{customdata[4]}<br>"
            "Contratos: %{customdata[5]:,.0f}  |  "
            "HHI: %{customdata[6]:,.0f}  |  "
            "Proveedores: %{customdata[7]:,.0f}"
            "<extra></extra>"
        )
    )
    fig_scatter_conc.update_layout(
        font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        xaxis=dict(tickprefix="$", tickformat=",.0f"),
        yaxis=dict(range=[-2, 107], ticksuffix=" %"),
        legend=dict(title="Concentración", orientation="h", y=-0.14, x=0),
        height=520
    )
    st.plotly_chart(fig_scatter_conc, use_container_width=True)

    # ── Gráfica 2: Barras — top 20 UCs por concentración ─────────────
    st.markdown("**Top 20 Unidades Compradoras con mayor concentración en el proveedor principal**")
    _top20_conc = (
        _conc_uc.nlargest(20, "Share_Top")
        .sort_values("Share_Top")
        .copy()
    )
    _top20_conc["UC_c"]      = _top20_conc["UC"].apply(lambda s: str(s)[:50] + "…" if len(str(s)) > 50 else str(s))
    _top20_conc["Share_Pct"] = (_top20_conc["Share_Top"] * 100).round(1)
    _top20_conc["Share_fmt"] = _top20_conc["Share_Pct"].apply(lambda x: f"{x:.1f} %")

    fig_bar_conc = px.bar(
        _top20_conc, x="Share_Pct", y="UC_c",
        orientation="h",
        color="Nivel_Conc",
        color_discrete_map=_color_conc,
        text="Share_fmt",
        title="Top 20 UCs — % del gasto concentrado en el proveedor principal",
        custom_data=["UC", "Top_Proveedor", "Monto_Top", "Monto_UC",
                     "N_Contratos", "HHI", "N_Proveedores", "Nivel_HHI"]
    )
    for _xval, _lcolor, _label in [
        (80, IMSS_ROJO, "80 %"),
        (50, "#E07B00", "50 %"),
    ]:
        fig_bar_conc.add_vline(
            x=_xval, line_dash="dash", line_color=_lcolor,
            annotation_text=_label, annotation_position="top right",
            annotation_font=dict(family="Noto Sans, sans-serif", color=_lcolor, size=11)
        )
    fig_bar_conc.update_layout(
        font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        xaxis_title="% del gasto de la UC concentrado en el proveedor principal",
        yaxis_title="", showlegend=True,
        legend=dict(title="Concentración", orientation="h", y=-0.10, x=0),
        xaxis=dict(range=[0, 115], ticksuffix=" %"),
        height=600
    )
    fig_bar_conc.update_traces(
        textposition="outside", cliponaxis=False,
        textfont=dict(family="Noto Sans, sans-serif"),
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "Proveedor principal: <b>%{customdata[1]}</b><br>"
            "% del gasto: %{x:.1f} %<br>"
            "Monto proveedor: $%{customdata[2]:,.0f}<br>"
            "Monto total UC: $%{customdata[3]:,.0f}<br>"
            "Contratos: %{customdata[4]:,.0f}  |  "
            "HHI: %{customdata[5]:,.0f}  |  "
            "Proveedores: %{customdata[6]:,.0f}<br>"
            "HHI nivel: %{customdata[7]}"
            "<extra></extra>"
        )
    )
    st.plotly_chart(fig_bar_conc, use_container_width=True)

    # Tabla resumen general (expandible)
    with st.expander("📋 Tabla de concentración — todas las UCs"):
        _tbl_conc_all = _conc_uc.sort_values("Share_Top", ascending=False).copy()
        _tbl_conc_all["% prov. principal"]     = _tbl_conc_all["Share_Top"].apply(lambda x: f"{x*100:.1f} %")
        _tbl_conc_all["Monto prov. principal"] = _tbl_conc_all["Monto_Top"].apply(lambda x: f"${x/1e6:,.1f} M")
        _tbl_conc_all["Monto total UC"]        = _tbl_conc_all["Monto_UC"].apply(lambda x: f"${x/1e6:,.1f} M")
        _tbl_conc_all["HHI"]                   = _tbl_conc_all["HHI"].apply(lambda x: f"{x:,.0f}")
        _tbl_conc_all = _tbl_conc_all.rename(columns={
            "UC":            "Unidad Compradora",
            "Top_Proveedor": "Proveedor principal",
            "Top_RFC":       "RFC",
            "N_Proveedores": "# Proveedores",
            "N_Contratos":   "# Contratos",
            "Nivel_Conc":    "Concentración",
            "Nivel_HHI":     "HHI — Nivel"
        })[["Unidad Compradora", "Proveedor principal", "RFC",
            "% prov. principal", "Monto prov. principal", "Monto total UC",
            "# Proveedores", "# Contratos", "HHI", "Concentración", "HHI — Nivel"]
        ].reset_index(drop=True)
        _tbl_conc_all.index += 1
        st.dataframe(_tbl_conc_all, use_container_width=True)

    st.divider()

    # ── Detalle por UC seleccionada ───────────────────────────────────
    st.markdown("**🔍 Distribución de proveedores en una Unidad Compradora**")

    _ucs_conc_det = sorted(_conc_uc["UC"].dropna().unique().tolist())
    _default_uc_conc = (
        _conc_uc.loc[_conc_uc["Share_Top"].idxmax(), "UC"]
        if len(_conc_uc) > 0 else _ucs_conc_det[0]
    )
    _uc_sel_conc = st.selectbox(
        "Seleccionar Unidad Compradora",
        _ucs_conc_det,
        index=_ucs_conc_det.index(_default_uc_conc) if _default_uc_conc in _ucs_conc_det else 0,
        key="conc_uc_sel",
        help="La UC preseleccionada es la de mayor concentración en el proveedor principal."
    )

    _df_prov_uc = _conc_grp[_conc_grp["UC"] == _uc_sel_conc].copy().sort_values("Monto_Prov", ascending=False)
    _info_uc    = _conc_uc[_conc_uc["UC"] == _uc_sel_conc]
    _info_uc    = _info_uc.iloc[0] if len(_info_uc) > 0 else None

    if _info_uc is not None:
        _cd1, _cd2, _cd3, _cd4 = st.columns(4)
        _cd1.metric("🏢 Núm. proveedores",          f"{int(_info_uc['N_Proveedores']):,}")
        _cd2.metric("💰 Monto total",               f"${_info_uc['Monto_UC']/1e6:,.1f} M MXN")
        _cd3.metric("📊 Índice HHI",                f"{_info_uc['HHI']:,.0f}",
                    delta=_info_uc["Nivel_HHI"], delta_color="off")
        _cd4.metric("🥇 % del proveedor principal", f"{_info_uc['Share_Top']*100:.1f} %")

    _col_pie_c, _col_bar_c = st.columns(2)

    # Donut: top 8 proveedores + "Otros"
    _pie_top8    = _df_prov_uc.head(8).copy()
    _monto_otros = _df_prov_uc.iloc[8:]["Monto_Prov"].sum()
    _monto_uc_t  = _df_prov_uc["Monto_UC"].iloc[0] if len(_df_prov_uc) > 0 else 1

    if _monto_otros > 0:
        _fila_otros = pd.DataFrame({
            "UC": [_uc_sel_conc], "Proveedor": ["Otros proveedores"],
            "RFC": ["—"], "Monto_Prov": [_monto_otros],
            "Monto_UC": [_monto_uc_t], "Share": [_monto_otros / _monto_uc_t]
        })
        _pie_data = pd.concat([_pie_top8, _fila_otros], ignore_index=True)
    else:
        _pie_data = _pie_top8.copy()

    _pie_data["Prov_c"]    = _pie_data["Proveedor"].apply(lambda s: str(s)[:42] + "…" if len(str(s)) > 42 else str(s))
    _pie_data["Share_Pct"] = (_pie_data["Share"] * 100).round(1)

    _pie_palette = (
        [IMSS_ROJO if len(_pie_data) > 0 and _pie_data["Share"].iloc[0] >= 0.50 else IMSS_VERDE]
        + [IMSS_ORO, "#5D8AA8", IMSS_ORO_CLARO, IMSS_VERDE_OSC,
           "#8B5CF6", "#6B7280", IMSS_GRIS, IMSS_NEGRO]
    )[:len(_pie_data)]

    fig_pie_conc = px.pie(
        _pie_data, names="Prov_c", values="Monto_Prov",
        color="Prov_c", color_discrete_sequence=_pie_palette,
        title=f"Distribución del gasto — {_uc_sel_conc}",
        hole=0.42
    )
    fig_pie_conc.update_traces(
        texttemplate="<b>%{label}</b><br>%{percent:.1%}",
        textposition="outside",
        hovertemplate="<b>%{label}</b><br>Monto: $%{value:,.0f}<br>%{percent:.1%}<extra></extra>"
    )
    fig_pie_conc.update_layout(
        font=plotly_font(), showlegend=False,
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        height=480
    )
    _col_pie_c.plotly_chart(fig_pie_conc, use_container_width=True)

    # Barras por proveedor (top 15, color continuo = share %)
    _bar_prov = _df_prov_uc.head(15).sort_values("Monto_Prov").copy()
    _bar_prov["Prov_c"]    = _bar_prov["Proveedor"].apply(lambda s: str(s)[:48] + "…" if len(str(s)) > 48 else str(s))
    _bar_prov["Share_Pct"] = (_bar_prov["Share"] * 100).round(1)
    _bar_prov["Share_fmt"] = _bar_prov["Share_Pct"].apply(lambda x: f"{x:.1f} %")

    fig_bar_prov = px.bar(
        _bar_prov, x="Monto_Prov", y="Prov_c",
        orientation="h",
        color="Share_Pct",
        color_continuous_scale=[[0, IMSS_VERDE], [0.5, IMSS_ORO_CLARO], [1, IMSS_ROJO]],
        range_color=[0, 100],
        text="Share_fmt",
        title="Monto por proveedor (top 15)",
        custom_data=["Proveedor", "RFC", "Share_Pct"]
    )
    fig_bar_prov.update_layout(
        font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        xaxis_title="Monto contratado (MXN)", yaxis_title="",
        xaxis=dict(tickprefix="$", tickformat=",.0f"),
        coloraxis_colorbar=dict(title="% del total<br>de la UC", ticksuffix=" %"),
        height=480
    )
    fig_bar_prov.update_traces(
        textposition="outside", cliponaxis=False,
        textfont=dict(family="Noto Sans, sans-serif"),
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "RFC: %{customdata[1]}<br>"
            "% de la UC: %{customdata[2]:.1f} %<br>"
            "Monto: $%{x:,.0f}<extra></extra>"
        )
    )
    _col_bar_c.plotly_chart(fig_bar_prov, use_container_width=True)

    # Tabla detallada de contratos de la UC seleccionada
    with st.expander(f"📋 Detalle de contratos — {_uc_sel_conc}"):
        _cols_uc_conc = [c for c in [
            "Proveedor o contratista", "rfc", "Tipo Simplificado",
            "Importe DRC", "Descripción del contrato",
            "Fecha de fallo", "Dirección del anuncio"
        ] if c in _dff2.columns]
        _tbl_uc_conc_det = (
            _dff2[_dff2["Nombre de la UC"] == _uc_sel_conc][_cols_uc_conc]
            .sort_values("Importe DRC", ascending=False)
            .reset_index(drop=True)
        )
        _tbl_uc_conc_det.index += 1
        _tbl_uc_conc_det["Importe DRC"] = _tbl_uc_conc_det["Importe DRC"].apply(
            lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
        )
        st.dataframe(
            _tbl_uc_conc_det,
            column_config={
                "Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                )
            },
            use_container_width=True, height=420
        )

# ───────────────────────────────────────────────────────────────
# PÁGINA 3: PAGINA_EXPLORADOR
# ───────────────────────────────────────────────────────────────
def pagina_explorador():

    st.header("🔍 Explorador de Gasto por UC y Adscripción")

    # ── Preparar datos: merge con Base_UC_2025_V2 (sin explotar) ──
    dff_uc = dff.merge(
        df_dir_uc[["Clave_UC", "Tipo UC", "Adscripción"]],
        left_on="Clave de la UC",
        right_on="Clave_UC",
        how="left"
    ).drop(columns=["Clave_UC"])
    dff_uc["Adscripción"] = dff_uc["Adscripción"].fillna("Sin clasificar")
    dff_uc["Tipo UC"]     = dff_uc["Tipo UC"].fillna("Sin clasificar")

    # ── Versión explodida + CUCoP para análisis de partidas ──
    _exp3 = dff_uc.copy()
    _exp3["_lista"] = _exp3["Partida específica"].str.split(",")
    _exp3["_n"]     = _exp3["_lista"].apply(len)
    _exp3["Importe DRC"] = _exp3["Importe DRC"] / _exp3["_n"]
    _exp3 = _exp3.explode("_lista")
    _exp3["Partida específica"] = _exp3["_lista"].str.strip().str.zfill(5)
    _exp3 = _exp3.drop(columns=["_lista", "_n"])

    dff_uc_cucop = _exp3.merge(
        df_cucop[["PARTIDA ESPECÍFICA", "DESC. PARTIDA ESPECÍFICA",
                  "PARTIDA GENÉRICA", "DESC. PARTIDA GENÉRICA",
                  "DESC. CAPÍTULO"]].drop_duplicates("PARTIDA ESPECÍFICA"),
        left_on="Partida específica",
        right_on="PARTIDA ESPECÍFICA",
        how="left"
    )

    # ── SELECTOR DE ÁMBITO: Nivel Central / OOAD / UMAE ──
    ambito_sel3 = st.radio(
        "📍 Ámbito",
        ["🏢  Nivel Central", "🗺️  OOAD", "🏥  UMAE"],
        horizontal=True,
        key="e3_ambito"
    )
    ambito_tipo      = ("Nivel Central" if "Nivel Central" in ambito_sel3
                        else "OOAD" if "OOAD" in ambito_sel3 else "UMAE")
    es_nivel_central = ambito_tipo == "Nivel Central"

    st.divider()

    # ── FILTROS INLINE (condicionales según ámbito) ──
    if es_nivel_central:
        # Nivel Central: una sola agrupación posible → por UC
        adsc_sel3  = "Nivel Central"
        agrupar3   = "Unidad Compradora"
        cf1, cf2 = st.columns([1, 1])
        with cf1:
            top_n_g3 = st.selectbox("Top N UCs", [10, 20, 30, 50],
                                     key="e3_topn", index=1)
        with cf2:
            pass  # espacio vacío para alineación
    else:
        # OOAD o UMAE: adscripciones acotadas al tipo seleccionado
        adscripciones_d = sorted([
            a for a in dff_uc[dff_uc["Tipo UC"] == ambito_tipo]["Adscripción"].dropna().unique()
            if a != "Sin clasificar"
        ])
        cf1, cf2, cf3 = st.columns([3, 1, 1])
        with cf1:
            adsc_sel3 = st.selectbox(
                "🏛️ Adscripción",
                ["Todas"] + adscripciones_d,
                key="e3_adsc"
            )
        with cf2:
            agrupar3 = st.selectbox("📊 Agrupar por",
                                    ["Adscripción", "Unidad Compradora"],
                                    key="e3_agrup")
        with cf3:
            top_n_g3 = st.selectbox("Top N", [10, 20, 30, 50],
                                     key="e3_topn", index=1)

    # Aplicar filtros a ambas versiones del DataFrame
    if es_nivel_central:
        mask_b = dff_uc["Tipo UC"] == "Nivel Central"
        mask_c = dff_uc_cucop["Tipo UC"] == "Nivel Central"
    else:
        mask_b = dff_uc["Tipo UC"] == ambito_tipo
        mask_c = dff_uc_cucop["Tipo UC"] == ambito_tipo
        if adsc_sel3 != "Todas":
            mask_b = mask_b & (dff_uc["Adscripción"] == adsc_sel3)
            mask_c = mask_c & (dff_uc_cucop["Adscripción"] == adsc_sel3)

    dff_g3  = dff_uc[mask_b].copy()
    dff_gc3 = dff_uc_cucop[mask_c].copy()

    if len(dff_g3) == 0:
        st.info("ℹ️ No hay contratos con los filtros actuales.")
    else:
        # ── KPIs ──
        n_adsc_g3 = dff_g3["Adscripción"].nunique()
        n_ucs_g3  = dff_g3["Nombre de la UC"].nunique()
        monto_g3  = dff_g3["Importe DRC"].sum()
        n_cont_g3 = len(dff_g3)

        k1, k2, k3, k4 = st.columns(4)
        if es_nivel_central:
            k1.metric("🏢 Ámbito", "Nivel Central")
        elif ambito_tipo == "OOAD":
            k1.metric("🗺️ Adscripciones (OOAD)", f"{n_adsc_g3:,}")
        else:
            k1.metric("🏥 Adscripciones (UMAE)", f"{n_adsc_g3:,}")
        k2.metric("🏥 Unidades Compradoras",  f"{n_ucs_g3:,}")
        k3.metric("💰 Monto total",
                  f"${monto_g3/1e9:,.2f} miles de millones MXN" if monto_g3 >= 1e9
                  else f"${monto_g3/1e6:,.1f} M MXN")
        k4.metric("📄 Contratos",             f"{n_cont_g3:,}")

        st.divider()

        # ── BLOQUE 1: Comparativa de gasto ──
        color_tipo_uc = {
            "Nivel Central":  IMSS_VERDE_OSC,
            "OOAD":           IMSS_VERDE,
            "UMAE":           IMSS_ORO,
            "Sin clasificar": IMSS_GRIS,
        }

        if agrupar3 == "Adscripción":
            st.subheader("Gasto total por Adscripción")
            top_g = (
                dff_g3.groupby("Adscripción")["Importe DRC"]
                .sum().sort_values(ascending=False).head(top_n_g3).reset_index()
            )
            top_g.columns = ["Grupo", "Monto"]
            top_g["Monto_M"]   = top_g["Monto"] / 1e6
            top_g["Monto_fmt"] = top_g["Monto"].apply(
                lambda x: f"${x/1e9:,.2f} mil M" if x >= 1e9 else f"${x/1e6:,.1f} M"
            )
            _h_g3 = max(340, len(top_g) * 40)
            fig_g3 = px.bar(
                top_g.sort_values("Monto_M"), x="Monto_M", y="Grupo",
                orientation="h", text="Monto_fmt",
                color_discrete_sequence=[IMSS_VERDE_OSC],
                title=f"Top {top_n_g3} adscripciones por monto contratado"
            )
            fig_g3.update_layout(
                font=plotly_font(),
                xaxis_title="Monto (M MXN)",
                yaxis_title="",
                showlegend=False,
                plot_bgcolor="#ffffff",
                paper_bgcolor="#ffffff",
                height=_h_g3,
                margin=dict(l=10, r=130, t=40, b=20),
                yaxis=dict(automargin=True, tickfont=dict(size=12)),
            )
            fig_g3.update_traces(
                marker_color=IMSS_VERDE_OSC,
                textfont=dict(family="Noto Sans, sans-serif"),
                textposition="outside",
                cliponaxis=False,
            )
            st.plotly_chart(fig_g3, use_container_width=True)

        else:  # Unidad Compradora
            st.subheader("Gasto total por Unidad Compradora")
            top_g = (
                dff_g3.groupby(["Nombre de la UC", "Adscripción", "Tipo UC"])["Importe DRC"]
                .sum().sort_values(ascending=False).head(top_n_g3).reset_index()
            )
            top_g.columns = ["UC", "Adscripción", "Tipo UC", "Monto"]
            top_g["Monto_M"]   = top_g["Monto"] / 1e6
            top_g["Monto_fmt"] = top_g["Monto"].apply(
                lambda x: f"${x/1e9:,.2f} mil M" if x >= 1e9 else f"${x/1e6:,.1f} M"
            )
            top_g["UC_corta"]  = top_g["UC"].apply(
                lambda s: s[:50] + "…" if len(s) > 50 else s
            )
            _h_g3_uc = max(340, len(top_g) * 40)
            fig_g3 = px.bar(
                top_g.sort_values("Monto_M"), x="Monto_M", y="UC_corta",
                orientation="h", text="Monto_fmt",
                color="Tipo UC", color_discrete_map=color_tipo_uc,
                title=f"Top {top_n_g3} unidades compradoras por monto contratado",
                custom_data=["UC", "Adscripción"]
            )
            fig_g3.update_layout(
                font=plotly_font(),
                xaxis_title="Monto (M MXN)",
                yaxis_title="",
                plot_bgcolor="#ffffff",
                paper_bgcolor="#ffffff",
                height=_h_g3_uc,
                margin=dict(l=10, r=130, t=50, b=20),
                yaxis=dict(automargin=True, tickfont=dict(size=11)),
                legend=dict(title="Tipo UC", orientation="h",
                            yanchor="bottom", y=1.01, xanchor="left", x=0),
            )
            fig_g3.update_traces(
                textfont=dict(family="Noto Sans, sans-serif"),
                textposition="outside", cliponaxis=False,
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Adscripción: %{customdata[1]}<br>"
                    "Monto: %{text}<extra></extra>"
                )
            )
            st.plotly_chart(fig_g3, use_container_width=True)

        st.divider()

        # ── BLOQUE 2: ¿En qué se gasta? (CUCoP) ──
        st.subheader("¿En qué se gasta? — Partidas Presupuestarias")

        cp1, cp2 = st.columns([2, 1])
        with cp1:
            nivel_cucop3 = st.selectbox(
                "Nivel CUCoP",
                ["Partida específica", "Partida genérica", "Capítulo"],
                key="e3_nivel_cucop"
            )
        with cp2:
            top_n_p3 = st.selectbox("Top N partidas", [10, 20, 50],
                                    key="e3_topn_part")

        if nivel_cucop3 == "Partida específica":
            dff_gc3["_etq"] = (dff_gc3["Partida específica"] + " — " +
                               dff_gc3["DESC. PARTIDA ESPECÍFICA"].fillna("Sin descripción"))
        elif nivel_cucop3 == "Partida genérica":
            dff_gc3["_etq"] = (dff_gc3["Partida específica"].str[:4] + " — " +
                               dff_gc3["DESC. PARTIDA GENÉRICA"].fillna("Sin descripción"))
        else:
            dff_gc3["_etq"] = (dff_gc3["Partida específica"].str[:1] + "000 — " +
                               dff_gc3["DESC. CAPÍTULO"].fillna("Sin descripción"))

        top_part3 = (
            dff_gc3.groupby("_etq")["Importe DRC"]
            .sum().sort_values(ascending=False).head(top_n_p3).reset_index()
        )
        top_part3.columns = ["Partida", "Monto"]
        top_part3["Monto_fmt"] = top_part3["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
        top_part3["Pct"] = (top_part3["Monto"] / top_part3["Monto"].sum() * 100).round(1)
        # Asignar colores cíclicos de la paleta IMSS a cada partida para diferenciación visual
        _paleta_part3 = [IMSS_VERDE, IMSS_ORO, IMSS_ROJO, IMSS_VERDE_OSC,
                         IMSS_GRIS, IMSS_ORO_CLARO, IMSS_NEGRO, "#3A7D6E", "#B8942E", "#C06080"]
        top_part3["Color"] = [
            _paleta_part3[i % len(_paleta_part3)] for i in range(len(top_part3))
        ]
        titulo_part3 = (
            f"Top {top_n_p3} {nivel_cucop3.lower()}s — "
            + (adsc_sel3 if adsc_sel3 != "Todas" else "todas las adscripciones")
        )
        # Treemap — área proporcional al monto, colores distintos por partida
        fig_part3 = go.Figure(go.Treemap(
            labels=top_part3["Partida"],
            values=top_part3["Monto"],
            parents=[""] * len(top_part3),
            marker=dict(colors=top_part3["Color"]),
            texttemplate=(
                "<b>%{label}</b><br>"
                "%{customdata[0]}"
            ),
            customdata=top_part3[["Monto_fmt"]].values,
            hovertemplate=(
                "<b>%{label}</b><br>"
                "Monto: %{customdata[0]}<br>"
                "Participación: %{percentRoot:.1%}<extra></extra>"
            ),
            textfont=dict(family="Noto Sans, sans-serif", size=12),
        ))
        fig_part3.update_layout(
            title=titulo_part3,
            font=plotly_font(),
            margin=dict(t=50, l=10, r=10, b=10),
            height=480,
        )
        st.plotly_chart(fig_part3, use_container_width=True)

        with st.expander("📋 Ver tabla de partidas"):
            tbl_p3 = top_part3[["Partida", "Monto"]].copy()
            tbl_p3["Monto"] = tbl_p3["Monto"].apply(lambda x: f"${x:,.0f}")
            tbl_p3.index = range(1, len(tbl_p3) + 1)
            st.dataframe(tbl_p3, use_container_width=True)

        st.divider()

        # ── BLOQUE HISTÓRICO: Evolución por partida ───────────────────────────
        if len(anios_sel) >= 2:
            st.subheader("📈 Evolución histórica por partida presupuestaria")
            st.caption(
                "Compara el monto contratado entre los años seleccionados. "
                "La variación % se calcula entre el penúltimo y el último año del período."
            )

            # ── Filtro de UC / Adscripción ────────────────────────────────────
            _tiene_adsc_hist = "Adscripción" in dff_uc_cucop.columns
            _fh_uc_sel   = None
            _fh_adsc_sel = None
            _fh_tipo_col, _fh_val_col = st.columns([1, 3])
            with _fh_tipo_col:
                _fh_tipo = st.radio(
                    "Filtrar por",
                    ["Todas las UCs", "Unidad Compradora", "Adscripción"],
                    key="fh_tipo_hist",
                    horizontal=False,
                )
            with _fh_val_col:
                if _fh_tipo == "Unidad Compradora":
                    _opts_uc_h = sorted(
                        dff_uc_cucop["Nombre de la UC"].dropna().unique().tolist()
                    )
                    _fh_uc_sel = st.selectbox(
                        "Unidad Compradora", _opts_uc_h, key="fh_uc_hist"
                    )
                elif _fh_tipo == "Adscripción" and _tiene_adsc_hist:
                    _opts_adsc_h = sorted(
                        dff_uc_cucop["Adscripción"].dropna().unique().tolist()
                    )
                    _fh_adsc_sel = st.selectbox(
                        "Adscripción", _opts_adsc_h, key="fh_adsc_hist"
                    )
                else:
                    st.caption(
                        "Mostrando todas las Unidades Compradoras del ámbito seleccionado."
                    )

            # ── Controles de nivel / top / modo ──────────────────────────────
            _ch1, _ch2, _ch3 = st.columns([2, 1, 1])
            with _ch1:
                _nivel_hist = st.selectbox(
                    "Nivel de agrupación",
                    ["Partida genérica", "Capítulo"],
                    key="nivel_hist_expl",
                )
            with _ch2:
                _top_n_hist = st.selectbox(
                    "Top partidas", [10, 15, 20], key="top_n_hist_expl"
                )
            with _ch3:
                _solo_variacion = st.checkbox(
                    "Solo variación %", value=False, key="solo_var_hist"
                )

            # Etiqueta según nivel
            _df_hist = dff_uc_cucop.copy()
            # Aplicar filtro de UC / Adscripción
            if _fh_tipo == "Unidad Compradora" and _fh_uc_sel:
                _df_hist = _df_hist[_df_hist["Nombre de la UC"] == _fh_uc_sel]
            elif _fh_tipo == "Adscripción" and _fh_adsc_sel and _tiene_adsc_hist:
                _df_hist = _df_hist[_df_hist["Adscripción"] == _fh_adsc_sel]
            if _nivel_hist == "Partida genérica":
                _df_hist["_etiq"] = (
                    _df_hist["PARTIDA GENÉRICA"].fillna("").str.strip()
                    + " — "
                    + _df_hist["DESC. PARTIDA GENÉRICA"].fillna("Sin descripción")
                )
            else:
                _df_hist["_etiq"] = (
                    _df_hist["Partida específica"].str[:1].fillna("") + "000 — "
                    + _df_hist["DESC. CAPÍTULO"].fillna("Sin descripción")
                )

            # Agrupar por año y partida
            _hist_grp = (
                _df_hist.groupby(["Año", "_etiq"])["Importe DRC"]
                .sum().reset_index()
            )

            # Top N partidas por monto total acumulado
            _top_etiq = (
                _hist_grp.groupby("_etiq")["Importe DRC"]
                .sum().sort_values(ascending=False)
                .head(_top_n_hist).index
            )
            _hist_grp = _hist_grp[_hist_grp["_etiq"].isin(_top_etiq)].copy()
            _hist_grp["Etiqueta"] = _hist_grp["_etiq"].apply(
                lambda s: s[:52] + "…" if len(str(s)) > 52 else s
            )
            _hist_grp["Monto_M"] = (_hist_grp["Importe DRC"] / 1e6).round(2)
            _hist_grp["Monto_fmt"] = _hist_grp["Importe DRC"].apply(
                lambda x: f"${x/1e6:,.1f} M"
            )

            _anios_ord = sorted(anios_sel)
            _paleta_hist = [IMSS_VERDE_OSC, IMSS_VERDE, IMSS_ORO_CLARO]

            if not _solo_variacion:
                # ── Barras agrupadas por año ──────────────────────────────────
                fig_hist = px.bar(
                    _hist_grp.sort_values("Importe DRC", ascending=False),
                    x="Etiqueta",
                    y="Monto_M",
                    color="Año",
                    barmode="group",
                    color_discrete_sequence=_paleta_hist,
                    labels={"Monto_M": "Monto (M MXN)", "Etiqueta": ""},
                    custom_data=["_etiq", "Año", "Monto_fmt"],
                )
                fig_hist.update_traces(
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Año: %{customdata[1]}<br>"
                        "Monto: %{customdata[2]}<extra></extra>"
                    )
                )
                fig_hist.update_layout(
                    font=plotly_font(),
                    xaxis_tickangle=-35,
                    plot_bgcolor="#ffffff",
                    paper_bgcolor="#ffffff",
                    height=430,
                    margin=dict(b=130, t=10),
                    legend_title="Año",
                    yaxis_title="Monto (M MXN)",
                    xaxis_title="",
                )
                st.plotly_chart(fig_hist, use_container_width=True)

            # ── Variación % penúltimo → último año ───────────────────────────
            _year_a = _anios_ord[-2]
            _year_b = _anios_ord[-1]

            _pivot_h = _hist_grp.pivot_table(
                index=["_etiq", "Etiqueta"],
                columns="Año",
                values="Importe DRC",
                aggfunc="sum",
            ).fillna(0).reset_index()

            if _year_a in _pivot_h.columns and _year_b in _pivot_h.columns:
                _pivot_h["Variación %"] = (
                    (_pivot_h[_year_b] - _pivot_h[_year_a])
                    / _pivot_h[_year_a].replace(0, pd.NA) * 100
                ).fillna(0).round(1)
                _pivot_h["Monto_a_fmt"] = _pivot_h[_year_a].apply(lambda x: f"${x/1e6:,.1f} M")
                _pivot_h["Monto_b_fmt"] = _pivot_h[_year_b].apply(lambda x: f"${x/1e6:,.1f} M")
                _pivot_h = _pivot_h.sort_values("Variación %")
                _pivot_h["Color"] = _pivot_h["Variación %"].apply(
                    lambda x: IMSS_ROJO if x < 0 else IMSS_VERDE
                )

                st.markdown(f"**Variación % de {_year_a} a {_year_b}**")
                fig_var = px.bar(
                    _pivot_h,
                    x="Variación %",
                    y="Etiqueta",
                    orientation="h",
                    color="Color",
                    color_discrete_map="identity",
                    text=_pivot_h["Variación %"].apply(lambda x: f"{x:+.1f}%"),
                    custom_data=["_etiq", "Monto_a_fmt", "Monto_b_fmt"],
                )
                fig_var.update_layout(
                    font=plotly_font(),
                    xaxis_title=f"Variación % {_year_a} → {_year_b}",
                    yaxis_title="",
                    showlegend=False,
                    plot_bgcolor="#ffffff",
                    paper_bgcolor="#ffffff",
                    height=max(360, len(_pivot_h) * 36),
                    margin=dict(l=10, r=20, t=10, b=40),
                    uniformtext=dict(mode="hide", minsize=9),
                )
                fig_var.update_traces(
                    textposition="inside",
                    insidetextanchor="middle",
                    textfont=dict(
                        family="Noto Sans, sans-serif",
                        size=11,
                        color="white",
                    ),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        f"{_year_a}: %{{customdata[1]}}<br>"
                        f"{_year_b}: %{{customdata[2]}}<br>"
                        "Variación: %{x:+.1f}%<extra></extra>"
                    ),
                )
                # Línea vertical en cero
                fig_var.add_vline(x=0, line_width=1, line_color=IMSS_GRIS)
                st.plotly_chart(fig_var, use_container_width=True)

            st.divider()

        # ── BLOQUE 3: Mapa de calor ──
        # Nivel Central o adscripción específica → filas = UC
        # Todas las delegaciones               → filas = Adscripción
        if es_nivel_central:
            st.subheader("Mapa de calor — UC (Nivel Central) × Partida Genérica (top 10)")
            row_col3 = "Nombre de la UC"
        elif adsc_sel3 == "Todas":
            st.subheader("Mapa de calor — Adscripción × Partida Genérica (top 10)")
            row_col3 = "Adscripción"
        else:
            st.subheader(f"Mapa de calor — UC × Partida Genérica (top 10)  ·  {adsc_sel3}")
            row_col3 = "Nombre de la UC"

        dff_gc3["_gen_etq"] = (
            dff_gc3["Partida específica"].str[:4] + " — " +
            dff_gc3["DESC. PARTIDA GENÉRICA"].fillna("Sin descripción")
        )
        top10_gen3 = (
            dff_gc3.groupby("_gen_etq")["Importe DRC"]
            .sum().sort_values(ascending=False).head(10).index.tolist()
        )
        pivot3 = (
            dff_gc3[dff_gc3["_gen_etq"].isin(top10_gen3)]
            .groupby([row_col3, "_gen_etq"])["Importe DRC"]
            .sum()
            .unstack(fill_value=0)
        )
        pivot3_M = pivot3 / 1e6

        if len(pivot3_M) > 0 and len(pivot3_M.columns) > 0:
            pivot3_M = pivot3_M.loc[
                pivot3_M.sum(axis=1).sort_values(ascending=False).index
            ]
            col_lbl3 = [c.split(" — ")[-1][:35] for c in pivot3_M.columns]
            row_lbl3 = [
                str(r)[:45] + "…" if len(str(r)) > 45 else str(r)
                for r in pivot3_M.index
            ]
            vmax3 = pivot3_M.values.max() if pivot3_M.values.max() > 0 else 1
            umbral3 = vmax3 * 0.02
            text3 = [
                [f"${v:,.0f} M" if v > umbral3 else ("—" if v == 0 else "") for v in row]
                for row in pivot3_M.values
            ]
            fig_heat3 = go.Figure(data=go.Heatmap(
                z=pivot3_M.values,
                x=col_lbl3,
                y=row_lbl3,
                colorscale=[
                    [0,    "#f7f7f7"],
                    [0.01, IMSS_ORO_CLARO],
                    [0.35, IMSS_ORO],
                    [0.7,  IMSS_VERDE],
                    [1.0,  IMSS_VERDE_OSC],
                ],
                text=text3,
                texttemplate="%{text}",
                hovertemplate=(
                    f"{row_col3}: %{{y}}<br>"
                    "Partida: %{x}<br>"
                    "Monto: $%{z:,.1f} M MXN<extra></extra>"
                ),
                showscale=True,
                colorbar=dict(
                    title="M MXN",
                    tickfont=dict(family="Noto Sans, sans-serif", size=11)
                )
            ))
            cell_h3 = max(28, min(50, 600 // max(len(pivot3_M), 1)))
            fig_heat3.update_layout(
                font=plotly_font(),
                xaxis=dict(tickangle=-35, tickfont=dict(size=10), side="bottom"),
                yaxis=dict(tickfont=dict(size=10), autorange="reversed"),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                height=max(420, len(pivot3_M) * cell_h3 + 220),
                margin=dict(l=10, r=10, t=40, b=130)
            )
            st.plotly_chart(fig_heat3, use_container_width=True)
        else:
            st.info("ℹ️ No hay suficientes datos para generar el mapa de calor.")

        st.divider()

        # ── BLOQUE 4: Tabla resumen descargable ──
        st.subheader("📋 Tabla resumen por Unidad Compradora")

        resumen_e3 = (
            dff_g3.groupby(["Adscripción", "Tipo UC", "Nombre de la UC"])
            .agg(Contratos=("Importe DRC", "count"),
                 Monto=("Importe DRC", "sum"))
            .reset_index()
            .sort_values("Monto", ascending=False)
        )
        resumen_e3["Monto"] = resumen_e3["Monto"].apply(lambda x: f"${x:,.0f}")
        resumen_e3.columns   = ["Adscripción", "Tipo UC", "Unidad Compradora",
                                 "Contratos", "Monto (MXN)"]
        resumen_e3 = resumen_e3.reset_index(drop=True)
        resumen_e3.index += 1

        st.dataframe(resumen_e3, use_container_width=True)

        csv_e3 = resumen_e3.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "📥 Descargar tabla resumen (CSV)",
            data=csv_e3,
            file_name=f"resumen_uc_{adsc_sel3.replace(' ', '_')}.csv",
            mime="text/csv",
            key="dl_uc_e3"
        )

    st.divider()

    # ── BLOQUE 5: Explorador por Partida Presupuestaria ──────────────
    st.subheader("📂 Explorador por Partida Presupuestaria")
    st.caption(
        "Selecciona una partida del CUCoP para ver todos los contratos asociados. "
        "Filtra opcionalmente por Adscripción y Unidad Compradora."
    )

    # Catálogo de partidas disponibles en los datos (con descripción)
    _part_cat = (
        dff_uc_cucop[["Partida específica", "DESC. PARTIDA ESPECÍFICA"]]
        .dropna(subset=["DESC. PARTIDA ESPECÍFICA"])
        .drop_duplicates("Partida específica")
        .sort_values("Partida específica")
    )
    _part_cat["_label"] = (
        _part_cat["Partida específica"] + " — " + _part_cat["DESC. PARTIDA ESPECÍFICA"].str.title()
    )
    _part_labels = _part_cat["_label"].tolist()
    _part_code_map = dict(zip(_part_cat["_label"], _part_cat["Partida específica"]))

    if not _part_labels:
        st.info("ℹ️ No hay partidas CUCoP disponibles con los filtros actuales.")
    else:
        # Filtros inline del Bloque 5
        _b5c1, _b5c2, _b5c3 = st.columns([3, 2, 2])
        with _b5c1:
            _part_sel = st.selectbox(
                "📑 Partida presupuestaria",
                options=_part_labels,
                index=0,
                key="e3_b5_partida"
            )
        _part_codigo = _part_code_map.get(_part_sel, "")

        # Adscripción (usa dff_uc completo para no depender del ámbito del radio)
        _adsc_opts_b5 = sorted(dff_uc["Adscripción"].dropna().unique().tolist())
        _adsc_opts_b5 = [a for a in _adsc_opts_b5 if a != "Sin clasificar"]
        with _b5c2:
            _adsc_b5 = st.selectbox(
                "🏛️ Adscripción",
                ["Todas"] + _adsc_opts_b5,
                key="e3_b5_adsc"
            )

        # UC (se actualiza según Adscripción elegida)
        _base_b5 = dff_uc.copy()
        if _adsc_b5 != "Todas":
            _base_b5 = _base_b5[_base_b5["Adscripción"] == _adsc_b5]
        _uc_opts_b5 = sorted(_base_b5["Nombre de la UC"].dropna().unique().tolist())
        with _b5c3:
            _uc_b5 = st.selectbox(
                "🏥 Unidad Compradora",
                ["Todas"] + _uc_opts_b5,
                key="e3_b5_uc"
            )

        # Filtrar contratos con la partida seleccionada
        # La columna "Partida específica" puede contener listas separadas por coma
        _mask_part = _base_b5["Partida específica"].fillna("").apply(
            lambda x: _part_codigo in [p.strip().zfill(5) for p in str(x).split(",")]
        )
        _df_b5 = _base_b5[_mask_part].copy()
        if _uc_b5 != "Todas":
            _df_b5 = _df_b5[_df_b5["Nombre de la UC"] == _uc_b5]

        # KPIs
        _n_b5     = len(_df_b5)
        _monto_b5 = _df_b5["Importe DRC"].sum()
        _uc_b5n   = _df_b5["Nombre de la UC"].nunique()
        _prov_b5  = _df_b5["Proveedor o contratista"].nunique()

        _kb1, _kb2, _kb3, _kb4 = st.columns(4)
        _kb1.metric("📄 Contratos",             f"{_n_b5:,}")
        _kb2.metric("💰 Monto total",
                    f"${_monto_b5/1e9:,.2f} miles de millones MXN" if _monto_b5 >= 1e9
                    else f"${_monto_b5/1e6:,.1f} M MXN")
        _kb3.metric("🏥 Unidades Compradoras",  f"{_uc_b5n:,}")
        _kb4.metric("🏭 Proveedores únicos",    f"{_prov_b5:,}")

        if _n_b5 == 0:
            st.info(f"ℹ️ No se encontraron contratos para la partida **{_part_sel}** con los filtros seleccionados.")
        else:
            # Tabla de contratos
            _cols_b5 = [c for c in [
                "Nombre de la UC", "Adscripción",
                "Proveedor o contratista", "rfc",
                "Tipo Simplificado",
                "Importe DRC",
                "Fecha de inicio del contrato", "Fecha de término del contrato",
                "Descripción del contrato",
                "Dirección del anuncio"
            ] if c in _df_b5.columns]
            _tbl_b5 = (
                _df_b5[_cols_b5]
                .sort_values("Importe DRC", ascending=False)
                .reset_index(drop=True)
            )
            _tbl_b5["Importe DRC"] = _tbl_b5["Importe DRC"].apply(
                lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
            )
            _tbl_b5.index += 1

            st.dataframe(
                _tbl_b5,
                column_config={
                    "Dirección del anuncio": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    ),
                    "Importe DRC": st.column_config.TextColumn("Importe DRC"),
                },
                use_container_width=True,
                height=500
            )

            # Descarga CSV (con montos numéricos originales)
            _tbl_b5_csv = (
                _df_b5[_cols_b5]
                .sort_values("Importe DRC", ascending=False)
                .reset_index(drop=True)
            )
            _nombre_archivo = (
                f"partida_{_part_codigo}_"
                + (_adsc_b5.replace(" ", "_")[:20] if _adsc_b5 != "Todas" else "todas_adsc")
                + ".csv"
            )
            st.download_button(
                "📥 Descargar contratos (CSV)",
                data=_tbl_b5_csv.to_csv(index=False).encode("utf-8-sig"),
                file_name=_nombre_archivo,
                mime="text/csv",
                key="dl_b5_partida"
            )

    st.divider()

    # ── BLOQUE 6: Comparativa de Gasto por OOAD ─────────────────────
    st.subheader("🗺️ Gasto por OOAD — Composición por Tipo de Procedimiento")
    st.caption(
        "Comparativa del monto contratado en los 35 OOAD, desglosado por tipo de procedimiento. "
        "Refleja los filtros activos en el sidebar (año, tipo de procedimiento, etc.)."
    )

    _dff_ooad6 = dff_uc[dff_uc["Tipo UC"] == "OOAD"].copy()

    if len(_dff_ooad6) == 0:
        st.info("ℹ️ No hay datos de OOAD con los filtros actuales.")
    else:
        # KPIs
        _n_ooad6   = _dff_ooad6["Adscripción"].nunique()
        _monto6    = _dff_ooad6["Importe DRC"].sum()
        _n_cont6   = len(_dff_ooad6)
        _pct_ad6   = (
            _dff_ooad6[_dff_ooad6["Tipo Simplificado"].str.contains("Adjudicación", na=False)]["Importe DRC"].sum()
            / _monto6 * 100 if _monto6 > 0 else 0
        )

        _o1, _o2, _o3, _o4 = st.columns(4)
        _o1.metric("🗺️ OOAD con contratos",   f"{_n_ooad6}")
        _o2.metric("💰 Monto total OOAD",
                   f"${_monto6/1e9:,.2f} miles de millones MXN" if _monto6 >= 1e9
                   else f"${_monto6/1e6:,.1f} M MXN")
        _o3.metric("📄 Contratos",             f"{_n_cont6:,}")
        _o4.metric("🔴 % Adjudicación Directa", f"{_pct_ad6:.1f}%")

        # Agrupar por Adscripción × Tipo Simplificado
        _ooad_grp6 = (
            _dff_ooad6.groupby(["Adscripción", "Tipo Simplificado"])["Importe DRC"]
            .sum().reset_index()
        )
        _ooad_grp6["Monto_M"] = _ooad_grp6["Importe DRC"] / 1e6

        # Renombrar "Adjudicación Directa — Fr. I" → "Adjudicación Directa — Patentes"
        _ooad_grp6["Tipo Simplificado"] = _ooad_grp6["Tipo Simplificado"].replace(
            "Adjudicación Directa — Fr. I", "Adjudicación Directa — Patentes"
        )
        _colores_ooad6 = {**COLORES_TIPO,
                          "Adjudicación Directa — Patentes": COLORES_TIPO.get("Adjudicación Directa — Fr. I", "#C05078")}

        # Texto dentro de las barras: mostrar monto solo si el segmento es ≥ 50 M
        _ooad_grp6["_texto"] = _ooad_grp6["Monto_M"].apply(
            lambda x: f"${x:,.0f} M" if x >= 50 else ""
        )

        # Orden: mayor a menor total (ascending en la lista → mayor queda arriba en plotly)
        _ooad_orden6 = (
            _ooad_grp6.groupby("Adscripción")["Monto_M"]
            .sum().sort_values(ascending=True).index.tolist()
        )

        fig_ooad6 = px.bar(
            _ooad_grp6,
            x="Monto_M", y="Adscripción",
            color="Tipo Simplificado",
            color_discrete_map=_colores_ooad6,
            orientation="h",
            barmode="stack",
            text="_texto",
            labels={"Monto_M": "Monto (M MXN)", "Adscripción": ""},
            custom_data=["Adscripción", "Tipo Simplificado", "Importe DRC"]
        )
        fig_ooad6.update_layout(
            font=plotly_font(),
            xaxis_title="Monto (millones MXN)",
            yaxis=dict(categoryorder="total ascending"),
            yaxis_title="",
            plot_bgcolor="#ffffff",
            paper_bgcolor="#ffffff",
            height=max(500, _n_ooad6 * 26),
            legend=dict(
                title="Tipo de procedimiento",
                orientation="h", yanchor="top", y=-0.08, xanchor="left", x=0
            ),
            margin=dict(l=220, r=80, t=20, b=100)
        )
        fig_ooad6.update_traces(
            textfont=dict(family="Noto Sans, sans-serif", size=11),
            textposition="inside",
            insidetextanchor="middle",
            cliponaxis=False,
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "%{customdata[1]}<br>"
                "Monto: $%{customdata[2]:,.0f}<extra></extra>"
            )
        )
        st.plotly_chart(fig_ooad6, use_container_width=True)

        # Tabla resumen con % por tipo
        with st.expander("📋 Ver tabla por OOAD"):
            _tbl6 = (
                _ooad_grp6.pivot_table(
                    index="Adscripción",
                    columns="Tipo Simplificado",
                    values="Monto_M",
                    aggfunc="sum",
                    fill_value=0
                ).reset_index()
            )
            _tbl6["Total (M MXN)"] = _tbl6.drop(columns="Adscripción").sum(axis=1)
            _tbl6 = _tbl6.sort_values("Total (M MXN)", ascending=False).reset_index(drop=True)
            _tbl6.index += 1

            # Renombrar Fr. I → Patentes también en la tabla
            _tbl6.columns = [
                c.replace("Adjudicación Directa — Fr. I", "Adjudicación Directa — Patentes")
                for c in _tbl6.columns
            ]

            # % AD
            _ad_cols6 = [c for c in _tbl6.columns if "Adjudicación" in c]
            if _ad_cols6:
                _tbl6["% AD"] = (
                    _tbl6[_ad_cols6].sum(axis=1) / _tbl6["Total (M MXN)"] * 100
                ).round(1).astype(str) + "%"

            # Formatear montos
            for _c in _tbl6.columns:
                if _c not in ("Adscripción", "% AD"):
                    _tbl6[_c] = _tbl6[_c].apply(lambda x: f"${x:,.1f} M")

            st.dataframe(_tbl6, use_container_width=True)

            st.download_button(
                "📥 Descargar tabla OOAD (CSV)",
                data=_tbl6.to_csv(index=False).encode("utf-8-sig"),
                file_name="gasto_por_ooad.csv",
                mime="text/csv",
                key="dl_b6_ooad"
            )


# ───────────────────────────────────────────────────────────────
# PÁGINA 4: PAGINA_HISTORICA
# ───────────────────────────────────────────────────────────────
def pagina_historica():

    st.header("📈 Evolución Histórica del Gasto por UC (2023–2026)")
    st.caption(
        "Aplica el mismo filtro de institución del sidebar. "
        "Los demás filtros del sidebar no se aplican para garantizar comparabilidad entre años."
    )

    # ── Catálogo completo de años disponibles ──
    _HIST_TODOS = {
        "2023": "contratos_compranet_2023.csv",
        "2024": "contratos_compranet_2024.csv",
        "2025": "contratos_comprasmx_2025.csv",
        "2026": "contratos_comprasmx_2026.csv",
    }

    _anios_hist_sel = st.multiselect(
        "Años a comparar",
        options=list(_HIST_TODOS.keys()),
        default=["2023", "2024", "2025", "2026"],
        key="hist_anios_sel",
    )

    if len(_anios_hist_sel) < 2:
        st.info("ℹ️ Selecciona al menos 2 años para comparar.")
        return

    _HIST = {a: _HIST_TODOS[a] for a in _anios_hist_sel}

    # ── Cargar años seleccionados (reutiliza caché de cargar_datos) ──
    _dfs_h = {}
    for _yr, _fn in _HIST.items():
        try:
            _dfa = cargar_datos(_fn)
            _dfa = _dfa.copy()
            _dfa["Año"] = _yr
            _dfs_h[_yr] = _dfa
        except Exception:
            pass

    _anios_h = sorted(_dfs_h.keys())

    if len(_anios_h) < 2:
        st.info("ℹ️ Se necesitan al menos 2 años de datos para comparar. "
                "Verifica que los archivos CSV estén en la carpeta del dashboard.")
    else:
        _df_h = pd.concat([_dfs_h[a] for a in _anios_h], ignore_index=True)

        # Aplicar nombres editados de UC (Base_UC_2025_V2.xlsx) al histórico multi-año
        if len(df_dir_uc) > 0 and "Clave_UC" in df_dir_uc.columns:
            _uc_map_h = (
                df_dir_uc.dropna(subset=["Nombre_editado"])
                .set_index("Clave_UC")["Nombre_editado"]
                .to_dict()
            )
            _df_h["Nombre de la UC"] = (
                _df_h["Clave de la UC"].map(_uc_map_h).fillna(_df_h["Nombre de la UC"])
            )

        # Aplicar filtro de institución del sidebar
        if inst_sel != "Todas":
            _df_h = _df_h[_df_h["Institución"] == inst_sel]

        # ── Pre-calcular columnas indicadoras ──
        _df_h["_imp_lp"] = _df_h["Importe DRC"].where(
            _df_h["Tipo Simplificado"] == "Licitación Pública", 0
        )
        _df_h["_imp_ad_ot"] = _df_h["Importe DRC"].where(
            _df_h["Tipo Simplificado"] == "Adjudicación Directa", 0
        )
        _df_h["_imp_cf"] = _df_h["Importe DRC"].where(
            _df_h["Descripción excepción"].str.upper().str.contains("CASO FORTUITO", na=False),
            0
        )
        _df_h["_imp_f1"] = _df_h["Importe DRC"].where(
            _df_h["Tipo Simplificado"] == "Adjudicación Directa — Fr. I", 0
        )

        # ── Métricas por año (global) ──
        _global_h = (
            _df_h.groupby("Año").agg(
                Contratos   =("Importe DRC",  "count"),
                Monto_total =("Importe DRC",  "sum"),
                Monto_LP    =("_imp_lp",      "sum"),
                Monto_AD_ot =("_imp_ad_ot",   "sum"),
                Monto_CF    =("_imp_cf",      "sum"),
                Monto_F1    =("_imp_f1",      "sum"),
            ).reset_index()
        )
        _global_h["Pct_LP"] = (
            _global_h["Monto_LP"] / _global_h["Monto_total"].replace(0, pd.NA) * 100
        ).fillna(0)
        _global_h["Pct_AD"] = (
            _global_h["Monto_AD_ot"] / _global_h["Monto_total"].replace(0, pd.NA) * 100
        ).fillna(0)
        _global_h["Pct_CF"] = (
            _global_h["Monto_CF"] / _global_h["Monto_total"].replace(0, pd.NA) * 100
        ).fillna(0)

        # ── Métricas por UC y año ──
        _uc_h = (
            _df_h.groupby(["Nombre de la UC", "Año"]).agg(
                Contratos   =("Importe DRC",  "count"),
                Monto_total =("Importe DRC",  "sum"),
                Monto_LP    =("_imp_lp",      "sum"),
                Monto_AD_ot =("_imp_ad_ot",   "sum"),
                Monto_CF    =("_imp_cf",      "sum"),
            ).reset_index()
        )
        _uc_h["Pct_LP"] = (
            _uc_h["Monto_LP"] / _uc_h["Monto_total"].replace(0, pd.NA) * 100
        ).fillna(0)
        _uc_h["Pct_AD"] = (
            _uc_h["Monto_AD_ot"] / _uc_h["Monto_total"].replace(0, pd.NA) * 100
        ).fillna(0)
        _uc_h["Pct_CF"] = (
            _uc_h["Monto_CF"] / _uc_h["Monto_total"].replace(0, pd.NA) * 100
        ).fillna(0)
        _uc_h["Monto_total_M"] = _uc_h["Monto_total"] / 1e6

        _yr_first_h = _anios_h[0]
        _yr_last_h  = _anios_h[-1]

        # ══════════════════════════════════════════════
        # SECCIÓN A: Panorama global comparativo
        # ══════════════════════════════════════════════
        st.subheader("Panorama global por año")

        # Métricas compactas de volumen por año
        _gcols = st.columns(len(_anios_h))
        for _i, _yr in enumerate(_anios_h):
            _r = _global_h[_global_h["Año"] == _yr]
            if len(_r) == 0:
                continue
            _r = _r.iloc[0]
            _prev = _global_h[_global_h["Año"] == str(int(_yr) - 1)]
            with _gcols[_i]:
                st.markdown(f"#### {_yr}")
                _delta_m = None
                if len(_prev) > 0:
                    _delta_m = f"{(_r['Monto_total'] / _prev.iloc[0]['Monto_total'] - 1) * 100:+.1f}% vs {str(int(_yr)-1)}"
                st.metric(
                    "💰 Monto total",
                    f"${_r['Monto_total']/1e9:,.2f} mm MXN" if _r["Monto_total"] >= 1e9
                    else f"${_r['Monto_total']/1e6:,.1f} M MXN",
                    delta=_delta_m, delta_color="off"
                )
                st.metric("📄 Contratos", f"{int(_r['Contratos']):,}")

        # Gráfica de líneas — evolución de indicadores clave (%)
        _global_pct_h = _global_h[["Año", "Pct_LP", "Pct_AD", "Pct_CF"]].melt(
            id_vars="Año",
            value_vars=["Pct_LP", "Pct_AD", "Pct_CF"],
            var_name="Indicador", value_name="Porcentaje"
        )
        _global_pct_h["Indicador"] = _global_pct_h["Indicador"].map({
            "Pct_LP": "% Licitación Pública",
            "Pct_AD": "% AD otras causales",
            "Pct_CF": "% Caso fortuito",
        })
        _colores_pan_h = {
            "% Licitación Pública": IMSS_VERDE,
            "% AD otras causales":  IMSS_ROJO,
            "% Caso fortuito":      IMSS_ORO,
        }
        fig_panorama_h = px.line(
            _global_pct_h, x="Año", y="Porcentaje", color="Indicador",
            markers=True,
            color_discrete_map=_colores_pan_h,
            labels={"Porcentaje": "% del monto contratado"},
        )
        fig_panorama_h.update_traces(line=dict(width=3), marker=dict(size=10))
        fig_panorama_h.update_layout(
            font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            xaxis=dict(type="category"),
            yaxis=dict(range=[0, 105], title="% del monto contratado"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
            margin=dict(t=60, b=40),
        )
        st.plotly_chart(fig_panorama_h, use_container_width=True)

        st.divider()

        # ══════════════════════════════════════════════
        # SECCIÓN B: Evolución del gasto por tipo
        # ══════════════════════════════════════════════
        st.subheader("Evolución del gasto y composición por tipo de procedimiento")

        _tipos_h = (
            _df_h.groupby(["Año", "Tipo Simplificado"])["Importe DRC"]
            .sum().reset_index()
        )
        _orden_tipos_h = [
            "Licitación Pública", "Invitación a 3 personas", "Entre Entes Públicos",
            "Adjudicación Directa — Fr. I", "Adjudicación Directa", "Sin clasificar"
        ]
        _tipos_h["Monto_M"] = _tipos_h["Importe DRC"] / 1e6
        fig_evo_t = px.bar(
            _tipos_h, x="Año", y="Monto_M", color="Tipo Simplificado",
            color_discrete_map=COLORES_TIPO, barmode="stack",
            category_orders={"Tipo Simplificado": _orden_tipos_h},
            title="Monto contratado por tipo de procedimiento y año",
            labels={"Monto_M": "Monto (M MXN)", "Tipo Simplificado": "Tipo"}
        )
        fig_evo_t.update_layout(
            font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            xaxis=dict(type="category"),
            yaxis_title="Monto (M MXN)",
            legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0)
        )
        fig_evo_t.update_traces(
            hovertemplate="<b>%{fullData.name}</b><br>Año: %{x}<br>Monto: $%{y:,.1f} M MXN<extra></extra>"
        )
        st.plotly_chart(fig_evo_t, use_container_width=True)

        st.divider()

        # ══════════════════════════════════════════════
        # SECCIÓN C: Evolución por UC seleccionada
        # ══════════════════════════════════════════════
        st.subheader("Evolución por Unidad Compradora")

        _ucs_h = sorted(_uc_h["Nombre de la UC"].unique().tolist())
        _ch1, _ch2 = st.columns([3, 1])
        with _ch1:
            _ucs_sel_h = st.multiselect(
                "Seleccionar UCs (máx. 8)",
                _ucs_h, max_selections=8, key="h4_ucs"
            )
        with _ch2:
            _met_sel_h = st.selectbox(
                "Métrica",
                ["Monto total", "% Licitación Pública",
                 "% AD otras causales", "% Caso fortuito", "N° contratos"],
                key="h4_met"
            )

        _met_col_h = {
            "Monto total":          "Monto_total_M",
            "% Licitación Pública": "Pct_LP",
            "% AD otras causales":  "Pct_AD",
            "% Caso fortuito":      "Pct_CF",
            "N° contratos":         "Contratos",
        }[_met_sel_h]
        _met_label_h = "Monto (M MXN)" if _met_sel_h == "Monto total" else _met_sel_h

        if _ucs_sel_h:
            _hist_sel_h = _uc_h[_uc_h["Nombre de la UC"].isin(_ucs_sel_h)]
            _pal_h = [IMSS_VERDE_OSC, IMSS_ROJO, IMSS_ORO, IMSS_VERDE,
                      "#4A90D9", "#8B5CF6", "#F59E0B", "#10B981"]
            _cmap_h = {uc: _pal_h[i % len(_pal_h)] for i, uc in enumerate(_ucs_sel_h)}

            fig_uc_h = px.line(
                _hist_sel_h, x="Año", y=_met_col_h,
                color="Nombre de la UC", markers=True,
                title=f"Evolución de '{_met_sel_h}' — UCs seleccionadas",
                color_discrete_map=_cmap_h,
                labels={_met_col_h: _met_label_h, "Nombre de la UC": "UC"}
            )
            if _met_sel_h == "% Licitación Pública":
                fig_uc_h.add_hline(
                    y=65, line_dash="dash", line_color=IMSS_ORO,
                    annotation_text="Referencia 65% LP", annotation_position="top right",
                    annotation_font=dict(family="Noto Sans, sans-serif",
                                         color=IMSS_ORO, size=11)
                )
            _es_pct_h = _met_sel_h.startswith("%")
            fig_uc_h.update_layout(
                font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis=dict(type="category"),
                yaxis=dict(range=[0, 105] if _es_pct_h else None,
                           title=_met_label_h),
                legend=dict(orientation="h", yanchor="top", y=-0.28,
                            xanchor="left", x=0)
            )
            fig_uc_h.update_traces(line=dict(width=2.5), marker=dict(size=9))
            st.plotly_chart(fig_uc_h, use_container_width=True)
        else:
            st.info("ℹ️ Selecciona una o más Unidades Compradoras para ver su evolución.")

        st.divider()

        # ══════════════════════════════════════════════
        # SECCIÓN D: Tabla comparativa pivot por UC
        # ══════════════════════════════════════════════
        st.subheader("📋 Tabla comparativa por UC")

        _met_tbl_h = st.selectbox(
            "Métrica a comparar",
            ["Monto total", "% Licitación Pública",
             "% AD otras causales", "% Caso fortuito", "N° contratos"],
            key="h4_met_tbl"
        )
        _col_tbl_h = {
            "Monto total":          "Monto_total",
            "% Licitación Pública": "Pct_LP",
            "% AD otras causales":  "Pct_AD",
            "% Caso fortuito":      "Pct_CF",
            "N° contratos":         "Contratos",
        }[_met_tbl_h]

        # Pivot numérico — ordenar antes de formatear
        _piv_h = _uc_h.pivot_table(
            index="Nombre de la UC", columns="Año",
            values=_col_tbl_h, aggfunc="sum"
        ).reset_index()

        if _yr_last_h in _piv_h.columns:
            _piv_h = _piv_h.sort_values(_yr_last_h, ascending=False)

        # Columna de variación (numérica)
        if _yr_first_h in _piv_h.columns and _yr_last_h in _piv_h.columns:
            if _met_tbl_h == "Monto total":
                _piv_h["↕ Var. %"] = (
                    (_piv_h[_yr_last_h] - _piv_h[_yr_first_h]) /
                    _piv_h[_yr_first_h].replace(0, pd.NA) * 100
                )
            else:
                _piv_h["↕ Var. pp"] = _piv_h[_yr_last_h] - _piv_h[_yr_first_h]

        # Formatear columnas de año
        _piv_fmt_h = _piv_h.copy()
        for _c in _anios_h:
            if _c not in _piv_fmt_h.columns:
                continue
            if _met_tbl_h == "Monto total":
                _piv_fmt_h[_c] = _piv_fmt_h[_c].apply(
                    lambda x: f"${x/1e6:,.1f} M" if pd.notna(x) else "—"
                )
            elif _met_tbl_h == "N° contratos":
                _piv_fmt_h[_c] = _piv_fmt_h[_c].apply(
                    lambda x: f"{int(x):,}" if pd.notna(x) else "—"
                )
            else:
                _piv_fmt_h[_c] = _piv_fmt_h[_c].apply(
                    lambda x: f"{x:.1f}%" if pd.notna(x) else "—"
                )

        if "↕ Var. %" in _piv_fmt_h.columns:
            _piv_fmt_h["↕ Var. %"] = _piv_fmt_h["↕ Var. %"].apply(
                lambda x: f"{x:+.1f}%" if pd.notna(x) else "—"
            )
        if "↕ Var. pp" in _piv_fmt_h.columns:
            _piv_fmt_h["↕ Var. pp"] = _piv_fmt_h["↕ Var. pp"].apply(
                lambda x: f"{x:+.1f} pp" if pd.notna(x) else "—"
            )

        _piv_fmt_h = _piv_fmt_h.rename(columns={"Nombre de la UC": "Unidad Compradora"})
        _piv_fmt_h = _piv_fmt_h.reset_index(drop=True)
        _piv_fmt_h.index += 1

        st.dataframe(_piv_fmt_h, use_container_width=True)

        _csv_h = _piv_fmt_h.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "📥 Descargar tabla comparativa (CSV)",
            data=_csv_h,
            file_name=f"evolucion_uc_{_yr_first_h}_{_yr_last_h}.csv",
            mime="text/csv",
            key="dl_hist4"
        )



# ───────────────────────────────────────────────────────────────
# PÁGINA: PAGINA_FRAGMENTACION
# ───────────────────────────────────────────────────────────────
def pagina_fragmentacion():
    import numpy as np

    st.header("🧩 Análisis de Fragmentación de Contratos")
    st.caption(
        "Identifica patrones en adjudicaciones directas que sugieren la división artificial "
        "de una necesidad en múltiples contratos menores para evadir umbrales de licitación "
        "(Art. 42 LAASSP)."
    )

    # ── Filtro de Unidad Compradora ──
    _ucs_frag = ["Todas"] + sorted(dff["Nombre de la UC"].dropna().unique().tolist())
    _uc_sel_frag = st.selectbox(
        "🏢 Filtrar por Unidad Compradora",
        _ucs_frag,
        key="uc_frag",
        help="Concentra el análisis de fragmentación en una Unidad Compradora específica."
    )
    _dff_base_f = (
        dff[dff["Nombre de la UC"] == _uc_sel_frag].copy()
        if _uc_sel_frag != "Todas" else dff.copy()
    )

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 0 — Contratos justo por debajo del umbral legal 🚦
    # ═══════════════════════════════════════════════════════════════
    st.subheader("0️⃣ Contratos justo por debajo del umbral legal 🚦")
    st.caption(
        "Contratos cuyo importe representa entre el **90 % y 100 %** del umbral legal aplicable "
        "(Art. 55 LAASSP / Art. 43 LOPSRM). Cuando el monto se mantiene deliberadamente por debajo "
        "del tope que exigiría licitación pública o invitación a tres personas, puede indicar "
        "fraccionamiento para evadir la competencia abierta."
    )

    _umbrales_pef0 = cargar_umbrales_pef()

    if not _umbrales_pef0:
        st.info("ℹ️ Archivo UmbralPEF.xlsx no disponible. No es posible calcular este indicador.")
    else:
        _TIPOS_AD_U0  = {"Adjudicación Directa", "Adjudicación Directa — Fr. I"}
        _TIPOS_I3P_U0 = {"Invitación a 3 personas"}
        _tipos_u0_all = list(_TIPOS_AD_U0 | _TIPOS_I3P_U0)

        _dff_u0 = _dff_base_f[
            _dff_base_f["Tipo Simplificado"].isin(_tipos_u0_all)
        ].copy()

        # Año del contrato
        _dff_u0["_fecha_u0"] = pd.to_datetime(
            _dff_u0["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
        )
        _dff_u0["_año_u0"] = _dff_u0["_fecha_u0"].dt.year

        # Ley y Tipo de contratación normalizados
        _dff_u0["_ley_u0"] = (
            _dff_u0["Ley"].astype(str).str.strip().str.upper()
            if "Ley" in _dff_u0.columns else "LAASSP"
        )
        _dff_u0["_contratacion_u0"] = (
            _dff_u0["Tipo de contratación"].astype(str).str.strip().str.upper()
            if "Tipo de contratación" in _dff_u0.columns else "ADQUISICIONES"
        )
        _dff_u0["_proc_u0"] = _dff_u0["Tipo Simplificado"]

        # Asignar umbral por fila (vectorizado por año)
        _dff_u0["_umbral_u0"] = float("nan")
        _TIPOS_SERV_U0 = {"SERVICIOS", "SERVICIOS RELACIONADOS CON LA OBRA", "ARRENDAMIENTOS"}
        for _año_u0, _th in _umbrales_pef0.items():
            _ma = _dff_u0["_año_u0"] == _año_u0
            # AD + LAASSP
            _dff_u0.loc[
                _ma & _dff_u0["_ley_u0"].eq("LAASSP") & _dff_u0["_proc_u0"].isin(_TIPOS_AD_U0),
                "_umbral_u0"
            ] = _th["ad_laassp"]
            # I3P + LAASSP
            _dff_u0.loc[
                _ma & _dff_u0["_ley_u0"].eq("LAASSP") & _dff_u0["_proc_u0"].isin(_TIPOS_I3P_U0),
                "_umbral_u0"
            ] = _th["i3p_laassp"]
            # AD + LOPSRM + OBRA PÚBLICA
            _dff_u0.loc[
                _ma & _dff_u0["_ley_u0"].eq("LOPSRM") & _dff_u0["_proc_u0"].isin(_TIPOS_AD_U0)
                & _dff_u0["_contratacion_u0"].eq("OBRA PÚBLICA"),
                "_umbral_u0"
            ] = _th["ad_obra_lopsrm"]
            # AD + LOPSRM + SERVICIOS
            _dff_u0.loc[
                _ma & _dff_u0["_ley_u0"].eq("LOPSRM") & _dff_u0["_proc_u0"].isin(_TIPOS_AD_U0)
                & _dff_u0["_contratacion_u0"].isin(_TIPOS_SERV_U0),
                "_umbral_u0"
            ] = _th["ad_serv_lopsrm"]
            # I3P + LOPSRM + OBRA PÚBLICA
            _dff_u0.loc[
                _ma & _dff_u0["_ley_u0"].eq("LOPSRM") & _dff_u0["_proc_u0"].isin(_TIPOS_I3P_U0)
                & _dff_u0["_contratacion_u0"].eq("OBRA PÚBLICA"),
                "_umbral_u0"
            ] = _th["i3p_obra_lopsrm"]
            # I3P + LOPSRM + SERVICIOS
            _dff_u0.loc[
                _ma & _dff_u0["_ley_u0"].eq("LOPSRM") & _dff_u0["_proc_u0"].isin(_TIPOS_I3P_U0)
                & _dff_u0["_contratacion_u0"].isin(_TIPOS_SERV_U0),
                "_umbral_u0"
            ] = _th["i3p_serv_lopsrm"]

        # Porcentaje respecto al umbral
        _dff_u0["_pct_u0"] = (_dff_u0["Importe DRC"] / _dff_u0["_umbral_u0"]) * 100

        # Zona sospechosa: 90 % ≤ pct < 100 %
        _mask_sosp_u0 = (
            _dff_u0["_pct_u0"] >= 90
        ) & (
            _dff_u0["_pct_u0"] < 100
        ) & _dff_u0["_umbral_u0"].notna()
        _dff_sosp0 = _dff_u0[_mask_sosp_u0].copy().sort_values("_pct_u0", ascending=False)

        # KPIs
        _k0a, _k0b, _k0c, _k0d = st.columns(4)
        _k0a.metric("🚦 Contratos en zona sospechosa", f"{len(_dff_sosp0):,}")
        _k0b.metric(
            "🏢 Proveedores únicos",
            f"{_dff_sosp0['rfc'].nunique():,}" if len(_dff_sosp0) > 0 else "0"
        )
        _k0c.metric(
            "💰 Monto total",
            f"${_dff_sosp0['Importe DRC'].sum() / 1e6:,.1f} M MXN"
        )
        _n_u0_con_umbral = int(_dff_u0["_umbral_u0"].notna().sum())
        _k0d.metric(
            "📊 % del universo AD/I3P",
            f"{len(_dff_sosp0) / _n_u0_con_umbral * 100:.1f}%"
            if _n_u0_con_umbral > 0 else "—"
        )

        if len(_dff_sosp0) == 0:
            st.success("✅ No se detectaron contratos en zona sospechosa de umbral legal.")
        else:
            st.warning(
                f"⚠️ **{len(_dff_sosp0):,}** contratos con importe entre el 90 % y 100 % "
                f"del umbral legal aplicable."
            )

            # Gráfica: Top 20 UCs con más contratos sospechosos
            _uc_u0 = (
                _dff_sosp0.groupby("Nombre de la UC")
                .agg(Contratos=("Importe DRC", "count"), Monto=("Importe DRC", "sum"))
                .reset_index()
                .sort_values("Contratos", ascending=False)
                .head(20)
            )
            # Truncar nombres de UC para dar más espacio visual a las barras
            _uc_u0["UC_corta"] = _uc_u0["Nombre de la UC"].apply(
                lambda s: s[:38] + "…" if len(str(s)) > 38 else s
            )
            _fig_u0 = go.Figure(go.Bar(
                x=_uc_u0["Contratos"],
                y=_uc_u0["UC_corta"],
                orientation="h",
                marker_color=IMSS_ROJO,
                text=_uc_u0["Contratos"].apply(lambda v: f"{v:,}"),
                textposition="outside",
                customdata=_uc_u0["Nombre de la UC"],
                hovertemplate=(
                    "<b>%{customdata}</b><br>"
                    "Contratos en zona sospechosa: <b>%{x:,}</b><br>"
                    "<extra></extra>"
                ),
            ))
            _fig_u0.update_layout(
                title="Top 20 UCs — contratos con importe en zona 90–100 % del umbral",
                xaxis_title="Número de contratos",
                yaxis=dict(autorange="reversed"),
                height=max(400, len(_uc_u0) * 34 + 120),
                margin=dict(l=10, r=80, t=60, b=40),
                font=plotly_font(),
            )
            st.plotly_chart(_fig_u0, use_container_width=True)

            # Tabla detallada
            with st.expander("📋 Ver contratos en zona sospechosa de umbral legal"):
                _det_u0 = _dff_sosp0[
                    [c for c in [
                        "Fecha de inicio del contrato",
                        "Nombre de la UC",
                        "Proveedor o contratista",
                        "Tipo Simplificado",
                        "Ley",
                        "Tipo de contratación",
                        "Importe DRC",
                        "_umbral_u0",
                        "_pct_u0",
                        "Dirección del anuncio",
                    ] if c in _dff_sosp0.columns]
                ].copy()
                _det_u0["Importe"] = _det_u0["Importe DRC"].apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _det_u0["Umbral legal"] = _det_u0["_umbral_u0"].apply(
                    lambda x: f"${x:,.0f}" if pd.notna(x) else ""
                )
                _det_u0["% del umbral"] = _det_u0["_pct_u0"].apply(
                    lambda x: f"{x:.1f}%" if pd.notna(x) else ""
                )
                _det_u0 = _det_u0.drop(
                    columns=["Importe DRC", "_umbral_u0", "_pct_u0"], errors="ignore"
                ).rename(columns={"Tipo Simplificado": "Tipo"})
                _det_u0.index = range(1, len(_det_u0) + 1)
                st.dataframe(
                    _det_u0,
                    use_container_width=True,
                    column_config={
                        "Dirección del anuncio": st.column_config.LinkColumn(
                            "🔗 ComprasMX", display_text="Ver contrato"
                        )
                    },
                )
                _csv_u0 = _det_u0.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "📥 Descargar tabla",
                    data=_csv_u0,
                    file_name="contratos_zona_umbral.csv",
                    mime="text/csv",
                    key="dl_umbral_u0",
                )

    st.divider()

    # Solo adjudicaciones directas (ambos subtipos)
    _dff_ad = _dff_base_f[
        _dff_base_f["Tipo Simplificado"].isin(
            ["Adjudicación Directa", "Adjudicación Directa — Fr. I"]
        )
    ].copy()

    if len(_dff_ad) == 0:
        st.info("ℹ️ No hay adjudicaciones directas con los filtros actuales.")
        return

    # Excluir contratos que comparten número de procedimiento:
    # cuando 2+ contratos AD tienen el mismo número, derivan de distintas partidas
    # del mismo proceso de contratación y no constituyen fragmentación.
    _col_proc = "Número de procedimiento"
    _n_excl_proc = 0
    if _col_proc in _dff_ad.columns:
        _proc_vals   = _dff_ad[_col_proc].astype(str).str.strip()
        _proc_counts = _proc_vals.value_counts()
        _procs_rep   = set(_proc_counts[_proc_counts >= 2].index) - {"", "nan", "NaN"}
        _mask_excl   = _proc_vals.isin(_procs_rep)
        _n_excl_proc = int(_mask_excl.sum())
        if _n_excl_proc > 0:
            _dff_ad = _dff_ad[~_mask_excl].copy()

    # Parsear fechas (Fecha de inicio del contrato DD/MM/YYYY)
    _dff_ad["_fecha"] = pd.to_datetime(
        _dff_ad["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
    )
    _dff_ad_dt = _dff_ad[_dff_ad["_fecha"].notna()].copy().reset_index(drop=True)

    # Columnas auxiliares calculadas una sola vez
    if len(_dff_ad_dt) > 0:
        _dff_ad_dt["_fecha_str"] = _dff_ad_dt["_fecha"].dt.date.astype(str)
        _dff_ad_dt["_bucket3"]   = _dff_ad_dt["_fecha"].apply(lambda d: d.toordinal() // 3)
        _iso_cal = _dff_ad_dt["_fecha"].dt.isocalendar()
        _dff_ad_dt["_semana"] = (
            _iso_cal["year"].astype(str) + "-W" +
            _iso_cal["week"].astype(int).astype(str).str.zfill(2)
        )

    _caption_universo = (
        f"Universo de análisis: **{len(_dff_ad):,} adjudicaciones directas** "
        f"· Monto total: **${_dff_ad['Importe DRC'].sum()/1e6:,.1f} M MXN** "
        f"· {_dff_ad_dt['rfc'].nunique() if len(_dff_ad_dt) > 0 else 0} proveedores únicos"
    )
    if _n_excl_proc > 0:
        _caption_universo += (
            f"  \n_Se excluyeron **{_n_excl_proc:,} contratos** que comparten número de "
            f"procedimiento (partidas distintas de un mismo proceso, no fragmentación)._"
        )
    st.caption(_caption_universo)
    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 1 — Concentración en el mismo día
    # ═══════════════════════════════════════════════════════════════
    st.subheader("1️⃣ Concentración en el mismo día")
    st.caption(
        "3+ contratos adjudicados al mismo proveedor en la misma fecha de inicio. "
        "Indicador clásico de fragmentación simultánea."
    )

    if len(_dff_ad_dt) > 0:
        _mismo_dia = (
            _dff_ad_dt
            .groupby(["rfc", "Proveedor o contratista", "Nombre de la UC", "_fecha_str"])
            .agg(
                Contratos   = ("Importe DRC", "count"),
                Monto_total = ("Importe DRC", "sum"),
                Monto_min   = ("Importe DRC", "min"),
                Monto_max   = ("Importe DRC", "max"),
            )
            .reset_index()
        )
        _mismo_dia = _mismo_dia[_mismo_dia["Contratos"] >= 3].sort_values(
            "Contratos", ascending=False
        )

        _k1a, _k1b, _k1c = st.columns(3)
        _k1a.metric("🔴 Grupos detectados",  f"{len(_mismo_dia):,}")
        _k1b.metric("🏢 Proveedores únicos", f"{_mismo_dia['rfc'].nunique():,}")
        _k1c.metric("💰 Monto total",        f"${_mismo_dia['Monto_total'].sum()/1e6:,.1f} M MXN")

        if len(_mismo_dia) > 0:
            st.warning(f"⚠️ **{len(_mismo_dia)}** casos con 3+ contratos al mismo proveedor en el mismo día.")
            with st.expander("📋 Ver detalle"):
                _md_fmt = _mismo_dia.rename(columns={
                    "Proveedor o contratista": "Proveedor",
                    "_fecha_str": "Fecha",
                    "Nombre de la UC": "Unidad Compradora",
                    "Monto_total": "Monto total",
                })[["Fecha", "Unidad Compradora", "Proveedor", "Contratos", "Monto total"]].copy()
                _md_fmt["Monto total"] = _md_fmt["Monto total"].apply(lambda x: f"${x:,.0f}")
                _md_fmt.index = range(1, len(_md_fmt) + 1)
                st.dataframe(_md_fmt, use_container_width=True)
                st.markdown("**Contratos individuales:**")
                _det1 = _dff_ad_dt.merge(
                    _mismo_dia[["rfc", "Nombre de la UC", "_fecha_str"]],
                    on=["rfc", "Nombre de la UC", "_fecha_str"]
                )[["Fecha de inicio del contrato", "Nombre de la UC",
                   "Proveedor o contratista", "Importe DRC",
                   "Dirección del anuncio"]].drop_duplicates()
                _det1 = _det1.sort_values(["Nombre de la UC", "Fecha de inicio del contrato"])
                _det1["Importe DRC"] = pd.to_numeric(_det1["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _det1 = _det1.rename(columns={"Importe DRC": "Importe"})
                _det1.index = range(1, len(_det1) + 1)
                st.dataframe(
                    _det1, use_container_width=True,
                    column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    )},
                )
        else:
            st.success("✅ Sin concentraciones en el mismo día (umbral: 3+ contratos).")
    else:
        st.info("ℹ️ Sin fechas de inicio disponibles para este indicador.")

    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 2 — Reparto entre proveedores
    # ═══════════════════════════════════════════════════════════════
    st.subheader("2️⃣ Reparto entre proveedores")
    st.caption(
        "Ventanas de 3 días dentro de la misma UC con 3+ contratos distribuidos a "
        "distintos proveedores con importes comparables (razón máx/mín ≤ 3). "
        "Fragmentación sofisticada entre adjudicatarios."
    )

    if len(_dff_ad_dt) > 0:
        _reparto_g = (
            _dff_ad_dt
            .groupby(["Nombre de la UC", "_bucket3"])
            .agg(
                n_contratos  = ("rfc", "count"),
                n_proveedores = ("rfc", "nunique"),
                Monto_min    = ("Importe DRC", "min"),
                Monto_max    = ("Importe DRC", "max"),
                Monto_total  = ("Importe DRC", "sum"),
            )
            .reset_index()
        )
        _reparto_g["ratio"] = (
            _reparto_g["Monto_max"] / _reparto_g["Monto_min"].replace(0, pd.NA)
        ).fillna(99)
        _reparto_alrt = _reparto_g[
            (_reparto_g["n_proveedores"] >= 3) & (_reparto_g["ratio"] <= 3.0)
        ].sort_values("n_contratos", ascending=False)

        _k2a, _k2b, _k2c = st.columns(3)
        _k2a.metric("🔴 Ventanas detectadas", f"{len(_reparto_alrt):,}")
        _k2b.metric("🏢 UCs involucradas",    f"{_reparto_alrt['Nombre de la UC'].nunique():,}")
        _k2c.metric("💰 Monto total",         f"${_reparto_alrt['Monto_total'].sum()/1e6:,.1f} M MXN")

        if len(_reparto_alrt) > 0:
            st.warning(
                f"⚠️ **{len(_reparto_alrt)}** ventanas de 3 días con 3+ proveedores distintos "
                "e importes comparables (posible reparto coordinado)."
            )
            with st.expander("📋 Ver detalle"):
                _rep_fmt = _reparto_alrt.rename(columns={
                    "Nombre de la UC": "Unidad Compradora",
                    "n_contratos": "Contratos",
                    "n_proveedores": "Proveedores distintos",
                    "Monto_total": "Monto total",
                    "ratio": "Razón máx/mín",
                })[["Unidad Compradora", "Contratos", "Proveedores distintos",
                    "Razón máx/mín", "Monto total"]].copy()
                _rep_fmt["Monto total"]   = _rep_fmt["Monto total"].apply(lambda x: f"${x:,.0f}")
                _rep_fmt["Razón máx/mín"] = _rep_fmt["Razón máx/mín"].apply(lambda x: f"{x:.2f}")
                _rep_fmt.index = range(1, len(_rep_fmt) + 1)
                st.dataframe(_rep_fmt, use_container_width=True)
                st.markdown("**Contratos individuales:**")
                _det2 = _dff_ad_dt.merge(
                    _reparto_alrt[["Nombre de la UC", "_bucket3"]],
                    on=["Nombre de la UC", "_bucket3"]
                )[["Fecha de inicio del contrato", "Nombre de la UC",
                   "Proveedor o contratista", "Importe DRC",
                   "Dirección del anuncio"]].drop_duplicates()
                _det2 = _det2.sort_values(["Nombre de la UC", "Fecha de inicio del contrato"])
                _det2["Importe DRC"] = pd.to_numeric(_det2["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _det2 = _det2.rename(columns={"Importe DRC": "Importe"})
                _det2.index = range(1, len(_det2) + 1)
                st.dataframe(
                    _det2, use_container_width=True,
                    column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    )},
                )
        else:
            st.success("✅ Sin patrones de reparto entre proveedores.")
    else:
        st.info("ℹ️ Sin fechas disponibles para este indicador.")

    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 3 — Concentración en 30 días
    # ═══════════════════════════════════════════════════════════════
    st.subheader("3️⃣ Concentración en 30 días")
    st.caption(
        "3+ contratos al mismo proveedor en cualquier ventana corrida de 30 días. "
        "Se reporta la ventana de mayor concentración por cada par proveedor-UC."
    )

    if len(_dff_ad_dt) > 0:
        _v30_rows = []
        for (_rfc_3, _uc_3), _grp_3 in _dff_ad_dt.groupby(["rfc", "Nombre de la UC"]):
            if len(_grp_3) < 3:
                continue
            _grp_3s    = _grp_3.sort_values("_fecha")
            _fechas3   = _grp_3s["_fecha"].tolist()
            _importes3 = _grp_3s["Importe DRC"].tolist()
            _max_n, _max_m = 0, 0.0
            for _i3 in range(len(_fechas3)):
                _idx_w = [
                    _j3 for _j3 in range(len(_fechas3))
                    if 0 <= (_fechas3[_j3] - _fechas3[_i3]).days <= 30
                ]
                if len(_idx_w) > _max_n:
                    _max_n = len(_idx_w)
                    _max_m = sum(_importes3[_j3] for _j3 in _idx_w)
            if _max_n >= 3:
                _v30_rows.append({
                    "rfc": _rfc_3,
                    "Proveedor": _grp_3s["Proveedor o contratista"].iloc[0],
                    "Unidad Compradora": _uc_3,
                    "Máx contratos (30 d)": _max_n,
                    "Monto máx ventana": _max_m,
                })

        _df_v30 = (
            pd.DataFrame(_v30_rows).sort_values("Máx contratos (30 d)", ascending=False)
            if _v30_rows else pd.DataFrame()
        )

        _k3a, _k3b, _k3c = st.columns(3)
        _k3a.metric("🔴 Proveedores detectados", f"{len(_df_v30):,}")
        _k3b.metric("🏢 UCs involucradas",
                    f"{_df_v30['Unidad Compradora'].nunique():,}" if len(_df_v30) > 0 else "0")
        _k3c.metric("💰 Mayor ventana",
                    f"${_df_v30['Monto máx ventana'].max()/1e6:,.1f} M MXN"
                    if len(_df_v30) > 0 else "N/D")

        if len(_df_v30) > 0:
            st.warning(f"⚠️ **{len(_df_v30)}** proveedores con 3+ contratos en alguna ventana de 30 días.")
            with st.expander("📋 Ver detalle"):
                _v30_fmt = _df_v30.drop(columns=["rfc"], errors="ignore").copy()
                _v30_fmt["Monto máx ventana"] = _v30_fmt["Monto máx ventana"].apply(
                    lambda x: f"${x:,.0f}"
                )
                _v30_fmt.index = range(1, len(_v30_fmt) + 1)
                st.dataframe(_v30_fmt, use_container_width=True)
                st.markdown("**Contratos individuales:**")
                _det3 = _dff_ad_dt.merge(
                    _df_v30[["rfc", "Unidad Compradora"]].rename(
                        columns={"Unidad Compradora": "Nombre de la UC"}
                    ),
                    on=["rfc", "Nombre de la UC"]
                )[["Fecha de inicio del contrato", "Nombre de la UC",
                   "Proveedor o contratista", "Importe DRC",
                   "Dirección del anuncio"]].drop_duplicates()
                _det3 = _det3.sort_values(["Nombre de la UC", "Fecha de inicio del contrato"])
                _det3["Importe DRC"] = pd.to_numeric(_det3["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _det3 = _det3.rename(columns={"Importe DRC": "Importe"})
                _det3.index = range(1, len(_det3) + 1)
                st.dataframe(
                    _det3, use_container_width=True,
                    column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    )},
                )
        else:
            st.success("✅ Sin concentraciones en ventanas de 30 días (umbral: 3+ contratos).")
    else:
        st.info("ℹ️ Sin fechas disponibles para este indicador.")

    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 4 — Objetos contractuales similares (NLP)
    # ═══════════════════════════════════════════════════════════════
    st.subheader("4️⃣ Objetos contractuales similares")
    st.caption(
        "Grupos de 3+ contratos dentro de la misma UC con descripciones ≥70% similares. "
        "Usa TF-IDF + similitud coseno (sklearn) o Jaccard como alternativa."
    )

    _desc_col4 = (
        "Descripción del contrato"
        if "Descripción del contrato" in _dff_ad.columns else None
    )

    if _desc_col4 is None:
        st.info("ℹ️ Columna 'Descripción del contrato' no encontrada.")
    else:
        _dff_nlp = _dff_ad[
            [_desc_col4, "Nombre de la UC", "Proveedor o contratista",
             "Importe DRC", "Fecha de inicio del contrato", "Dirección del anuncio"]
        ].copy()
        _dff_nlp = _dff_nlp[
            _dff_nlp[_desc_col4].notna() & (_dff_nlp[_desc_col4].str.strip() != "")
        ].reset_index(drop=True)

        _STOP_ES = [
            "de","la","el","en","y","a","del","los","las","con","para","por","se","que",
            "un","una","al","su","no","es","lo","como","más","o","sus","le","ya","entre",
            "cuando","todo","esta","ser","son","dos","también","fue","este","ha","si",
            "sobre","han","les","hay","donde","me","hasta","e","al","sin","ni","muy"
        ]

        _grupos_sim4       = []
        _contratos_grupos4 = []
        _metodo_nlp        = "TF-IDF (sklearn)"

        try:
            from sklearn.feature_extraction.text import TfidfVectorizer
            from sklearn.metrics.pairwise import cosine_similarity as _cos_sim

            for _uc_nlp, _g_nlp in _dff_nlp.groupby("Nombre de la UC"):
                if len(_g_nlp) < 3:
                    continue
                _g_nlp = _g_nlp.reset_index(drop=True)
                try:
                    _vect4 = TfidfVectorizer(
                        stop_words=_STOP_ES, ngram_range=(1, 2),
                        min_df=1, max_df=0.95, max_features=3000
                    )
                    _mat4 = _vect4.fit_transform(_g_nlp[_desc_col4].astype(str))
                    _sim4 = _cos_sim(_mat4)
                    np.fill_diagonal(_sim4, 0)
                    _visto4 = set()
                    for _i4 in range(len(_g_nlp)):
                        if _i4 in _visto4:
                            continue
                        _idx_sim4 = [
                            _j4 for _j4 in range(len(_g_nlp))
                            if _j4 != _i4 and _sim4[_i4, _j4] >= 0.70
                        ]
                        if len(_idx_sim4) >= 2:
                            _grupo4 = frozenset([_i4] + _idx_sim4)
                            _visto4.update(_grupo4)
                            _rows4 = _g_nlp.iloc[list(_grupo4)]
                            _grupos_sim4.append({
                                "UC": _uc_nlp,
                                "Contratos similares": len(_rows4),
                                "Monto total": _rows4["Importe DRC"].sum(),
                                "Descripción (ejemplo)": _g_nlp.iloc[_i4][_desc_col4][:130],
                            })
                            _contratos_grupos4.append(_rows4[[
                                "Fecha de inicio del contrato", "Nombre de la UC",
                                "Proveedor o contratista", "Importe DRC",
                                _desc_col4, "Dirección del anuncio"
                            ]])
                except Exception:
                    continue

        except ImportError:
            _metodo_nlp = "Jaccard (fallback — instala scikit-learn para TF-IDF)"

            def _jac_words(texto):
                return set(_re.findall(r'\b\w{4,}\b', str(texto).lower()))

            for _uc_nlp, _g_nlp in _dff_nlp.groupby("Nombre de la UC"):
                if len(_g_nlp) < 3:
                    continue
                _g_nlp = _g_nlp.reset_index(drop=True)
                _wsets = [_jac_words(d) for d in _g_nlp[_desc_col4]]
                _visto4 = set()
                for _i4 in range(len(_g_nlp)):
                    if _i4 in _visto4 or not _wsets[_i4]:
                        continue
                    _idx_sim4 = [
                        _j4 for _j4 in range(len(_g_nlp))
                        if _j4 != _i4 and _wsets[_j4] and
                        len(_wsets[_i4] & _wsets[_j4]) / max(len(_wsets[_i4] | _wsets[_j4]), 1) >= 0.70
                    ]
                    if len(_idx_sim4) >= 2:
                        _grupo4 = frozenset([_i4] + _idx_sim4)
                        _visto4.update(_grupo4)
                        _rows4 = _g_nlp.iloc[list(_grupo4)]
                        _grupos_sim4.append({
                            "UC": _uc_nlp,
                            "Contratos similares": len(_rows4),
                            "Monto total": _rows4["Importe DRC"].sum(),
                            "Descripción (ejemplo)": _g_nlp.iloc[_i4][_desc_col4][:130],
                        })
                        _contratos_grupos4.append(_rows4[[
                            "Fecha de inicio del contrato", "Nombre de la UC",
                            "Proveedor o contratista", "Importe DRC",
                            _desc_col4, "Dirección del anuncio"
                        ]])

        _k4a, _k4b, _k4c = st.columns(3)
        _k4a.metric("🔴 Grupos detectados", f"{len(_grupos_sim4):,}")
        _k4b.metric("🏢 UCs involucradas",
                    f"{len({g['UC'] for g in _grupos_sim4}):,}" if _grupos_sim4 else "0")
        _k4c.metric("💰 Monto involucrado",
                    f"${sum(g['Monto total'] for g in _grupos_sim4)/1e6:,.1f} M MXN"
                    if _grupos_sim4 else "N/D")
        st.caption(f"_Método: {_metodo_nlp}_")

        if _grupos_sim4:
            st.warning(f"⚠️ **{len(_grupos_sim4)}** grupos con descripciones ≥70% similares.")
            with st.expander("📋 Ver grupos similares"):
                _df_sim4 = (
                    pd.DataFrame(_grupos_sim4)
                    .sort_values("Contratos similares", ascending=False)
                )
                _df_sim4["Monto total"] = _df_sim4["Monto total"].apply(lambda x: f"${x:,.0f}")
                _df_sim4.index = range(1, len(_df_sim4) + 1)
                st.dataframe(_df_sim4, use_container_width=True)
                if _contratos_grupos4:
                    st.markdown("**Contratos individuales:**")
                    _det4 = (
                        pd.concat(_contratos_grupos4, ignore_index=True)
                        .drop_duplicates()
                        .sort_values(["Nombre de la UC", "Proveedor o contratista"])
                    )
                    _det4["Importe DRC"] = pd.to_numeric(_det4["Importe DRC"], errors="coerce").apply(
                        lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                    )
                    _det4 = _det4.rename(columns={_desc_col4: "Descripción", "Importe DRC": "Importe"})
                    _det4.index = range(1, len(_det4) + 1)
                    st.dataframe(
                        _det4, use_container_width=True,
                        column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                            "🔗 ComprasMX", display_text="Ver contrato"
                        )},
                    )
        else:
            st.success("✅ Sin grupos de contratos con objetos similares (umbral ≥70%).")

    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 5 — Concentración por Partida
    # ═══════════════════════════════════════════════════════════════
    st.subheader("5️⃣ Concentración por Partida")
    st.caption(
        "Proveedor con 4+ contratos en la misma partida CUCoP que representan >15% "
        "del total de contratos adjudicados en esa partida. Alta concentración en una "
        "categoría puede indicar fragmentación por tipo de producto o servicio."
    )

    _dff_part5 = _dff_ad[
        ["rfc", "Proveedor o contratista", "Nombre de la UC", "Partida específica", "Importe DRC"]
    ].copy()
    # Excluir partidas nulas o rellenas de ceros
    _dff_part5 = _dff_part5[
        _dff_part5["Partida específica"].notna() &
        (_dff_part5["Partida específica"].str.replace("0", "", regex=False).str.strip() != "")
    ]

    if len(_dff_part5) > 0:
        _total_p5 = (
            _dff_part5.groupby("Partida específica")
            .agg(total_contratos=("rfc", "count"), total_monto=("Importe DRC", "sum"))
            .reset_index()
        )
        _prov_p5 = (
            _dff_part5
            .groupby(["rfc", "Proveedor o contratista", "Nombre de la UC", "Partida específica"])
            .agg(n_contratos=("Importe DRC", "count"), monto_proveedor=("Importe DRC", "sum"))
            .reset_index()
        )
        _conc_p5 = _prov_p5.merge(_total_p5, on="Partida específica")
        _conc_p5["pct"] = _conc_p5["n_contratos"] / _conc_p5["total_contratos"] * 100
        _conc_p5_alrt = _conc_p5[
            (_conc_p5["n_contratos"] >= 4) & (_conc_p5["pct"] > 15)
        ].sort_values("n_contratos", ascending=False)

        # Enriquecer con descripción CUCoP
        _conc_p5_alrt = _conc_p5_alrt.merge(
            df_cucop[["PARTIDA ESPECÍFICA", "DESC. PARTIDA ESPECÍFICA"]]
            .drop_duplicates("PARTIDA ESPECÍFICA"),
            left_on="Partida específica", right_on="PARTIDA ESPECÍFICA", how="left"
        )

        _k5a, _k5b, _k5c = st.columns(3)
        _k5a.metric("🔴 Alertas detectadas", f"{len(_conc_p5_alrt):,}")
        _k5b.metric("🏢 Proveedores únicos", f"{_conc_p5_alrt['rfc'].nunique():,}")
        _k5c.metric("💰 Monto total",        f"${_conc_p5_alrt['monto_proveedor'].sum()/1e6:,.1f} M MXN")

        if len(_conc_p5_alrt) > 0:
            st.warning(
                f"⚠️ **{len(_conc_p5_alrt)}** combinaciones proveedor-partida con 4+ contratos "
                "que representan >15% del total adjudicado en esa partida."
            )
            with st.expander("📋 Ver detalle"):
                _cp_fmt5 = _conc_p5_alrt.rename(columns={
                    "Proveedor o contratista": "Proveedor",
                    "Nombre de la UC": "UC",
                    "Partida específica": "Partida",
                    "DESC. PARTIDA ESPECÍFICA": "Descripción partida",
                    "n_contratos": "Contratos",
                    "pct": "% de la partida",
                    "monto_proveedor": "Monto proveedor",
                })[["Partida", "Descripción partida", "UC", "Proveedor",
                    "Contratos", "% de la partida", "Monto proveedor"]].copy()
                _cp_fmt5["% de la partida"] = _cp_fmt5["% de la partida"].apply(
                    lambda x: f"{x:.1f}%"
                )
                _cp_fmt5["Monto proveedor"] = _cp_fmt5["Monto proveedor"].apply(
                    lambda x: f"${x:,.0f}"
                )
                _cp_fmt5.index = range(1, len(_cp_fmt5) + 1)
                st.dataframe(_cp_fmt5, use_container_width=True)
                st.markdown("**Contratos individuales:**")
                _det5 = _dff_ad.merge(
                    _conc_p5_alrt[["rfc", "Nombre de la UC", "Partida específica"]],
                    on=["rfc", "Nombre de la UC", "Partida específica"]
                )[["Fecha de inicio del contrato", "Nombre de la UC",
                   "Proveedor o contratista", "Importe DRC",
                   "Dirección del anuncio"]].drop_duplicates()
                _det5 = _det5.sort_values(["Nombre de la UC", "Proveedor o contratista"])
                _det5["Importe DRC"] = pd.to_numeric(_det5["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _det5 = _det5.rename(columns={"Importe DRC": "Importe"})
                _det5.index = range(1, len(_det5) + 1)
                st.dataframe(
                    _det5, use_container_width=True,
                    column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    )},
                )
        else:
            st.success("✅ Sin concentraciones anómalas por partida (umbral: 4+ contratos y >15%).")
    else:
        st.info("ℹ️ Sin datos de partida disponibles.")

    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 6 — Patrón recurrente
    # ═══════════════════════════════════════════════════════════════
    st.subheader("6️⃣ Patrón recurrente")
    st.caption(
        "3+ contratos al mismo proveedor y UC con importes similares (±20% de la mediana) "
        "a intervalos regulares (coeficiente de variación de los intervalos < 0.40). "
        "Sugiere un servicio continuado disfrazado de contratos menores independientes."
    )

    if len(_dff_ad_dt) > 0:
        _patron_rows6 = []
        for (_rfc_6, _uc_6), _grp_6 in _dff_ad_dt.groupby(["rfc", "Nombre de la UC"]):
            if len(_grp_6) < 3:
                continue
            _grp_6s  = _grp_6.sort_values("_fecha")
            _imp6    = _grp_6s["Importe DRC"].dropna().tolist()
            if not _imp6:
                continue
            _med6 = float(np.median(_imp6))
            if _med6 <= 0:
                continue
            # Similitud de importes: todos dentro del ±20% de la mediana
            if not all(abs(v - _med6) / _med6 <= 0.20 for v in _imp6):
                continue
            _fechas6 = _grp_6s["_fecha"].tolist()
            _intv6   = [(_fechas6[_i + 1] - _fechas6[_i]).days for _i in range(len(_fechas6) - 1)]
            if not _intv6:
                continue
            _mean_i6 = float(np.mean(_intv6))
            if _mean_i6 < 1:
                continue
            _cv6 = float(np.std(_intv6)) / _mean_i6
            if _cv6 < 0.40:   # intervalos regulares
                _patron_rows6.append({
                    "rfc": _rfc_6,
                    "Proveedor": _grp_6s["Proveedor o contratista"].iloc[0],
                    "Unidad Compradora": _uc_6,
                    "Contratos": len(_grp_6s),
                    "Intervalo medio (días)": round(_mean_i6, 1),
                    "CV intervalos": round(_cv6, 3),
                    "Importe mediano": _med6,
                    "Monto total": _grp_6s["Importe DRC"].sum(),
                })

        _df_p6 = (
            pd.DataFrame(_patron_rows6).sort_values("Contratos", ascending=False)
            if _patron_rows6 else pd.DataFrame()
        )

        _k6a, _k6b, _k6c = st.columns(3)
        _k6a.metric("🔴 Patrones detectados", f"{len(_df_p6):,}")
        _k6b.metric("🏢 UCs involucradas",
                    f"{_df_p6['Unidad Compradora'].nunique():,}" if len(_df_p6) > 0 else "0")
        _k6c.metric("💰 Monto total",
                    f"${_df_p6['Monto total'].sum()/1e6:,.1f} M MXN" if len(_df_p6) > 0 else "N/D")

        if len(_df_p6) > 0:
            st.warning(f"⚠️ **{len(_df_p6)}** patrones recurrentes detectados.")
            with st.expander("📋 Ver detalle"):
                _df_p6_fmt = _df_p6.drop(columns=["rfc"], errors="ignore").copy()
                _df_p6_fmt["Importe mediano"] = _df_p6_fmt["Importe mediano"].apply(
                    lambda x: f"${x:,.0f}"
                )
                _df_p6_fmt["Monto total"] = _df_p6_fmt["Monto total"].apply(
                    lambda x: f"${x:,.0f}"
                )
                _df_p6_fmt.index = range(1, len(_df_p6_fmt) + 1)
                st.dataframe(_df_p6_fmt, use_container_width=True)
                st.markdown("**Contratos individuales:**")
                _det6 = _dff_ad_dt.merge(
                    _df_p6[["rfc", "Unidad Compradora"]].rename(
                        columns={"Unidad Compradora": "Nombre de la UC"}
                    ),
                    on=["rfc", "Nombre de la UC"]
                )[["Fecha de inicio del contrato", "Nombre de la UC",
                   "Proveedor o contratista", "Importe DRC",
                   "Dirección del anuncio"]].drop_duplicates()
                _det6 = _det6.sort_values(["Nombre de la UC", "Fecha de inicio del contrato"])
                _det6["Importe DRC"] = pd.to_numeric(_det6["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _det6 = _det6.rename(columns={"Importe DRC": "Importe"})
                _det6.index = range(1, len(_det6) + 1)
                st.dataframe(
                    _det6, use_container_width=True,
                    column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    )},
                )
        else:
            st.success("✅ Sin patrones recurrentes (importes ±20%, intervalos regulares).")
    else:
        st.info("ℹ️ Sin fechas disponibles para este indicador.")

    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # INDICADOR 7 — Pico de frecuencia semanal
    # ═══════════════════════════════════════════════════════════════
    st.subheader("7️⃣ Pico de frecuencia semanal")
    st.caption(
        "Semanas con volumen de adjudicaciones directas estadísticamente anómalo "
        "(Z-score ≥ 2.5, mínimo 10 contratos en la semana). "
        "Concentraciones que se desvían del patrón habitual."
    )

    if len(_dff_ad_dt) >= 10:
        _weekly7 = (
            _dff_ad_dt.groupby("_semana")
            .agg(Contratos=("Importe DRC", "count"), Monto=("Importe DRC", "sum"))
            .reset_index()
            .sort_values("_semana")
        )
        if len(_weekly7) >= 4:
            _mean_w7 = _weekly7["Contratos"].mean()
            _std_w7  = _weekly7["Contratos"].std()

            if _std_w7 > 0:
                _weekly7["Z-score"] = (_weekly7["Contratos"] - _mean_w7) / _std_w7

                # Etiqueta legible: fecha del lunes que inicia la semana ISO
                # Ej. "2025-W03" → "13 Ene '25"
                import datetime as _dt7
                _MESES_C7 = ["Ene","Feb","Mar","Abr","May","Jun",
                             "Jul","Ago","Sep","Oct","Nov","Dic"]
                def _sem_label(s):
                    try:
                        _lun = _dt7.datetime.strptime(s + "-1", "%G-W%V-%u").date()
                        return (f"{_lun.day:02d} {_MESES_C7[_lun.month - 1]}"
                                f" '{str(_lun.year)[2:]}")
                    except Exception:
                        return s
                _weekly7["_semana_label"] = _weekly7["_semana"].apply(_sem_label)

                _picos7 = _weekly7[
                    (_weekly7["Z-score"] >= 2.5) & (_weekly7["Contratos"] >= 10)
                ].sort_values("Z-score", ascending=False)

                _k7a, _k7b, _k7c = st.columns(3)
                _k7a.metric("🔴 Semanas anómalas", f"{len(_picos7):,}")
                _k7b.metric("📊 Media semanal",    f"{_mean_w7:.1f} contratos")
                _k7c.metric("💰 Monto en picos",
                            f"${_picos7['Monto'].sum()/1e6:,.1f} M MXN"
                            if len(_picos7) > 0 else "N/D")

                # Gráfica de barras coloreadas por Z-score
                _weekly7["Color"] = _weekly7["Z-score"].apply(
                    lambda z: IMSS_ROJO if z >= 2.5 else (IMSS_ORO if z >= 1.5 else IMSS_VERDE)
                )
                _fig_sem7 = go.Figure()
                _fig_sem7.add_bar(
                    x=_weekly7["_semana_label"],
                    y=_weekly7["Contratos"],
                    marker_color=_weekly7["Color"].tolist(),
                    customdata=_weekly7["_semana"].tolist(),
                    hovertemplate=(
                        "<b>%{x}</b><br>Contratos: %{y}"
                        "<br><span style='color:gray;font-size:11px'>%{customdata}</span>"
                        "<extra></extra>"
                    )
                )
                _umbral_z7 = _mean_w7 + 2.5 * _std_w7
                _fig_sem7.add_hline(
                    y=_umbral_z7, line_dash="dash", line_color=IMSS_ROJO,
                    annotation_text=f"Umbral Z=2.5  ({_umbral_z7:.1f} contratos)",
                    annotation_font_color=IMSS_ROJO,
                )
                _fig_sem7.update_layout(
                    font=plotly_font(),
                    title="Adjudicaciones directas por semana  ·  rojo = Z-score ≥ 2.5",
                    xaxis_title="Semana (lunes de inicio)", yaxis_title="Nº contratos",
                    plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                    xaxis_tickangle=-45, showlegend=False,
                )
                st.plotly_chart(_fig_sem7, use_container_width=True)

                if len(_picos7) > 0:
                    st.warning(f"⚠️ **{len(_picos7)}** semanas con actividad estadísticamente anómala.")
                    with st.expander("📋 Ver semanas anómalas"):
                        _picos7_fmt = _picos7.rename(
                            columns={"_semana_label": "Semana", "Monto": "Monto total"}
                        ).copy()
                        _picos7_fmt["Monto total"] = _picos7_fmt["Monto total"].apply(
                            lambda x: f"${x:,.0f}"
                        )
                        _picos7_fmt["Z-score"] = _picos7_fmt["Z-score"].apply(
                            lambda x: f"{x:.2f}"
                        )
                        _picos7_fmt = _picos7_fmt[["Semana", "Contratos", "Monto total", "Z-score"]]
                        _picos7_fmt.index = range(1, len(_picos7_fmt) + 1)
                        st.dataframe(_picos7_fmt, use_container_width=True)
                        st.markdown("**Contratos individuales en semanas anómalas:**")
                        _semanas_anom7 = set(_picos7["_semana"].tolist())
                        _det7 = _dff_ad_dt[
                            _dff_ad_dt["_semana"].isin(_semanas_anom7)
                        ][["Fecha de inicio del contrato", "Nombre de la UC",
                           "Proveedor o contratista", "Importe DRC",
                           "Dirección del anuncio"]].drop_duplicates()
                        _det7 = _det7.sort_values(["Nombre de la UC", "Fecha de inicio del contrato"])
                        _det7["Importe DRC"] = pd.to_numeric(_det7["Importe DRC"], errors="coerce").apply(
                            lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                        )
                        _det7 = _det7.rename(columns={"Importe DRC": "Importe"})
                        _det7.index = range(1, len(_det7) + 1)
                        st.dataframe(
                            _det7, use_container_width=True,
                            column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                                "🔗 ComprasMX", display_text="Ver contrato"
                            )},
                        )
                else:
                    st.success("✅ Sin semanas con actividad estadísticamente anómala.")
            else:
                st.info("ℹ️ Variabilidad semanal nula — no se puede calcular Z-score.")
        else:
            st.info("ℹ️ Se necesitan al menos 4 semanas distintas de datos para este análisis.")
    else:
        st.info("ℹ️ Se necesitan al menos 10 contratos con fecha para este indicador.")

    st.divider()


# ───────────────────────────────────────────────────────────────
# PÁGINA 5: PAGINA_MAPA_RIESGO
# ───────────────────────────────────────────────────────────────
def pagina_mapa_riesgo():
    import re as _re_mr
    from datetime import date as _date_mr

    st.header("🗺️ Perfil UC")
    st.caption(
        "Selecciona una Unidad Compradora o Adscripción para obtener su numeralia, "
        "los riesgos específicos detectados por los indicadores del dashboard "
        "y el listado completo de contratos. "
        "Solo se muestran las categorías de riesgo donde se encontraron hallazgos."
    )

    # \u2500\u2500 1. Selector \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    _s1, _s2 = st.columns([1, 3])
    _tipo_vista_mr = _s1.radio(
        "Ver por:", ["UC espec\u00edfica", "Adscripci\u00f3n"],
        horizontal=False, key="mr_tipo_vista"
    )

    # Join dff con metadatos de UC
    _dff_base_mr  = dff.copy()
    _tiene_meta_mr = len(df_dir_uc) > 0 and "Clave de la UC" in _dff_base_mr.columns
    if _tiene_meta_mr:
        _uc_meta_mr = (
            df_dir_uc[["Clave_UC", "Tipo UC", "Adscripci\u00f3n"]]
            .rename(columns={"Clave_UC": "Clave de la UC"})
        )
        _dff_base_mr = _dff_base_mr.merge(_uc_meta_mr, on="Clave de la UC", how="left")

    if _tipo_vista_mr == "UC espec\u00edfica":
        _opts_mr = sorted(_dff_base_mr["Nombre de la UC"].dropna().unique().tolist())
        if not _opts_mr:
            st.info("\u2139\ufe0f Sin UCs disponibles con los filtros actuales.")
            return
        _sel_mr   = _s2.selectbox("Seleccionar Unidad Compradora:", _opts_mr, key="mr_uc_sel")
        _dff_sel  = _dff_base_mr[_dff_base_mr["Nombre de la UC"] == _sel_mr].copy()
        _label_mr = _sel_mr
        _n_ucs_sel = 1
        _meta_tipo_mr = (
            _dff_sel["Tipo UC"].dropna().iloc[0]
            if "Tipo UC" in _dff_sel.columns and len(_dff_sel["Tipo UC"].dropna()) > 0 else "\u2014"
        )
        _meta_adsc_mr = (
            _dff_sel["Adscripci\u00f3n"].dropna().iloc[0]
            if "Adscripci\u00f3n" in _dff_sel.columns and len(_dff_sel["Adscripci\u00f3n"].dropna()) > 0 else "\u2014"
        )
    else:
        if not _tiene_meta_mr or "Adscripci\u00f3n" not in _dff_base_mr.columns:
            st.info("\u2139\ufe0f Informaci\u00f3n de Adscripci\u00f3n no disponible (requiere Base_UC_2025_V2.xlsx).")
            return
        _opts_mr = sorted(_dff_base_mr["Adscripci\u00f3n"].dropna().unique().tolist())
        if not _opts_mr:
            st.info("\u2139\ufe0f Sin Adscripciones disponibles.")
            return
        _sel_mr   = _s2.selectbox("Seleccionar Adscripci\u00f3n:", _opts_mr, key="mr_adsc_sel")
        _dff_sel  = _dff_base_mr[_dff_base_mr["Adscripci\u00f3n"] == _sel_mr].copy()
        _label_mr  = _sel_mr
        _n_ucs_sel = _dff_sel["Nombre de la UC"].nunique()
        _meta_tipo_mr = None
        _meta_adsc_mr = _sel_mr

    if len(_dff_sel) == 0:
        st.info("\u2139\ufe0f Sin contratos para la selecci\u00f3n actual.")
        return

    st.divider()

    # \u2500\u2500 2. Numeralia r\u00e1pida \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    st.subheader(f"\U0001f4ca Numeralia \u2014 {_label_mr}")

    _total_mr  = len(_dff_sel)
    _monto_mr  = _dff_sel["Importe DRC"].sum()
    _n_prov_mr = _dff_sel["rfc"].nunique()
    # Ambas métricas sobre el mismo denominador: % del monto total
    _m_lp_mr   = _dff_sel.loc[
        _dff_sel["Tipo Simplificado"] == "Licitaci\u00f3n P\u00fablica", "Importe DRC"
    ].sum()
    _m_ad_mr   = _dff_sel.loc[
        _dff_sel["Tipo Simplificado"].isin(
            ["Adjudicaci\u00f3n Directa", "Adjudicaci\u00f3n Directa \u2014 Fr. I"]
        ), "Importe DRC"
    ].sum()
    _pct_lp_mr = _m_lp_mr / _monto_mr * 100 if _monto_mr > 0 else 0
    _pct_ad_mr = _m_ad_mr / _monto_mr * 100 if _monto_mr > 0 else 0

    _nm1, _nm2, _nm3, _nm4, _nm5 = st.columns(5)
    _nm1.metric("\U0001f4c4 Contratos", f"{_total_mr:,}")
    _nm2.metric(
        "\U0001f4b0 Monto total",
        f"${_monto_mr/1e9:,.2f} miles de millones MXN" if _monto_mr >= 1e9 else f"${_monto_mr/1e6:,.1f} M MXN"
    )
    _nm3.metric("\U0001f3e2 Proveedores", f"{_n_prov_mr:,}")
    _nm4.metric(
        "\U0001f7e2 % Monto LP",
        f"{_pct_lp_mr:.1f}%"
    )
    _nm5.metric("\U0001f534 % Monto AD", f"{_pct_ad_mr:.1f}%")

    # Fila de informaci\u00f3n contextual
    _info_parts = []
    if _tipo_vista_mr == "UC espec\u00edfica":
        if _meta_tipo_mr and _meta_tipo_mr != "\u2014":
            _info_parts.append(f"**Tipo UC:** {_meta_tipo_mr}")
        if _meta_adsc_mr and _meta_adsc_mr != "\u2014":
            _info_parts.append(f"**Adscripci\u00f3n:** {_meta_adsc_mr}")
    else:
        _info_parts.append(f"**UCs incluidas:** {_n_ucs_sel}")

    _top_prov_ser = (
        _dff_sel.groupby(["Proveedor o contratista", "rfc"])["Importe DRC"].sum()
        .sort_values(ascending=False)
    )
    if len(_top_prov_ser) > 0 and _monto_mr > 0:
        _top_prov_nm  = _top_prov_ser.index[0][0]
        _top_prov_pct = _top_prov_ser.iloc[0] / _monto_mr * 100
        _info_parts.append(f"**Proveedor principal:** {_top_prov_nm} ({_top_prov_pct:.1f}% del monto)")

    _f_parsed_mr = pd.to_datetime(
        _dff_sel["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
    )
    if _f_parsed_mr.notna().any():
        _info_parts.append(
            f"**Rango de contratos:** {_f_parsed_mr.min().strftime('%d/%m/%Y')} \u2013 "
            f"{_f_parsed_mr.max().strftime('%d/%m/%Y')}"
        )

    if _info_parts:
        st.markdown("  \u00a0|\u00a0  ".join(_info_parts))

    # ─────────────────────────────────────────────────────────────────────
    # PDF Export — button (top; generation triggered at end of section)
    # ─────────────────────────────────────────────────────────────────────
    _pdf_state_key = f"pdf_bytes_mr_{str(_sel_mr)[:30]}"
    _pdf_flag_key  = f"pdf_flag_mr_{str(_sel_mr)[:30]}"
    _c_pdf1, _c_pdf2 = st.columns([1, 3])
    if _c_pdf1.button(
        "\U0001f4c4 Generar PDF del Perfil UC",
        use_container_width=True,
        key="btn_gen_pdf_mr",
        help="Genera un PDF A4 horizontal con numeralia y visualizaciones",
    ):
        st.session_state[_pdf_flag_key] = True
        st.session_state.pop(_pdf_state_key, None)
    if st.session_state.get(_pdf_state_key):
        _c_pdf2.download_button(
            "\u2b07\ufe0f Descargar PDF del Perfil UC",
            data=st.session_state[_pdf_state_key],
            file_name=f"perfil_uc_{str(_label_mr).replace(' ', '_')[:30]}.pdf",
            mime="application/pdf",
            use_container_width=True,
            key="btn_dl_pdf_mr",
        )
    elif st.session_state.get(_pdf_flag_key):
        _c_pdf2.info("\u23f3 Generando PDF\u2026 espera un momento.")

    st.divider()

    # \u2500\u2500 3. Riesgos espec\u00edficos \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    # ─────────────────────────────────────────────────────────────────────
    # VIZ A — Distribución por Tipo de Procedimiento
    # ─────────────────────────────────────────────────────────────────────
    st.subheader("📊 Distribución por Tipo de Procedimiento")
    _TIPO_RENAME_MR = {"Adjudicación Directa — Fr. I": "Adjudicación Directa — Patentes"}

    _col_pA1, _col_pA2 = st.columns(2)
    with _col_pA1:
        _dist_both_mr = (
            _dff_sel.copy()
            .assign(Tipo=lambda d: d["Tipo Simplificado"].replace(_TIPO_RENAME_MR))
            .groupby("Tipo")
            .agg(Contratos=("Importe DRC", "count"), Monto=("Importe DRC", "sum"))
            .reset_index()
        )
        _dist_both_mr["Monto_fmt"] = _dist_both_mr["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
        _fig_pA1 = px.pie(
            _dist_both_mr, names="Tipo", values="Contratos",
            color="Tipo", color_discrete_map=COLORES_TIPO,
            title="Por número de contratos", hole=0.35,
            custom_data=["Monto_fmt"],
        )
        _fig_pA1.update_traces(
            textinfo="percent", textposition="inside",
            insidetextorientation="horizontal",
            textfont=dict(family="Noto Sans, sans-serif", size=13),
            hovertemplate="<b>%{label}</b><br>Contratos: %{value:,}<br>Monto: %{customdata[0]}<br>%{percent}<extra></extra>",
        )
        _fig_pA1.update_layout(
            font=plotly_font(), title_font_color=IMSS_VERDE_OSC,
            legend=dict(orientation="v", font=dict(family="Noto Sans, sans-serif", size=11),
                        x=1.01, y=0.5, xanchor="left"),
            margin=dict(r=160),
        )
        _col_pA1.plotly_chart(_fig_pA1, use_container_width=True)

    with _col_pA2:
        _dist_monto_mr = _dff_sel.copy()
        _dist_monto_mr["Tipo Simplificado"] = _dist_monto_mr["Tipo Simplificado"].replace(_TIPO_RENAME_MR)
        _dist_monto_mr = (
            _dist_monto_mr.groupby("Tipo Simplificado")["Importe DRC"]
            .sum().reset_index()
        )
        _dist_monto_mr.columns = ["Tipo", "Monto"]
        _dist_monto_mr["Monto_fmt"] = _dist_monto_mr["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
        _fig_pA2 = px.pie(
            _dist_monto_mr, names="Tipo", values="Monto",
            color="Tipo", color_discrete_map=COLORES_TIPO,
            title="Por monto total", hole=0.35,
            custom_data=["Monto_fmt"],
        )
        _fig_pA2.update_traces(
            texttemplate="<b>%{percent:.1%}</b><br>%{customdata[0]}",
            textposition="outside",
            textfont=dict(family="Noto Sans, sans-serif", size=12),
            hovertemplate="<b>%{label}</b><br>Monto: $%{value:,.0f}<br>%{percent}<extra></extra>",
        )
        _fig_pA2.update_layout(
            font=plotly_font(), title_font_color=IMSS_VERDE_OSC,
            legend=dict(orientation="v", font=dict(family="Noto Sans, sans-serif", size=11),
                        x=1.01, y=0.5, xanchor="left"),
            margin=dict(r=160),
        )
        _col_pA2.plotly_chart(_fig_pA2, use_container_width=True)

    st.divider()

    # ─────────────────────────────────────────────────────────────────────
    # VIZ B — Proveedores por Monto Contratado
    # ─────────────────────────────────────────────────────────────────────
    st.subheader("🏢 Proveedores por Monto Contratado")
    _prov_mr = (
        _dff_sel.groupby(["Proveedor o contratista", "rfc"])["Importe DRC"]
        .sum().reset_index()
        .sort_values("Importe DRC", ascending=False)
        .head(15)
    )
    _prov_mr["Share_mr"]   = (_prov_mr["Importe DRC"] / _monto_mr * 100) if _monto_mr > 0 else 0
    _prov_mr["Share_fmt"]  = _prov_mr["Share_mr"].apply(lambda x: f"{x:.1f}%")
    _prov_mr["Prov_c"]     = _prov_mr["Proveedor o contratista"].apply(
        lambda s: str(s)[:48] + "…" if len(str(s)) > 48 else str(s)
    )
    _prov_mr = _prov_mr.sort_values("Importe DRC", ascending=True)
    _prov_mr["Monto_fmt"] = _prov_mr["Importe DRC"].apply(lambda x: f"${x/1e6:,.1f} M")

    _fig_pB = px.treemap(
        _prov_mr,
        path=["Prov_c"],
        values="Importe DRC",
        color="Share_mr",
        color_continuous_scale=[[0, IMSS_VERDE_OSC], [0.5, IMSS_VERDE], [1, IMSS_ORO_CLARO]],
        range_color=[0, 100],
        custom_data=["Proveedor o contratista", "rfc", "Share_fmt", "Monto_fmt"],
    )
    _fig_pB.update_traces(
        texttemplate="<b>%{label}</b><br>%{customdata[3]}<br>%{customdata[2]}",
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "RFC: %{customdata[1]}<br>"
            "% de la UC: %{customdata[2]}<br>"
            "Monto: %{customdata[3]}<extra></extra>"
        ),
        textfont=dict(family="Noto Sans, sans-serif", size=12),
    )
    _fig_pB.update_layout(
        font=plotly_font(),
        height=460,
        margin=dict(l=10, r=10, t=20, b=10),
        coloraxis_colorbar=dict(title="% del total<br>de la UC", ticksuffix="%"),
    )
    st.plotly_chart(_fig_pB, use_container_width=True)

    st.divider()

    # ─────────────────────────────────────────────────────────────────────
    # VIZ C — Gasto por Partida Presupuestaria (CUCoP)
    # ─────────────────────────────────────────────────────────────────────
    st.subheader("💰 Gasto por Partida Presupuestaria (CUCoP)")
    _fig_pC = None  # initialized for PDF export; set below if CUCoP data available
    if len(df_cucop) > 0 and "Partida específica" in _dff_sel.columns:
        _exp_mr = _dff_sel.copy()
        _exp_mr["_lista_mr"] = _exp_mr["Partida específica"].str.split(",")
        _exp_mr["_n_mr"]     = _exp_mr["_lista_mr"].apply(len)
        _exp_mr["Importe DRC"] = _exp_mr["Importe DRC"] / _exp_mr["_n_mr"]
        _exp_mr = _exp_mr.explode("_lista_mr")
        _exp_mr["Partida específica"] = _exp_mr["_lista_mr"].str.strip().str.zfill(5)
        _exp_mr = _exp_mr.drop(columns=["_lista_mr", "_n_mr"])

        _cucop_cols = [
            c for c in ["PARTIDA ESPECÍFICA", "DESC. PARTIDA ESPECÍFICA",
                         "PARTIDA GENÉRICA", "DESC. PARTIDA GENÉRICA",
                         "CONCEPTO", "DESC. CONCEPTO", "CAPÍTULO", "DESC. CAPÍTULO"]
            if c in df_cucop.columns
        ]
        _cucop_mr = _exp_mr.merge(
            df_cucop[_cucop_cols].drop_duplicates(subset=["PARTIDA ESPECÍFICA"]),
            left_on="Partida específica", right_on="PARTIDA ESPECÍFICA", how="left",
        )

        _nivel_mr = st.radio(
            "Agrupar por:",
            ["Partida específica", "Partida genérica", "Capítulo"],
            horizontal=True, key="mr_nivel_cucop",
        )

        if _nivel_mr == "Partida específica":
            _cucop_mr["_etq_mr"] = (
                _cucop_mr["Partida específica"] + " — " +
                _cucop_mr.get("DESC. PARTIDA ESPECÍFICA", pd.Series()).fillna("Sin descripción")
            )
        elif _nivel_mr == "Partida genérica":
            _cucop_mr["_etq_mr"] = (
                _cucop_mr["Partida específica"].str[:4] + " — " +
                _cucop_mr.get("DESC. PARTIDA GENÉRICA", pd.Series()).fillna("Sin descripción")
            )
        else:
            _cucop_mr["_etq_mr"] = (
                _cucop_mr["Partida específica"].str[:1] + "000 — " +
                _cucop_mr.get("DESC. CAPÍTULO", pd.Series()).fillna("Sin descripción")
            )

        _gasto_mr = (
            _cucop_mr.groupby("_etq_mr")["Importe DRC"]
            .sum().sort_values(ascending=False).head(15).reset_index()
        )
        _gasto_mr.columns = ["Partida", "Monto"]
        _gasto_mr["Etq_c"] = _gasto_mr["Partida"].apply(
            lambda s: s if len(str(s)) <= 55 else str(s)[:55] + "…"
        )
        _gasto_mr = _gasto_mr.sort_values("Monto", ascending=True)

        if len(_gasto_mr) > 0:
            _gasto_mr["Monto_fmt"] = _gasto_mr["Monto"].apply(lambda v: f"${v/1e6:.1f} M")
            _fig_pC = px.treemap(
                _gasto_mr,
                path=["Etq_c"],
                values="Monto",
                color="Monto",
                color_continuous_scale=[[0, IMSS_VERDE_OSC], [1, IMSS_VERDE]],
                custom_data=["Partida", "Monto_fmt"],
            )
            _fig_pC.update_traces(
                texttemplate="<b>%{label}</b><br>%{customdata[1]}",
                hovertemplate="<b>%{customdata[0]}</b><br>Monto: %{customdata[1]}<extra></extra>",
                textfont=dict(family="Noto Sans, sans-serif", size=12),
            )
            _fig_pC.update_layout(
                font=plotly_font(),
                height=460,
                margin=dict(l=10, r=10, t=20, b=10),
                coloraxis_showscale=False,
            )
            st.plotly_chart(_fig_pC, use_container_width=True)
        else:
            st.info("ℹ️ Sin datos de partidas disponibles para esta UC.")
    else:
        st.info("ℹ️ Catálogo CUCoP no disponible o sin datos de partidas.")

    st.divider()

    # ─────────────────────────────────────────────────────────────────────
    # VIZ D — Distribución de Proveedores en la UC
    # ─────────────────────────────────────────────────────────────────────
    st.subheader("🔍 Distribución de Proveedores en la UC")

    _prov_dist_mr = (
        _dff_sel.groupby(["Proveedor o contratista", "rfc"])
        .agg(Monto=("Importe DRC", "sum"), Contratos=("Importe DRC", "count"))
        .reset_index()
        .sort_values("Monto", ascending=False)
        .reset_index(drop=True)
    )
    _prov_dist_mr["Share"] = (
        _prov_dist_mr["Monto"] / _monto_mr if _monto_mr > 0 else 0
    )

    # HHI
    _hhi_mr = float((_prov_dist_mr["Share"] ** 2).sum() * 10000)
    if _hhi_mr < 1500:
        _nivel_hhi_mr = "Competitivo (< 1,500)"
    elif _hhi_mr < 2500:
        _nivel_hhi_mr = "Moderadamente concentrado"
    else:
        _nivel_hhi_mr = "Altamente concentrado (> 2,500)"

    _kh1, _kh2, _kh3 = st.columns(3)
    _kh1.metric("Núm. proveedores",        f"{len(_prov_dist_mr):,}")
    _kh2.metric("Índice HHI",              f"{_hhi_mr:,.0f}",
                delta=_nivel_hhi_mr, delta_color="off")
    _kh3.metric("% del proveedor principal",
                f"{_prov_dist_mr['Share'].iloc[0]*100:.1f}%"
                if len(_prov_dist_mr) > 0 else "—")

    # Donut full-width: top 8 + Otros
    _pie_top8_mr  = _prov_dist_mr.head(8).copy()
    _monto_otros_mr = _prov_dist_mr.iloc[8:]["Monto"].sum()
    if _monto_otros_mr > 0:
        _otros_mr = pd.DataFrame({
            "Proveedor o contratista": ["Otros proveedores"], "rfc": ["—"],
            "Monto": [_monto_otros_mr], "Contratos": [0],
            "Share": [_monto_otros_mr / _monto_mr if _monto_mr > 0 else 0],
        })
        _pie_data_mr = pd.concat([_pie_top8_mr, _otros_mr], ignore_index=True)
    else:
        _pie_data_mr = _pie_top8_mr.copy()

    _pie_data_mr["Prov_c"] = _pie_data_mr["Proveedor o contratista"].apply(
        lambda s: str(s)[:42] + "…" if len(str(s)) > 42 else str(s)
    )
    _pie_palette_mr = (
        [IMSS_ROJO if len(_prov_dist_mr) > 0 and _prov_dist_mr["Share"].iloc[0] >= 0.50 else IMSS_VERDE]
        + [IMSS_ORO, "#5D8AA8", IMSS_ORO_CLARO, IMSS_VERDE_OSC,
           "#8B5CF6", "#6B7280", IMSS_GRIS, IMSS_NEGRO]
    )[:len(_pie_data_mr)]

    _fig_donut_mr = px.pie(
        _pie_data_mr, names="Prov_c", values="Monto",
        color="Prov_c", color_discrete_sequence=_pie_palette_mr,
        title="Distribución del gasto por proveedor", hole=0.42,
    )
    _fig_donut_mr.update_traces(
        texttemplate="<b>%{label}</b><br>%{percent:.1%}",
        textposition="outside",
        hovertemplate="<b>%{label}</b><br>Monto: $%{value:,.0f}<br>%{percent:.1%}<extra></extra>",
    )
    _fig_donut_mr.update_layout(
        font=plotly_font(), showlegend=False,
        plot_bgcolor="white", paper_bgcolor="white", height=520,
    )
    st.plotly_chart(_fig_donut_mr, use_container_width=True)


    # ─────────────────────────────────────────────────────────────────────
    # PDF generation (triggered by flag set at top of section)
    # ─────────────────────────────────────────────────────────────────────
    _rfc_norm_mr    = _dff_sel["rfc"].astype(str).str.strip().str.upper()
    _riesgos_activos = []  # list of (tipo_str, data)
    _riesgos_limpios = []  # list of label strings

    # \u2500\u2500 A. Inhabilitados SABG (criterio Art. 46 LAASSP: fecha de fallo) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    try:
        _df_san_mr   = cargar_sancionados()
        _rfcs_san_mr = set(_df_san_mr["RFC"].astype(str).str.strip().str.upper())
        _hits_san    = _dff_sel[_rfc_norm_mr.isin(_rfcs_san_mr)].copy()
        if len(_hits_san) > 0:
            _hits_san["_rfc_up"] = _hits_san["rfc"].astype(str).str.strip().str.upper()
            # Incluir Inicio inhabilitaci\u00f3n para aplicar el criterio de fecha de fallo
            _san_lkp = _df_san_mr[["RFC", "Empresa", "Nivel de Riesgo", "Inicio inhabilitaci\u00f3n"]].copy()
            _san_lkp["RFC"] = _san_lkp["RFC"].astype(str).str.strip().str.upper()
            _hits_san = _hits_san.merge(
                _san_lkp.rename(columns={"RFC": "_rfc_up"}), on="_rfc_up", how="left"
            )
            # Criterio Art. 46 LAASSP: fecha_fallo >= inicio_inhabilitaci\u00f3n \u2192 violaci\u00f3n
            # (misma l\u00f3gica que Secci\u00f3n 4 de Indicadores de Riesgo)
            if "Fecha de fallo" in _hits_san.columns:
                _hits_san["_fallo_d"] = pd.to_datetime(
                    _hits_san["Fecha de fallo"], errors="coerce"
                ).dt.date
            else:
                _hits_san["_fallo_d"] = None
            _hits_san["_inicio_d"] = pd.to_datetime(
                _hits_san["Inicio inhabilitaci\u00f3n"], errors="coerce"
            ).dt.date

            def _ajustar_nivel_mr(row):
                nivel = row["Nivel de Riesgo"]
                if nivel != "\U0001f534 Riesgo cr\u00edtico \u2014 Inhabilitaci\u00f3n vigente":
                    return nivel
                ff, ini = row["_fallo_d"], row["_inicio_d"]
                if pd.isna(ff) or pd.isna(ini):
                    return "\u26aa Sin fecha de fallo (verificar manualmente)"
                return (
                    "\U0001f534 Riesgo cr\u00edtico \u2014 Inhabilitaci\u00f3n vigente"
                    if ff >= ini else
                    "\u26ab Fallo anterior a inhabilitaci\u00f3n (sin violaci\u00f3n)"
                )

            _hits_san["Nivel de Riesgo"] = _hits_san.apply(_ajustar_nivel_mr, axis=1)
            # Excluir los no accionables, igual que la tabla de la Secci\u00f3n 4
            _hits_san = _hits_san[
                ~_hits_san["Nivel de Riesgo"].isin([
                    "\u26ab Fallo anterior a inhabilitaci\u00f3n (sin violaci\u00f3n)",
                    "\u26aa Sin fecha de fallo (verificar manualmente)",
                ])
            ].copy()

            if len(_hits_san) > 0:
                _riesgos_activos.append(("san", _hits_san))
            else:
                _riesgos_limpios.append(
                    "Inhabilitados SABG (fallos previos a la inhabilitaci\u00f3n \u2014 sin violaci\u00f3n activa)"
                )
        else:
            _riesgos_limpios.append("Inhabilitados SABG")
    except Exception:
        _riesgos_limpios.append("Inhabilitados SABG")

    # \u2500\u2500 B. EFOS Art. 69-B \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    try:
        _df_efos_mr   = cargar_efos()
        _rfcs_edef_mr = set(_df_efos_mr[_df_efos_mr["Situaci\u00f3n del contribuyente"] == "Definitivo"]["RFC"])
        _rfcs_epre_mr = set(_df_efos_mr[_df_efos_mr["Situaci\u00f3n del contribuyente"] == "Presunto"]["RFC"])
        _hits_edef = _dff_sel[_rfc_norm_mr.isin(_rfcs_edef_mr)].copy()
        _hits_epre = _dff_sel[_rfc_norm_mr.isin(_rfcs_epre_mr)].copy()
        if len(_hits_edef) > 0:
            _riesgos_activos.append(("efos_def", _hits_edef))
        else:
            _riesgos_limpios.append("EFOS Definitivo")
        if len(_hits_epre) > 0:
            _riesgos_activos.append(("efos_pre", _hits_epre))
        else:
            _riesgos_limpios.append("EFOS Presunto")
    except Exception:
        _riesgos_limpios.extend(["EFOS Definitivo", "EFOS Presunto"])

    # \u2500\u2500 C. Empresas de reciente creaci\u00f3n \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    def _parse_rfc_mr(s):
        s2 = str(s).strip().upper()
        if not _re_mr.match(r'^[A-Z\u00d1&]{3}[0-9]{6}[A-Z0-9]{3}$', s2):
            return None
        yy, mm, dd = int(s2[3:5]), int(s2[5:7]), int(s2[7:9])
        if not (1 <= mm <= 12 and 1 <= dd <= 31):
            return None
        yr = 2000 + yy if yy <= 30 else 1900 + yy
        try:
            return _date_mr(yr, mm, dd)
        except Exception:
            return None

    _dff_rc_mr = _dff_sel.copy()
    _dff_rc_mr["_f_ini_rc"] = pd.to_datetime(
        _dff_rc_mr["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
    ).dt.date
    _dff_rc_mr["_f_rfc_mr"] = _dff_rc_mr["rfc"].apply(_parse_rfc_mr)
    _dff_rc_mr = _dff_rc_mr.dropna(subset=["_f_ini_rc", "_f_rfc_mr"])
    _dff_rc_mr["_edad_mr"] = (
        _dff_rc_mr["_f_ini_rc"] - _dff_rc_mr["_f_rfc_mr"]
    ).apply(lambda x: x.days if hasattr(x, "days") else None)
    _rc_hits = _dff_rc_mr[
        (_dff_rc_mr["_edad_mr"] >= 0) & (_dff_rc_mr["_edad_mr"] < 365)
    ].copy()
    if len(_rc_hits) > 0:
        _riesgos_activos.append(("reciente", _rc_hits))
    else:
        _riesgos_limpios.append("Reciente creaci\u00f3n")

    # \u2500\u2500 D. Contratos cerca del umbral legal \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    _umbrales_mr = cargar_umbrales_pef()
    if _umbrales_mr:
        _TIPOS_AD_MR   = {"Adjudicaci\u00f3n Directa", "Adjudicaci\u00f3n Directa \u2014 Fr. I"}
        _TIPOS_I3P_MR  = {"Invitaci\u00f3n a 3 personas"}
        _TIPOS_SERV_MR = {"SERVICIOS", "SERVICIOS RELACIONADOS CON LA OBRA", "ARRENDAMIENTOS"}
        _dff_umr = _dff_sel[
            _dff_sel["Tipo Simplificado"].isin(_TIPOS_AD_MR | _TIPOS_I3P_MR)
        ].copy()
        _dff_umr["_f_umr"]    = pd.to_datetime(_dff_umr["Fecha de inicio del contrato"], dayfirst=True, errors="coerce")
        _dff_umr["_a_umr"]    = _dff_umr["_f_umr"].dt.year
        _dff_umr["_ley_umr"]  = _dff_umr["Ley"].astype(str).str.strip().str.upper() if "Ley" in _dff_umr.columns else "LAASSP"
        _dff_umr["_cont_umr"] = _dff_umr["Tipo de contrataci\u00f3n"].astype(str).str.strip().str.upper() if "Tipo de contrataci\u00f3n" in _dff_umr.columns else "ADQUISICIONES"
        _dff_umr["_proc_umr"] = _dff_umr["Tipo Simplificado"]
        _dff_umr["_umbral_v"] = float("nan")
        for _a_u, _th_u in _umbrales_mr.items():
            _ma_u = _dff_umr["_a_umr"] == _a_u
            _dff_umr.loc[_ma_u & _dff_umr["_ley_umr"].eq("LAASSP") & _dff_umr["_proc_umr"].isin(_TIPOS_AD_MR),   "_umbral_v"] = _th_u["ad_laassp"]
            _dff_umr.loc[_ma_u & _dff_umr["_ley_umr"].eq("LAASSP") & _dff_umr["_proc_umr"].isin(_TIPOS_I3P_MR),  "_umbral_v"] = _th_u["i3p_laassp"]
            _dff_umr.loc[_ma_u & _dff_umr["_ley_umr"].eq("LOPSRM") & _dff_umr["_proc_umr"].isin(_TIPOS_AD_MR)  & _dff_umr["_cont_umr"].eq("OBRA P\u00daBLICA"),   "_umbral_v"] = _th_u["ad_obra_lopsrm"]
            _dff_umr.loc[_ma_u & _dff_umr["_ley_umr"].eq("LOPSRM") & _dff_umr["_proc_umr"].isin(_TIPOS_AD_MR)  & _dff_umr["_cont_umr"].isin(_TIPOS_SERV_MR),       "_umbral_v"] = _th_u["ad_serv_lopsrm"]
            _dff_umr.loc[_ma_u & _dff_umr["_ley_umr"].eq("LOPSRM") & _dff_umr["_proc_umr"].isin(_TIPOS_I3P_MR) & _dff_umr["_cont_umr"].eq("OBRA P\u00daBLICA"),   "_umbral_v"] = _th_u["i3p_obra_lopsrm"]
            _dff_umr.loc[_ma_u & _dff_umr["_ley_umr"].eq("LOPSRM") & _dff_umr["_proc_umr"].isin(_TIPOS_I3P_MR) & _dff_umr["_cont_umr"].isin(_TIPOS_SERV_MR),       "_umbral_v"] = _th_u["i3p_serv_lopsrm"]
        _dff_umr["_pct_v"] = (_dff_umr["Importe DRC"] / _dff_umr["_umbral_v"]) * 100
        _umr_hits = _dff_umr[
            (_dff_umr["_pct_v"] >= 90) & (_dff_umr["_pct_v"] < 100) & _dff_umr["_umbral_v"].notna()
        ].copy().sort_values("_pct_v", ascending=False)
        if len(_umr_hits) > 0:
            _riesgos_activos.append(("umbral", _umr_hits))
        else:
            _riesgos_limpios.append("Contratos cerca del umbral")
    else:
        _riesgos_limpios.append("Contratos cerca del umbral (sin datos PEF)")

    # \u2500\u2500 E. Fragmentaci\u00f3n \u2014 mismo d\u00eda \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    _dff_ad_fr = _dff_sel[_dff_sel["Tipo Simplificado"].isin(
        ["Adjudicaci\u00f3n Directa", "Adjudicaci\u00f3n Directa \u2014 Fr. I"]
    )].copy()
    _dff_ad_fr["_f_fr"] = pd.to_datetime(
        _dff_ad_fr["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
    )
    _dff_ad_fr = _dff_ad_fr[_dff_ad_fr["_f_fr"].notna()].copy()
    _dff_ad_fr["_fstr_fr"] = _dff_ad_fr["_f_fr"].dt.date.astype(str)
    # Excluir contratos que comparten número de procedimiento (mismas partidas, no fragmentación)
    _col_proc_fr = "Número de procedimiento"
    if _col_proc_fr in _dff_ad_fr.columns:
        _proc_vals_fr   = _dff_ad_fr[_col_proc_fr].astype(str).str.strip()
        _proc_counts_fr = _proc_vals_fr.value_counts()
        _procs_rep_fr   = set(_proc_counts_fr[_proc_counts_fr >= 2].index) - {"", "nan", "NaN"}
        if _procs_rep_fr:
            _dff_ad_fr = _dff_ad_fr[~_proc_vals_fr.isin(_procs_rep_fr)].copy()
    if len(_dff_ad_fr) > 0:
        _has_proc_col_fr = _col_proc_fr in _dff_ad_fr.columns
        _agg_spec_fr = {
            "Contratos":   ("Importe DRC", "count"),
            "Monto_total": ("Importe DRC", "sum"),
            "Monto_min":   ("Importe DRC", "min"),
            "Monto_max":   ("Importe DRC", "max"),
        }
        if _has_proc_col_fr:
            _agg_spec_fr["Procs_distintos"] = (_col_proc_fr, "nunique")
        _g_fr = (
            _dff_ad_fr.groupby(["rfc", "Proveedor o contratista", "Nombre de la UC", "_fstr_fr"])
            .agg(**_agg_spec_fr)
            .reset_index()
        )
        _mask_frag_fr = _g_fr["Contratos"] >= 3
        if _has_proc_col_fr and "Procs_distintos" in _g_fr.columns:
            _mask_frag_fr = _mask_frag_fr & (_g_fr["Procs_distintos"] >= 2)
        _g_fr = _g_fr[_mask_frag_fr].sort_values("Contratos", ascending=False)
        if len(_g_fr) > 0:
            # Guardar también _dff_ad_fr para poder mostrar contratos individuales con links
            _riesgos_activos.append(("frag_dia", (_g_fr, _dff_ad_fr)))
        else:
            _riesgos_limpios.append("Fragmentaci\u00f3n (mismo d\u00eda)")
    else:
        _riesgos_limpios.append("Fragmentaci\u00f3n (mismo d\u00eda)")

    # \u2500\u2500 F. Brecha LP \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    if _monto_mr > 0 and _pct_lp_mr < 50:
        _riesgos_activos.append(("brecha_lp", _pct_lp_mr))
    elif _monto_mr > 0:
        _riesgos_limpios.append(f"LP {_pct_lp_mr:.1f}% (mayoritario)")

    # \u2500\u2500 G. Alta concentraci\u00f3n en AD \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    _dff_ad_cg = _dff_sel[_dff_sel["Tipo Simplificado"].isin(
        ["Adjudicaci\u00f3n Directa", "Adjudicaci\u00f3n Directa \u2014 Fr. I"]
    )]
    if len(_dff_ad_cg) > 0 and _monto_mr > 0:
        _ser_cg = _dff_ad_cg.groupby(["rfc", "Proveedor o contratista"])["Importe DRC"].sum()
        _top_cg_pct = _ser_cg.max() / _monto_mr * 100
        if _top_cg_pct >= 40:
            _riesgos_activos.append(("conc_ad", (_top_cg_pct, _ser_cg)))
        else:
            _riesgos_limpios.append("Concentraci\u00f3n en proveedor AD")
    else:
        _riesgos_limpios.append("Concentraci\u00f3n en proveedor AD")

    # \u2500\u2500 H. Excepci\u00f3n Art. 55 > 30 % \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    if "Descripci\u00f3n excepci\u00f3n" in _dff_sel.columns:
        _mask_exc_mr = (
            _dff_sel["Descripci\u00f3n excepci\u00f3n"].notna()
            & _dff_sel["Descripci\u00f3n excepci\u00f3n"].astype(str).str.strip().ne("")
        )
        _pct_exc_mr = _mask_exc_mr.mean() * 100
        if _pct_exc_mr > 30:
            _riesgos_activos.append(("exc_alto", _pct_exc_mr))
        else:
            _riesgos_limpios.append(f"Excepci\u00f3n Art. 55 ({_pct_exc_mr:.1f}% < 30%)")
    else:
        _riesgos_limpios.append("Excepci\u00f3n Art. 55")

    if st.session_state.get(_pdf_flag_key):
        with st.spinner("Generando PDF con visualizaciones\u2026 \u23f3"):
            try:
                from fpdf import FPDF as _FPDF_MR
                import io as _io_mr

                def _s(text):
                    """Sanitize string for fpdf Helvetica (latin-1 safe)."""
                    t = str(text)
                    t = (t.replace('\u2014', '--').replace('\u2013', '-')
                          .replace('\u2026', '...').replace('\u201c', '"')
                          .replace('\u201d', '"')
                          .replace('\u2018', "'")
                          .replace('\u2019', "'"))
                    return t.encode('latin-1', errors='replace').decode('latin-1')

                _pdf_obj = _FPDF_MR(orientation="L", unit="mm", format="A4")
                _pdf_obj.set_auto_page_break(auto=True, margin=12)
                _pdf_obj.set_margins(12, 12, 12)
                _pdf_obj.add_page()

                # ── Header ──
                _pdf_obj.set_font("Helvetica", "B", 15)
                _pdf_obj.set_text_color(11, 84, 69)
                _pdf_obj.cell(
                    0, 10, _s(f"Perfil UC -- {str(_label_mr)[:75]}"),
                    new_x="LMARGIN", new_y="NEXT",
                )
                _pdf_obj.set_font("Helvetica", "", 9)
                _pdf_obj.set_text_color(23, 27, 25)
                if _tipo_vista_mr == "UC espec\u00edfica":
                    _pdf_meta_str = _s(
                        f"Tipo UC: {_meta_tipo_mr or '--'}   |   "
                        f"Adscripcion: {_meta_adsc_mr or '--'}"
                    )
                else:
                    _pdf_meta_str = _s(f"Adscripcion: {_meta_adsc_mr}")
                _pdf_obj.cell(0, 6, _pdf_meta_str, new_x="LMARGIN", new_y="NEXT")
                _pdf_obj.ln(3)

                # ── KPIs ──
                _pdf_obj.set_font("Helvetica", "B", 11)
                _pdf_obj.set_text_color(11, 84, 69)
                _pdf_obj.cell(0, 7, "Numeralia General", new_x="LMARGIN", new_y="NEXT")
                _pdf_obj.set_text_color(23, 27, 25)
                _monto_pdf_str = (
                    f"${_monto_mr/1e9:,.2f} miles de millones MXN"
                    if _monto_mr >= 1e9
                    else f"${_monto_mr/1e6:,.1f} M MXN"
                )
                _kpi_labels_pdf = [
                    "Contratos", "Monto total",
                    "Proveedores unicos", "% Monto LP", "% Monto AD",
                ]
                _kpi_values_pdf = [
                    f"{_total_mr:,}", _monto_pdf_str,
                    f"{_n_prov_mr:,}", f"{_pct_lp_mr:.1f}%", f"{_pct_ad_mr:.1f}%",
                ]
                _cw_pdf = _pdf_obj.epw / 5
                for _lbl in _kpi_labels_pdf:
                    _pdf_obj.set_font("Helvetica", "B", 8)
                    _pdf_obj.cell(_cw_pdf, 5, _s(_lbl))
                _pdf_obj.ln(5)
                for _val in _kpi_values_pdf:
                    _pdf_obj.set_font("Helvetica", "", 10)
                    _pdf_obj.cell(_cw_pdf, 6, _s(_val))
                _pdf_obj.ln(10)

                _pdf_obj.ln(4)

                # ── Riesgos detectados en el PDF ──
                _pdf_obj.set_font("Helvetica", "B", 11)
                _pdf_obj.set_text_color(11, 84, 69)
                _pdf_obj.cell(0, 7, _s("Riesgos especificos detectados"), new_x="LMARGIN", new_y="NEXT")
                _pdf_obj.set_text_color(23, 27, 25)
                if not _riesgos_activos:
                    _pdf_obj.set_font("Helvetica", "", 9)
                    _pdf_obj.cell(0, 6, _s("Sin alertas de riesgo detectadas."), new_x="LMARGIN", new_y="NEXT")
                else:
                    _RISK_LABELS_PDF = {
                        "san":       "Proveedores sancionados (SABG)",
                        "efos_def":  "EFOS definitivo (Art. 69-B CFF)",
                        "efos_pre":  "EFOS presunto (Art. 69-B CFF)",
                        "reciente":  "Empresa de reciente creacion",
                        "umbral":    "Contratos cerca del umbral legal",
                        "frag_dia":  "Fragmentacion -- mismo dia",
                        "brecha_lp": "Baja proporcion de LP",
                        "conc_ad":   "Alta concentracion en AD",
                        "exc_alto":  "Excepcion Art. 55 (>30%)",
                    }
                    for _tr_p, _dr_p in _riesgos_activos:
                        _lbl_p = _RISK_LABELS_PDF.get(_tr_p, _tr_p)
                        if isinstance(_dr_p, tuple):
                            _g_p = _dr_p[0]
                            _line_p = f"  * {_lbl_p}: {len(_g_p)} grupo(s)"
                        elif isinstance(_dr_p, pd.DataFrame):
                            _m_p = pd.to_numeric(_dr_p.get("Importe DRC", pd.Series(dtype=float)), errors="coerce").sum()
                            _line_p = f"  * {_lbl_p}: {len(_dr_p)} contrato(s), ${_m_p/1e6:,.1f} M MXN"
                        elif isinstance(_dr_p, (int, float)):
                            _line_p = f"  * {_lbl_p}: {_dr_p:.1f}%"
                        else:
                            _line_p = f"  * {_lbl_p}"
                        _pdf_obj.set_font("Helvetica", "", 9)
                        _pdf_obj.cell(0, 5, _s(_line_p), new_x="LMARGIN", new_y="NEXT")
                _pdf_obj.ln(4)

                # ── Charts ──
                _charts_export = [
                    ("Distribucion por Tipo de Procedimiento (contratos)", _fig_pA1),
                    ("Distribucion por Tipo de Procedimiento (monto)", _fig_pA2),
                    ("Proveedores por Monto Contratado", _fig_pB),
                    ("Gasto por Partida Presupuestaria (CUCoP)", _fig_pC),
                    ("Distribucion de Proveedores en la UC (HHI)", _fig_donut_mr),
                ]
                for _chart_title, _chart_fig in _charts_export:
                    if _chart_fig is None:
                        continue
                    try:
                        _img_bytes = _chart_fig.to_image(
                            format="png", width=1100, height=440, scale=1.5
                        )
                        _pdf_obj.set_font("Helvetica", "B", 10)
                        _pdf_obj.set_text_color(11, 84, 69)
                        _pdf_obj.cell(
                            0, 6, _s(_chart_title),
                            new_x="LMARGIN", new_y="NEXT",
                        )
                        _pdf_obj.image(_io_mr.BytesIO(_img_bytes), w=_pdf_obj.epw)
                        _pdf_obj.ln(4)
                    except Exception:
                        pass

                # ── Footer ──
                _pdf_obj.set_font("Helvetica", "I", 8)
                _pdf_obj.set_text_color(134, 134, 136)
                _pdf_obj.cell(
                    0, 5,
                    _s("Division de Monitoreo de la Integridad Institucional -- IMSS | ComprasMX 2026"),
                    align="C",
                )

                st.session_state[_pdf_state_key] = bytes(_pdf_obj.output())
            except Exception as _e_pdf:
                st.error(f"\u26a0\ufe0f Error generando PDF: {_e_pdf}")
        st.session_state[_pdf_flag_key] = False
        st.rerun()

    st.divider()

    st.subheader("\U0001f6a8 Riesgos espec\u00edficos detectados")

    # \u2500\u2500 Banner resumen \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    _n_act = len(_riesgos_activos)
    if _n_act == 0:
        st.success(f"\u2705 Sin alertas de riesgo detectadas para **{_label_mr}**.")
    elif _n_act <= 2:
        st.warning(f"\U0001f7e0 **{_n_act} categor\u00eda(s) de riesgo activa(s)** para {_label_mr}.")
    else:
        st.error(f"\U0001f534 **{_n_act} categor\u00edas de riesgo activas** para {_label_mr}.")

    # \u2500\u2500 Renderizar riesgos activos \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    for _tr, _dr in _riesgos_activos:

        if _tr == "san":
            _m_r = pd.to_numeric(_dr["Importe DRC"], errors="coerce").sum()
            _n_crit_s = (_dr["Nivel de Riesgo"] == "\U0001f534 Riesgo cr\u00edtico \u2014 Inhabilitaci\u00f3n vigente").sum()
            _n_alto_s = (_dr["Nivel de Riesgo"] == "\U0001f7e0 Riesgo alto \u2014 Inhabilitaci\u00f3n suspendida judicialmente").sum()
            _icon_san = "\U0001f534" if _n_crit_s > 0 else ("\U0001f7e0" if _n_alto_s > 0 else "\U0001f7e1")
            with st.expander(
                f"{_icon_san} Proveedores inhabilitados (SABG) \u2014 {len(_dr):,} contrato(s) \u00b7 "
                f"${_m_r/1e6:,.1f} M MXN", expanded=True
            ):
                if _n_crit_s > 0:
                    st.error(
                        f"**{_n_crit_s} contrato(s) con inhabilitaci\u00f3n vigente** \u2014 posible violaci\u00f3n "
                        f"a la LAASSP (Art. 46): el fallo se emiti\u00f3 despu\u00e9s del inicio de la inhabilitaci\u00f3n."
                    )
                if _n_alto_s > 0:
                    st.warning(
                        f"**{_n_alto_s} contrato(s) con inhabilitaci\u00f3n suspendida judicialmente** \u2014 "
                        f"el proceso sancionador sigue abierto."
                    )
                if _n_crit_s == 0 and _n_alto_s == 0:
                    st.info(
                        "Contratos con proveedores que tienen historial de inhabilitaci\u00f3n "
                        "(la sanci\u00f3n ya concluy\u00f3 al momento del fallo)."
                    )
                st.caption(
                    "\u2139\ufe0f Criterio aplicado: Art. 46 LAASSP \u2014 solo se consideran violaciones "
                    "los contratos cuyo fallo fue emitido durante la vigencia de la inhabilitaci\u00f3n. "
                    "Contratos con fallo previo a la inhabilitaci\u00f3n han sido excluidos."
                )
                _cols_s = [c for c in [
                    "Fecha de inicio del contrato", "Nombre de la UC",
                    "Proveedor o contratista", "Empresa", "Nivel de Riesgo",
                    "Importe DRC", "Direcci\u00f3n del anuncio"
                ] if c in _dr.columns]
                _disp_s = _dr[_cols_s].copy()
                _disp_s["Importe"] = pd.to_numeric(_disp_s["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _disp_s = _disp_s.drop(columns=["Importe DRC"], errors="ignore").rename(
                    columns={"Proveedor o contratista": "Proveedor"}
                )
                _disp_s.index = range(1, len(_disp_s) + 1)
                st.dataframe(_disp_s, use_container_width=True, column_config={
                    "Direcci\u00f3n del anuncio": st.column_config.LinkColumn("\U0001f517 ComprasMX", display_text="Ver contrato")
                })

        elif _tr == "efos_def":
            _m_r = pd.to_numeric(_dr["Importe DRC"], errors="coerce").sum()
            with st.expander(
                f"\U0001f534 EFOS definitivo (Art. 69-B CFF) \u2014 {len(_dr):,} contrato(s) \u00b7 "
                f"${_m_r/1e6:,.1f} M MXN", expanded=True
            ):
                st.error(
                    "Contratos con proveedores clasificados como EFOS **definitivos** por el SAT "
                    "(operaciones simuladas confirmadas). Mayor nivel de riesgo fiscal."
                )
                _cols_ed = [c for c in [
                    "Fecha de inicio del contrato", "Nombre de la UC",
                    "Proveedor o contratista", "Importe DRC", "Direcci\u00f3n del anuncio"
                ] if c in _dr.columns]
                _disp_ed = _dr[_cols_ed].copy()
                _disp_ed["Importe"] = pd.to_numeric(_disp_ed["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _disp_ed = _disp_ed.drop(columns=["Importe DRC"], errors="ignore").rename(
                    columns={"Proveedor o contratista": "Proveedor"}
                )
                _disp_ed.index = range(1, len(_disp_ed) + 1)
                st.dataframe(_disp_ed, use_container_width=True, column_config={
                    "Direcci\u00f3n del anuncio": st.column_config.LinkColumn("\U0001f517 ComprasMX", display_text="Ver contrato")
                })

        elif _tr == "efos_pre":
            _m_r = pd.to_numeric(_dr["Importe DRC"], errors="coerce").sum()
            with st.expander(
                f"\U0001f7e1 EFOS presunto (Art. 69-B CFF) \u2014 {len(_dr):,} contrato(s) \u00b7 "
                f"${_m_r/1e6:,.1f} M MXN"
            ):
                st.warning(
                    "Contratos con proveedores clasificados como EFOS **presuntos** por el SAT "
                    "(proceso en curso, a\u00fan no definitivo)."
                )
                _cols_ep = [c for c in [
                    "Fecha de inicio del contrato", "Nombre de la UC",
                    "Proveedor o contratista", "Importe DRC", "Direcci\u00f3n del anuncio"
                ] if c in _dr.columns]
                _disp_ep = _dr[_cols_ep].copy()
                _disp_ep["Importe"] = pd.to_numeric(_disp_ep["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _disp_ep = _disp_ep.drop(columns=["Importe DRC"], errors="ignore").rename(
                    columns={"Proveedor o contratista": "Proveedor"}
                )
                _disp_ep.index = range(1, len(_disp_ep) + 1)
                st.dataframe(_disp_ep, use_container_width=True, column_config={
                    "Direcci\u00f3n del anuncio": st.column_config.LinkColumn("\U0001f517 ComprasMX", display_text="Ver contrato")
                })

        elif _tr == "reciente":
            _m_r = pd.to_numeric(_dr["Importe DRC"], errors="coerce").sum()
            with st.expander(
                f"\U0001f7e0 Empresa de reciente creaci\u00f3n \u2014 {len(_dr):,} contrato(s) \u00b7 "
                f"${_m_r/1e6:,.1f} M MXN"
            ):
                st.warning(
                    "Contratos adjudicados a empresas con **menos de un a\u00f1o** de constituci\u00f3n "
                    "al inicio del contrato (riesgo de empresa creada ad hoc)."
                )
                _cols_rc = [c for c in [
                    "Fecha de inicio del contrato", "Nombre de la UC",
                    "Proveedor o contratista", "_edad_mr", "Importe DRC", "Direcci\u00f3n del anuncio"
                ] if c in _dr.columns]
                _disp_rc = _dr[_cols_rc].copy()
                _disp_rc["Antig\u00fcedad (d\u00edas)"] = _disp_rc["_edad_mr"].apply(
                    lambda x: f"{int(x):,}" if pd.notna(x) else ""
                )
                _disp_rc["Importe"] = pd.to_numeric(_disp_rc["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _disp_rc = _disp_rc.drop(columns=["Importe DRC", "_edad_mr"], errors="ignore").rename(
                    columns={"Proveedor o contratista": "Proveedor"}
                )
                _disp_rc.index = range(1, len(_disp_rc) + 1)
                st.dataframe(_disp_rc, use_container_width=True, column_config={
                    "Direcci\u00f3n del anuncio": st.column_config.LinkColumn("\U0001f517 ComprasMX", display_text="Ver contrato")
                })

        elif _tr == "umbral":
            _m_r = pd.to_numeric(_dr["Importe DRC"], errors="coerce").sum()
            with st.expander(
                f"\U0001f6a6 Contratos cerca del umbral legal \u2014 {len(_dr):,} contrato(s) \u00b7 "
                f"${_m_r/1e6:,.1f} M MXN"
            ):
                st.warning(
                    "Contratos con importe entre el **90 % y 100 %** del umbral legal "
                    "(Art. 55 LAASSP / Art. 43 LOPSRM). Posible fraccionamiento para "
                    "evadir licitaci\u00f3n p\u00fablica."
                )
                _cols_um = [c for c in [
                    "Fecha de inicio del contrato", "Nombre de la UC",
                    "Proveedor o contratista", "Tipo Simplificado",
                    "Importe DRC", "_umbral_v", "_pct_v", "Direcci\u00f3n del anuncio"
                ] if c in _dr.columns]
                _disp_um = _dr[_cols_um].copy()
                _disp_um["Importe"] = pd.to_numeric(_disp_um["Importe DRC"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _disp_um["Umbral legal"] = _disp_um["_umbral_v"].apply(
                    lambda x: f"${x:,.0f}" if pd.notna(x) else ""
                )
                _disp_um["% del umbral"] = _disp_um["_pct_v"].apply(
                    lambda x: f"{x:.1f}%" if pd.notna(x) else ""
                )
                _disp_um = _disp_um.drop(
                    columns=["Importe DRC", "_umbral_v", "_pct_v"], errors="ignore"
                ).rename(columns={"Proveedor o contratista": "Proveedor", "Tipo Simplificado": "Tipo"})
                _disp_um.index = range(1, len(_disp_um) + 1)
                st.dataframe(_disp_um, use_container_width=True, column_config={
                    "Direcci\u00f3n del anuncio": st.column_config.LinkColumn("\U0001f517 ComprasMX", display_text="Ver contrato")
                })

        elif _tr == "frag_dia":
            _g_fr_d, _dff_ad_fr_d = _dr  # desempacar (grupos, contratos_individuales)
            with st.expander(
                f"\U0001f9e9 Fragmentaci\u00f3n \u2014 concentraci\u00f3n en el mismo d\u00eda \u2014 {len(_g_fr_d)} grupo(s)"
            ):
                st.warning(
                    "El mismo proveedor recibi\u00f3 **3 o m\u00e1s contratos** de adjudicaci\u00f3n directa "
                    "en la misma fecha \u2014 patr\u00f3n cl\u00e1sico de fragmentaci\u00f3n simult\u00e1nea."
                )
                _disp_fd = _g_fr_d.rename(columns={
                    "_fstr_fr": "Fecha",
                    "Proveedor o contratista": "Proveedor",
                    "Contratos": "# Contratos",
                    "Monto_total": "Monto total",
                    "Monto_min": "Monto m\u00edn",
                    "Monto_max": "Monto m\u00e1x",
                }).copy()
                for _mc in ["Monto total", "Monto m\u00edn", "Monto m\u00e1x"]:
                    if _mc in _disp_fd.columns:
                        _disp_fd[_mc] = pd.to_numeric(_disp_fd[_mc], errors="coerce").apply(
                            lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                        )
                _disp_fd = _disp_fd[[c for c in [
                    "Fecha", "Nombre de la UC", "Proveedor", "# Contratos",
                    "Monto total", "Monto m\u00edn", "Monto m\u00e1x"
                ] if c in _disp_fd.columns]]
                _disp_fd.index = range(1, len(_disp_fd) + 1)
                st.dataframe(_disp_fd, use_container_width=True)
                # Contratos individuales con links a ComprasMX
                st.markdown("**Contratos individuales:**")
                _cols_frd = ["rfc", "Nombre de la UC", "_fstr_fr"]
                _det_fr_d = _dff_ad_fr_d.merge(
                    _g_fr_d[_cols_frd], on=_cols_frd, how="inner"
                )
                _cols_det_fr = [c for c in [
                    "Fecha de inicio del contrato", "Nombre de la UC",
                    "Proveedor o contratista", "N\u00famero de procedimiento",
                    "Importe DRC", "Direcci\u00f3n del anuncio"
                ] if c in _det_fr_d.columns]
                _det_fr_d = _det_fr_d[_cols_det_fr].drop_duplicates().sort_values(
                    ["Nombre de la UC", "Proveedor o contratista",
                     "Fecha de inicio del contrato"]
                )
                _det_fr_d["Importe"] = pd.to_numeric(
                    _det_fr_d["Importe DRC"], errors="coerce"
                ).apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "")
                _det_fr_d = _det_fr_d.drop(columns=["Importe DRC"], errors="ignore")
                _det_fr_d.index = range(1, len(_det_fr_d) + 1)
                st.dataframe(
                    _det_fr_d, use_container_width=True,
                    column_config={
                        "Direcci\u00f3n del anuncio": st.column_config.LinkColumn(
                            "\U0001f517 ComprasMX", display_text="Ver contrato"
                        )
                    },
                )

        elif _tr == "brecha_lp":
            with st.expander(
                f"\U0001f4c9 Baja proporci\u00f3n de LP \u2014 {_dr:.1f}% del monto por licitaci\u00f3n p\u00fablica"
            ):
                st.warning(
                    f"El porcentaje del monto contratado mediante licitaci\u00f3n p\u00fablica es "
                    f"**{_dr:.1f}%**. La licitaci\u00f3n p\u00fablica es la regla general (Art. 35 LAASSP 2025); "
                    f"un porcentaje reducido puede indicar uso intensivo de procedimientos de excepci\u00f3n."
                )
                _tipo_dist_b = (
                    _dff_sel.groupby("Tipo Simplificado")["Importe DRC"].sum()
                    .reset_index().sort_values("Importe DRC", ascending=True)
                )
                _fig_b = go.Figure(go.Bar(
                    x=_tipo_dist_b["Importe DRC"] / 1e6,
                    y=_tipo_dist_b["Tipo Simplificado"],
                    orientation="h",
                    marker_color=[
                        IMSS_VERDE if t == "Licitaci\u00f3n P\u00fablica"
                        else IMSS_ROJO if "Adjudicaci\u00f3n" in t
                        else IMSS_ORO
                        for t in _tipo_dist_b["Tipo Simplificado"]
                    ],
                    text=(_tipo_dist_b["Importe DRC"] / 1e6).apply(lambda x: f"${x:,.1f} M"),
                    textposition="outside",
                ))
                _fig_b.update_layout(
                    title="Distribuci\u00f3n de monto por tipo de procedimiento",
                    xaxis_title="Monto (M MXN)", height=300,
                    margin=dict(l=200, r=80, t=40, b=30), font=plotly_font(),
                )
                st.plotly_chart(_fig_b, use_container_width=True)

        elif _tr == "conc_ad":
            _pct_cg, _ser_cg = _dr
            _top_cg_name = _ser_cg.sort_values(ascending=False).index[0][1] if len(_ser_cg) > 0 else "\u2014"
            with st.expander(
                f"\u26a0\ufe0f Alta concentraci\u00f3n en proveedor AD \u2014 {_pct_cg:.1f}% del monto total"
            ):
                st.warning(
                    f"El proveedor principal (**{_top_cg_name}**) concentra el "
                    f"**{_pct_cg:.1f}%** del monto total en adjudicaciones directas. "
                    f"Alta concentraci\u00f3n sugiere posible falta de competencia."
                )
                _top10_cg = _ser_cg.sort_values(ascending=False).head(10).reset_index()
                _top10_cg.columns = ["rfc", "Proveedor", "Monto AD"]
                _top10_cg["% del monto total"] = (_top10_cg["Monto AD"] / _monto_mr * 100).apply(
                    lambda x: f"{x:.1f}%"
                )
                _top10_cg["Monto AD"] = _top10_cg["Monto AD"].apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )
                _top10_cg = _top10_cg.drop(columns=["rfc"])
                _top10_cg.index = range(1, len(_top10_cg) + 1)
                st.dataframe(_top10_cg, use_container_width=True)

        elif _tr == "exc_alto":
            with st.expander(
                f"\u26a0\ufe0f Alto uso de excepciones (Art. 55 LAASSP) \u2014 {_dr:.1f}% de los contratos"
            ):
                st.warning(
                    f"El **{_dr:.1f}%** de los contratos se tramitaron mediante alguna excepci\u00f3n "
                    f"a la licitaci\u00f3n p\u00fablica (Art. 55 LAASSP / Art. 42 anterior). "
                    f"El umbral de alerta es **30 %**."
                )

    # \u2500\u2500 Lista compacta de indicadores sin hallazgos \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    if _riesgos_limpios:
        st.caption("\u2705 Sin hallazgos en: " + " \u00b7 ".join(_riesgos_limpios))

    st.divider()

    # ── Listado completo de contratos ─────────────────────────────────────────
    with st.expander(
        f"📄 Listado completo de contratos — {_label_mr} ({_total_mr:,} contratos)",
        expanded=False,
    ):
        _COLS_TABLA_MR = [c for c in [
            "Número de procedimiento",
            "Proveedor o contratista",
            "Tipo Simplificado",
            "Importe DRC",
            "Fecha de inicio del contrato",
            "Descripción del contrato",
            "Dirección del anuncio",
        ] if c in _dff_sel.columns]

        _tabla_mr = _dff_sel[_COLS_TABLA_MR].copy()

        # Formato moneda
        if "Importe DRC" in _tabla_mr.columns:
            _tabla_mr["Importe DRC"] = _tabla_mr["Importe DRC"].apply(
                lambda x: f"${x:,.2f}" if pd.notna(x) else "—"
            )

        # Botón de descarga
        _csv_mr = _dff_sel[_COLS_TABLA_MR].to_csv(index=False, encoding="utf-8")
        st.download_button(
            label="⬇️ Descargar CSV",
            data=_csv_mr,
            file_name=f"contratos_{_label_mr.replace(' ', '_')}.csv",
            mime="text/csv",
            key="dl_tabla_mr",
        )

        _col_cfg_mr = {}
        if "Dirección del anuncio" in _tabla_mr.columns:
            _col_cfg_mr["Dirección del anuncio"] = st.column_config.LinkColumn(
                "🔗 ComprasMX", display_text="Ver contrato"
            )

        st.dataframe(
            _tabla_mr,
            use_container_width=True,
            column_config=_col_cfg_mr,
            hide_index=True,
        )


# ───────────────────────────────────────────────────────────────
# PÁGINA 6: PAGINA_PRECIOS
# ───────────────────────────────────────────────────────────────
def pagina_precios():
    st.header("💊 Analítica de Precios Unitarios — Medicamentos")
    st.caption(
        "Análisis estadístico de precios unitarios de medicamentos contratados por el IMSS (2024–2026). "
        "Los precios atípicos se identifican mediante el método de cercas de Tukey (IQR). "
        "Fuente: herramienta analítica de la DMII / ComprasMX."
    )

    try:
        df_pu = cargar_precios_unitarios()

        # Aplicar nombres editados de UC desde Base_UC_2025_V2.xlsx
        if len(df_dir_uc) > 0 and "Clave_UC" in df_dir_uc.columns:
            _pu_uc_map = (
                df_dir_uc.dropna(subset=["Nombre_editado"])
                .set_index("Clave_UC")["Nombre_editado"]
                .to_dict()
            )
            df_pu["UC"] = df_pu["Clave UC"].map(_pu_uc_map).fillna(df_pu["UC"])

        # ── Filtros inline ──────────────────────────────────────────────
        _puf1, _puf2, _puf3, _puf4 = st.columns([2, 3, 1, 1])

        _anos_pu = sorted(df_pu["Fuente Compras MX"].dropna().unique().tolist())
        _ano_sel_pu = _puf1.multiselect(
            "📅 Año de origen", _anos_pu, default=_anos_pu, key="pu_anos"
        )

        _casos_pu_opts = sorted(df_pu["Caso de atención crítico"].dropna().unique().tolist())
        _caso_sel_pu = _puf2.multiselect(
            "🚨 Caso de atención", _casos_pu_opts, default=_casos_pu_opts, key="pu_caso"
        )

        _solo_sig_pu  = _puf3.checkbox("Solo muestra significativa", value=False, key="pu_sig")
        _excl_cons_pu = _puf4.checkbox("Excluir consolidadas",        value=False, key="pu_cons")

        # Aplicar filtros
        df_pu_f = df_pu.copy()
        if _ano_sel_pu:
            df_pu_f = df_pu_f[df_pu_f["Fuente Compras MX"].isin(_ano_sel_pu)]
        if _caso_sel_pu:
            df_pu_f = df_pu_f[df_pu_f["Caso de atención crítico"].isin(_caso_sel_pu)]
        if _solo_sig_pu:
            df_pu_f = df_pu_f[df_pu_f["Muestra significativa"] == 1]
        if _excl_cons_pu:
            df_pu_f = df_pu_f[df_pu_f["Consolidada"].str.upper() != "SI"]

        # ── KPIs ───────────────────────────────────────────────────────
        _n_part_pu    = len(df_pu_f)
        _n_atip_pu    = (df_pu_f["Precio atípico"] == "SI").sum()
        _pct_atip_pu  = _n_atip_pu / _n_part_pu * 100 if _n_part_pu > 0 else 0
        _monto_atip_pu = df_pu_f.loc[df_pu_f["Precio atípico"] == "SI", "Monto partida"].sum()
        _n_uc_atip_pu  = df_pu_f.loc[df_pu_f["Precio atípico"] == "SI", "UC"].nunique()

        _kp1, _kp2, _kp3, _kp4 = st.columns(4)
        _kp1.metric("📋 Partidas analizadas",    f"{_n_part_pu:,}")
        _kp2.metric("⚠️ Con precio atípico",
                    f"{_n_atip_pu:,}",
                    delta=f"{_pct_atip_pu:.1f}% del total",
                    delta_color="off")
        _kp3.metric("💰 Monto en sobreprecios",
                    f"${_monto_atip_pu/1e6:,.1f} M MXN" if _monto_atip_pu > 0 else "N/D")
        _kp4.metric("🏥 UCs con precio atípico", f"{_n_uc_atip_pu:,}")

        # subconjunto atípicos para usar en las secciones siguientes
        _atip_pu = df_pu_f[df_pu_f["Precio atípico"] == "SI"].copy()

        st.divider()

        # ── SECCIÓN 1: SOBREPRECIOS POR UNIDAD COMPRADORA ──────────────
        st.subheader("1️⃣ Sobreprecios por Unidad Compradora")

        if len(_atip_pu) == 0:
            st.info("ℹ️ No hay partidas con precio atípico con los filtros actuales.")
        else:
            # Tabla base por UC
            _uc_tot = (
                df_pu_f.groupby("UC")
                .agg(Partidas=("Precio atípico", "count"),
                     Atipicas=("Precio atípico", lambda x: (x == "SI").sum()),
                     Monto_Total=("Monto partida", "sum"),
                     Z_Prom=("Precio estandarizado", "mean"))
                .reset_index()
            )
            _uc_atip_monto = (
                _atip_pu.groupby("UC")["Monto partida"]
                .sum().rename("Monto_Atipico").reset_index()
            )
            _uc_tbl = _uc_tot.merge(_uc_atip_monto, on="UC", how="left").fillna({"Monto_Atipico": 0})
            _uc_tbl["Pct_Atipicas"] = _uc_tbl["Atipicas"] / _uc_tbl["Partidas"] * 100

            _col_bar1, _col_bar2 = st.columns(2)

            # Gráfica A — monto en atípicos por UC (top 20)
            _top_monto_uc = (
                _uc_tbl.nlargest(20, "Monto_Atipico")
                .sort_values("Monto_Atipico")
            )
            _top_monto_uc["UC_c"]      = _top_monto_uc["UC"].apply(lambda s: str(s)[:45] + "…" if len(str(s)) > 45 else str(s))
            _top_monto_uc["Monto_fmt"] = _top_monto_uc["Monto_Atipico"].apply(lambda x: f"${x/1e6:,.1f} M")

            fig_pu_monto = px.bar(
                _top_monto_uc, x="Monto_Atipico", y="UC_c",
                orientation="h",
                color_discrete_sequence=[IMSS_ROJO],
                text="Monto_fmt",
                title="Top 20 UCs — Monto en partidas con precio atípico",
                custom_data=["UC", "Atipicas", "Partidas"]
            )
            fig_pu_monto.update_layout(
                font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis_title="Monto MXN", yaxis_title="", height=520,
                xaxis=dict(tickprefix="$", tickformat=",.0f")
            )
            fig_pu_monto.update_traces(
                textposition="outside", cliponaxis=False,
                textfont=dict(family="Noto Sans, sans-serif"),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Monto atípico: $%{x:,.0f}<br>"
                    "Partidas atípicas: %{customdata[1]:.0f} / %{customdata[2]:.0f}"
                    "<extra></extra>"
                )
            )
            _col_bar1.plotly_chart(fig_pu_monto, use_container_width=True)

            # Gráfica B — % partidas atípicas por UC (mínimo 3 partidas, top 20)
            _top_pct_uc = (
                _uc_tbl[_uc_tbl["Partidas"] >= 3]
                .nlargest(20, "Pct_Atipicas")
                .sort_values("Pct_Atipicas")
            )
            _top_pct_uc["UC_c"]    = _top_pct_uc["UC"].apply(lambda s: str(s)[:45] + "…" if len(str(s)) > 45 else str(s))
            _top_pct_uc["Pct_fmt"] = _top_pct_uc["Pct_Atipicas"].apply(lambda x: f"{x:.1f}%")

            fig_pu_pct = px.bar(
                _top_pct_uc, x="Pct_Atipicas", y="UC_c",
                orientation="h",
                color_discrete_sequence=["#E07B00"],
                text="Pct_fmt",
                title="Top 20 UCs — % de partidas con precio atípico (mín. 3 partidas)",
                custom_data=["UC", "Atipicas", "Partidas"]
            )
            fig_pu_pct.add_vline(
                x=50, line_dash="dash", line_color=IMSS_GRIS,
                annotation_text="50%", annotation_position="top right",
                annotation_font=dict(family="Noto Sans, sans-serif", color=IMSS_GRIS, size=11)
            )
            fig_pu_pct.update_layout(
                font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis_title="% de partidas con precio atípico", yaxis_title="",
                xaxis=dict(range=[0, 110]), height=520
            )
            fig_pu_pct.update_traces(
                textposition="outside", cliponaxis=False,
                textfont=dict(family="Noto Sans, sans-serif"),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Atípicas: %{customdata[1]:.0f} / %{customdata[2]:.0f}<br>"
                    "%{x:.1f}%<extra></extra>"
                )
            )
            _col_bar2.plotly_chart(fig_pu_pct, use_container_width=True)

            # Tabla resumen por UC
            with st.expander("📋 Tabla resumen por Unidad Compradora"):
                _tbl_uc_show = _uc_tbl.sort_values("Monto_Atipico", ascending=False).copy()
                _tbl_uc_show["% Atípicas"]   = _tbl_uc_show["Pct_Atipicas"].apply(lambda x: f"{x:.1f}%")
                _tbl_uc_show["Monto total"]   = _tbl_uc_show["Monto_Total"].apply(lambda x: f"${x/1e6:,.1f} M")
                _tbl_uc_show["Monto atípico"] = _tbl_uc_show["Monto_Atipico"].apply(lambda x: f"${x/1e6:,.1f} M")
                _tbl_uc_show["Z-score prom."] = _tbl_uc_show["Z_Prom"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/D")
                _tbl_uc_show = _tbl_uc_show[
                    ["UC", "Partidas", "Atipicas", "% Atípicas",
                     "Monto total", "Monto atípico", "Z-score prom."]
                ].rename(columns={"Atipicas": "Atípicas"}).reset_index(drop=True)
                _tbl_uc_show.index += 1
                st.dataframe(_tbl_uc_show, use_container_width=True)

            # ── Detalle de medicamentos con sobreprecio por UC seleccionada ──
            st.markdown("#### 🔍 Medicamentos con mayor sobreprecio en una UC")

            _ucs_det_s1 = sorted(_atip_pu["UC"].dropna().unique().tolist())
            _uc_sel_s1  = st.selectbox(
                "Seleccionar Unidad Compradora",
                _ucs_det_s1,
                key="pu_uc_det_s1",
                help="Elige una UC para ver qué medicamentos tienen los precios más atípicos en ella."
            )

            _df_uc_s1     = _atip_pu[_atip_pu["UC"] == _uc_sel_s1].copy()
            _df_uc_s1_all = df_pu_f[df_pu_f["UC"] == _uc_sel_s1].copy()

            # KPIs de la UC seleccionada
            _n_atip_s1  = len(_df_uc_s1)
            _n_total_s1 = len(_df_uc_s1_all)
            _monto_s1   = _df_uc_s1["Monto partida"].sum()
            _z_prom_s1  = _df_uc_s1["Precio estandarizado"].mean()
            _pct_s1     = _n_atip_s1 / _n_total_s1 * 100 if _n_total_s1 > 0 else 0

            _u1, _u2, _u3, _u4 = st.columns(4)
            _u1.metric("📋 Partidas atípicas",    f"{_n_atip_s1:,} de {_n_total_s1:,}")
            _u2.metric("📊 % partidas atípicas",  f"{_pct_s1:.1f}%")
            _u3.metric("💰 Monto en sobreprecios", f"${_monto_s1/1e6:,.1f} M MXN")
            _u4.metric("📈 Z-score promedio",      f"{_z_prom_s1:.2f}" if pd.notna(_z_prom_s1) else "N/D")

            if _n_atip_s1 == 0:
                st.info(f"ℹ️ No hay partidas con precio atípico para {_uc_sel_s1}.")
            else:
                _col_s1a, _col_s1b = st.columns(2)

                # ── Gráfica A: top 15 medicamentos por MONTO atípico ──
                _med_monto_s1 = (
                    _df_uc_s1.groupby("Descripción")
                    .agg(
                        Monto_Atipico=("Monto partida", "sum"),
                        N_Atipicas=("Precio atípico", "count"),
                        Z_Prom=("Precio estandarizado", "mean")
                    )
                    .reset_index()
                    .sort_values("Monto_Atipico", ascending=False)
                    .head(15)
                )
                _med_monto_s1["Desc_c"]    = _med_monto_s1["Descripción"].apply(
                    lambda s: str(s)[:52] + "…" if len(str(s)) > 52 else str(s)
                )
                _med_monto_s1["Monto_fmt"] = _med_monto_s1["Monto_Atipico"].apply(
                    lambda x: f"${x/1e6:,.2f} M"
                )

                fig_s1_monto = px.bar(
                    _med_monto_s1.sort_values("Monto_Atipico"),
                    x="Monto_Atipico", y="Desc_c",
                    orientation="h",
                    color_discrete_sequence=[IMSS_ROJO],
                    text="Monto_fmt",
                    title="Top 15 — Monto en sobreprecios",
                    custom_data=["Descripción", "N_Atipicas", "Z_Prom"]
                )
                fig_s1_monto.update_layout(
                    font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                    xaxis_title="Monto MXN", yaxis_title="", height=500,
                    xaxis=dict(tickprefix="$", tickformat=",.0f"),
                    title_font_color=IMSS_VERDE_OSC
                )
                fig_s1_monto.update_traces(
                    textposition="outside", cliponaxis=False,
                    textfont=dict(family="Noto Sans, sans-serif"),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Monto atípico: $%{x:,.0f}<br>"
                        "Partidas: %{customdata[1]:.0f}  |  Z-score prom.: %{customdata[2]:.2f}"
                        "<extra></extra>"
                    )
                )
                _col_s1a.plotly_chart(fig_s1_monto, use_container_width=True)

                # ── Gráfica B: top 15 medicamentos por Z-score MÁXIMO ──
                _med_z_s1 = (
                    _df_uc_s1.groupby("Descripción")
                    .agg(
                        Z_Max=("Precio estandarizado", "max"),
                        Z_Prom=("Precio estandarizado", "mean"),
                        Monto_Atipico=("Monto partida", "sum"),
                        N_Atipicas=("Precio atípico", "count")
                    )
                    .reset_index()
                    .sort_values("Z_Max", ascending=False)
                    .head(15)
                )
                _med_z_s1["Desc_c"] = _med_z_s1["Descripción"].apply(
                    lambda s: str(s)[:52] + "…" if len(str(s)) > 52 else str(s)
                )
                _med_z_s1["Z_fmt"] = _med_z_s1["Z_Max"].apply(
                    lambda x: f"{x:.2f}σ" if pd.notna(x) else "N/D"
                )

                fig_s1_z = px.bar(
                    _med_z_s1.sort_values("Z_Max"),
                    x="Z_Max", y="Desc_c",
                    orientation="h",
                    color_discrete_sequence=["#E07B00"],
                    text="Z_fmt",
                    title="Top 15 — Mayor desviación del precio (Z-score máximo)",
                    custom_data=["Descripción", "Z_Prom", "Monto_Atipico", "N_Atipicas"]
                )
                fig_s1_z.add_vline(
                    x=0, line_dash="solid", line_color=IMSS_GRIS, line_width=1
                )
                fig_s1_z.update_layout(
                    font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                    xaxis_title="Z-score máximo (σ sobre la mediana del mercado)",
                    yaxis_title="", height=500,
                    title_font_color=IMSS_VERDE_OSC
                )
                fig_s1_z.update_traces(
                    textposition="outside", cliponaxis=False,
                    textfont=dict(family="Noto Sans, sans-serif"),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Z-score máx.: %{x:.2f}σ  |  Z-score prom.: %{customdata[1]:.2f}σ<br>"
                        "Monto: $%{customdata[2]:,.0f}  |  Partidas: %{customdata[3]:.0f}"
                        "<extra></extra>"
                    )
                )
                _col_s1b.plotly_chart(fig_s1_z, use_container_width=True)

                # ── Tabla detallada de contratos atípicos de esta UC ──
                with st.expander(
                    f"📋 Contratos con precio atípico — {_uc_sel_s1} ({_n_atip_s1:,} partidas)"
                ):
                    _cols_s1_tbl = [c for c in [
                        "Descripción", "Proveedor", "RFC del proveedor adjudicado",
                        "Precio unitario", "Mediana (P)", "Precio estandarizado",
                        "Cantidad", "Monto partida", "Fuente Compras MX",
                        "Caso de atención crítico", "Tipo de proveedor por historial",
                        "Vínculo"
                    ] if c in _df_uc_s1.columns]
                    _tbl_s1 = (
                        _df_uc_s1[_cols_s1_tbl]
                        .sort_values("Precio estandarizado", ascending=False)
                        .reset_index(drop=True)
                    )
                    _tbl_s1.index += 1
                    st.dataframe(
                        _tbl_s1,
                        column_config={
                            "Vínculo": st.column_config.LinkColumn(
                                "🔗 ComprasMX", display_text="Ver contrato"
                            )
                        },
                        use_container_width=True,
                        height=420
                    )
                    _csv_s1 = (
                        _tbl_s1.drop(columns=["Vínculo"], errors="ignore")
                        .to_csv(index=False)
                        .encode("utf-8-sig")
                    )
                    st.download_button(
                        "📥 Descargar contratos de esta UC (CSV)",
                        data=_csv_s1,
                        file_name=f"precios_atipicos_{str(_uc_sel_s1)[:30].replace(' ', '_')}.csv",
                        mime="text/csv",
                        key="dl_s1_uc_det"
                    )

        st.divider()

        # ── SECCIÓN 2: MAPA DE CALOR UC × MEDICAMENTO ──────────────────
        st.subheader("2️⃣ Mapa de Calor — UC × Medicamento (Z-score del precio)")
        st.caption(
            "Z-score promedio del precio unitario para combinaciones UC × medicamento "
            "con al menos un precio atípico. Rojo = precio muy por encima de la mediana del mercado."
        )

        _heat_src = df_pu_f[
            (df_pu_f["Precio atípico"] == "SI") &
            df_pu_f["Precio estandarizado"].notna() &
            df_pu_f["UC"].notna() &
            df_pu_f["Descripción"].notna()
        ].copy()

        if len(_heat_src) < 4:
            st.info("ℹ️ Datos insuficientes para el mapa de calor con los filtros actuales.")
        else:
            _top_uc_h  = _heat_src.groupby("UC")["Monto partida"].sum().nlargest(15).index.tolist()
            _top_med_h = (
                _heat_src.groupby("Descripción")["Monto partida"]
                .sum().nlargest(12).index.tolist()
            )
            _pivot_h = (
                _heat_src[
                    _heat_src["UC"].isin(_top_uc_h) &
                    _heat_src["Descripción"].isin(_top_med_h)
                ]
                .groupby(["UC", "Descripción"])["Precio estandarizado"]
                .mean()
                .unstack(fill_value=0)
            )
            # Etiquetas cortas para medicamentos
            _med_labels_h = [
                str(c)[:55] + "…" if len(str(c)) > 55 else str(c)
                for c in _pivot_h.columns
            ]
            fig_heat_pu = go.Figure(data=go.Heatmap(
                z=_pivot_h.values,
                x=_med_labels_h,
                y=_pivot_h.index.tolist(),
                colorscale=[[0, IMSS_VERDE], [0.45, "#FFFFFF"], [1, IMSS_ROJO]],
                zmid=0,
                colorbar=dict(
                    title=dict(text="Z-score", font=dict(family="Noto Sans, sans-serif")),
                    tickfont=dict(family="Noto Sans, sans-serif")
                ),
                hovertemplate="<b>%{y}</b><br>%{x}<br>Z-score prom.: %{z:.2f}<extra></extra>"
            ))
            fig_heat_pu.update_layout(
                font=plotly_font(),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis=dict(tickangle=-40, tickfont=dict(size=9, family="Noto Sans, sans-serif")),
                yaxis=dict(tickfont=dict(size=9, family="Noto Sans, sans-serif")),
                height=max(400, 30 * len(_pivot_h) + 120),
                margin=dict(l=260, b=260, t=60, r=80)
            )
            st.plotly_chart(fig_heat_pu, use_container_width=True)

        st.divider()

        # ── SECCIÓN 3: DISTRIBUCIÓN DE PRECIOS POR MEDICAMENTO ─────────
        st.subheader("3️⃣ Distribución de Precios por Medicamento")
        st.caption(
            "Selecciona un medicamento para ver cómo varían los precios unitarios "
            "entre Unidades Compradoras. La línea punteada indica la mediana de mercado "
            "y el límite superior de Tukey."
        )

        _meds_list = sorted(df_pu_f["Descripción"].dropna().unique().tolist())
        _med_sel_pu = st.selectbox(
            "🔍 Seleccionar medicamento (búsqueda por clave o nombre)",
            _meds_list,
            key="pu_med_sel"
        )

        if _med_sel_pu:
            _df_med = (
                df_pu_f[df_pu_f["Descripción"] == _med_sel_pu]
                .copy()
                .sort_values("Precio unitario")
            )
            _mediana_med  = _df_med["Mediana (P)"].dropna().iloc[0] if len(_df_med) > 0 else None
            _lim_sup_med  = pd.to_numeric(
                _df_med["Límite superior (P)"].dropna().iloc[0] if len(_df_med) > 0 else None,
                errors="coerce"
            )
            _lim_inf_med  = _df_med["Límite inferior (P)"].dropna().iloc[0] if len(_df_med) > 0 else None
            _n_med        = len(_df_med)
            _n_atip_med   = (_df_med["Precio atípico"] == "SI").sum()
            _pct_atip_med = _n_atip_med / _n_med * 100 if _n_med > 0 else 0

            _m1, _m2, _m3, _m4 = st.columns(4)
            _m1.metric("Partidas",        f"{_n_med:,}")
            _m2.metric("Precio atípico",  f"{_n_atip_med:,} ({_pct_atip_med:.1f}%)")
            _m3.metric("Mediana precio",  f"${_mediana_med:,.2f}" if pd.notna(_mediana_med) else "N/D")
            _m4.metric("Límite superior", f"${_lim_sup_med:,.2f}" if pd.notna(_lim_sup_med) else "N/D")

            # Strip plot: cada punto = un contrato, eje X = UC, eje Y = precio unitario
            _df_med["_color"] = _df_med["Precio atípico"].map(
                {"SI": "🔴 Precio atípico", "NO": "🟢 Precio normal"}
            ).fillna("⚪ Sin clasificar")

            _hover_cols = [c for c in [
                "Proveedor", "Cantidad", "Monto partida",
                "Precio estandarizado", "Fuente Compras MX", "Caso de atención crítico"
            ] if c in _df_med.columns]

            fig_med = px.strip(
                _df_med,
                x="UC", y="Precio unitario",
                color="_color",
                color_discrete_map={
                    "🔴 Precio atípico": IMSS_ROJO,
                    "🟢 Precio normal":  IMSS_VERDE,
                    "⚪ Sin clasificar": IMSS_GRIS,
                },
                title=f"Precios unitarios por UC — {str(_med_sel_pu)[:90]}",
                hover_data=_hover_cols
            )
            if pd.notna(_mediana_med):
                fig_med.add_hline(
                    y=_mediana_med, line_dash="dash", line_color=IMSS_ORO, line_width=2,
                    annotation_text=f"Mediana: ${_mediana_med:,.2f}",
                    annotation_position="top left",
                    annotation_font=dict(family="Noto Sans, sans-serif", color=IMSS_ORO, size=11)
                )
            if pd.notna(_lim_sup_med):
                fig_med.add_hline(
                    y=_lim_sup_med, line_dash="dot", line_color=IMSS_ROJO, line_width=1.5,
                    annotation_text=f"Límite superior: ${_lim_sup_med:,.2f}",
                    annotation_position="top right",
                    annotation_font=dict(family="Noto Sans, sans-serif", color=IMSS_ROJO, size=11)
                )
            fig_med.update_layout(
                font=plotly_font(),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis=dict(tickangle=-45, tickfont=dict(size=9, family="Noto Sans, sans-serif")),
                yaxis_title="Precio unitario (MXN)",
                legend=dict(title="", orientation="h", y=1.04, x=0),
                height=480
            )
            st.plotly_chart(fig_med, use_container_width=True)

            with st.expander(f"📋 Todos los contratos de este medicamento ({_n_med:,})"):
                _cols_med_tbl = [c for c in [
                    "UC", "Proveedor", "Precio unitario", "Mediana (P)",
                    "Precio estandarizado", "Precio atípico", "Cantidad",
                    "Monto partida", "Fuente Compras MX",
                    "Caso de atención crítico", "Vínculo"
                ] if c in _df_med.columns]
                _tbl_med = (
                    _df_med[_cols_med_tbl]
                    .sort_values("Precio estandarizado", ascending=False)
                    .reset_index(drop=True)
                )
                _tbl_med.index += 1
                st.dataframe(
                    _tbl_med,
                    column_config={
                        "Vínculo": st.column_config.LinkColumn(
                            "🔗 ComprasMX", display_text="Ver contrato"
                        )
                    },
                    use_container_width=True,
                    height=400
                )

        st.divider()

        # ── SECCIÓN 4: TABLA DETALLADA — PARTIDAS ATÍPICAS ─────────────
        st.subheader("4️⃣ Detalle de Partidas con Precio Atípico")

        _det_c1, _det_c2, _det_c3 = st.columns([3, 2, 2])
        _ucs_det = ["Todas"] + sorted(_atip_pu["UC"].dropna().unique().tolist())
        _uc_sel_det = _det_c1.selectbox("🏢 Unidad Compradora", _ucs_det, key="pu_det_uc")
        _casos_det  = ["Todos"] + sorted(_atip_pu["Caso de atención crítico"].dropna().unique().tolist())
        _caso_sel_det = _det_c2.selectbox("🚨 Caso de atención", _casos_det, key="pu_det_caso")
        _tipo_prov_det = ["Todos"] + sorted(
            _atip_pu["Tipo de proveedor por historial"].dropna().unique().tolist()
        )
        _tipoprov_sel = _det_c3.selectbox(
            "🏢 Tipo de proveedor", _tipo_prov_det, key="pu_det_tipo"
        )

        _det_df = _atip_pu.copy()
        if _uc_sel_det   != "Todas": _det_df = _det_df[_det_df["UC"] == _uc_sel_det]
        if _caso_sel_det != "Todos": _det_df = _det_df[_det_df["Caso de atención crítico"] == _caso_sel_det]
        if _tipoprov_sel != "Todos": _det_df = _det_df[_det_df["Tipo de proveedor por historial"] == _tipoprov_sel]

        if len(_det_df) == 0:
            st.info("ℹ️ No hay partidas con los filtros seleccionados.")
        else:
            st.caption(f"Mostrando **{len(_det_df):,}** partidas con precio atípico")

            _cols_det = [c for c in [
                "Caso de atención crítico",
                "UC", "Descripción",
                "Proveedor", "RFC del proveedor adjudicado",
                "Precio unitario", "Mediana (P)", "Precio estandarizado",
                "Cantidad", "Monto partida",
                "Fuente Compras MX", "Consolidada",
                "Tipo de proveedor por historial",
                "Vínculo"
            ] if c in _det_df.columns]

            _tbl_det = (
                _det_df[_cols_det]
                .sort_values("Precio estandarizado", ascending=False)
                .reset_index(drop=True)
            )
            _tbl_det.index += 1

            st.dataframe(
                _tbl_det,
                column_config={
                    "Vínculo": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    )
                },
                use_container_width=True,
                height=460
            )

            _csv_pu_det = (
                _tbl_det
                .drop(columns=["Vínculo"], errors="ignore")
                .to_csv(index=False)
                .encode("utf-8-sig")
            )
            st.download_button(
                "📥 Descargar partidas atípicas (CSV)",
                data=_csv_pu_det,
                file_name="precios_atipicos_medicamentos.csv",
                mime="text/csv",
                key="dl_pu5"
            )

    except FileNotFoundError:
        st.info(
            "ℹ️ Para activar esta sección, coloca el archivo "
            "`AnaliticaPreciosUnitarios.xlsx` en la misma carpeta que el dashboard."
        )



# ───────────────────────────────────────────────────────────────
# PÁGINA 6: PAGINA_EXPEDIENTE
# ───────────────────────────────────────────────────────────────
def pagina_expediente():
    st.subheader("🔎 Expediente de Contrato")
    st.caption(
        "Busca un contrato por proveedor, RFC, descripción o unidad compradora. "
        "Consulta su ficha detallada y sus indicadores de riesgo individuales."
    )

    # ── Buscador ──────────────────────────────────────────────
    _col_busq6, _col_campo6 = st.columns([4, 2])
    with _col_busq6:
        _busqueda_t6 = st.text_input(
            "Buscar contrato",
            placeholder="RFC, nombre del proveedor, descripción del contrato…",
            key="busq_expediente_t6"
        )
    with _col_campo6:
        _campo_t6 = st.selectbox(
            "Buscar en",
            ["Proveedor / RFC", "Número de procedimiento", "Descripción del contrato", "Unidad Compradora"],
            key="campo_busq_exp_t6"
        )

    if not _busqueda_t6.strip():
        st.info(
            "🔍 Ingresa un término de búsqueda para localizar contratos. "
            "Puedes usar el RFC, el nombre del proveedor, el número de procedimiento o palabras clave de la descripción."
        )
    else:
        _q6 = _busqueda_t6.strip().upper()
        if _campo_t6 == "Proveedor / RFC":
            _mask_t6 = (
                dff["Proveedor o contratista"].str.upper().str.contains(_q6, na=False)
                | dff["rfc"].str.upper().str.contains(_q6, na=False)
            )
        elif _campo_t6 == "Número de procedimiento":
            _mask_t6 = dff["Número de procedimiento"].str.upper().str.contains(_q6, na=False)
        elif _campo_t6 == "Descripción del contrato":
            _mask_t6 = dff["Descripción del contrato"].str.upper().str.contains(_q6, na=False)
        else:
            _mask_t6 = dff["Nombre de la UC"].str.upper().str.contains(_q6, na=False)

        _res_t6   = dff[_mask_t6].reset_index(drop=True)
        _n_res_t6 = len(_res_t6)

        if _n_res_t6 == 0:
            st.warning("⚠️ No se encontraron contratos con ese criterio de búsqueda.")
        elif _n_res_t6 > 300:
            st.warning(
                f"⚠️ La búsqueda arrojó **{_n_res_t6:,} contratos**. "
                "Refina el término para acotar los resultados (máx. 300 para seleccionar)."
            )
        else:
            st.caption(f"✅ {_n_res_t6:,} contrato(s) encontrado(s). Selecciona uno para ver su expediente.")

            def _fmt_label_t6(r):
                prov  = str(r.get("Proveedor o contratista", "N/D"))[:55]
                uc    = str(r.get("Nombre de la UC", ""))[:38]
                imp   = r.get("Importe DRC")
                imp_s = f"${float(imp)/1e6:.1f} M" if pd.notna(imp) else "N/D"
                return f"{prov}  ·  {uc}  ·  {imp_s}"

            _labels_t6 = [_fmt_label_t6(_res_t6.iloc[i]) for i in range(_n_res_t6)]
            _idx_t6 = st.selectbox(
                "Seleccionar contrato",
                range(_n_res_t6),
                format_func=lambda i: _labels_t6[i],
                key="sel_contrato_exp_t6"
            )

            _c      = _res_t6.iloc[_idx_t6]
            _rfc_c6 = str(_c.get("rfc", "")).strip().upper()

            st.divider()

            # ══════════════════════════════════════════════════════════
            # FICHA  |  INDICADORES DE RIESGO
            # ══════════════════════════════════════════════════════════
            _col_ficha6, _col_riesgo6 = st.columns([3, 2], gap="large")

            # ── Ficha del contrato ─────────────────────────────────
            with _col_ficha6:
                st.markdown(
                    f"<h4 style='color:{IMSS_VERDE_OSC}; margin-bottom:0.3em'>📄 Ficha del Contrato</h4>",
                    unsafe_allow_html=True
                )

                def _r6(label, val):
                    """Renderiza fila label: valor solo si val no es vacío."""
                    if val is not None and pd.notna(val) and str(val).strip() not in ("", "nan", "None"):
                        st.markdown(f"**{label}** &nbsp; {val}", unsafe_allow_html=True)

                # Partida + descripción CUCoP
                _partida_c6  = str(_c.get("Partida específica", "")).strip().zfill(5)
                _cucop_desc6 = ""
                if len(df_cucop) > 0:
                    _cr6 = df_cucop[df_cucop["PARTIDA ESPECÍFICA"] == _partida_c6]
                    if len(_cr6) > 0:
                        _cucop_desc6 = str(_cr6.iloc[0].get("DESC. PARTIDA ESPECÍFICA", ""))

                _imp_c6   = _c.get("Importe DRC")
                _imp_str6 = f"${float(_imp_c6):,.2f} MXN" if pd.notna(_imp_c6) else "N/D"

                _r6("🏢 Institución",            _c.get("Institución"))
                _r6("🏥 Unidad Compradora",      _c.get("Nombre de la UC"))
                _r6("🏭 Proveedor",              _c.get("Proveedor o contratista"))
                _r6("🪪 RFC",                    _rfc_c6 or None)
                _r6("🔢 Núm. de procedimiento",  _c.get("Número de procedimiento"))
                _r6("📋 Tipo de procedimiento",  _c.get("Tipo Procedimiento"))
                _r6("📦 Tipo de contratación",  _c.get("Tipo de contratación"))
                _r6("📊 Estratificación",       _c.get("Estratificación"))
                _r6("🔄 Compra consolidada",    _c.get("Compra consolidada"))
                _r6("💰 Importe",              _imp_str6)
                _r6("📅 Fecha de inicio",       _c.get("Fecha de inicio del contrato"))
                _r6("⚖️ Fecha de fallo",        _c.get("Fecha de fallo"))
                _partida_fmt6 = _partida_c6 + (f" — {_cucop_desc6}" if _cucop_desc6 else "")
                _r6("🗂️ Partida presupuestaria", _partida_fmt6)
                _r6("📝 Descripción",           _c.get("Descripción del contrato"))

                _art_c6  = _c.get("Artículo de excepción", "")
                _desc_c6 = _c.get("Descripción excepción", "")
                if pd.notna(_art_c6) and str(_art_c6).strip() not in ("", "nan"):
                    _r6("⚠️ Art. de excepción", _art_c6)
                    _r6("📃 Desc. excepción",   _desc_c6)

                _url_c6 = str(_c.get("Dirección del anuncio", "")).strip()
                if _url_c6.startswith("http"):
                    st.markdown(
                        f"<br>🔗 <a href='{_url_c6}' target='_blank'>"
                        f"<strong>Ver contrato en ComprasMX</strong></a>",
                        unsafe_allow_html=True
                    )

            # ── Indicadores de Riesgo ──────────────────────────────
            with _col_riesgo6:
                st.markdown(
                    f"<h4 style='color:{IMSS_VERDE_OSC}; margin-bottom:0.3em'>🚨 Indicadores de Riesgo</h4>",
                    unsafe_allow_html=True
                )

                # ── Pre-cómputo del score de riesgo compuesto ─────────
                # (misma lógica que el Top-20; aquí mostramos el score_base
                #  sin normalizar ya que corresponde a un solo contrato)
                _pts_c6 = 0

                # Tipo procedimiento
                _tipo_c6_sc = str(_c.get("Tipo Simplificado", "")).strip()
                if _tipo_c6_sc == "Adjudicación Directa":            _pts_c6 += 20
                elif _tipo_c6_sc == "Adjudicación Directa — Fr. I":  _pts_c6 += 5
                elif _tipo_c6_sc == "Invitación a 3 personas":        _pts_c6 += 5

                # SABG (fecha de fallo o firma como referencia)
                try:
                    _df_san6_sc = cargar_sancionados()
                    _san6_sc = _df_san6_sc[_df_san6_sc["RFC"].str.strip() == _rfc_c6]
                    if len(_san6_sc) > 0:
                        _ff6_sc = pd.to_datetime(_c.get("Fecha de fallo"), dayfirst=True, errors="coerce")
                        _fm6_sc = pd.to_datetime(_c.get("Fecha de firma del contrato"), dayfirst=True, errors="coerce")
                        _fr6_sc = _ff6_sc if not pd.isna(_ff6_sc) else _fm6_sc
                        # Evaluar TODOS los registros del RFC (no solo el primero)
                        _sc6 = _sa6 = _sm6 = False
                        for _, _row6 in _san6_sc.iterrows():
                            _niv6_l = str(_row6.get("Nivel de Riesgo", "")).lower()
                            _ini6   = pd.to_datetime(_row6.get("Inicio inhabilitación"), errors="coerce")
                            if "crítico" in _niv6_l:
                                if not pd.isna(_fr6_sc) and not pd.isna(_ini6) and _fr6_sc >= _ini6:
                                    _sc6 = True
                            elif "alto" in _niv6_l:
                                _sa6 = True
                            elif "medio" in _niv6_l:
                                _sm6 = True
                        if _sc6:   _pts_c6 += 100
                        elif _sa6: _pts_c6 += 60
                        elif _sm6: _pts_c6 += 30
                except Exception:
                    pass

                # EFOS
                try:
                    _df_efos6_sc = cargar_efos()
                    _ef6_sc = _df_efos6_sc[_df_efos6_sc["RFC"] == _rfc_c6]
                    if len(_ef6_sc) > 0:
                        _sit6_sc = _ef6_sc.iloc[0]["Situación del contribuyente"]
                        if _sit6_sc == "Definitivo":  _pts_c6 += 80
                        elif _sit6_sc == "Presunto":  _pts_c6 += 40
                except Exception:
                    pass

                # Umbral legal (90–100 % del tope PEF)
                _f_umb_c6 = False
                _umb_disponible_c6 = False   # True solo cuando hay umbral aplicable para este contrato
                _pct_sc = None
                _umb_sc = None
                try:
                    _umbs6_sc = cargar_umbrales_pef()
                    if _umbs6_sc:
                        _TIPOS_AD_sc   = {"Adjudicación Directa", "Adjudicación Directa — Fr. I"}
                        _TIPOS_I3P_sc  = {"Invitación a 3 personas"}
                        _TIPOS_SERV_sc = {"SERVICIOS", "SERVICIOS RELACIONADOS CON LA OBRA", "ARRENDAMIENTOS"}
                        _proc_sc = str(_c.get("Tipo Simplificado", "")).strip()
                        if _proc_sc in (_TIPOS_AD_sc | _TIPOS_I3P_sc):
                            _ley_sc  = str(_c.get("Ley", "LAASSP")).strip().upper()
                            _cont_sc = str(_c.get("Tipo de contratación", "ADQUISICIONES")).strip().upper()
                            _finic_sc = pd.to_datetime(
                                _c.get("Fecha de inicio del contrato"), dayfirst=True, errors="coerce"
                            )
                            _año_sc = int(_finic_sc.year) if not pd.isna(_finic_sc) else None
                            if _año_sc and _año_sc in _umbs6_sc:
                                _th_sc = _umbs6_sc[_año_sc]
                                if _ley_sc == "LAASSP":
                                    _umb_sc = _th_sc["ad_laassp"] if _proc_sc in _TIPOS_AD_sc else _th_sc["i3p_laassp"]
                                elif _ley_sc == "LOPSRM":
                                    if _proc_sc in _TIPOS_AD_sc:
                                        _umb_sc = _th_sc["ad_obra_lopsrm"] if _cont_sc == "OBRA PÚBLICA" else (
                                            _th_sc["ad_serv_lopsrm"] if _cont_sc in _TIPOS_SERV_sc else None)
                                    else:
                                        _umb_sc = _th_sc["i3p_obra_lopsrm"] if _cont_sc == "OBRA PÚBLICA" else (
                                            _th_sc["i3p_serv_lopsrm"] if _cont_sc in _TIPOS_SERV_sc else None)
                                if _umb_sc:
                                    _imp_sc = _c.get("Importe DRC")
                                    if pd.notna(_imp_sc):
                                        _pct_sc = float(_imp_sc) / _umb_sc * 100
                                        _umb_disponible_c6 = True
                                        if 90 <= _pct_sc < 100:
                                            _pts_c6 += 45
                                            _f_umb_c6 = True
                except Exception:
                    pass

                # Reciente creación
                _rc_dias_c6 = None
                _m_rfc_sc = _re.match(r'^[A-ZÑ&]{3}(\d{2})(\d{2})(\d{2})[A-Z0-9]{3}$', _rfc_c6)
                if _m_rfc_sc:
                    _yy_sc, _mm_sc, _dd_sc = int(_m_rfc_sc.group(1)), int(_m_rfc_sc.group(2)), int(_m_rfc_sc.group(3))
                    if 1 <= _mm_sc <= 12 and 1 <= _dd_sc <= 31:
                        _yr_sc = 2000 + _yy_sc if _yy_sc <= 30 else 1900 + _yy_sc
                        try:
                            _fecha_rfc_sc = _date(_yr_sc, _mm_sc, _dd_sc)
                            _fi_str_sc = _c.get("Fecha de inicio del contrato", "")
                            if pd.notna(_fi_str_sc):
                                _fi_sc = pd.to_datetime(str(_fi_str_sc), dayfirst=True, errors="coerce").date()
                                if not pd.isna(pd.Timestamp(_fi_sc)):
                                    _rc_dias_c6 = (_fi_sc - _fecha_rfc_sc).days
                                    if 0 <= _rc_dias_c6 < 365:
                                        _pts_c6 += 30
                        except Exception:
                            pass

                # ── Mostrar score ──────────────────────────────────────
                _score_color_c6 = (
                    IMSS_ROJO   if _pts_c6 >= 100 else
                    "#E07B00"   if _pts_c6 >= 50  else
                    IMSS_ORO    if _pts_c6 >= 20  else
                    IMSS_VERDE
                )
                _score_label_c6 = (
                    "Riesgo crítico"  if _pts_c6 >= 100 else
                    "Riesgo alto"     if _pts_c6 >= 50  else
                    "Riesgo medio"    if _pts_c6 >= 20  else
                    "Riesgo bajo"
                )
                st.markdown(
                    f"<div style='background:{_score_color_c6}18; border-left:4px solid {_score_color_c6}; "
                    f"padding:0.6em 1em; border-radius:4px; margin-bottom:1em'>"
                    f"<span style='font-size:0.75em; color:{_score_color_c6}; font-weight:700; "
                    f"text-transform:uppercase; letter-spacing:0.06em'>Score de riesgo compuesto</span><br>"
                    f"<span style='font-size:2.2em; font-weight:800; color:{_score_color_c6}'>{_pts_c6}</span>"
                    f"<span style='font-size:0.9em; color:{_score_color_c6}'> pts &nbsp;·&nbsp; {_score_label_c6}</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )

                # ① Tipo de procedimiento
                _tipo_c6 = str(_c.get("Tipo Simplificado", "")).strip()
                if _tipo_c6 == "Adjudicación Directa":
                    st.error("🔴 **Adjudicación Directa** — Procedimiento no competitivo")
                elif _tipo_c6 == "Adjudicación Directa — Fr. I":
                    st.warning("🟡 **Adjudicación Directa — Fr. I (Patente / Exclusividad)** — Excepción estructural al Art. 54 Fr. I LAASSP")
                elif _tipo_c6 == "Invitación a 3 personas":
                    st.warning("🟡 **Invitación a 3 personas**")
                elif _tipo_c6 == "Licitación Pública":
                    st.success("🟢 **Licitación Pública**")
                else:
                    st.info(f"🔵 **{_tipo_c6}**")

                # ② Proceso de excepción
                _art_exc_c6 = str(_c.get("Artículo de excepción", "")).strip().upper()
                if any(_art_exc_c6.startswith(a) for a in ["ART. 55", "ART. 42", "ART. 43"]):
                    st.warning(f"🟠 **Proceso de excepción:** {_c.get('Artículo de excepción', '')}")
                else:
                    st.success("✅ **Proceso de excepción:** No aplica")

                # ③ SABG — Inhabilitados
                st.markdown("---")
                try:
                    _df_san6 = cargar_sancionados()
                    _san6    = _df_san6[_df_san6["RFC"].str.strip() == _rfc_c6]
                    if len(_san6) > 0:
                        # Usar el registro de MAYOR severidad (no solo el primero)
                        _NIVO_RANK = {"crítico": 3, "alto": 2, "medio": 1}
                        _best_san6 = _san6.iloc[0]
                        for _, _r6x in _san6.iterrows():
                            _nl_new = str(_r6x.get("Nivel de Riesgo", "")).lower()
                            _nl_best = str(_best_san6.get("Nivel de Riesgo", "")).lower()
                            if (max((_NIVO_RANK.get(k, 0) for k in _NIVO_RANK if k in _nl_new), default=0)
                                    > max((_NIVO_RANK.get(k, 0) for k in _NIVO_RANK if k in _nl_best), default=0)):
                                _best_san6 = _r6x
                        _niv_san6 = str(_best_san6.get("Nivel de Riesgo", ""))
                        _ini_san6 = _best_san6.get("Inicio inhabilitación", "")
                        if "crítico" in _niv_san6.lower():
                            # Fecha de referencia: fallo si existe, firma si no (AD no genera fallo)
                            _ff_c6   = pd.to_datetime(_c.get("Fecha de fallo"), dayfirst=True, errors="coerce")
                            _fm_c6   = pd.to_datetime(_c.get("Fecha de firma del contrato"), dayfirst=True, errors="coerce")
                            _fref_c6 = _ff_c6 if not pd.isna(_ff_c6) else _fm_c6
                            _ini_dt  = pd.to_datetime(_ini_san6, errors="coerce")
                            if pd.isna(_fref_c6):
                                st.warning("⚪ **SABG:** Inhabilitado — Sin fecha de referencia (verificar manualmente)")
                            elif _fref_c6 >= _ini_dt:
                                st.error(f"🔴 **SABG:** {_niv_san6}")
                            else:
                                st.info("⚫ **SABG:** Inhabilitado — Fecha anterior a la sanción (sin violación LAASSP)")
                        elif "alto" in _niv_san6.lower():
                            st.warning(f"🟠 **SABG:** {_niv_san6}")
                        else:
                            st.warning(f"🟡 **SABG:** {_niv_san6}")
                        st.caption(f"Inhabilitación desde: {_ini_san6}")
                    else:
                        st.success("✅ **SABG:** No aparece en lista de inhabilitados")
                except FileNotFoundError:
                    st.info("ℹ️ Archivo SABG no disponible")

                # ④ EFOS Art. 69-B
                st.markdown("---")
                try:
                    _df_efos6 = cargar_efos()
                    _efos6    = _df_efos6[_df_efos6["RFC"] == _rfc_c6]
                    if len(_efos6) > 0:
                        _sit6 = _efos6.iloc[0]["Situación del contribuyente"]
                        _niv6 = nivel_efos(_sit6)
                        if "definitivo" in _niv6.lower():
                            st.error(f"🔴 **EFOS:** {_niv6}")
                        elif "presunto" in _niv6.lower():
                            st.warning(f"🟡 **EFOS:** {_niv6}")
                        else:
                            st.success(f"✅ **EFOS:** {_niv6}")
                    else:
                        st.success("✅ **EFOS:** No aparece en lista Art. 69-B CFF")
                except FileNotFoundError:
                    st.info("ℹ️ Archivo EFOS no disponible")

                # ⑤ Zona umbral legal
                st.markdown("---")
                if _f_umb_c6:
                    st.warning(
                        f"🚦 **Zona umbral legal:** El importe representa el "
                        f"**{_pct_sc:.1f} %** del tope máximo para este tipo de contratación "
                        f"(umbral: ${_umb_sc/1e6:.2f} M MXN). "
                        "Posible fraccionamiento para evadir licitación pública."
                    )
                elif _umb_disponible_c6:
                    st.success(
                        f"✅ **Zona umbral legal:** Importe al **{_pct_sc:.1f} %** del tope "
                        f"(umbral: ${_umb_sc/1e6:.2f} M MXN) — fuera de zona de sospecha"
                    )
                elif str(_c.get("Tipo Simplificado", "")).strip() in {
                    "Adjudicación Directa", "Adjudicación Directa — Fr. I", "Invitación a 3 personas"
                }:
                    st.info("ℹ️ **Zona umbral legal:** Sin datos de umbral PEF para este año/tipo de contratación")
                else:
                    st.success("✅ **Zona umbral legal:** No aplica (licitación pública)")

                # ⑥ Empresa de reciente creación
                st.markdown("---")

                def _parse_rfc_t6(rfc_str):
                    s = str(rfc_str).strip().upper()
                    if not _re.match(r'^[A-ZÑ&]{3}[0-9]{6}[A-Z0-9]{3}$', s):
                        return None
                    yy, mm, dd = int(s[3:5]), int(s[5:7]), int(s[7:9])
                    if mm < 1 or mm > 12 or dd < 1 or dd > 31:
                        return None
                    yr = 2000 + yy if yy <= 30 else 1900 + yy
                    try:
                        return _date(yr, mm, dd)
                    except ValueError:
                        return None

                _fecha_rfc_t6 = _parse_rfc_t6(_rfc_c6)
                _fecha_ini_t6 = None
                _fi_str6 = _c.get("Fecha de inicio del contrato", "")
                if pd.notna(_fi_str6):
                    try:
                        _fecha_ini_t6 = pd.to_datetime(str(_fi_str6), dayfirst=True).date()
                    except Exception:
                        pass

                if _fecha_rfc_t6 is None:
                    st.info("ℹ️ **Reciente creación:** RFC de persona física o formato no analizable")
                elif _fecha_ini_t6 is None:
                    st.info("ℹ️ **Reciente creación:** Fecha de inicio del contrato no disponible")
                else:
                    _edad_t6 = (_fecha_ini_t6 - _fecha_rfc_t6).days
                    if 0 <= _edad_t6 < 365:
                        st.warning(
                            f"🟡 **Empresa de reciente creación:** "
                            f"{_edad_t6} días de antigüedad al inicio del contrato "
                            f"(constituida el {_fecha_rfc_t6})"
                        )
                    else:
                        _anios_t6 = _edad_t6 // 365
                        st.success(
                            f"✅ **Reciente creación:** No aplica "
                            f"({_anios_t6} año(s) de antigüedad al inicio del contrato)"
                        )

# ─────────────────────────────────────────────


# ───────────────────────────────────────────────────────────────
# HELPER: GENERADOR DE PDF PARA FICHA DE EMPRESA
# ───────────────────────────────────────────────────────────────
def _generar_pdf_empresa(nombre_emp, rfc_emp, kpis, df_contratos, anios_label):
    """
    Genera un PDF con la ficha de la empresa:
    KPIs generales + tabla de contratos (top 50 por monto).
    Devuelve bytes del PDF.
    """
    from fpdf import FPDF, XPos, YPos
    import datetime as _dt
    import os as _os

    # ── Selección de fuente con soporte Unicode ─────────────────
    # Orden de preferencia: macOS → Linux → Windows → fallback Latin-1
    _FONT_CANDIDATES = [
        # (regular,                                                   bold,                                                     italic)
        ("/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
         "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
         "/System/Library/Fonts/Supplemental/Arial Italic.ttf"),
        ("/Library/Fonts/Arial Unicode.ttf",
         "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
         "/System/Library/Fonts/Supplemental/Arial Italic.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf"),
        ("/usr/share/fonts/truetype/freefont/FreeSans.ttf",
         "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
         "/usr/share/fonts/truetype/freefont/FreeSansOblique.ttf"),
        ("C:/Windows/Fonts/arialuni.ttf",
         "C:/Windows/Fonts/arialbd.ttf",
         "C:/Windows/Fonts/ariali.ttf"),
    ]

    _font_r = _font_b = _font_i = None
    for _r, _b, _i in _FONT_CANDIDATES:
        if _os.path.isfile(_r):
            _font_r = _r
            _font_b = _b if _os.path.isfile(_b) else _r   # bold fallback = regular
            _font_i = _i if _os.path.isfile(_i) else _r   # italic fallback = regular
            break

    _use_unicode = _font_r is not None

    # ── Sanitizador de texto ─────────────────────────────────────
    # Con fuente Unicode: pasa el texto tal cual.
    # Sin fuente Unicode: reemplaza caracteres fuera de Latin-1.
    _SUBS = {
        "\u2014": "-",   # em dash —
        "\u2013": "-",   # en dash –
        "\u2019": "'",   "\u2018": "'",   # comillas simples curvas
        "\u201c": '"',   "\u201d": '"',   # comillas dobles curvas
        "\u2026": "...", "\u00a0": " ",   # elipsis, espacio no separable
        "\u00ab": '"',   "\u00bb": '"',   # «»
        "\u00b7": ".",   "\u2022": "*",   # punto medio, viñeta
    }

    def _safe(text):
        s = str(text) if text is not None else ""
        if _use_unicode:
            return s
        for orig, repl in _SUBS.items():
            s = s.replace(orig, repl)
        return s.encode("latin-1", errors="replace").decode("latin-1")

    # Alias de posicionamiento fpdf2
    _NL  = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}
    _STY = {"new_x": XPos.RIGHT,   "new_y": YPos.TOP}

    # Nombre del font registrado en el PDF
    _FN = "F" if _use_unicode else "Helvetica"

    class _FichaPDF(FPDF):
        pass

    pdf = _FichaPDF(orientation="P", unit="mm", format="A4")
    if _use_unicode:
        pdf.add_font(_FN, style="",  fname=_font_r)
        pdf.add_font(_FN, style="B", fname=_font_b)
        pdf.add_font(_FN, style="I", fname=_font_i)

    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    pdf.set_margins(12, 12, 12)

    # ── CABECERA INSTITUCIONAL ─────────────────────────────────
    pdf.set_fill_color(11, 84, 69)
    pdf.set_xy(12, 12)
    pdf.set_font(_FN, "B", 9)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(186, 9,
             _safe("  DIVISIÓN DE MONITOREO DE LA INTEGRIDAD INSTITUCIONAL — IMSS"),
             fill=True, **_NL)
    pdf.ln(3)

    # Nombre empresa
    pdf.set_text_color(15, 38, 23)
    pdf.set_font(_FN, "B", 13)
    nombre_disp = nombre_emp if len(nombre_emp) <= 90 else nombre_emp[:87] + "..."
    pdf.cell(186, 8, _safe(nombre_disp), align="C", **_NL)

    pdf.set_font(_FN, "", 9)
    pdf.set_text_color(134, 134, 136)
    pdf.cell(186, 5,
             _safe(f"RFC: {rfc_emp}   |   ComprasMX {anios_label}   |   "
                   f"Generado: {_dt.date.today().strftime('%d/%m/%Y')}"),
             align="C", **_NL)
    pdf.ln(5)

    # ── KPIs ──────────────────────────────────────────────────
    pdf.set_text_color(15, 38, 23)
    pdf.set_font(_FN, "B", 10)
    pdf.cell(186, 6, _safe("ESTADÍSTICAS GENERALES"), **_NL)
    _y_sep = pdf.get_y()
    pdf.set_draw_color(11, 84, 69)
    pdf.line(12, _y_sep, 198, _y_sep)
    pdf.ln(2)

    _kpi_lbls = ["Contratos totales", "Monto total", "Adj. directas", "Instituciones"]
    _kpi_vals = [kpis["n_contratos"], kpis["monto_fmt"], kpis["pct_ad"], kpis["n_inst"]]
    _cw = 46
    for _k in _kpi_lbls:
        pdf.set_font(_FN, "", 7)
        pdf.set_text_color(134, 134, 136)
        pdf.cell(_cw, 4, _safe(_k), align="C", **_STY)
    pdf.ln()
    for _v in _kpi_vals:
        pdf.set_font(_FN, "B", 10)
        pdf.set_text_color(11, 84, 69)
        pdf.cell(_cw, 8, _safe(str(_v)), align="C", **_STY)
    pdf.ln()
    pdf.ln(5)

    # ── TABLA DE CONTRATOS ────────────────────────────────────
    pdf.set_text_color(15, 38, 23)
    pdf.set_font(_FN, "B", 10)
    _n_tot = len(df_contratos)
    _nota = (f" (Top 50 por monto — total: {_n_tot:,})"
             if _n_tot > 50 else f" ({_n_tot:,} contratos)")
    pdf.cell(186, 6, _safe(f"CONTRATOS{_nota}"), **_NL)
    _y_sep2 = pdf.get_y()
    pdf.line(12, _y_sep2, 198, _y_sep2)
    pdf.ln(2)

    # Cabecera tabla
    _cols_h = [
        ("N°",                9,  "C"),
        ("Tipo",              30, "L"),
        ("Importe",           28, "R"),
        ("Unidad Compradora", 54, "L"),
        ("Descripción",       65, "L"),
    ]
    pdf.set_fill_color(11, 84, 69)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(_FN, "B", 7)
    for _hdr, _hw, _ha in _cols_h:
        pdf.cell(_hw, 5.5, _safe(_hdr), fill=True, align=_ha, **_STY)
    pdf.ln()

    # Filas (top 50 por monto)
    _df_show = (
        df_contratos.sort_values("Importe DRC", ascending=False)
        .head(50)
        .reset_index(drop=True)
    )
    for _i, _r in _df_show.iterrows():
        _bg = (245, 250, 248) if _i % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*_bg)
        pdf.set_text_color(23, 27, 25)
        pdf.set_font(_FN, "", 6.5)

        _tipo = _safe(str(_r.get("Tipo Simplificado", ""))[:24])
        _imp_raw = _r.get("Importe DRC", 0)
        try:
            _imp_v = float(_imp_raw) if pd.notna(_imp_raw) else 0.0
        except (ValueError, TypeError):
            _imp_v = 0.0
        _imp  = (f"${_imp_v/1e9:,.2f} mil M" if _imp_v >= 1e9
                 else f"${_imp_v/1e6:,.1f} M")
        _uc   = _safe(str(_r.get("Nombre de la UC", ""))[:42])
        _desc = _safe(str(_r.get("Descripción del contrato", ""))[:56])

        pdf.cell(9,  4.5, str(_i + 1), fill=True, align="C", **_STY)
        pdf.cell(30, 4.5, _tipo,        fill=True,             **_STY)
        pdf.cell(28, 4.5, _imp,         fill=True, align="R",  **_STY)
        pdf.cell(54, 4.5, _uc,          fill=True,             **_STY)
        pdf.cell(65, 4.5, _desc,        fill=True,             **_NL)

    # ── FOOTER ─────────────────────────────────────────────────
    if pdf.get_y() < 268:
        pdf.set_y(268)
    pdf.set_draw_color(11, 84, 69)
    _y_ft = pdf.get_y()
    pdf.line(12, _y_ft, 198, _y_ft)
    pdf.ln(1)
    pdf.set_font(_FN, "I", 7)
    pdf.set_text_color(134, 134, 136)
    pdf.cell(186, 5,
             _safe(f"División de Monitoreo de la Integridad Institucional"
                   f" — IMSS | ComprasMX {anios_label}"),
             align="C")

    return bytes(pdf.output())


# ───────────────────────────────────────────────────────────────
# PÁGINA 7: PAGINA_EMPRESA
# ───────────────────────────────────────────────────────────────
def pagina_empresa():
    st.subheader("🏭 Ficha de la Empresa")
    st.caption(
        "Busca una empresa por nombre o RFC para ver su perfil de riesgo, "
        "numeralia de contratos con todas las instituciones y el detalle de "
        "sus contratos con el IMSS."
    )

    # ── Buscador ──────────────────────────────────────────────
    _busqueda_emp = st.text_input(
        "Buscar empresa",
        placeholder="RFC o nombre del proveedor…",
        key="busq_empresa_pg7"
    )

    if not _busqueda_emp.strip():
        st.info(
            "🔍 Ingresa el nombre o RFC de la empresa para consultar su ficha. "
            "La búsqueda abarca todos los contratos del año seleccionado, "
            "sin importar la institución."
        )
        return

    _q_emp = _busqueda_emp.strip().upper()
    _mask_emp = (
        df["Proveedor o contratista"].str.upper().str.contains(_q_emp, na=False)
        | df["rfc"].str.upper().str.contains(_q_emp, na=False)
    )
    _empresas_df = (
        df[_mask_emp][["Proveedor o contratista", "rfc"]]
        .dropna(subset=["rfc"])
        .drop_duplicates(subset=["rfc"])
        .sort_values("Proveedor o contratista")
        .reset_index(drop=True)
    )

    if len(_empresas_df) == 0:
        st.warning(f"⚠️ No se encontraron empresas que coincidan con «{_busqueda_emp}».")
        return

    _labels_emp = [
        f"{row['Proveedor o contratista']}  —  {str(row['rfc']).strip().upper()}"
        for _, row in _empresas_df.iterrows()
    ]
    _sel_idx = st.selectbox(
        f"Se encontraron {len(_empresas_df):,} empresa(s). Selecciona una:",
        range(len(_labels_emp)),
        format_func=lambda i: _labels_emp[i],
        key="sel_empresa_pg7"
    )
    _rfc_emp    = str(_empresas_df.iloc[_sel_idx]["rfc"]).strip().upper()
    _nombre_emp = _empresas_df.iloc[_sel_idx]["Proveedor o contratista"]

    # Todos los contratos de esta empresa (todas las instituciones, año seleccionado)
    df_emp = df[df["rfc"].str.strip().str.upper() == _rfc_emp].copy()

    st.markdown(f"### 🏭 {_nombre_emp}")
    st.caption(f"RFC: **{_rfc_emp}** · {len(df_emp):,} contratos en el año seleccionado")

    # ── Botón de descarga PDF ──────────────────────────────────
    _n_c_pdf = len(df_emp)
    _mo_pdf  = df_emp["Importe DRC"].sum()
    _nad_pdf = df_emp["Tipo Simplificado"].isin(
        ["Adjudicación Directa", "Adjudicación Directa — Fr. I"]
    ).sum()
    _pct_pdf = (f"{(_nad_pdf/_n_c_pdf*100 if _n_c_pdf > 0 else 0):.1f}%"
                f"  ({_nad_pdf:,} contratos)")
    _ins_pdf = df_emp["Institución"].nunique()
    _kpis_pdf = {
        "n_contratos": f"{_n_c_pdf:,}",
        "monto_fmt":   (f"${_mo_pdf/1e9:,.2f} miles de M MXN"
                        if _mo_pdf >= 1e9 else f"${_mo_pdf/1e6:,.1f} M MXN"),
        "pct_ad":      _pct_pdf,
        "n_inst":      f"{_ins_pdf:,}",
    }
    try:
        _pdf_bytes = _generar_pdf_empresa(
            _nombre_emp, _rfc_emp, _kpis_pdf, df_emp, _anios_label
        )
        st.download_button(
            label="⬇️ Descargar ficha PDF",
            data=_pdf_bytes,
            file_name=f"ficha_empresa_{_rfc_emp}.pdf",
            mime="application/pdf",
            help="Descarga una ficha PDF con estadísticas y los contratos de esta empresa.",
        )
    except Exception as _e_pdf:
        st.caption(f"_No se pudo generar el PDF: {_e_pdf}_")

    # ════════════════════════════════════════════════════════════
    # BLOQUE 1 — PERFIL DE RIESGO
    # ════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 🚨 Perfil de Riesgo")

    _col_sabg_e, _col_efos_e = st.columns(2)

    # ── SABG ──────────────────────────────────────────────────
    with _col_sabg_e:
        st.markdown("**Empresas inhabilitadas (SABG)**")
        try:
            _df_san_e = cargar_sancionados()
            _san_e    = _df_san_e[_df_san_e["RFC"].str.strip() == _rfc_emp]
            if len(_san_e) > 0:
                _niv_san_e  = _san_e.iloc[0]["Nivel de Riesgo"]
                _ini_san_e  = _san_e.iloc[0]["Inicio inhabilitación"]
                _meses_san_e = _san_e.iloc[0]["Meses"]
                if "crítico" in _niv_san_e.lower():
                    st.error(f"🔴 {_niv_san_e}")
                elif "alto" in _niv_san_e.lower():
                    st.warning(f"🟠 {_niv_san_e}")
                elif "medio" in _niv_san_e.lower():
                    st.warning(f"🟡 {_niv_san_e}")
                else:
                    st.info(f"ℹ️ {_niv_san_e}")
                st.caption(
                    f"Inicio de inhabilitación: {_ini_san_e} · "
                    f"Duración: {int(_meses_san_e) if pd.notna(_meses_san_e) else 'N/D'} meses. "
                    "Para verificar violación LAASSP consulta el Expediente de Contrato (criterio Art. 46)."
                )
            else:
                st.success("✅ No aparece en el listado de inhabilitados (SABG)")
        except FileNotFoundError:
            st.info("ℹ️ Archivo SABG no disponible")

    # ── EFOS Art. 69-B CFF ────────────────────────────────────
    with _col_efos_e:
        st.markdown("**Lista Art. 69-B CFF (EFOS — SAT)**")
        try:
            _df_efos_e = cargar_efos()
            _efos_e    = _df_efos_e[_df_efos_e["RFC"] == _rfc_emp]
            if len(_efos_e) > 0:
                _sit_e  = _efos_e.iloc[0]["Situación del contribuyente"]
                _niv_e  = nivel_efos(_sit_e)
                if "definitivo" in _niv_e.lower():
                    st.error(f"🔴 {_niv_e}")
                elif "presunto" in _niv_e.lower():
                    st.warning(f"🟡 {_niv_e}")
                else:
                    st.success(f"✅ {_niv_e}")
                st.caption(f"Situación SAT: {_sit_e}")
            else:
                st.success("✅ No aparece en la lista Art. 69-B CFF (SAT)")
        except FileNotFoundError:
            st.info("ℹ️ Archivo EFOS no disponible")

    # ── Empresa de reciente creación ──────────────────────────
    def _parse_rfc_empresa(rfc_str):
        s = str(rfc_str).strip().upper()
        if not _re.match(r'^[A-ZÑ&]{3}[0-9]{6}[A-Z0-9]{3}$', s):
            return None
        yy, mm, dd = int(s[3:5]), int(s[5:7]), int(s[7:9])
        if mm < 1 or mm > 12 or dd < 1 or dd > 31:
            return None
        yr = 2000 + yy if yy <= 30 else 1900 + yy
        try:
            return _date(yr, mm, dd)
        except ValueError:
            return None

    _fecha_rfc_e = _parse_rfc_empresa(_rfc_emp)
    if _fecha_rfc_e is not None:
        _fechas_ini_e = pd.to_datetime(
            df_emp["Fecha de inicio del contrato"], format="%d/%m/%Y", errors="coerce"
        ).dt.date
        _fecha_primer_e = _fechas_ini_e.dropna().min() if _fechas_ini_e.notna().any() else None
        if _fecha_primer_e is not None:
            _edad_emp = (_fecha_primer_e - _fecha_rfc_e).days
            if 0 <= _edad_emp < 365:
                st.warning(
                    f"🟡 **Empresa de reciente creación:** {_edad_emp} días de antigüedad "
                    f"al inicio de su primer contrato registrado "
                    f"(constituida el {_fecha_rfc_e}, primer contrato: {_fecha_primer_e})"
                )
            else:
                _anios_e = _edad_emp // 365
                st.success(
                    f"✅ **Reciente creación:** No aplica "
                    f"({_anios_e} año(s) de antigüedad al inicio de su primer contrato registrado)"
                )
        else:
            st.info("ℹ️ **Reciente creación:** Fecha de inicio del contrato no disponible")
    else:
        st.info("ℹ️ **Reciente creación:** RFC de persona física o formato no analizable (aplica solo a personas morales)")

    # ════════════════════════════════════════════════════════════
    # BLOQUE 2 — ESTADÍSTICAS GENERALES (TODAS LAS INSTITUCIONES)
    # ════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 📊 Estadísticas Generales (todas las instituciones)")

    _n_contratos_e  = len(df_emp)
    _monto_total_e  = df_emp["Importe DRC"].sum()
    _n_ad_e         = df_emp["Tipo Simplificado"].isin(["Adjudicación Directa", "Adjudicación Directa — Fr. I"]).sum()
    _pct_ad_e       = (_n_ad_e / _n_contratos_e * 100) if _n_contratos_e > 0 else 0
    _n_inst_e       = df_emp["Institución"].nunique()

    _ke1, _ke2, _ke3, _ke4 = st.columns(4)
    _ke1.metric("📄 Contratos totales",        f"{_n_contratos_e:,}")
    _ke2.metric("💰 Monto total",
                f"${_monto_total_e/1e9:,.2f} miles de millones MXN" if _monto_total_e >= 1e9
                else f"${_monto_total_e/1e6:,.1f} M MXN")
    _ke3.metric("🔴 % Adj. Directas",          f"{_pct_ad_e:.1f}%  ({_n_ad_e:,} contratos)")
    _ke4.metric("🏛️ Instituciones",            f"{_n_inst_e:,}")

    _col_pie_e, _col_top5_e = st.columns(2)

    with _col_pie_e:
        _dist_tipo_e = df_emp["Tipo Simplificado"].value_counts().reset_index()
        _dist_tipo_e.columns = ["Tipo", "Contratos"]
        _fig_pie_e = px.pie(
            _dist_tipo_e, names="Tipo", values="Contratos",
            color="Tipo", color_discrete_map=COLORES_TIPO,
            title="Distribución por tipo de procedimiento",
            hole=0.35
        )
        _fig_pie_e.update_traces(
            textinfo="percent",
            textposition="inside",
            insidetextorientation="horizontal",
            textfont=dict(family="Noto Sans, sans-serif", size=13),
            hovertemplate="<b>%{label}</b><br>Contratos: %{value:,}<br>%{percent}<extra></extra>"
        )
        _fig_pie_e.update_layout(
            font=plotly_font(),
            title_font_color=IMSS_VERDE_OSC,
            legend=dict(orientation="v", font=dict(family="Noto Sans, sans-serif", size=11),
                        x=1.01, y=0.5, xanchor="left"),
            margin=dict(r=160)
        )
        st.plotly_chart(_fig_pie_e, use_container_width=True)

    with _col_top5_e:
        # Top 5 contratos individuales (sin agrupar por descripción)
        # para garantizar exactamente 5 barras y evitar colapso de registros.
        _top5_ind = (
            df_emp.nlargest(5, "Importe DRC")
            .reset_index(drop=True)
        )
        _top5_ind["Monto_fmt"] = _top5_ind["Importe DRC"].apply(
            lambda x: (f"${x/1e9:,.2f} miles de M"
                       if pd.notna(x) and x >= 1e9
                       else f"${x/1e6:,.1f} M") if pd.notna(x) else "N/D"
        )
        _top5_ind["Desc_corta"] = _top5_ind["Descripción del contrato"].apply(
            lambda s: str(s)[:50] + "…" if pd.notna(s) and len(str(s)) > 50 else str(s)
        )
        # Usar millones en el eje X para evitar que Plotly muestre "B"
        _top5_ind["Monto_M"] = _top5_ind["Importe DRC"] / 1e6
        _top5_s   = _top5_ind.sort_values("Monto_M")
        _max_m    = _top5_s["Monto_M"].max() if len(_top5_s) > 0 else 1

        _fig_top5_e = go.Figure(go.Bar(
            x=_top5_s["Monto_M"],
            y=_top5_s["Desc_corta"],
            orientation="h",
            marker_color=IMSS_VERDE,
            text=_top5_s["Monto_fmt"],
            textposition="outside",
            cliponaxis=False,
            customdata=_top5_s[["Descripción del contrato"]].values,
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Monto: %{text}<extra></extra>"
            ),
        ))
        _fig_top5_e.update_layout(
            title="Top 5 contratos por monto",
            title_font_color=IMSS_VERDE_OSC,
            font=plotly_font(),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            xaxis=dict(
                title="Monto (millones MXN)",
                # Extender el rango para que el texto quede fuera de las barras
                range=[0, _max_m * 1.55],
                tickformat=",.0f",
                ticksuffix=" M",
                tickfont=dict(family="Noto Sans, sans-serif", size=9),
            ),
            yaxis=dict(tickfont=dict(size=9, family="Noto Sans, sans-serif")),
            height=max(280, len(_top5_s) * 70 + 100),
            margin=dict(l=10, r=30, t=40, b=30),
        )
        st.plotly_chart(_fig_top5_e, use_container_width=True)

    # ── Desglose por institución ──────────────────────────────
    st.markdown("#### 🏛️ Desglose por Institución")
    _inst_group = (
        df_emp.groupby("Institución")
        .agg(Contratos=("Importe DRC", "count"), Monto=("Importe DRC", "sum"))
        .sort_values("Monto", ascending=False)
        .reset_index()
    )
    _inst_group["Monto_fmt"] = _inst_group["Monto"].apply(
        lambda x: f"${x/1e6:,.1f} M"
    )
    _inst_group["Inst_corta"] = _inst_group["Institución"].apply(
        lambda s: s if len(str(s)) <= 55 else str(s)[:55] + "…"
    )
    _fig_inst_e = px.bar(
        _inst_group.sort_values("Monto"),
        x="Monto", y="Inst_corta",
        orientation="h", text="Monto_fmt",
        color_discrete_sequence=[IMSS_VERDE],
        custom_data=["Institución", "Contratos"],
        title="Monto contratado por institución"
    )
    _fig_inst_e.update_traces(
        textfont=dict(family="Noto Sans, sans-serif"),
        textposition="outside", cliponaxis=False,
        hovertemplate="<b>%{customdata[0]}</b><br>Monto: %{text}<br>Contratos: %{customdata[1]:,}<extra></extra>"
    )
    _fig_inst_e.update_layout(
        font=plotly_font(),
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        title_font_color=IMSS_VERDE_OSC,
        xaxis_title="Monto (MXN)", yaxis_title="",
        height=max(320, 35 * len(_inst_group)),
        yaxis=dict(tickfont=dict(size=10, family="Noto Sans, sans-serif"))
    )
    st.plotly_chart(_fig_inst_e, use_container_width=True)

    # ════════════════════════════════════════════════════════════
    # BLOQUE 3 — EVOLUCIÓN HISTÓRICA (MULTI-AÑO)
    # ════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 📈 Evolución Histórica de Contratos")
    st.caption(
        "Carga automáticamente todos los años disponibles (2024–2026). "
        "Los demás filtros del sidebar no se aplican para garantizar comparabilidad."
    )

    _HIST_ARCH_E = {
        "2024": "contratos_compranet_2024.csv",
        "2025": "contratos_comprasmx_2025.csv",
        "2026": "contratos_comprasmx_2026.csv",
    }
    _hist_emp_d = {}
    for _yr_h, _fn_h in _HIST_ARCH_E.items():
        try:
            _dfa_h = cargar_datos(_fn_h)
            _rows_h = _dfa_h[_dfa_h["rfc"].str.strip().str.upper() == _rfc_emp].copy()
            if len(_rows_h) > 0:
                _rows_h["Año"] = _yr_h
                _hist_emp_d[_yr_h] = _rows_h
        except Exception:
            pass

    _IMSS_LABEL = "INSTITUTO MEXICANO DEL SEGURO SOCIAL"

    if len(_hist_emp_d) < 2:
        st.info(
            "ℹ️ Solo se encontraron datos para un año. "
            "El histórico requiere al menos dos años con contratos registrados."
        )
    else:
        _df_hist_e = pd.concat(_hist_emp_d.values(), ignore_index=True)

        # Clasificar IMSS vs Otras instituciones
        _df_hist_e["Ente"] = _df_hist_e["Institución"].apply(
            lambda x: "IMSS" if x == _IMSS_LABEL else "Otras instituciones"
        )

        # ── Gráfica: monto por año, agrupado IMSS vs Otras ────
        _hist_chart_e = (
            _df_hist_e.groupby(["Año", "Ente"])["Importe DRC"]
            .sum().reset_index()
        )
        _hist_chart_e["Monto_fmt"] = _hist_chart_e["Importe DRC"].apply(
            lambda x: f"${x/1e6:,.1f} M"
        )
        _color_ente = {"IMSS": IMSS_VERDE, "Otras instituciones": IMSS_ORO}
        _fig_hist_e = px.bar(
            _hist_chart_e,
            x="Año", y="Importe DRC",
            color="Ente", barmode="group",
            text="Monto_fmt",
            color_discrete_map=_color_ente,
            title="Monto contratado por año (IMSS vs otras instituciones)",
            category_orders={"Año": sorted(_hist_emp_d.keys())}
        )
        _fig_hist_e.update_traces(
            textfont=dict(family="Noto Sans, sans-serif"),
            textposition="outside", cliponaxis=False
        )
        _fig_hist_e.update_layout(
            font=plotly_font(),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            title_font_color=IMSS_VERDE_OSC,
            xaxis=dict(type="category", title="Año"), yaxis_title="Monto (MXN)",
            legend_title_text="",
            legend=dict(font=dict(family="Noto Sans, sans-serif", size=11))
        )
        st.plotly_chart(_fig_hist_e, use_container_width=True)

        # ── Tabla resumen por año ──────────────────────────────
        _resumen_hist_e = (
            _df_hist_e.groupby("Año")
            .agg(
                Contratos   = ("Importe DRC",      "count"),
                Monto       = ("Importe DRC",      "sum"),
                AD          = ("Tipo Simplificado", lambda x: x.isin(["Adjudicación Directa", "Adjudicación Directa — Fr. I"]).sum()),
                IMSS_Monto  = ("Importe DRC",      lambda x: x[_df_hist_e.loc[x.index, "Ente"] == "IMSS"].sum()),
                Inst        = ("Institución",       "nunique")
            )
            .reset_index()
            .sort_values("Año")
        )
        _resumen_hist_e["% AD"]        = (_resumen_hist_e["AD"] / _resumen_hist_e["Contratos"] * 100).round(1).astype(str) + "%"
        _resumen_hist_e["Monto total"] = _resumen_hist_e["Monto"].apply(lambda x: f"${x/1e6:,.1f} M")
        _resumen_hist_e["Monto IMSS"]  = _resumen_hist_e["IMSS_Monto"].apply(lambda x: f"${x/1e6:,.1f} M" if x > 0 else "—")
        _resumen_hist_e = _resumen_hist_e.rename(columns={
            "Contratos": "Contratos", "Inst": "Instituciones"
        })[["Año", "Contratos", "Monto total", "% AD", "Monto IMSS", "Instituciones"]]
        _resumen_hist_e.index = range(1, len(_resumen_hist_e) + 1)
        st.dataframe(_resumen_hist_e, use_container_width=True)

    # ════════════════════════════════════════════════════════════
    # BLOQUE 4 — CONTRATOS CON EL IMSS (AÑO SELECCIONADO)
    # ════════════════════════════════════════════════════════════
    st.divider()
    df_imss_e   = df_emp[df_emp["Institución"] == _IMSS_LABEL].copy()

    if len(df_imss_e) == 0:
        st.info("ℹ️ Esta empresa no tiene contratos registrados con el IMSS en el año seleccionado.")
        return

    st.markdown("#### 🏥 Contratos con el IMSS")

    _n_imss_e     = len(df_imss_e)
    _monto_imss_e = df_imss_e["Importe DRC"].sum()
    _n_ad_imss_e  = df_imss_e["Tipo Simplificado"].isin(["Adjudicación Directa", "Adjudicación Directa — Fr. I"]).sum()
    _pct_ad_imss  = (_n_ad_imss_e / _n_imss_e * 100) if _n_imss_e > 0 else 0
    _n_uc_imss_e  = df_imss_e["Nombre de la UC"].nunique()

    _ki1, _ki2, _ki3, _ki4 = st.columns(4)
    _ki1.metric("📄 Contratos IMSS",     f"{_n_imss_e:,}")
    _ki2.metric("💰 Monto IMSS",
                f"${_monto_imss_e/1e9:,.2f} miles de millones MXN" if _monto_imss_e >= 1e9
                else f"${_monto_imss_e/1e6:,.1f} M MXN")
    _ki3.metric("🔴 % Adj. Directas",   f"{_pct_ad_imss:.1f}%  ({_n_ad_imss_e:,} contratos)")
    _ki4.metric("🏢 UCs contratantes",  f"{_n_uc_imss_e:,}")

    # ── Top-10 UCs por monto ──────────────────────────────────
    _top_uc_e = (
        df_imss_e.groupby("Nombre de la UC")["Importe DRC"]
        .sum().nlargest(10).sort_values()
        .reset_index()
    )
    _top_uc_e["Monto_fmt"] = _top_uc_e["Importe DRC"].apply(lambda x: f"${x/1e6:,.1f} M")
    _top_uc_e["UC_corta"]  = _top_uc_e["Nombre de la UC"].apply(
        lambda s: s if len(str(s)) <= 50 else str(s)[:50] + "…"
    )
    _fig_uc_e = px.bar(
        _top_uc_e, x="Importe DRC", y="UC_corta",
        orientation="h", text="Monto_fmt",
        color_discrete_sequence=[IMSS_VERDE],
        custom_data=["Nombre de la UC"],
        title="Top 10 Unidades Compradoras del IMSS por monto contratado"
    )
    _fig_uc_e.update_traces(
        textfont=dict(family="Noto Sans, sans-serif"),
        textposition="outside", cliponaxis=False,
        hovertemplate="<b>%{customdata[0]}</b><br>Monto: %{text}<extra></extra>"
    )
    _fig_uc_e.update_layout(
        font=plotly_font(),
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        title_font_color=IMSS_VERDE_OSC,
        xaxis_title="Monto (MXN)", yaxis_title="",
        height=max(360, 40 * len(_top_uc_e)),
        yaxis=dict(tickfont=dict(size=10, family="Noto Sans, sans-serif"))
    )
    st.plotly_chart(_fig_uc_e, use_container_width=True)

    # ── Tabla detalle contratos IMSS ──────────────────────────
    with st.expander("📋 Ver detalle de contratos con el IMSS"):
        # Añadir nivel SABG como columna si está disponible
        try:
            _df_san_det = cargar_sancionados()
            _san_det    = _df_san_det[_df_san_det["RFC"].str.strip() == _rfc_emp]
            if len(_san_det) > 0:
                df_imss_e["Nivel SABG"] = _san_det.iloc[0]["Nivel de Riesgo"]
        except (FileNotFoundError, Exception):
            pass

        _cols_det_e = [c for c in [
            "Nivel SABG",
            "Nombre de la UC", "Tipo Procedimiento",
            "Importe DRC", "Descripción del contrato",
            "Fecha de inicio del contrato", "Fecha de fallo",
            "Artículo de excepción",
            "Dirección del anuncio"
        ] if c in df_imss_e.columns]

        _tabla_det_e = (
            df_imss_e[_cols_det_e]
            .sort_values("Importe DRC", ascending=False)
            .reset_index(drop=True)
        )
        _tabla_det_e["Importe DRC"] = _tabla_det_e["Importe DRC"].apply(
            lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
        )
        _tabla_det_e.index += 1
        st.dataframe(
            _tabla_det_e,
            column_config={
                "Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                )
            },
            use_container_width=True
        )


# ═══════════════════════════════════════════════════════════════
# PÁGINA: SIMULACIÓN DE COMPETENCIA / RIESGO DE COLUSIÓN
# ═══════════════════════════════════════════════════════════════

def _construir_grafo_colusion(filas_proc_prov, min_cooc=2):
    """Devuelve (Graph, cooc_Counter) dado list de (proc, prov)."""
    import networkx as _nx
    from itertools import combinations as _comb
    from collections import Counter as _Ctr
    _exp = {}
    for proc, prov in filas_proc_prov:
        _exp.setdefault(proc, set()).add(prov)
    _cooc = _Ctr()
    for provs in _exp.values():
        if len(provs) >= 2:
            for pair in _comb(sorted(provs), 2):
                _cooc[pair] += 1
    G = _nx.Graph()
    for (a, b), w in _cooc.items():
        if w >= min_cooc:
            G.add_edge(a, b, weight=w)
    return G, _cooc, _exp


# ═══════════════════════════════════════════════════════════════
# HELPER: GRAFO INTERACTIVO PyVis
# ═══════════════════════════════════════════════════════════════
# Paleta de 20 colores distintivos para comunidades Louvain
_COLORES_COMUNIDAD = [
    "#E63946", "#457B9D", "#2A9D8F", "#E9C46A", "#F4A261",
    "#264653", "#8338EC", "#06D6A0", "#118AB2", "#FFB703",
    "#FB8500", "#8B2FC9", "#6B9E78", "#D62828", "#3D405B",
    "#81B29A", "#C77DFF", "#4CC9F0", "#F72585", "#7209B7",
]


def _render_grafo_pyvis(G, partition, max_nodos=200,
                        comunidades_sospechosas=None, show_labels=False):
    """
    Genera el HTML de un grafo interactivo de red de proveedores usando PyVis.

    Args:
        G               : networkx Graph (nodos=proveedores, aristas ponderadas)
        partition       : dict {nodo: community_id} resultado de Louvain
        max_nodos       : limitar a los N proveedores con mayor grado ponderado
        comunidades_sospechosas : set de community_id a resaltar (None = todas)
        show_labels     : mostrar etiquetas de texto en los nodos

    Returns:
        str con el HTML completo del grafo PyVis
    """
    import tempfile
    import os
    from pyvis.network import Network

    # Filtrar a los top-N nodos por grado ponderado para rendimiento
    _deg_w_all = dict(G.degree(weight="weight"))
    if len(G.nodes()) > max_nodos:
        _top_nodes = set(
            sorted(_deg_w_all, key=_deg_w_all.get, reverse=True)[:max_nodos]
        )
        G_vis = G.subgraph(_top_nodes).copy()
    else:
        G_vis = G

    if len(G_vis.nodes()) == 0:
        return "<html><body><p>No hay nodos para visualizar.</p></body></html>"

    _deg   = dict(G_vis.degree())
    _deg_w = dict(G_vis.degree(weight="weight"))
    _max_d = max(_deg.values()) if _deg else 1

    net = Network(
        height="580px",
        width="100%",
        bgcolor="#ffffff",
        font_color="#171B19",
        directed=False,
    )
    # Física ForceAtlas2 — similar a Gephi
    net.set_options("""
    {
      "nodes": {
        "font": {"size": 9, "face": "Noto Sans, sans-serif", "color": "#171B19"},
        "borderWidth": 1,
        "borderWidthSelected": 3,
        "shadow": false,
        "shape": "dot"
      },
      "edges": {
        "color": {"inherit": false, "color": "#bbbbbb", "opacity": 0.40},
        "smooth": {"enabled": false},
        "scaling": {"min": 1, "max": 7, "label": {"enabled": false}},
        "hoverWidth": 2
      },
      "physics": {
        "forceAtlas2Based": {
          "gravitationalConstant": -65,
          "centralGravity": 0.005,
          "springLength": 180,
          "springConstant": 0.06,
          "damping": 0.4,
          "avoidOverlap": 0.8
        },
        "maxVelocity": 50,
        "solver": "forceAtlas2Based",
        "timestep": 0.35,
        "stabilization": {
          "enabled": true,
          "iterations": 300,
          "updateInterval": 25,
          "fit": true
        }
      },
      "interaction": {
        "hover": true,
        "tooltipDelay": 80,
        "navigationButtons": false,
        "keyboard": false,
        "zoomView": true
      }
    }
    """)

    # Añadir nodos
    for node in G_vis.nodes():
        comm_id = partition.get(node, 0)
        color   = _COLORES_COMUNIDAD[comm_id % len(_COLORES_COMUNIDAD)]

        # Atenuar si no pertenece a comunidades sospechosas
        if comunidades_sospechosas is not None and comm_id not in comunidades_sospechosas:
            color = "#dddddd"

        deg_v  = _deg.get(node, 1)
        # Tamaño sub-lineal: evita que los hubs dominen visualmente
        size   = 7 + 28 * (deg_v / _max_d) ** 0.55

        label  = (node[:22] + "…" if len(node) > 22 else node) if show_labels else ""
        title  = (
            f"<b>{node}</b><br>"
            f"Conexiones: {deg_v}<br>"
            f"Procedimientos acum.: {int(_deg_w.get(node, 0))}<br>"
            f"Comunidad: {comm_id}"
        )

        net.add_node(
            node,
            label=label,
            title=title,
            color=color,
            size=float(size),
        )

    # Añadir aristas
    for u, v, data in G_vis.edges(data=True):
        w = data.get("weight", 1)
        net.add_edge(u, v, value=float(w),
                     title=f"Procedimientos compartidos: {w}")

    # Generar HTML desde archivo temporal
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".html", delete=False, encoding="utf-8"
    ) as _fh:
        _tmp_path = _fh.name
    net.save_graph(_tmp_path)
    with open(_tmp_path, encoding="utf-8") as _fh:
        html_out = _fh.read()
    os.unlink(_tmp_path)

    return html_out


def pagina_colusion():
    try:
        import networkx as nx
        import community as community_louvain
    except ImportError:
        st.error(
            "⚠️ Librerías requeridas no instaladas. "
            "Ejecuta: pip install networkx python-louvain"
        )
        return

    from itertools import combinations as _comb
    from collections import Counter as _Counter

    st.header("🕸️ Simulación de Competencia — Riesgo de Colusión")

    with st.expander("ℹ️ Metodología y fundamento legal", expanded=False):
        st.markdown(
            """
            **Fundamento legal — Art. 71 Fracción VII LAASSP:** Es causa de impedimento para
            participar en una licitación estar vinculado o pertenecer al mismo grupo empresarial
            que otro licitante del mismo proceso de contratación.

            Se construye un **grafo de co-aparición** donde los nodos son proveedores y las
            aristas representan procedimientos compartidos (peso = número de procedimientos
            en común). Sobre este grafo se calculan cuatro indicadores:

            | Indicador | Descripción | Umbral de alerta |
            |---|---|---|
            | **Pares de co-aparición** | Pares que comparten más procedimientos | Configurable |
            | **Colusión (Louvain)** | Comunidades por algoritmo Louvain | Densidad > 0.6, 3+ miembros |
            | **Concentración en red** | Proveedores con alta centralidad de grado | Degree centrality > 0.20 |
            | **Comunidades inusuales** | Subcomunidades densas y cerradas | Densidad > 0.6, 5+ contratos |
            """
        )

    col_proc = "Número de procedimiento"
    if col_proc not in dff.columns:
        st.warning("⚠️ La columna 'Número de procedimiento' no está disponible.")
        return

    # ── Parámetros globales ───────────────────────────────────────────
    _p1, _p2, _p3, _p4 = st.columns(4)
    with _p1:
        _min_cooc = st.number_input(
            "Mín. procedimientos compartidos (arista)",
            min_value=1, max_value=30, value=2, step=1,
            help="Umbral mínimo para incluir un par en el grafo."
        )
    with _p2:
        _min_density = st.slider("Umbral densidad comunidad", 0.1, 1.0, 0.60, 0.05,
                                  help="Densidad interna mínima para marcar comunidad sospechosa.")
    with _p3:
        _min_dc = st.slider("Umbral degree centrality", 0.05, 1.0, 0.20, 0.05,
                             help="Centralidad de grado mínima para marcar hub dominante.")
    with _p4:
        _uc_opts = ["Todas"] + sorted(dff["Nombre de la UC"].dropna().unique().tolist())
        _uc_f = st.selectbox("🏢 Filtrar por UC", _uc_opts, key="colusion_uc")

    _dff_c = dff.copy() if _uc_f == "Todas" else dff[dff["Nombre de la UC"] == _uc_f].copy()
    _dff_c = _dff_c[[col_proc, "rfc", "Proveedor o contratista", "Nombre de la UC",
                     "Importe DRC", "Tipo Simplificado",
                     "Dirección del anuncio", "Descripción del contrato"]].copy()
    _dff_c["Proveedor o contratista"] = _dff_c["Proveedor o contratista"].astype(str).str.strip()
    _dff_c[col_proc] = _dff_c[col_proc].astype(str).str.strip()
    _dff_c = _dff_c[
        (_dff_c[col_proc] != "") & (_dff_c[col_proc] != "nan") &
        (_dff_c["Proveedor o contratista"] != "") &
        (_dff_c["Proveedor o contratista"] != "nan")
    ]

    if len(_dff_c) == 0:
        st.info("ℹ️ No hay contratos con los filtros actuales.")
        return

    # ── Construir grafo ─────────────────────────────────────────────
    with st.spinner("Construyendo grafo de red de proveedores…"):
        _filas = list(zip(
            _dff_c[col_proc].tolist(),
            _dff_c["Proveedor o contratista"].tolist()
        ))
        G, _cooc, _exp_dict = _construir_grafo_colusion(_filas, min_cooc=int(_min_cooc))

    _n_procs_total = _dff_c[col_proc].nunique()
    _n_procs_multi = sum(1 for v in _exp_dict.values() if len(v) >= 2)
    _n_nodes = G.number_of_nodes()
    _n_edges = G.number_of_edges()
    _pct_multi = _n_procs_multi / _n_procs_total * 100 if _n_procs_total > 0 else 0

    _gk1, _gk2, _gk3, _gk4 = st.columns(4)
    _gk1.metric("📋 Procedimientos analizados", f"{_n_procs_total:,}")
    _gk2.metric("👥 Procs. con 2+ proveedores", f"{_n_procs_multi:,}",
                delta=f"{_pct_multi:.1f}% del total", delta_color="off")
    _gk3.metric("🔵 Nodos en grafo", f"{_n_nodes:,}")
    _gk4.metric("🔗 Aristas activas", f"{_n_edges:,}")

    if _n_nodes == 0:
        st.info("ℹ️ No se encontraron pares. Reduce el mínimo de procedimientos compartidos.")
        return

    # Calcular comunidades Louvain UNA VEZ (se reutilizan en Tab0, Tab2 y Tab4)
    _comm_data = []    # inicializar antes de tabs para que nunca falle
    _partition = {}    # fallback: sin comunidades detectadas
    try:
        _partition = community_louvain.best_partition(G, weight="weight", random_state=42)
        _communities_raw = {}
        for _node, _cid in _partition.items():
            _communities_raw.setdefault(_cid, []).append(_node)

        for _cid, _members in _communities_raw.items():
            if len(_members) < 3:
                continue
            _subg = G.subgraph(_members)
            _density = nx.density(_subg)
            _total_w = sum(d["weight"] for _, _, d in _subg.edges(data=True))
            _n_cont_c = int(_dff_c["Proveedor o contratista"].isin(_members).sum())
            _monto_c  = _dff_c.loc[_dff_c["Proveedor o contratista"].isin(_members), "Importe DRC"].sum()
            # Nombre descriptivo: empresa con más contratos en la comunidad
            _cnt_by_m = {
                m: int(_dff_c["Proveedor o contratista"].eq(m).sum())
                for m in _members
            }
            _top_m = max(_cnt_by_m, key=_cnt_by_m.get)
            _nombre_c = (_top_m[:28] + "…") if len(_top_m) > 28 else _top_m
            _comm_data.append({
                "Comunidad": f"C-{_cid:03d}",
                "Nombre": _nombre_c,
                "N° proveedores": len(_members),
                "Densidad interna": round(_density, 4),
                "Procs. compartidos": _total_w,
                "Contratos": _n_cont_c,
                "Monto (MXN)": _monto_c,
                "🚨 Alerta": "🔴 Sospechosa" if _density >= _min_density else "⚪ Normal",
                "_members": _members,
            })
        _comm_data.sort(key=lambda x: (-x["Densidad interna"], -x["N° proveedores"]))
    except Exception as _louvain_err:
        st.warning(f"⚠️ Error en detección de comunidades Louvain: {_louvain_err}")

    st.divider()

    # ══ TABS ══════════════════════════════════════════════════════════════
    _tab0, _tab1, _tab2, _tab3, _tab4 = st.tabs([
        "🕸️ Grafo Interactivo",
        "👥 Co-aparición de pares",
        "🔴 Colusión — Comunidades Louvain",
        "📡 Concentración en red",
        "🔵 Comunidades inusuales",
    ])

    # ───────────────────────────────────────────────────────────────────
    # TAB 0 — Grafo Interactivo PyVis
    # ───────────────────────────────────────────────────────────────────
    with _tab0:
        st.subheader("🕸️ Grafo Interactivo — Red de Co-aparición")
        st.caption(
            "Cada **punto (nodo)** es un proveedor. Cada **línea (arista)** conecta dos "
            "proveedores que coincidieron en el mismo procedimiento de contratación; "
            "a mayor grosor, más procedimientos en común. "
            "El **color** agrupa a los proveedores en comunidades detectadas. "
            "Arrastra, haz zoom y pasa el cursor para explorar."
        )
        st.info(
            "💡 **Consejo:** Activa *Solo grupos sospechosos* para enfocarte en los "
            "proveedores que más se repiten juntos y representan mayor riesgo de colusión."
        )

        # ── Controles ──────────────────────────────────────────
        _cg1, _cg2, _cg3 = st.columns([2, 1, 1])
        with _cg1:
            _max_n_vis = st.slider(
                "Máx. proveedores en el grafo",
                min_value=50, max_value=min(600, _n_nodes),
                value=min(150, _n_nodes),
                step=25,
                key="col_g_maxn",
                help="Se toman los N proveedores con mayor número de conexiones. Reducir mejora la legibilidad.",
            )
        with _cg2:
            _solo_sosp_g = st.checkbox(
                "Solo grupos sospechosos",
                value=True,   # activo por defecto: filtra el ruido visual
                key="col_g_sosp",
                help=(
                    "Muestra únicamente los grupos de proveedores con alta densidad de conexiones "
                    f"(≥ {_min_density:.2f}). Atenúa el resto para facilitar la lectura."
                ),
            )
        with _cg3:
            _show_lbl_g = st.checkbox(
                "Mostrar etiquetas",
                value=False,
                key="col_g_labels",
                help="Muestra el nombre de cada proveedor. Puede dificultar la lectura en redes grandes.",
            )

        # Determinar comunidades sospechosas para resaltar
        _comms_sosp_g = None
        if _solo_sosp_g and _comm_data and _partition:
            _comms_sosp_g = set()
            for _cr_g in _comm_data:
                if _cr_g["🚨 Alerta"] == "🔴 Sospechosa":
                    for _nd_g in _cr_g["_members"]:
                        _cid_g = _partition.get(_nd_g)
                        if _cid_g is not None:
                            _comms_sosp_g.add(_cid_g)

        with st.spinner("Construyendo grafo interactivo… (puede tardar unos segundos)"):
            _html_grafo = _render_grafo_pyvis(
                G,
                _partition,
                max_nodos=_max_n_vis,
                comunidades_sospechosas=_comms_sosp_g,
                show_labels=_show_lbl_g,
            )

        st.components.v1.html(_html_grafo, height=620, scrolling=False)

        # ── Leyenda de comunidades ──────────────────────────────
        if _comm_data and _partition:
            st.markdown("**Grupos detectados (top 15 por densidad):**")
            _ley_items = _comm_data[:15]
            # Calcular community_id para cada entrada
            _ley_cids = []
            for _lr in _ley_items:
                _cid_l = _partition.get(_lr["_members"][0], 0) if _lr["_members"] else 0
                _ley_cids.append(_cid_l)

            _n_cols_l = min(5, len(_ley_items))
            _ley_cols = st.columns(_n_cols_l)
            for _j, (_lr, _cid_l) in enumerate(zip(_ley_items, _ley_cids)):
                _color_l = _COLORES_COMUNIDAD[_cid_l % len(_COLORES_COMUNIDAD)]
                _icon_l  = "🔴" if _lr["🚨 Alerta"] == "🔴 Sospechosa" else "⚪"
                _label_l = _lr.get("Nombre", _lr["Comunidad"])
                _ley_cols[_j % _n_cols_l].markdown(
                    f'<span style="color:{_color_l};font-size:18px;">●</span> '
                    f'{_icon_l} {_label_l} '
                    f'({_lr["N° proveedores"]} prov., '
                    f'dens. {_lr["Densidad interna"]:.2f})',
                    unsafe_allow_html=True,
                )
        elif not _partition:
            st.info(
                "ℹ️ La detección de comunidades Louvain no está disponible. "
                "Los nodos se muestran sin colorear por comunidad."
            )

    # ───────────────────────────────────────────────────────────────────
    # TAB 1 — Pares de co-aparición
    # ───────────────────────────────────────────────────────────────────
    with _tab1:
        st.subheader("👥 Pares de proveedores — co-aparición en procedimientos")
        st.caption(
            "Proveedores que comparten el mayor número de procedimientos de contratación. "
            "Alta frecuencia de co-aparición sugiere coordinación o pertenencia a una red."
        )

        _pairs_list = [
            {"Proveedor A": a, "Proveedor B": b, "Procedimientos compartidos": n}
            for (a, b), n in _cooc.most_common()
            if n >= int(_min_cooc)
        ]
        _df_pairs = pd.DataFrame(_pairs_list) if _pairs_list else pd.DataFrame(
            columns=["Proveedor A", "Proveedor B", "Procedimientos compartidos"]
        )

        if len(_df_pairs) == 0:
            st.info("No se encontraron pares con el umbral actual.")
        else:
            _n_pares = len(_df_pairs)
            _max_sh  = int(_df_pairs["Procedimientos compartidos"].max())
            st.error(
                f"🔴 **{_n_pares:,}** par(es) con ≥{int(_min_cooc)} procedimientos compartidos. "
                f"El par más frecuente coincide en **{_max_sh:,}** procedimientos. "
                f"Posible **simulación de competencia (Art. 71 Fr. VII LAASSP)**."
            )

            _top_n_t1 = st.selectbox("Mostrar top N pares", [20, 50, 100, 200], index=1, key="col_t1_topn")
            _df_top   = _df_pairs.head(_top_n_t1).copy()
            _df_top["Par"] = (
                _df_top["Proveedor A"].apply(lambda s: s[:28] + "…" if len(s) > 28 else s)
                + "  ↔  "
                + _df_top["Proveedor B"].apply(lambda s: s[:28] + "…" if len(s) > 28 else s)
            )
            _df_top_s = _df_top.sort_values("Procedimientos compartidos")

            _fig_p = go.Figure(go.Bar(
                x=_df_top_s["Procedimientos compartidos"],
                y=_df_top_s["Par"],
                orientation="h",
                marker_color=IMSS_ROJO,
                text=_df_top_s["Procedimientos compartidos"].apply(lambda v: f"{v:,}"),
                textposition="outside",
                customdata=_df_top_s[["Proveedor A", "Proveedor B", "Procedimientos compartidos"]].values,
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "    ↔  <b>%{customdata[1]}</b><br>"
                    "Procedimientos compartidos: <b>%{customdata[2]:,}</b><extra></extra>"
                ),
            ))
            _fig_p.update_layout(
                xaxis_title="Procedimientos compartidos",
                yaxis_title="",
                height=max(400, len(_df_top_s) * 22 + 120),
                margin=dict(l=10, r=80, t=30, b=30),
                font=plotly_font(),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            )
            st.plotly_chart(_fig_p, use_container_width=True)

            st.subheader("📋 Tabla de pares")
            _tbl_pairs = _df_pairs.head(_top_n_t1).copy()
            _tbl_pairs.index = range(1, len(_tbl_pairs) + 1)
            st.dataframe(_tbl_pairs, use_container_width=True)

            st.subheader("🔍 Detalle de procedimientos compartidos")
            _opciones_par = [
                f"{row['Proveedor A']}  ↔  {row['Proveedor B']}"
                for _, row in _df_pairs.head(100).iterrows()
            ]
            _par_sel = st.selectbox("Selecciona un par:", _opciones_par, key="col_t1_par")
            _idx     = _opciones_par.index(_par_sel)
            _pa      = _df_pairs.iloc[_idx]["Proveedor A"]
            _pb      = _df_pairs.iloc[_idx]["Proveedor B"]

            _procs_a = set(_dff_c.loc[_dff_c["Proveedor o contratista"] == _pa, col_proc])
            _procs_b = set(_dff_c.loc[_dff_c["Proveedor o contratista"] == _pb, col_proc])
            _sh      = _procs_a & _procs_b

            _mask_sh = (
                _dff_c[col_proc].isin(_sh) &
                _dff_c["Proveedor o contratista"].isin([_pa, _pb])
            )
            _det_t1 = (
                _dff_c[_mask_sh]
                [[col_proc, "Proveedor o contratista", "Nombre de la UC",
                  "Tipo Simplificado", "Importe DRC",
                  "Descripción del contrato", "Dirección del anuncio"]]
                .sort_values([col_proc, "Proveedor o contratista"])
                .rename(columns={
                    col_proc: "Procedimiento",
                    "Proveedor o contratista": "Proveedor",
                    "Nombre de la UC": "UC",
                    "Tipo Simplificado": "Tipo",
                    "Importe DRC": "Importe",
                    "Descripción del contrato": "Descripción",
                })
            )
            _det_t1["Importe"] = _det_t1["Importe"].apply(
                lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
            )
            _monto_par = _dff_c[_mask_sh]["Importe DRC"].sum()
            _cm1, _cm2 = st.columns(2)
            _cm1.metric(
                "💰 Monto total involucrado",
                f"${_monto_par/1e9:,.2f} miles de millones MXN" if _monto_par >= 1e9
                else f"${_monto_par/1e6:,.1f} M MXN"
            )
            _cm2.metric("📋 Contratos en procedimientos compartidos", f"{len(_det_t1):,}")
            _det_t1.index = range(1, len(_det_t1) + 1)
            st.dataframe(
                _det_t1,
                column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                )},
                use_container_width=True,
            )

    # ───────────────────────────────────────────────────────────────────
    # TAB 2 — Colusión por Louvain
    # ───────────────────────────────────────────────────────────────────
    with _tab2:
        st.subheader("🔴 Grupos de proveedores con vínculos inusuales")
        st.caption(
            "Un **grupo (comunidad)** es un conjunto de proveedores que coinciden repetidamente "
            "en los mismos procedimientos de contratación. Cuando varios proveedores que "
            "deberían competir entre sí aparecen juntos de forma sistemática, puede indicar "
            "coordinación o simulación de competencia (Art. 71 Fr. VII LAASSP)."
        )

        with st.expander("ℹ️ ¿Qué es la densidad de una comunidad?"):
            st.markdown(
                """
                La **densidad** mide qué tan "cerrado" es un grupo de proveedores:

                - **Densidad = 1.0** → todos los proveedores del grupo coincidieron entre sí en algún procedimiento.
                - **Densidad = 0.0** → ningún par de proveedores dentro del grupo comparte procedimientos.
                - **Densidad ≥ 0.60** → al menos el 60% de los posibles pares del grupo han coincidido. Esto es **atípico** en una licitación abierta donde los proveedores deberían competir de forma independiente.

                Un grupo con alta densidad sugiere que sus miembros no compiten realmente entre sí, sino que participan juntos de forma coordinada.
                """
            )

        if not _comm_data:
            st.info("No se detectaron grupos de 3+ proveedores con el umbral actual.")
        else:
            _df_comm = pd.DataFrame([{k: v for k, v in r.items() if k != "_members"}
                                      for r in _comm_data])
            _n_alert_c  = (_df_comm["🚨 Alerta"] == "🔴 Sospechosa").sum()
            _prov_alert = set(m for r in _comm_data if r["🚨 Alerta"] == "🔴 Sospechosa"
                              for m in r["_members"])

            _lk1, _lk2, _lk3 = st.columns(3)
            _lk1.metric("🔵 Grupos detectados", f"{len(_comm_data):,}")
            _lk2.metric("🔴 Grupos sospechosos", f"{_n_alert_c:,}",
                        help=f"Densidad interna ≥ {_min_density:.2f} — vínculos muy frecuentes entre sus miembros")
            _lk3.metric("👥 Proveedores en alerta", f"{len(_prov_alert):,}")

            if _n_alert_c > 0:
                st.error(
                    f"🔴 **{_n_alert_c}** grupo(s) con densidad ≥ {_min_density:.2f}: "
                    f"sus proveedores coinciden juntos de forma sistemática, "
                    f"lo que es inusual en un mercado con competencia real."
                )
            else:
                st.success(
                    f"✅ Ningún grupo supera el umbral de densidad {_min_density:.2f}. "
                    f"Ajusta el umbral o el mínimo de procedimientos compartidos si deseas explorar más."
                )

            st.dataframe(
                _df_comm[["Comunidad", "Nombre", "N° proveedores", "Densidad interna",
                           "Procs. compartidos", "Contratos", "🚨 Alerta"]],
                use_container_width=True,
                column_config={
                    "Nombre": st.column_config.TextColumn(
                        "Empresa principal",
                        help="Empresa con más contratos dentro del grupo (referencia descriptiva)"
                    ),
                    "Densidad interna": st.column_config.NumberColumn(
                        "Densidad", format="%.2f",
                        help="Proporción de pares que coincidieron en ≥1 procedimiento (0=ninguno, 1=todos)"
                    ),
                },
            )

            # Detalle por comunidad
            st.subheader("🔍 Detalle de grupo")
            _comm_opts = [
                f"{r['Comunidad']} — {r['Nombre']} y {r['N° proveedores']-1} más "
                f"(densidad {r['Densidad interna']:.2f}) {r['🚨 Alerta']}"
                for r in _comm_data
            ]
            _cs = st.selectbox("Selecciona un grupo:", _comm_opts, key="col_t2_comm")
            _ci = _comm_opts.index(_cs)
            _cr = _comm_data[_ci]
            _mems = _cr["_members"]

            st.markdown(
                f"**Proveedores del grupo {_cr['Comunidad']}** "
                f"(densidad {_cr['Densidad interna']:.2f} — {_cr['🚨 Alerta']}):"
            )
            for _m in sorted(_mems):
                _nc = int(_dff_c["Proveedor o contratista"].eq(_m).sum())
                st.markdown(f"- {_m} — {_nc:,} contratos")

            with st.expander("📋 Ver contratos de esta comunidad"):
                _det_c = (
                    _dff_c[_dff_c["Proveedor o contratista"].isin(_mems)]
                    [[col_proc, "Proveedor o contratista", "Nombre de la UC",
                      "Tipo Simplificado", "Importe DRC",
                      "Descripción del contrato", "Dirección del anuncio"]]
                    .sort_values([col_proc, "Proveedor o contratista"])
                    .rename(columns={col_proc: "Procedimiento",
                                      "Proveedor o contratista": "Proveedor",
                                      "Nombre de la UC": "UC",
                                      "Tipo Simplificado": "Tipo",
                                      "Importe DRC": "Importe",
                                      "Descripción del contrato": "Descripción"})
                )
                _det_c["Importe"] = _det_c["Importe"].apply(
                    lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
                )
                _det_c.index = range(1, len(_det_c) + 1)
                st.dataframe(
                    _det_c,
                    column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                        "🔗 ComprasMX", display_text="Ver contrato"
                    )},
                    use_container_width=True,
                )

    # ───────────────────────────────────────────────────────────────────
    # TAB 3 — Concentración en red (Degree Centrality)
    # ───────────────────────────────────────────────────────────────────
    with _tab3:
        st.subheader("📡 Proveedores con demasiadas conexiones")
        st.caption(
            "Identifica proveedores que coinciden con un número inusualmente alto de "
            "otros proveedores a lo largo de los procedimientos de contratación. "
            "En un mercado competitivo, cada empresa debería ganar licitaciones "
            "de forma independiente — si una empresa aparece repetidamente junto a "
            "muchos otros proveedores distintos, puede estar actuando como coordinador "
            "de una red de simulación de competencia."
        )
        with st.expander("ℹ️ ¿Qué es el Degree Centrality?"):
            st.markdown(
                """
                El **Degree Centrality** (centralidad de grado) es un número entre 0 y 1 que indica
                qué fracción del total de proveedores en la red tiene vínculos directos con
                el proveedor analizado:

                - **Centralidad = 0.0** → el proveedor no comparte ningún procedimiento con otros.
                - **Centralidad = 1.0** → el proveedor coincidió con *todos* los demás proveedores de la red.
                - **Centralidad ≥ 0.20** (umbral predeterminado) → el proveedor tiene vínculos con al
                  menos el 20% del universo de proveedores, lo que es atípicamente alto y merece revisión.

                Un proveedor con centralidad muy alta podría estar funcionando como "hub" o nodo
                central de una red coordinada de adjudicaciones.
                """
            )

        _dc_dict   = nx.degree_centrality(G)
        _deg_w     = dict(G.degree(weight="weight"))
        _n_vecinos = dict(G.degree())

        _dc_rows = []
        for _prov, _dval in _dc_dict.items():
            _nc_p = int(_dff_c["Proveedor o contratista"].eq(_prov).sum())
            _mo_p = _dff_c.loc[_dff_c["Proveedor o contratista"] == _prov, "Importe DRC"].sum()
            _dc_rows.append({
                "Proveedor": _prov,
                "Degree centrality": round(_dval, 4),
                "Vecinos distintos": _n_vecinos[_prov],
                "Procs. acumulados": _deg_w[_prov],
                "Contratos": _nc_p,
                "Monto (MXN)": _mo_p,
                "Alerta": "🔴 Hub dominante" if _dval >= _min_dc else "⚪ Normal",
            })

        _df_dc = pd.DataFrame(_dc_rows).sort_values("Degree centrality", ascending=False)
        _df_dc_alerta = _df_dc[_df_dc["Alerta"] == "🔴 Hub dominante"]

        _dk1, _dk2, _dk3 = st.columns(3)
        _dk1.metric("🔴 Hubs dominantes", f"{len(_df_dc_alerta):,}",
                    help=f"Degree centrality ≥ {_min_dc:.2f}")
        if len(_df_dc) > 0:
            _top_hub = _df_dc.iloc[0]
            _dk2.metric("🏆 Mayor degree centrality",
                        f"{float(_top_hub['Degree centrality']):.3f}",
                        delta=_top_hub["Proveedor"][:35], delta_color="off")
            _dk3.metric("🔗 Mayor nº de vecinos",
                        f"{int(_df_dc['Vecinos distintos'].max()):,}")

        if len(_df_dc_alerta) > 0:
            st.error(
                f"🔴 **{len(_df_dc_alerta)}** proveedor(es) con centralidad ≥ {_min_dc:.2f}: "
                f"coinciden con un porcentaje inusualmente alto de los demás proveedores de la red. "
                f"Posible nodo coordinador de simulación de competencia."
            )

        _top_n_dc = st.selectbox("Top N proveedores", [20, 30, 50], index=1, key="col_t3_topn")
        _df_dc_top = _df_dc.head(_top_n_dc).sort_values("Degree centrality")
        _dc_colors = [
            IMSS_ROJO if r == "🔴 Hub dominante" else IMSS_VERDE
            for r in _df_dc_top["Alerta"]
        ]
        _prov_corto = _df_dc_top["Proveedor"].apply(lambda s: s[:35] + "…" if len(s) > 35 else s)

        _fig_dc = go.Figure(go.Bar(
            x=_df_dc_top["Degree centrality"],
            y=_prov_corto,
            orientation="h",
            marker_color=_dc_colors,
            text=_df_dc_top["Degree centrality"].apply(lambda v: f"{v:.3f}"),
            textposition="outside",
            customdata=_df_dc_top[["Proveedor", "Vecinos distintos", "Contratos"]].values,
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Degree centrality: <b>%{x:.4f}</b><br>"
                "Vecinos: %{customdata[1]:,} | Contratos: %{customdata[2]:,}<extra></extra>"
            ),
        ))
        _fig_dc.add_vline(
            x=_min_dc, line_dash="dash", line_color=IMSS_ORO, line_width=2,
            annotation_text=f"Umbral {_min_dc:.2f}",
            annotation_position="top right",
            annotation_font=dict(color=IMSS_ORO),
        )
        _fig_dc.update_layout(
            xaxis_title="Degree centrality",
            yaxis_title="",
            height=max(400, _top_n_dc * 22 + 120),
            margin=dict(l=10, r=80, t=40, b=30),
            font=plotly_font(),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        )
        st.plotly_chart(_fig_dc, use_container_width=True)

        _df_dc_show = _df_dc_alerta.copy() if len(_df_dc_alerta) > 0 else _df_dc.head(30).copy()
        _df_dc_show["Monto (MXN)"] = _df_dc_show["Monto (MXN)"].apply(
            lambda x: f"${x/1e6:,.1f} M" if pd.notna(x) else "N/D"
        )
        _df_dc_show = _df_dc_show.drop(columns=["Alerta"])
        _df_dc_show.index = range(1, len(_df_dc_show) + 1)
        st.dataframe(_df_dc_show, use_container_width=True)

    # ───────────────────────────────────────────────────────────────────
    # TAB 4 — Comunidades inusuales
    # ───────────────────────────────────────────────────────────────────
    with _tab4:
        st.subheader("🔵 Comunidades inusuales — Subcomunidades densas y cerradas")
        st.caption(
            "Subcomunidades con alta densidad interna (3+ proveedores, densidad > umbral, "
            "5+ contratos): los miembros se repiten sistemáticamente entre sí "
            "en lugar de competir abiertamente."
        )

        if not _comm_data:
            st.info("No se detectaron comunidades. Reduce el umbral de procedimientos compartidos.")
        else:
            _min_cont_t4 = st.number_input(
                "Mínimo de contratos en la comunidad",
                min_value=1, max_value=500, value=5, step=1,
                key="col_t4_mincont",
            )
            _inusuales = [
                r for r in _comm_data
                if r["Densidad interna"] >= _min_density and r["Contratos"] >= _min_cont_t4
            ]

            if not _inusuales:
                st.success(
                    f"✅ No se encontraron comunidades con densidad ≥ {_min_density:.2f} "
                    f"y ≥ {int(_min_cont_t4)} contratos con el umbral actual."
                )
            else:
                st.error(
                    f"🔴 **{len(_inusuales)}** comunidad(es) inusual(es): "
                    f"alta densidad interna y {int(_min_cont_t4)}+ contratos. "
                    f"Posible grupo cerrado de proveedores coordinados."
                )

                _df_inu = pd.DataFrame([
                    {k: v for k, v in r.items() if k != "_members"} for r in _inusuales
                ])

                # Burbuja: densidad vs. tamaño
                _fig_bub = px.scatter(
                    _df_inu,
                    x="Densidad interna", y="N° proveedores",
                    size="Contratos", color="Procs. compartidos",
                    hover_name="Comunidad",
                    color_continuous_scale=[[0, IMSS_ORO_CLARO], [0.5, IMSS_ORO], [1.0, IMSS_ROJO]],
                    title="Comunidades inusuales — densidad vs. tamaño",
                    labels={"Densidad interna": "Densidad interna",
                            "N° proveedores": "N° de proveedores",
                            "Procs. compartidos": "Procs. compartidos"},
                )
                _fig_bub.update_layout(
                    font=plotly_font(), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                    coloraxis_showscale=False,
                )
                _fig_bub.add_vline(x=_min_density, line_dash="dash",
                                    line_color=IMSS_ROJO, line_width=1.5)
                st.plotly_chart(_fig_bub, use_container_width=True)

                for _row_u in _inusuales:
                    _mems_u  = _row_u["_members"]
                    _monto_u = _row_u["Monto (MXN)"]
                    with st.expander(
                        f"**{_row_u['Comunidad']}** — "
                        f"{_row_u['N° proveedores']} proveedores, "
                        f"densidad {_row_u['Densidad interna']:.2f}, "
                        f"{_row_u['Contratos']:,} contratos, "
                        f"${_monto_u/1e6:,.1f} M MXN"
                    ):
                        for _mu in sorted(_mems_u):
                            _nc_u = int(_dff_c["Proveedor o contratista"].eq(_mu).sum())
                            st.markdown(f"- {_mu} — {_nc_u:,} contratos")

                        _procs_u = (
                            _dff_c[_dff_c["Proveedor o contratista"].isin(_mems_u)]
                            .groupby(col_proc)["Proveedor o contratista"].nunique()
                        )
                        _procs_int = _procs_u[_procs_u >= 2].index
                        if len(_procs_int) > 0:
                            st.markdown(
                                f"**{len(_procs_int)} procedimiento(s) con 2+ miembros de esta comunidad:**"
                            )
                            _det_u = (
                                _dff_c[
                                    _dff_c[col_proc].isin(_procs_int) &
                                    _dff_c["Proveedor o contratista"].isin(_mems_u)
                                ]
                                [[col_proc, "Proveedor o contratista", "Nombre de la UC",
                                  "Importe DRC", "Dirección del anuncio"]]
                                .sort_values([col_proc, "Proveedor o contratista"])
                                .rename(columns={
                                    col_proc: "Procedimiento",
                                    "Proveedor o contratista": "Proveedor",
                                    "Nombre de la UC": "UC",
                                    "Importe DRC": "Importe",
                                })
                            )
                            _det_u["Importe"] = _det_u["Importe"].apply(
                                lambda x: f"${x:,.0f}" if pd.notna(x) else "N/D"
                            )
                            _det_u.index = range(1, len(_det_u) + 1)
                            st.dataframe(
                                _det_u,
                                column_config={"Dirección del anuncio": st.column_config.LinkColumn(
                                    "🔗 ComprasMX", display_text="Ver contrato"
                                )},
                                use_container_width=True,
                            )

    st.divider()

# ───────────────────────────────────────────────────────────────
# PÁGINA: RANKING DE RIESGO COMPUESTO
# ───────────────────────────────────────────────────────────────
def pagina_ranking_riesgo():
    import re as _re_rk
    from datetime import date as _date_rk

    st.header("🏆 Ranking de Riesgo Compuesto")
    st.caption(
        "Contratos ordenados por su acumulación de indicadores de riesgo. "
        "El score combina inhabilitaciones SABG (Art. 46 LAASSP), EFOS Art. 69-B CFF, "
        "empresas de reciente creación, zona umbral legal y tipo de procedimiento, "
        "ponderado por el monto del contrato."
    )

    _c_rk1, _c_rk2 = st.columns([3, 1])
    _ucs_rk = ["Todas"] + sorted(dff["Nombre de la UC"].dropna().unique().tolist())
    _uc_rk  = _c_rk1.selectbox(
        "🏢 Filtrar por Unidad Compradora (opcional):",
        _ucs_rk, key="rk_uc_sel",
    )
    _top_n_rk = _c_rk2.selectbox(
        "Mostrar top:", [20, 50, 100, 200], index=0, key="rk_top_n",
        help="Número de contratos a mostrar en el ranking",
    )
    _dff_rk = (
        dff[dff["Nombre de la UC"] == _uc_rk].copy()
        if _uc_rk != "Todas" else dff.copy()
    )
    if len(_dff_rk) == 0:
        st.info("ℹ️ Sin contratos para la selección actual.")
        return

    with st.expander("ℹ️ Metodología — Score de riesgo compuesto", expanded=False):
        st.markdown(
            """
            Cada contrato recibe un **score de riesgo base** calculado como la suma de los
            indicadores de alerta que activan, y luego se **pondera por el monto** del contrato
            (contratos de mayor valor amplifican el riesgo):

            | Indicador | Puntos | Fundamento |
            |---|---|---|
            | 🔴 SABG — Inhabilitación vigente | 100 | Contratación legalmente prohibida. Fecha de fallo (o firma en AD) ≥ inicio inhabilitación (Art. 46 LAASSP) |
            | 🟠 SABG — Inhabilitación suspendida | 60 | Resolución judicial pendiente |
            | 🟡 SABG — Historial de inhabilitación | 30 | Antecedente de sanción |
            | 🔴 EFOS definitivo (Art. 69-B CFF) | 80 | Operaciones simuladas confirmadas por el SAT |
            | 🟡 EFOS presunto | 40 | Investigación SAT en curso |
            | 🚦 Zona umbral legal (90–100 % del tope) | 45 | Monto justo por debajo del límite que exigiría licitación |
            | 🟡 Empresa < 1 año de constitución | 30 | Posible empresa creada *ad hoc* |
            | 🔴 Adjudicación directa (sin concurso) | 20 | Proceso sin competencia abierta |
            | 🟡 Invitación a 3 personas | 5 | Competencia restringida |

            > **SABG — fecha de referencia:** Para licitaciones e invitaciones a tres personas se usa
            la *fecha de fallo*; para adjudicaciones directas (que no generan fallo formal) se usa la
            *fecha de firma del contrato*. Fundamento: Art. 46 LAASSP.

            **Ponderación por monto:**
            `Score_final = Score_base × (1 + 0.5 × percentil_monto)` — contratos en el percentil 99
            de monto reciben hasta un 50 % más de peso.
            El score se normaliza a **0–100**. Un contrato puede acumular alertas de múltiples indicadores.
            """
        )

    # ════════════════════════════════════════════════════════════
    # CÁLCULO DE SCORES
    # ════════════════════════════════════════════════════════════
    with st.spinner("Calculando scores de riesgo…"):

        # ── 1. Indicadores externos ────────────────────────────
        # SABG: dict RFC → LISTA de (inicio_dt, nivel_str)
        # BUG FIX: almacenar TODOS los registros por RFC, no solo el primero.
        # Una empresa puede tener múltiples inhabilitaciones en distintos períodos.
        _sabg_records_rk = {}
        try:
            _df_sabg_rk = cargar_sancionados()
            for _, _sr in _df_sabg_rk.iterrows():
                _ru2 = str(_sr["RFC"]).strip().upper()
                _ini_d = pd.to_datetime(_sr.get("Inicio inhabilitación"), errors="coerce")
                _niv2  = str(_sr.get("Nivel de Riesgo", ""))
                if _ru2:
                    _sabg_records_rk.setdefault(_ru2, []).append((_ini_d, _niv2))
        except Exception:
            pass

        try:
            _df_efos_rk = cargar_efos()
            _rfcs_ed_rk = set(_df_efos_rk[
                _df_efos_rk["Situación del contribuyente"] == "Definitivo"]["RFC"])
            _rfcs_ep_rk = set(_df_efos_rk[
                _df_efos_rk["Situación del contribuyente"] == "Presunto"]["RFC"])
        except Exception:
            _rfcs_ed_rk = _rfcs_ep_rk = set()

        # ── 2. Flags vectorizados ──────────────────────────────
        _rfc_n_rk = _dff_rk["rfc"].astype(str).str.strip().str.upper()

        # dayfirst=True sin format= para manejar tanto DD/MM/YYYY (2026)
        # como ISO datetime (2025: "2025-10-09 00:00:00")
        _fecha_fallo_rk = pd.to_datetime(
            _dff_rk["Fecha de fallo"] if "Fecha de fallo" in _dff_rk.columns
            else pd.Series(pd.NaT, index=_dff_rk.index),
            dayfirst=True, errors="coerce",
        )
        _fecha_firma_rk = pd.to_datetime(
            _dff_rk["Fecha de firma del contrato"]
            if "Fecha de firma del contrato" in _dff_rk.columns
            else pd.Series(pd.NaT, index=_dff_rk.index),
            dayfirst=True, errors="coerce",
        )
        _fecha_ref_rk = _fecha_fallo_rk.combine_first(_fecha_firma_rk)

        # SABG flags — evalúa TODOS los registros por RFC (fix del bug original)
        _f_sc_l, _f_sa_l, _f_sm_l, _f_sb_prior_l, _f_sb_unk_l = [], [], [], [], []
        for _ix2 in _dff_rk.index:
            _ru2   = _rfc_n_rk.loc[_ix2]
            _fref2 = _fecha_ref_rk.loc[_ix2]
            if _ru2 not in _sabg_records_rk:
                _f_sc_l.append(False); _f_sa_l.append(False); _f_sm_l.append(False)
                _f_sb_prior_l.append(False); _f_sb_unk_l.append(False)
                continue
            if pd.isna(_fref2):
                _f_sc_l.append(False); _f_sa_l.append(False); _f_sm_l.append(False)
                _f_sb_prior_l.append(False); _f_sb_unk_l.append(True)
                continue
            # Evaluar TODOS los registros de inhabilitación del RFC
            _sc2 = _sa2 = _sm2 = False
            _any_viol = False
            for _ini_dt2, _niv2 in _sabg_records_rk[_ru2]:
                if not pd.isna(_ini_dt2) and _fref2 < _ini_dt2:
                    continue  # contrato anterior a este período de inhabilitación
                # Violación: fecha dentro del período (o fecha de inicio desconocida)
                _any_viol = True
                _niv2_l = _niv2.lower()
                if "crítico" in _niv2_l:
                    _sc2 = True
                elif "alto" in _niv2_l:
                    _sa2 = True
                elif "medio" in _niv2_l:
                    _sm2 = True
            if not _any_viol:
                _f_sc_l.append(False); _f_sa_l.append(False); _f_sm_l.append(False)
                _f_sb_prior_l.append(True); _f_sb_unk_l.append(False)
            else:
                _f_sc_l.append(_sc2)
                _f_sa_l.append(_sa2 and not _sc2)
                _f_sm_l.append(_sm2 and not _sc2 and not _sa2)
                _f_sb_prior_l.append(False); _f_sb_unk_l.append(False)

        _f_sc_rk       = pd.Series(_f_sc_l,       index=_dff_rk.index)
        _f_sa_rk       = pd.Series(_f_sa_l,       index=_dff_rk.index)
        _f_sm_rk       = pd.Series(_f_sm_l,       index=_dff_rk.index)
        _f_sb_prior_rk = pd.Series(_f_sb_prior_l, index=_dff_rk.index)
        _f_sb_unk_rk   = pd.Series(_f_sb_unk_l,   index=_dff_rk.index)

        _f_ed_rk = _rfc_n_rk.isin(_rfcs_ed_rk)
        _f_ep_rk = _rfc_n_rk.isin(_rfcs_ep_rk)

        # Zona umbral legal
        _f_umb_rk = pd.Series(False, index=_dff_rk.index)
        try:
            _umbrales_pef_rk = cargar_umbrales_pef()
            if _umbrales_pef_rk:
                _TIPOS_AD_RK   = {"Adjudicación Directa", "Adjudicación Directa — Fr. I"}
                _TIPOS_I3P_RK  = {"Invitación a 3 personas"}
                _TIPOS_SERV_RK = {"SERVICIOS", "SERVICIOS RELACIONADOS CON LA OBRA", "ARRENDAMIENTOS"}
                _mask_proc_rk  = _dff_rk["Tipo Simplificado"].isin(_TIPOS_AD_RK | _TIPOS_I3P_RK)
                _dff_u_rk = _dff_rk[_mask_proc_rk].copy()
                _dff_u_rk["_fecha_ur"] = pd.to_datetime(
                    _dff_u_rk["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
                )
                _dff_u_rk["_año_ur"] = _dff_u_rk["_fecha_ur"].dt.year
                _dff_u_rk["_ley_ur"] = (
                    _dff_u_rk["Ley"].astype(str).str.strip().str.upper()
                    if "Ley" in _dff_u_rk.columns else "LAASSP"
                )
                _dff_u_rk["_cont_ur"] = (
                    _dff_u_rk["Tipo de contratación"].astype(str).str.strip().str.upper()
                    if "Tipo de contratación" in _dff_u_rk.columns else "ADQUISICIONES"
                )
                _dff_u_rk["_proc_ur"]   = _dff_u_rk["Tipo Simplificado"]
                _dff_u_rk["_umbral_ur"] = float("nan")
                for _año_rk2, _th_rk2 in _umbrales_pef_rk.items():
                    _ma_rk2 = _dff_u_rk["_año_ur"] == _año_rk2
                    _dff_u_rk.loc[
                        _ma_rk2 & _dff_u_rk["_ley_ur"].eq("LAASSP")
                        & _dff_u_rk["_proc_ur"].isin(_TIPOS_AD_RK),
                        "_umbral_ur"] = _th_rk2["ad_laassp"]
                    _dff_u_rk.loc[
                        _ma_rk2 & _dff_u_rk["_ley_ur"].eq("LAASSP")
                        & _dff_u_rk["_proc_ur"].isin(_TIPOS_I3P_RK),
                        "_umbral_ur"] = _th_rk2["i3p_laassp"]
                    _dff_u_rk.loc[
                        _ma_rk2 & _dff_u_rk["_ley_ur"].eq("LOPSRM")
                        & _dff_u_rk["_proc_ur"].isin(_TIPOS_AD_RK)
                        & _dff_u_rk["_cont_ur"].eq("OBRA PÚBLICA"),
                        "_umbral_ur"] = _th_rk2["ad_obra_lopsrm"]
                    _dff_u_rk.loc[
                        _ma_rk2 & _dff_u_rk["_ley_ur"].eq("LOPSRM")
                        & _dff_u_rk["_proc_ur"].isin(_TIPOS_AD_RK)
                        & _dff_u_rk["_cont_ur"].isin(_TIPOS_SERV_RK),
                        "_umbral_ur"] = _th_rk2["ad_serv_lopsrm"]
                    _dff_u_rk.loc[
                        _ma_rk2 & _dff_u_rk["_ley_ur"].eq("LOPSRM")
                        & _dff_u_rk["_proc_ur"].isin(_TIPOS_I3P_RK)
                        & _dff_u_rk["_cont_ur"].eq("OBRA PÚBLICA"),
                        "_umbral_ur"] = _th_rk2["i3p_obra_lopsrm"]
                    _dff_u_rk.loc[
                        _ma_rk2 & _dff_u_rk["_ley_ur"].eq("LOPSRM")
                        & _dff_u_rk["_proc_ur"].isin(_TIPOS_I3P_RK)
                        & _dff_u_rk["_cont_ur"].isin(_TIPOS_SERV_RK),
                        "_umbral_ur"] = _th_rk2["i3p_serv_lopsrm"]
                _dff_u_rk["_pct_ur"] = _dff_u_rk["Importe DRC"] / _dff_u_rk["_umbral_ur"] * 100
                _ixs_umb_rk = _dff_u_rk[
                    _dff_u_rk["_umbral_ur"].notna()
                    & (_dff_u_rk["_pct_ur"] >= 90)
                    & (_dff_u_rk["_pct_ur"] < 100)
                ].index
                _f_umb_rk.loc[_ixs_umb_rk] = True
        except Exception:
            pass

        _tipo_rk = _dff_rk["Tipo Simplificado"]
        _f_ad_rk  = _tipo_rk == "Adjudicación Directa"
        _f_fri_rk = _tipo_rk == "Adjudicación Directa — Fr. I"
        _f_i3p_rk = _tipo_rk == "Invitación a 3 personas"

        # Reciente creación
        _RFC_PAT_RK = _re_rk.compile(r'^[A-ZÑ&]{3}(\d{2})(\d{2})(\d{2})[A-Z0-9]{3}$')

        def _parse_rfc_rk(rfc):
            m = _RFC_PAT_RK.match(str(rfc).strip().upper())
            if not m:
                return None
            yy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if not (1 <= mm <= 12 and 1 <= dd <= 31):
                return None
            yr = 2000 + yy if yy <= 30 else 1900 + yy
            try:
                return _date_rk(yr, mm, dd)
            except ValueError:
                return None

        _fecha_rfc_rk = _dff_rk["rfc"].map(_parse_rfc_rk)
        _fecha_ini_rk = pd.to_datetime(
            _dff_rk["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
        ).dt.date

        _edad_rk = pd.Series([
            (fi - fr).days
            if (fi is not None and fr is not None
                and not (isinstance(fi, float) and pd.isna(fi)))
            else None
            for fi, fr in zip(_fecha_ini_rk, _fecha_rfc_rk)
        ], index=_dff_rk.index)

        _f_rc_rk = _edad_rk.apply(lambda x: bool(x is not None and 0 <= x < 365))

        # ── 3. Score base ─────────────────────────────────────
        _score_base_rk = (
            _f_sc_rk.astype(int)  * 100 +
            _f_sa_rk.astype(int)  * 60  +
            _f_sm_rk.astype(int)  * 30  +
            _f_ed_rk.astype(int)  * 80  +
            _f_ep_rk.astype(int)  * 40  +
            _f_umb_rk.astype(int) * 45  +
            _f_rc_rk.astype(int)  * 30  +
            _f_ad_rk.astype(int)  * 20  +
            _f_fri_rk.astype(int) * 5   +
            _f_i3p_rk.astype(int) * 5
        )

        # ── 4. Ponderación por monto ──────────────────────────
        _monto_pct_rk = _dff_rk["Importe DRC"].rank(pct=True, na_option="bottom")
        _score_w_rk   = (_score_base_rk * (1.0 + 0.5 * _monto_pct_rk)).round(2)
        _score_norm_rk = _score_w_rk.clip(upper=100).round(1)

        # ── 5. Etiquetas de alertas ────────────────────────────
        _alerta_parts_rk = []
        for _ix in _dff_rk.index:
            _parts = []
            _ru = _rfc_n_rk.loc[_ix]
            if _f_sc_rk.loc[_ix]:              _parts.append("🔴 SABG inhabilitado")
            elif _f_sa_rk.loc[_ix]:            _parts.append("🟠 SABG suspendido")
            elif _f_sm_rk.loc[_ix]:            _parts.append("🟡 SABG historial")
            elif _f_sb_prior_rk.loc[_ix]:      _parts.append("🟤 Fallo previo a inhabilitación")
            if _ru in _rfcs_ed_rk:             _parts.append("🔴 EFOS definitivo")
            elif _ru in _rfcs_ep_rk:           _parts.append("🟡 EFOS presunto")
            if _f_umb_rk.loc[_ix]:             _parts.append("🚦 Zona umbral legal")
            if _f_rc_rk.loc[_ix]:             _parts.append("🟡 Reciente creación")
            _tp = _tipo_rk.loc[_ix]
            if _tp == "Adjudicación Directa":          _parts.append("🔴 Adj. directa")
            elif _tp == "Adjudicación Directa — Fr. I": _parts.append("⚪ AD — Patentes")
            elif _tp == "Invitación a 3 personas":      _parts.append("🟡 Inv. 3 personas")
            _alerta_parts_rk.append(" · ".join(_parts) if _parts else "—")

        _alertas_s_rk = pd.Series(_alerta_parts_rk, index=_dff_rk.index)

        # ── 6. Tabla top N ─────────────────────────────────────
        _top_rk = (
            _dff_rk.assign(
                Score   = _score_norm_rk,
                Score_raw = _score_w_rk,
                Alertas = _alertas_s_rk,
            )
            .query("Score_raw > 0")
            .sort_values("Score_raw", ascending=False)
            .head(_top_n_rk)
            [["Score", "Alertas",
              "Proveedor o contratista", "Importe DRC",
              "Tipo Simplificado", "Nombre de la UC",
              "Descripción del contrato", "Dirección del anuncio"]]
            .rename(columns={
                "Proveedor o contratista": "Proveedor",
                "Importe DRC":            "Importe",
                "Tipo Simplificado":      "Tipo",
                "Nombre de la UC":        "UC",
                "Descripción del contrato": "Descripción",
            })
            .reset_index(drop=True)
        )
        _top_rk.index += 1

    # ── KPIs ──────────────────────────────────────────────────
    _n_con_riesgo_rk = int((_score_w_rk > 0).sum())
    _n_criticos_rk   = int((_score_norm_rk >= 60).sum())
    _monto_riesgo_rk = _dff_rk.loc[_score_w_rk > 0, "Importe DRC"].sum()

    _kr1, _kr2, _kr3 = st.columns(3)
    _kr1.metric(
        "⚠️ Contratos con algún indicador de riesgo",
        f"{_n_con_riesgo_rk:,}",
        delta=(f"{_n_con_riesgo_rk / len(_dff_rk) * 100:.1f}% del total"
               if len(_dff_rk) > 0 else ""),
        delta_color="off",
    )
    _kr2.metric(
        "🔴 Contratos con score ≥ 60 (riesgo alto)",
        f"{_n_criticos_rk:,}",
    )
    _kr3.metric(
        "💰 Monto en contratos con riesgo",
        (f"${_monto_riesgo_rk/1e9:,.2f} miles de millones MXN"
         if _monto_riesgo_rk >= 1e9 else f"${_monto_riesgo_rk/1e6:,.1f} M MXN"),
    )

    # ── Tabla ranking ─────────────────────────────────────────
    if len(_top_rk) == 0:
        st.success("✅ No se detectaron contratos con indicadores de riesgo activos.")
    else:
        _top_rk["Importe_fmt"] = _top_rk["Importe"].apply(
            lambda x: (f"${x/1e9:,.2f} mil M" if pd.notna(x) and x >= 1e9
                       else f"${x/1e6:,.1f} M") if pd.notna(x) else "N/D"
        )
        _top_rk["Score"] = _top_rk["Score"].astype(float)
        _top_rk_show = _top_rk.drop(columns=["Importe"]).rename(
            columns={"Importe_fmt": "Importe"}
        )
        _top_rk_show = _top_rk_show[[
            "Score", "Alertas", "Proveedor", "Importe",
            "Tipo", "UC", "Descripción", "Dirección del anuncio"
        ]]
        st.dataframe(
            _top_rk_show,
            column_config={
                "Score": st.column_config.ProgressColumn(
                    "🎯 Score",
                    help="Score de riesgo ponderado por monto (0–100).",
                    min_value=0, max_value=100, format="%.1f",
                ),
                "Alertas":     st.column_config.TextColumn("🚨 Alertas detectadas", width="large"),
                "Proveedor":   st.column_config.TextColumn("Proveedor",           width="medium"),
                "Importe":     st.column_config.TextColumn("💰 Importe",   width="small"),
                "Tipo":        st.column_config.TextColumn("Tipo",                 width="medium"),
                "UC":          st.column_config.TextColumn("Unidad Compradora",    width="medium"),
                "Descripción": st.column_config.TextColumn("Descripción", width="large"),
                "Dirección del anuncio": st.column_config.LinkColumn(
                    "🔗 ComprasMX", display_text="Ver contrato"
                ),
            },
            use_container_width=True,
            height=min(750, 60 + len(_top_rk_show) * 55),
        )

    st.divider()

    # ══════════════════════════════════════════════════════════════════
    # RANKING DE UCS — score compuesto multidimensional (Opción B)
    # Cálculo independiente sobre `dff` global; filtros propios de sección
    # ══════════════════════════════════════════════════════════════════
    st.subheader("🏢 Ranking de Unidades Compradoras por Riesgo Compuesto")
    st.caption(
        "Clasifica las Unidades Compradoras según un score que combina cuatro dimensiones: "
        "integridad del proveedor (D1), amplitud del riesgo (D2), exposición económica (D3) "
        "y prácticas anticompetitivas (D4). "
        "Los filtros de esta sección son independientes del selector de UC del ranking de contratos."
    )

    with st.expander("ℹ️ Metodología — Score compuesto UC", expanded=False):
        st.markdown(
            """
            El score final de cada UC es la suma ponderada de cuatro dimensiones (escala 0–100):

            | Dimensión | Peso | Qué mide |
            |---|---|---|
            | **D1 — Integridad del proveedor** | 30 % | Contratos con inhabilitados SABG o EFOS |
            | **D2 — Amplitud del riesgo** | 25 % | % de contratos con algún indicador de alerta |
            | **D3 — Exposición económica** | 25 % | % del monto de la UC en contratos con riesgo |
            | **D4 — Prácticas anticompetitivas** | 20 % | Caso fortuito · HHI · Reciente creación · Zona umbral · Fraccionamiento |

            **D1 — regla de disparo inmediato:** cualquier contrato con SABG inhabilitación vigente o EFOS
            definitivo eleva D1 = 100 de forma automática (incumplimiento normativo). Para SABG
            suspendido/historial y EFOS presunto, D1 es proporcional al volumen de contratos afectados
            (máximo alcanzable sin violación confirmada: 60).

            **D4** combina cinco sub-indicadores en partes iguales:
            - **Caso fortuito**: % del monto adjudicado bajo esta excepción
            - **HHI**: concentración de proveedores (Herfindahl-Hirschman normalizado)
            - **Reciente creación**: % del monto a empresas < 1 año de constitución
            - **Zona umbral**: % de contratos en 90–100 % del tope legal PEF
            - **Fraccionamiento**: % de proveedores con ≥ 3 AD a la misma UC

            `Score_UC = 0.30·D1 + 0.25·D2 + 0.25·D3 + 0.20·D4`
            """
        )

    # ── Filtros locales (independientes del ranking de contratos) ──
    _uc_rk2_c1, _uc_rk2_c2, _uc_rk2_c3 = st.columns([2, 1, 1])

    _adsc_opts = ["Todas"]
    if len(df_dir_uc) > 0 and "Adscripción" in df_dir_uc.columns:
        _adsc_opts += sorted(df_dir_uc["Adscripción"].dropna().unique().tolist())
    _adsc_sel_uc = _uc_rk2_c1.selectbox(
        "🏛️ Adscripción / OOAD", _adsc_opts, key="uc_rk2_adsc"
    )
    _top_n_uc  = _uc_rk2_c2.selectbox(
        "Top N UCs", [10, 20, 50, "Todas"], index=0, key="uc_rk2_topn"
    )
    _score_min_uc = _uc_rk2_c3.selectbox(
        "Score mínimo", [0, 10, 25, 50], index=0, key="uc_rk2_smin",
        help="Oculta UCs con score menor al umbral seleccionado."
    )

    # Filtrar dff por adscripción si se eligió una
    _dff_uc2 = dff.copy()
    if _adsc_sel_uc != "Todas" and len(df_dir_uc) > 0:
        _claves_adsc = df_dir_uc.loc[
            df_dir_uc["Adscripción"] == _adsc_sel_uc, "Clave_UC"
        ].tolist()
        _dff_uc2 = _dff_uc2[_dff_uc2["Clave de la UC"].isin(_claves_adsc)]

    if len(_dff_uc2) == 0:
        st.info("ℹ️ Sin contratos para la selección actual.")
    else:
        import re as _re_uc2
        from datetime import date as _date_uc2

        with st.spinner("Calculando scores de riesgo por UC…"):

            # ── Indicadores externos (cargados una vez) ────────────
            _sabg_rec_uc = {}
            try:
                _df_sabg_uc = cargar_sancionados()
                for _, _sr in _df_sabg_uc.iterrows():
                    _ru = str(_sr["RFC"]).strip().upper()
                    _ini = pd.to_datetime(_sr.get("Inicio inhabilitación"), errors="coerce")
                    _niv = str(_sr.get("Nivel de Riesgo", ""))
                    if _ru:
                        _sabg_rec_uc.setdefault(_ru, []).append((_ini, _niv))
            except Exception:
                pass

            _rfcs_edef_uc = set()
            _rfcs_epres_uc = set()
            try:
                _df_efos_uc = cargar_efos()
                _rfcs_edef_uc  = set(_df_efos_uc[
                    _df_efos_uc["Situación del contribuyente"] == "Definitivo"]["RFC"])
                _rfcs_epres_uc = set(_df_efos_uc[
                    _df_efos_uc["Situación del contribuyente"] == "Presunto"]["RFC"])
            except Exception:
                pass

            # ── Flags por contrato (vectorizados) ─────────────────
            _rfc_n_uc = _dff_uc2["rfc"].astype(str).str.strip().str.upper()

            _ff_uc = pd.to_datetime(
                _dff_uc2["Fecha de fallo"] if "Fecha de fallo" in _dff_uc2.columns
                else pd.Series(pd.NaT, index=_dff_uc2.index),
                dayfirst=True, errors="coerce",
            )
            _ffm_uc = pd.to_datetime(
                _dff_uc2["Fecha de firma del contrato"]
                if "Fecha de firma del contrato" in _dff_uc2.columns
                else pd.Series(pd.NaT, index=_dff_uc2.index),
                dayfirst=True, errors="coerce",
            )
            _fref_uc = _ff_uc.combine_first(_ffm_uc)

            # SABG por contrato
            _fc_uc, _fa_uc, _fm_uc = [], [], []
            for _ix in _dff_uc2.index:
                _ru  = _rfc_n_uc.loc[_ix]
                _fr  = _fref_uc.loc[_ix]
                if _ru not in _sabg_rec_uc:
                    _fc_uc.append(False); _fa_uc.append(False); _fm_uc.append(False)
                    continue
                if pd.isna(_fr):
                    _fc_uc.append(False); _fa_uc.append(False); _fm_uc.append(False)
                    continue
                _sc2 = _sa2 = _sm2 = _any = False
                for _ini2, _niv2 in _sabg_rec_uc[_ru]:
                    if not pd.isna(_ini2) and _fr < _ini2:
                        continue
                    _any = True
                    _nl = _niv2.lower()
                    if "crítico" in _nl:   _sc2 = True
                    elif "alto" in _nl:    _sa2 = True
                    elif "medio" in _nl:   _sm2 = True
                if not _any:
                    _fc_uc.append(False); _fa_uc.append(False); _fm_uc.append(False)
                else:
                    _fc_uc.append(_sc2)
                    _fa_uc.append(_sa2 and not _sc2)
                    _fm_uc.append(_sm2 and not _sc2 and not _sa2)

            _f_sc_uc  = pd.Series(_fc_uc, index=_dff_uc2.index)
            _f_sa_uc  = pd.Series(_fa_uc, index=_dff_uc2.index)
            _f_sm_uc  = pd.Series(_fm_uc, index=_dff_uc2.index)
            _f_ed_uc  = _rfc_n_uc.isin(_rfcs_edef_uc)
            _f_ep_uc  = _rfc_n_uc.isin(_rfcs_epres_uc)

            # Flag genérico "riesgo D1" para D2/D3
            _f_crit_uc = _f_sc_uc | _f_ed_uc

            # Zona umbral
            _f_umb_uc = pd.Series(False, index=_dff_uc2.index)
            try:
                _umbs_uc = cargar_umbrales_pef()
                if _umbs_uc:
                    _TIPOS_AD_U   = {"Adjudicación Directa", "Adjudicación Directa — Fr. I"}
                    _TIPOS_I3P_U  = {"Invitación a 3 personas"}
                    _TIPOS_SERV_U = {"SERVICIOS", "SERVICIOS RELACIONADOS CON LA OBRA", "ARRENDAMIENTOS"}
                    _mask_u  = _dff_uc2["Tipo Simplificado"].isin(_TIPOS_AD_U | _TIPOS_I3P_U)
                    _dff_uu  = _dff_uc2[_mask_u].copy()
                    _dff_uu["_fecha_u"] = pd.to_datetime(
                        _dff_uu["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
                    )
                    _dff_uu["_año_u"] = _dff_uu["_fecha_u"].dt.year
                    _dff_uu["_ley_u"] = (
                        _dff_uu["Ley"].astype(str).str.strip().str.upper()
                        if "Ley" in _dff_uu.columns else "LAASSP"
                    )
                    _dff_uu["_cont_u"] = (
                        _dff_uu["Tipo de contratación"].astype(str).str.strip().str.upper()
                        if "Tipo de contratación" in _dff_uu.columns else "ADQUISICIONES"
                    )
                    _dff_uu["_proc_u"]   = _dff_uu["Tipo Simplificado"]
                    _dff_uu["_umbral_u"] = float("nan")
                    for _año_u2, _th_u2 in _umbs_uc.items():
                        _ma_u2 = _dff_uu["_año_u"] == _año_u2
                        _dff_uu.loc[_ma_u2 & _dff_uu["_ley_u"].eq("LAASSP")
                                    & _dff_uu["_proc_u"].isin(_TIPOS_AD_U),
                                    "_umbral_u"] = _th_u2["ad_laassp"]
                        _dff_uu.loc[_ma_u2 & _dff_uu["_ley_u"].eq("LAASSP")
                                    & _dff_uu["_proc_u"].isin(_TIPOS_I3P_U),
                                    "_umbral_u"] = _th_u2["i3p_laassp"]
                        _dff_uu.loc[_ma_u2 & _dff_uu["_ley_u"].eq("LOPSRM")
                                    & _dff_uu["_proc_u"].isin(_TIPOS_AD_U)
                                    & _dff_uu["_cont_u"].eq("OBRA PÚBLICA"),
                                    "_umbral_u"] = _th_u2["ad_obra_lopsrm"]
                        _dff_uu.loc[_ma_u2 & _dff_uu["_ley_u"].eq("LOPSRM")
                                    & _dff_uu["_proc_u"].isin(_TIPOS_AD_U)
                                    & _dff_uu["_cont_u"].isin(_TIPOS_SERV_U),
                                    "_umbral_u"] = _th_u2["ad_serv_lopsrm"]
                        _dff_uu.loc[_ma_u2 & _dff_uu["_ley_u"].eq("LOPSRM")
                                    & _dff_uu["_proc_u"].isin(_TIPOS_I3P_U)
                                    & _dff_uu["_cont_u"].eq("OBRA PÚBLICA"),
                                    "_umbral_u"] = _th_u2["i3p_obra_lopsrm"]
                        _dff_uu.loc[_ma_u2 & _dff_uu["_ley_u"].eq("LOPSRM")
                                    & _dff_uu["_proc_u"].isin(_TIPOS_I3P_U)
                                    & _dff_uu["_cont_u"].isin(_TIPOS_SERV_U),
                                    "_umbral_u"] = _th_u2["i3p_serv_lopsrm"]
                    _dff_uu["_pct_u"] = _dff_uu["Importe DRC"] / _dff_uu["_umbral_u"] * 100
                    _ixs_umb_uc = _dff_uu[
                        _dff_uu["_umbral_u"].notna()
                        & (_dff_uu["_pct_u"] >= 90)
                        & (_dff_uu["_pct_u"] < 100)
                    ].index
                    _f_umb_uc.loc[_ixs_umb_uc] = True
            except Exception:
                pass

            # Caso fortuito
            _f_cf_uc = _dff_uc2["Descripción excepción"].astype(str).str.upper().str.contains(
                "CASO FORTUITO", na=False
            )

            # Reciente creación
            _RFC_PAT_UC = _re_uc2.compile(r'^[A-ZÑ&]{3}(\d{2})(\d{2})(\d{2})[A-Z0-9]{3}$')

            def _parse_rfc_uc2(r):
                m = _RFC_PAT_UC.match(str(r).strip().upper())
                if not m:
                    return None
                yy, mm2, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if not (1 <= mm2 <= 12 and 1 <= dd <= 31):
                    return None
                yr = 2000 + yy if yy <= 30 else 1900 + yy
                try:
                    return _date_uc2(yr, mm2, dd)
                except ValueError:
                    return None

            _fecha_rfc_uc = _dff_uc2["rfc"].map(_parse_rfc_uc2)
            _fecha_ini_uc = pd.to_datetime(
                _dff_uc2["Fecha de inicio del contrato"], dayfirst=True, errors="coerce"
            ).dt.date
            _edad_uc = pd.Series([
                (fi - fr).days
                if fi is not None and fr is not None and not (isinstance(fi, float) and pd.isna(fi))
                else None
                for fi, fr in zip(_fecha_ini_uc, _fecha_rfc_uc)
            ], index=_dff_uc2.index)
            _f_rc_uc = _edad_uc.apply(lambda x: bool(x is not None and 0 <= x < 365))

            # Flag "algún riesgo" (para D2 / D3)
            _f_tipo_uc   = _dff_uc2["Tipo Simplificado"]
            _f_ad_uc     = _f_tipo_uc.isin(
                {"Adjudicación Directa", "Adjudicación Directa — Fr. I"})
            _score_base_uc = (
                _f_sc_uc.astype(int) * 100 +
                _f_ed_uc.astype(int) * 100 +
                _f_sa_uc.astype(int) * 60  +
                _f_ep_uc.astype(int) * 40  +
                _f_sm_uc.astype(int) * 30  +
                _f_umb_uc.astype(int) * 45 +
                _f_rc_uc.astype(int) * 30  +
                _f_ad_uc.astype(int) * 20
            )
            _f_any_riesgo_uc = _score_base_uc > 0

            # ── Agregar por UC ─────────────────────────────────────
            _uc_col = "Nombre de la UC"
            _imp    = _dff_uc2["Importe DRC"]

            _uc2 = _dff_uc2.copy()
            _uc2["_sc"]   = _f_sc_uc
            _uc2["_ed"]   = _f_ed_uc
            _uc2["_sa"]   = _f_sa_uc
            _uc2["_ep"]   = _f_ep_uc
            _uc2["_sm"]   = _f_sm_uc
            _uc2["_umb"]  = _f_umb_uc
            _uc2["_cf"]   = _f_cf_uc
            _uc2["_rc"]   = _f_rc_uc
            _uc2["_ad"]   = _f_ad_uc
            _uc2["_any"]  = _f_any_riesgo_uc
            _uc2["_imp"]  = _imp

            _g = _uc2.groupby(_uc_col)

            _agg_uc = _g.agg(
                n_total    =("_imp",  "count"),
                monto_total=("_imp",  "sum"),
                n_sc       =("_sc",   "sum"),
                n_ed       =("_ed",   "sum"),
                n_sa       =("_sa",   "sum"),
                n_ep       =("_ep",   "sum"),
                n_sm       =("_sm",   "sum"),
                n_umb      =("_umb",  "sum"),
                n_cf       =("_cf",   "sum"),
                n_rc       =("_rc",   "sum"),
                n_ad       =("_ad",   "sum"),
                n_any      =("_any",  "sum"),
                monto_any  =("_imp",  lambda s: s[_uc2.loc[s.index, "_any"]].sum()),
                monto_cf   =("_imp",  lambda s: s[_uc2.loc[s.index, "_cf"]].sum()),
                monto_rc   =("_imp",  lambda s: s[_uc2.loc[s.index, "_rc"]].sum()),
            ).reset_index()

            # HHI por UC
            _shares = (
                _uc2.groupby([_uc_col, "rfc"])["_imp"].sum()
                / _uc2.groupby(_uc_col)["_imp"].sum()
            )
            _hhi_uc = (_shares ** 2).groupby(level=0).sum() * 10_000
            _agg_uc = _agg_uc.merge(
                _hhi_uc.rename("HHI").reset_index(), on=_uc_col, how="left"
            )
            _agg_uc["HHI"] = _agg_uc["HHI"].fillna(0)

            # Fraccionamiento: % proveedores AD con ≥ 3 contratos a la misma UC
            _ad_mask = _uc2["_ad"]
            _frag_prov = (
                _uc2[_ad_mask]
                .groupby([_uc_col, "rfc"])
                .size()
                .reset_index(name="_cnt")
            )
            _frag_multi = (
                _frag_prov[_frag_prov["_cnt"] >= 3]
                .groupby(_uc_col)["rfc"].nunique()
                .rename("n_prov_frag")
            )
            _frag_total = (
                _frag_prov.groupby(_uc_col)["rfc"].nunique()
                .rename("n_prov_ad")
            )
            _agg_uc = _agg_uc.merge(
                _frag_multi.reset_index(), on=_uc_col, how="left"
            ).merge(
                _frag_total.reset_index(), on=_uc_col, how="left"
            )
            _agg_uc["n_prov_frag"] = _agg_uc["n_prov_frag"].fillna(0)
            _agg_uc["n_prov_ad"]   = _agg_uc["n_prov_ad"].fillna(0)

            # ── Calcular dimensiones ───────────────────────────────
            _N  = _agg_uc["n_total"].replace(0, pd.NA)
            _MT = _agg_uc["monto_total"].replace(0, pd.NA)

            # D1 — Integridad
            _crit_flag = (_agg_uc["n_sc"] + _agg_uc["n_ed"]) > 0
            _d1_prop = (
                (_agg_uc["n_sa"] / _N * 60 +
                 _agg_uc["n_ep"] / _N * 40 +
                 _agg_uc["n_sm"] / _N * 30)
                .clip(upper=60)
                .fillna(0)
            )
            _agg_uc["D1"] = _crit_flag.astype(float) * 100 + (~_crit_flag).astype(float) * _d1_prop

            # D2 — Amplitud
            _agg_uc["D2"] = (_agg_uc["n_any"] / _N * 100).clip(upper=100).fillna(0)

            # D3 — Exposición económica
            _agg_uc["D3"] = (_agg_uc["monto_any"] / _MT * 100).clip(upper=100).fillna(0)

            # D4 — Prácticas anticompetitivas
            _sub_cf   = (_agg_uc["monto_cf"] / _MT * 100).clip(upper=100).fillna(0)
            _sub_hhi  = (_agg_uc["HHI"] / 10_000 * 100).clip(upper=100).fillna(0)
            _sub_rc   = (_agg_uc["monto_rc"] / _MT * 100).clip(upper=100).fillna(0)
            _sub_umb  = (_agg_uc["n_umb"] / _N * 100).clip(upper=100).fillna(0)
            _n_prov_ad_safe = _agg_uc["n_prov_ad"].replace(0, pd.NA)
            _sub_frag = (_agg_uc["n_prov_frag"] / _n_prov_ad_safe * 100).clip(upper=100).fillna(0)
            _agg_uc["D4"] = (_sub_cf + _sub_hhi + _sub_rc + _sub_umb + _sub_frag) / 5

            # Score final
            _agg_uc["Score_UC"] = (
                0.30 * _agg_uc["D1"] +
                0.25 * _agg_uc["D2"] +
                0.25 * _agg_uc["D3"] +
                0.20 * _agg_uc["D4"]
            ).round(1)

            # Alerta dominante
            def _alerta_dom(row):
                if row["n_sc"] > 0:   return "🔴 SABG inhabilitado"
                if row["n_ed"] > 0:   return "🔴 EFOS definitivo"
                if row["n_sa"] > 0:   return "🟠 SABG suspendido"
                if row["n_ep"] > 0:   return "🟡 EFOS presunto"
                if row["n_sm"] > 0:   return "🟡 SABG historial"
                if row["n_cf"] > 0:   return "⚠️ Caso fortuito"
                if row["HHI"] > 2500: return "⚠️ Alta concentración"
                if row["n_rc"] > 0:   return "🟡 Reciente creación"
                if row["n_umb"] > 0:  return "🚦 Zona umbral"
                if row["n_prov_frag"] > 0: return "⚠️ Fraccionamiento"
                return "—"

            _agg_uc["Alerta dominante"] = _agg_uc.apply(_alerta_dom, axis=1)

            # Aplicar score mínimo y ordenar
            _agg_uc_filt = (
                _agg_uc[_agg_uc["Score_UC"] >= _score_min_uc]
                .sort_values("Score_UC", ascending=False)
            )
            _agg_uc_f = (
                _agg_uc_filt if _top_n_uc == "Todas"
                else _agg_uc_filt.head(_top_n_uc)
            ).reset_index(drop=True)
            _agg_uc_f.index += 1

        # ── KPIs ──────────────────────────────────────────────────
        _n_uc_total    = _agg_uc["Nombre de la UC"].nunique()
        _n_uc_crit     = int(_crit_flag.sum())
        _n_uc_alto_rk  = int((_agg_uc["Score_UC"] >= 60).sum())
        _top1_row      = _agg_uc.sort_values("Score_UC", ascending=False).iloc[0] if len(_agg_uc) > 0 else None
        _top5_monto    = (
            _agg_uc.sort_values("Score_UC", ascending=False)
            .head(5)["monto_any"].sum()
        )
        _monto_total_g = _agg_uc["monto_any"].sum()
        _pct_top5      = (_top5_monto / _monto_total_g * 100) if _monto_total_g > 0 else 0

        _kuc1, _kuc2, _kuc3 = st.columns(3)
        _kuc1.metric(
            "🔴 UCs con incumplimiento normativo (D1 = 100)",
            f"{_n_uc_crit:,}",
            delta=f"de {_n_uc_total} UCs con contratos",
            delta_color="off",
        )
        _kuc2.metric(
            "⚠️ UCs con Score ≥ 60",
            f"{_n_uc_alto_rk:,}",
        )
        _kuc3.metric(
            "💰 Concentración top 5 UCs",
            f"{_pct_top5:.1f}%",
            delta="del monto en riesgo",
            delta_color="off",
        )
        if _top1_row is not None:
            st.caption(
                f"UC con mayor score: **{_top1_row['Nombre de la UC']}** — "
                f"Score {_top1_row['Score_UC']:.1f} · "
                f"{_top1_row['Alerta dominante']}"
            )

        if len(_agg_uc_f) == 0:
            st.success("✅ No hay UCs con indicadores de riesgo en el umbral seleccionado.")
        else:
            # Helper: truncar nombres largos para ejes de gráficas
            def _trunc(name, n=28):
                s = str(name)
                return s if len(s) <= n else s[:n - 1] + "…"

            # ── Barras apiladas: contribuciones ponderadas (suman = Score_UC) ──
            st.markdown("#### Distribución del score por dimensión")
            _bar_data = _agg_uc_f[[_uc_col, "D1", "D2", "D3", "D4", "Score_UC"]].copy()
            # Contribuciones ponderadas → su suma = Score_UC exacto
            _bar_data["D1_w"] = (_bar_data["D1"] * 0.30).round(2)
            _bar_data["D2_w"] = (_bar_data["D2"] * 0.25).round(2)
            _bar_data["D3_w"] = (_bar_data["D3"] * 0.25).round(2)
            _bar_data["D4_w"] = (_bar_data["D4"] * 0.20).round(2)
            # En Plotly horizontal bar, el ÚLTIMO elemento de category_orders queda ARRIBA.
            # Ordenamos ascending=True → el de mayor score es el último → aparece arriba.
            _uc_order_bar = (
                _bar_data
                .sort_values("Score_UC", ascending=True)[_uc_col]
                .tolist()
            )
            # Etiquetas truncadas para el eje y (solo display; la clave sigue siendo el nombre completo)
            _ticktext_bar = [_trunc(n) for n in _uc_order_bar]
            _bar_data_m = _bar_data.melt(
                id_vars=[_uc_col, "Score_UC"],
                value_vars=["D1_w", "D2_w", "D3_w", "D4_w"],
                var_name="Dimensión", value_name="Valor"
            )
            _dim_labels = {
                "D1_w": "D1 — Integridad (×0.30)",
                "D2_w": "D2 — Amplitud (×0.25)",
                "D3_w": "D3 — Exposición econ. (×0.25)",
                "D4_w": "D4 — Prácticas anticmp. (×0.20)",
            }
            _dim_colors = {
                "D1 — Integridad (×0.30)":          IMSS_ROJO,
                "D2 — Amplitud (×0.25)":            IMSS_ORO,
                "D3 — Exposición econ. (×0.25)":    "#E07B00",
                "D4 — Prácticas anticmp. (×0.20)":  IMSS_GRIS,
            }
            _bar_data_m["Dimensión"] = _bar_data_m["Dimensión"].map(_dim_labels)
            fig_uc_rank = px.bar(
                _bar_data_m,
                x="Valor", y=_uc_col,          # nombre completo como clave → sin duplicados
                color="Dimensión",
                orientation="h",
                color_discrete_map=_dim_colors,
                category_orders={_uc_col: _uc_order_bar},
                labels={"Valor": "Score UC (0–100)", _uc_col: ""},
            )
            fig_uc_rank.update_layout(
                font=plotly_font(),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                xaxis=dict(range=[0, 100], title="Score UC (0–100)"),
                yaxis=dict(
                    tickvals=_uc_order_bar,     # claves (nombres completos)
                    ticktext=_ticktext_bar,     # etiquetas truncadas para display
                    tickfont=dict(size=10),
                ),
                legend=dict(orientation="h", yanchor="bottom", y=1.01,
                            xanchor="left", x=0),
                margin=dict(t=20, b=40, l=190),
                height=max(350, len(_agg_uc_f) * 32 + 80),
            )
            st.plotly_chart(fig_uc_rank, use_container_width=True)

            st.divider()

            # ── Tabla detallada ────────────────────────────────────
            st.markdown("#### Tabla detallada por UC")
            _tbl_uc = _agg_uc_f[[
                _uc_col, "Score_UC", "D1", "D2", "D3", "D4",
                "Alerta dominante", "n_total", "n_any", "monto_any"
            ]].copy()
            _tbl_uc["D2_fmt"]  = _tbl_uc.apply(
                lambda r: f"{int(r['n_any'])} de {int(r['n_total'])} ({r['D2']:.0f}%)", axis=1
            )
            _tbl_uc["D3_fmt"] = _tbl_uc["monto_any"].apply(
                lambda x: f"${x/1e6:,.1f} M" if pd.notna(x) else "—"
            )
            _tbl_uc["D1"] = _tbl_uc["D1"].round(1)
            _tbl_uc["D3"] = _tbl_uc["D3"].round(1)
            _tbl_uc["D4"] = _tbl_uc["D4"].round(1)
            _tbl_show = _tbl_uc[[
                _uc_col, "Score_UC", "D1", "D2_fmt", "D3_fmt", "D3", "D4", "Alerta dominante"
            ]].rename(columns={
                _uc_col:    "Unidad Compradora",
                "Score_UC": "Score",
                "D1":       "D1 Integridad",
                "D2_fmt":   "D2 Amplitud",
                "D3_fmt":   "D3 Monto en riesgo",
                "D3":       "D3 %",
                "D4":       "D4 Anticmp.",
                "Alerta dominante": "Alerta principal",
            })
            st.dataframe(
                _tbl_show,
                column_config={
                    "Score": st.column_config.ProgressColumn(
                        "🎯 Score UC",
                        min_value=0, max_value=100, format="%.1f",
                    ),
                    "D1 Integridad": st.column_config.ProgressColumn(
                        "D1 Integridad", min_value=0, max_value=100, format="%.1f",
                    ),
                    "D2 Amplitud":       st.column_config.TextColumn("D2 Amplitud",    width="medium"),
                    "D3 Monto en riesgo": st.column_config.TextColumn("D3 Monto",       width="small"),
                    "D3 %": st.column_config.ProgressColumn(
                        "D3 %", min_value=0, max_value=100, format="%.1f",
                    ),
                    "D4 Anticmp.": st.column_config.ProgressColumn(
                        "D4 Anticmp.", min_value=0, max_value=100, format="%.1f",
                    ),
                    "Alerta principal": st.column_config.TextColumn("Alerta principal", width="medium"),
                },
                use_container_width=True,
                height=min(700, 60 + len(_tbl_show) * 55),
            )

            # Descarga CSV
            _csv_uc2 = _tbl_show.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "📥 Descargar ranking de UCs (CSV)",
                data=_csv_uc2,
                file_name="ranking_uc_riesgo.csv",
                mime="text/csv",
                key="dl_uc_ranking",
            )

    st.divider()


# ─────────────────────────────────────────────────────────────────────────────
# OCDS — Open Contracting Data Standard — helpers y página
# ─────────────────────────────────────────────────────────────────────────────

_OCDS_PREFIX = "ocds-comprasmx"   # prefijo provisional; no registrado oficialmente

_OCDS_METHOD_MAP = {
    "Licitación Pública":              "open",
    "Invitación a 3 personas":         "selective",
    "Adjudicación Directa":            "limited",
    "Adjudicación Directa — Fr. I":    "limited",
    "Adjudicación Directa — Patentes": "limited",
    "Entre Entes Públicos":            "limited",
}

_OCDS_CATEGORY_MAP = {
    "Adquisiciones":                      "goods",
    "Arrendamientos":                     "services",
    "Servicios":                          "services",
    "Obra pública":                       "works",
    "Servicios relacionados con la obra": "works",
}


def _fmt_date_ocds(val) -> str | None:
    """DD/MM/YYYY → YYYY-MM-DDT00:00:00Z. Devuelve None si no aplica."""
    if pd.isna(val) or not str(val).strip():
        return None
    try:
        d = _datetime.strptime(str(val).strip()[:10], "%d/%m/%Y")
        return d.strftime("%Y-%m-%dT00:00:00Z")
    except Exception:
        return None


def _build_ocds_release(row: dict) -> dict:
    """Construye un OCDS Release (tag: award + contract) a partir de una fila de contratos."""
    num_proc = str(row.get("Número de procedimiento") or "").strip()
    cod_cont = str(row.get("Código del contrato")     or "").strip()
    raw_id   = num_proc or cod_cont or str(row.get("_idx", ""))
    safe_id  = _re.sub(r"[^A-Za-z0-9\-_]", "-", raw_id) if raw_id else f"row-{row.get('_idx','0')}"
    ocid     = f"{_OCDS_PREFIX}-{safe_id}"

    # ── Partes ──────────────────────────────────────────────────────────────
    buyer_clave = str(row.get("Clave de la UC")            or "").strip()
    buyer_name  = str(row.get("Nombre de la UC")           or "").strip()
    inst_name   = str(row.get("Institución")               or "").strip()
    sup_rfc     = str(row.get("rfc")                       or "").strip().upper()
    sup_name    = str(row.get("Proveedor o contratista")   or "").strip()

    parties = []
    if buyer_clave or buyer_name:
        parties.append({
            "id":   f"MX-UC-{buyer_clave}" if buyer_clave else buyer_name,
            "name": buyer_name,
            "identifier": {"scheme": "MX-UC", "id": buyer_clave, "legalName": inst_name},
            "roles": ["buyer", "procuringEntity"],
        })
    if sup_rfc or sup_name:
        sup_party = {
            "id":   f"MX-RFC-{sup_rfc}" if sup_rfc else sup_name,
            "name": sup_name,
            "roles": ["supplier", "tenderer"],
        }
        if sup_rfc:
            sup_party["identifier"] = {"scheme": "MX-RFC", "id": sup_rfc, "legalName": sup_name}
        parties.append(sup_party)

    # ── Monto ────────────────────────────────────────────────────────────────
    raw_amount = row.get("Importe DRC")
    amount     = float(raw_amount) if pd.notna(raw_amount) else None
    currency   = str(row.get("Moneda") or "MXN").strip()
    if currency not in ("MXN", "USD", "EUR"):
        currency = "MXN"
    value_obj  = {"amount": amount, "currency": currency} if amount is not None else None

    # ── Clasificación CUCoP ──────────────────────────────────────────────────
    partida = str(row.get("Partida específica") or "").strip().zfill(5)
    items = []
    if partida and partida != "00000":
        items = [{
            "id": "1",
            "description": str(row.get("Descripción del contrato") or "").strip(),
            "classification": {"scheme": "CUCoP", "id": partida},
        }]

    # ── Método de contratación ───────────────────────────────────────────────
    tipo_simp = str(row.get("Tipo Simplificado") or row.get("Tipo Procedimiento") or "").strip()
    method    = _OCDS_METHOD_MAP.get(tipo_simp, "limited")
    category  = _OCDS_CATEGORY_MAP.get(str(row.get("Tipo de contratación") or ""), "goods")

    # ── Fechas ───────────────────────────────────────────────────────────────
    d_pub    = _fmt_date_ocds(row.get("Fecha de publicación"))
    d_fallo  = _fmt_date_ocds(row.get("Fecha de fallo"))
    d_inicio = _fmt_date_ocds(row.get("Fecha de inicio del contrato"))
    d_fin    = _fmt_date_ocds(row.get("Fecha de fin del contrato"))
    d_firma  = _fmt_date_ocds(row.get("Fecha de firma del contrato"))
    d_release = d_firma or d_fallo or d_pub or _datetime.utcnow().strftime("%Y-%m-%dT00:00:00Z")

    # ── Tender ───────────────────────────────────────────────────────────────
    tender = {
        "id":     num_proc or safe_id,
        "title":  str(row.get("Título del contrato") or row.get("Descripción del contrato") or "").strip(),
        "description": str(row.get("Descripción del contrato") or "").strip(),
        "status": "complete",
        "procurementMethod":       method,
        "mainProcurementCategory": category,
        "procuringEntity": {
            "id":   f"MX-UC-{buyer_clave}" if buyer_clave else buyer_name,
            "name": buyer_name,
        },
    }
    if value_obj:
        tender["value"] = value_obj
    if items:
        tender["items"] = items
    exc_art  = str(row.get("Artículo de excepción") or "").strip()
    exc_desc = str(row.get("Descripción excepción") or "").strip()
    if exc_art:
        tender["procurementMethodRationale"] = exc_art
    if exc_desc:
        tender["procurementMethodDetails"] = exc_desc
    if d_pub:
        tender["tenderPeriod"] = {"startDate": d_pub}
    if d_fallo:
        tender["awardPeriod"] = {"endDate": d_fallo}
    period_t = {}
    if d_inicio: period_t["startDate"] = d_inicio
    if d_fin:    period_t["endDate"]   = d_fin
    if period_t: tender["contractPeriod"] = period_t

    # ── Award ────────────────────────────────────────────────────────────────
    award_id = f"{ocid}-award-1"
    award    = {"id": award_id, "status": "active"}
    if d_fallo:   award["date"]  = d_fallo
    if value_obj: award["value"] = value_obj
    if sup_rfc or sup_name:
        award["suppliers"] = [{
            "id":   f"MX-RFC-{sup_rfc}" if sup_rfc else sup_name,
            "name": sup_name,
        }]

    # ── Contract ─────────────────────────────────────────────────────────────
    contract_id = cod_cont or f"{ocid}-contract-1"
    contract = {"id": contract_id, "awardID": award_id, "status": "active"}
    if value_obj: contract["value"]      = value_obj
    if d_firma:   contract["dateSigned"] = d_firma
    period_c = {}
    if d_inicio: period_c["startDate"] = d_inicio
    if d_fin:    period_c["endDate"]   = d_fin
    if period_c: contract["period"] = period_c
    if items:    contract["items"]  = items
    url = str(row.get("Dirección del anuncio") or "").strip()
    if url:
        contract["documents"] = [{
            "id": "1",
            "documentType": "contractSigned",
            "title": "Publicación en ComprasMX",
            "url": url,
            "format": "text/html",
        }]

    # ── Release ──────────────────────────────────────────────────────────────
    release = {
        "ocid":           ocid,
        "id":             f"{ocid}-release-1",
        "date":           d_release,
        "tag":            ["award", "contract"],
        "initiationType": "tender",
        "language":       "es",
    }
    if parties:
        release["parties"] = parties
    if buyer_clave or buyer_name:
        release["buyer"] = {
            "id":   f"MX-UC-{buyer_clave}" if buyer_clave else buyer_name,
            "name": buyer_name,
        }
    release["tender"]    = tender
    release["awards"]    = [award]
    release["contracts"] = [contract]
    return release


def _build_ocds_package(df_exp: pd.DataFrame, publisher_name: str) -> dict:
    """Genera un OCDS Release Package completo a partir del DataFrame filtrado."""
    records  = df_exp.reset_index(drop=True).copy()
    records["_idx"] = records.index
    releases = [_build_ocds_release(r) for r in records.to_dict("records")]
    return {
        "uri":           "https://comprasmx.buengobierno.gob.mx/ocds/package",
        "version":       "1.1",
        "publishedDate": _datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "releases":      releases,
        "publisher": {
            "name": publisher_name,
            "uri":  "https://comprasmx.buengobierno.gob.mx",
        },
        "license":           "https://creativecommons.org/licenses/by/4.0/",
        "publicationPolicy": "https://standard.open-contracting.org/latest/en/",
    }


def _build_ocds_flat_df(df_exp: pd.DataFrame) -> pd.DataFrame:
    """Genera un DataFrame aplanado con nomenclatura de rutas OCDS para exportar como CSV."""
    records = df_exp.reset_index(drop=True).copy()
    records["_idx"] = records.index
    rows = []
    for r in records.to_dict("records"):
        num_proc = str(r.get("Número de procedimiento") or "").strip()
        cod_cont = str(r.get("Código del contrato")     or "").strip()
        raw_id   = num_proc or cod_cont or str(r.get("_idx", ""))
        safe_id  = _re.sub(r"[^A-Za-z0-9\-_]", "-", raw_id) if raw_id else f"row-{r.get('_idx','0')}"
        ocid     = f"{_OCDS_PREFIX}-{safe_id}"
        buyer_clave = str(r.get("Clave de la UC")          or "").strip()
        buyer_name  = str(r.get("Nombre de la UC")         or "").strip()
        sup_rfc     = str(r.get("rfc")                     or "").strip().upper()
        sup_name    = str(r.get("Proveedor o contratista") or "").strip()
        tipo_simp   = str(r.get("Tipo Simplificado") or r.get("Tipo Procedimiento") or "").strip()
        amount      = r.get("Importe DRC")
        amount_val  = float(amount) if pd.notna(amount) else ""
        currency    = str(r.get("Moneda") or "MXN").strip()
        if currency not in ("MXN", "USD", "EUR"):
            currency = "MXN"
        partida = str(r.get("Partida específica") or "").strip().zfill(5)
        rows.append({
            "ocid":                                  ocid,
            "release.id":                            f"{ocid}-release-1",
            "release.date":                          _fmt_date_ocds(r.get("Fecha de firma del contrato")) or "",
            "release.tag":                           "award,contract",
            "release.language":                      "es",
            "buyer.id":                              f"MX-UC-{buyer_clave}" if buyer_clave else buyer_name,
            "buyer.name":                            buyer_name,
            "buyer.identifier.scheme":               "MX-UC",
            "buyer.identifier.id":                   buyer_clave,
            "tender.id":                             num_proc,
            "tender.title":                          str(r.get("Título del contrato") or r.get("Descripción del contrato") or "").strip(),
            "tender.description":                    str(r.get("Descripción del contrato") or "").strip(),
            "tender.status":                         "complete",
            "tender.procurementMethod":              _OCDS_METHOD_MAP.get(tipo_simp, "limited"),
            "tender.mainProcurementCategory":        _OCDS_CATEGORY_MAP.get(str(r.get("Tipo de contratación") or ""), "goods"),
            "tender.procurementMethodRationale":     str(r.get("Artículo de excepción")  or "").strip(),
            "tender.procurementMethodDetails":       str(r.get("Descripción excepción")   or "").strip(),
            "tender.value.amount":                   amount_val,
            "tender.value.currency":                 currency,
            "tender.tenderPeriod.startDate":         _fmt_date_ocds(r.get("Fecha de publicación"))             or "",
            "tender.awardPeriod.endDate":            _fmt_date_ocds(r.get("Fecha de fallo"))                   or "",
            "tender.contractPeriod.startDate":       _fmt_date_ocds(r.get("Fecha de inicio del contrato"))     or "",
            "tender.contractPeriod.endDate":         _fmt_date_ocds(r.get("Fecha de fin del contrato"))        or "",
            "tender.items[0].classification.scheme": "CUCoP" if partida != "00000" else "",
            "tender.items[0].classification.id":     partida  if partida != "00000" else "",
            "award.id":                              f"{ocid}-award-1",
            "award.date":                            _fmt_date_ocds(r.get("Fecha de fallo")) or "",
            "award.status":                          "active",
            "award.value.amount":                    amount_val,
            "award.value.currency":                  currency,
            "award.suppliers[0].id":                 f"MX-RFC-{sup_rfc}" if sup_rfc else sup_name,
            "award.suppliers[0].name":               sup_name,
            "award.suppliers[0].identifier.scheme":  "MX-RFC" if sup_rfc else "",
            "award.suppliers[0].identifier.id":      sup_rfc,
            "contract.id":                           cod_cont,
            "contract.awardID":                      f"{ocid}-award-1",
            "contract.status":                       "active",
            "contract.dateSigned":                   _fmt_date_ocds(r.get("Fecha de firma del contrato")) or "",
            "contract.period.startDate":             _fmt_date_ocds(r.get("Fecha de inicio del contrato")) or "",
            "contract.period.endDate":               _fmt_date_ocds(r.get("Fecha de fin del contrato"))    or "",
            "contract.value.amount":                 amount_val,
            "contract.value.currency":               currency,
            "contract.documents[0].url":             str(r.get("Dirección del anuncio") or "").strip(),
            "contract.documents[0].documentType":    "contractSigned" if str(r.get("Dirección del anuncio") or "").strip() else "",
        })
    return pd.DataFrame(rows)


def pagina_ocds():
    st.header("📤 Exportación OCDS — Open Contracting Data Standard")
    st.markdown(
        "El **Open Contracting Data Standard (OCDS)** es el estándar internacional de datos abiertos "
        "para contrataciones públicas, desarrollado por la [Open Contracting Partnership](https://www.open-contracting.org/). "
        "Permite publicar y reutilizar datos de contrataciones en un formato estructurado y comparable entre países."
    )

    st.info(
        "**¿Qué se exporta?** Los contratos actualmente visibles con los filtros del sidebar, "
        "transformados al esquema OCDS v1.1. Cada contrato se representa como un *Release* con "
        "etiquetas `award` + `contract`, ya que los datos de ComprasMX corresponden a la etapa "
        "de adjudicación/firma (no al proceso de licitación completo).",
        icon="ℹ️",
    )

    with st.expander("⚠️ Limitaciones y notas técnicas"):
        st.markdown("""
**Prefijo OCID provisional**
El identificador `ocds-comprasmx-` es provisional. Para publicación oficial se requiere registrar
un prefijo ante la Open Contracting Partnership en
[standard.open-contracting.org/infrastructure/latest/en/guidance/identifiers/](https://standard.open-contracting.org/infrastructure/latest/en/guidance/identifiers/).

**Cobertura de campos**
| Disponible | No disponible |
|---|---|
| Identificadores RFC (proveedor) | Dirección del proveedor |
| Monto del contrato (`Importe DRC`) | Número de participantes en la licitación |
| Tipo de procedimiento → `procurementMethod` | Documentos del proceso (convocatoria, bases) |
| Clasificación CUCoP → `items[].classification` | Modificaciones contractuales detalladas |
| Fechas: publicación, fallo, firma, inicio, fin | Precio unitario por ítem |
| URL del anuncio en ComprasMX | |

**Etapa del proceso**
Los releases generados cubren únicamente la etapa `award + contract`. No se incluye
la etapa `tender` previa (publicación de la convocatoria) por no estar disponible en el CSV.

**Identificadores de la UC compradora**
Se usa el esquema `MX-UC` con la clave interna de la unidad compradora. No existe un scheme
OCDS registrado oficialmente para las UCs mexicanas.
        """)

    # ── Métricas del lote a exportar ─────────────────────────────────────────
    n_contratos  = len(dff)
    monto_total  = dff["Importe DRC"].sum()
    n_proveedores = dff["rfc"].nunique()
    n_ucs         = dff["Nombre de la UC"].nunique()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📄 Contratos a exportar",  f"{n_contratos:,}")
    c2.metric("💰 Monto total",           f"${monto_total/1e6:,.1f} M MXN")
    c3.metric("🏢 Unidades compradoras",  f"{n_ucs:,}")
    c4.metric("🏭 Proveedores únicos",    f"{n_proveedores:,}")

    st.divider()

    # ── Controles de exportación ─────────────────────────────────────────────
    st.subheader("Generar archivos")

    # Aviso de volumen
    _CAP_JSON = 5_000
    if n_contratos > _CAP_JSON:
        st.warning(
            f"El lote actual tiene **{n_contratos:,} contratos**. "
            f"La exportación JSON se limitará a los primeros **{_CAP_JSON:,}** para evitar archivos excesivamente grandes. "
            f"Aplica más filtros en el sidebar para exportar un subconjunto específico. "
            f"La exportación CSV no tiene límite.",
            icon="⚠️",
        )

    # Nombre del publisher para el paquete
    inst_label = inst_sel if inst_sel != "Todas" else "APF — Administración Pública Federal"
    publisher_name = f"DMII-IMSS / {inst_label} — ComprasMX {_anios_label}"

    col_json, col_csv = st.columns(2)

    # ── JSON (OCDS Release Package) ──────────────────────────────────────────
    with col_json:
        st.markdown("#### JSON — OCDS Release Package")
        st.caption(
            "Formato nativo OCDS. Compatible con validadores como "
            "[OCDS Kit](https://ocdskit.readthedocs.io/) y [lib-cove-ocds](https://github.com/open-contracting/lib-cove-ocds)."
        )
        if st.button("Generar JSON", key="btn_gen_json", use_container_width=True):
            df_exp = dff.head(_CAP_JSON) if n_contratos > _CAP_JSON else dff
            with st.spinner(f"Generando paquete OCDS con {len(df_exp):,} releases…"):
                pkg      = _build_ocds_package(df_exp, publisher_name)
                indent   = 2 if len(df_exp) <= 1_000 else None
                json_str = json.dumps(pkg, ensure_ascii=False, indent=indent)
            fname = f"ocds_comprasmx_{_anios_label.replace(', ','-')}_{len(df_exp)}contratos.json"
            st.download_button(
                label=f"⬇️ Descargar JSON ({len(df_exp):,} releases, {len(json_str)/1024:.0f} KB)",
                data=json_str.encode("utf-8"),
                file_name=fname,
                mime="application/json",
                key="dl_ocds_json",
                use_container_width=True,
            )

    # ── CSV aplanado ─────────────────────────────────────────────────────────
    with col_csv:
        st.markdown("#### CSV — Campos con nomenclatura OCDS")
        st.caption(
            "Versión tabular con columnas nombradas según las rutas OCDS (ej. `tender.procurementMethod`). "
            "Útil para análisis en Excel, R o Python sin necesidad de parsear JSON."
        )
        if st.button("Generar CSV", key="btn_gen_csv", use_container_width=True):
            with st.spinner(f"Generando CSV con {n_contratos:,} filas…"):
                df_flat  = _build_ocds_flat_df(dff)
                csv_flat = df_flat.to_csv(index=False, encoding="utf-8")
            fname_csv = f"ocds_flat_comprasmx_{_anios_label.replace(', ','-')}_{n_contratos}contratos.csv"
            st.download_button(
                label=f"⬇️ Descargar CSV ({n_contratos:,} filas, {len(csv_flat)/1024:.0f} KB)",
                data=csv_flat.encode("utf-8"),
                file_name=fname_csv,
                mime="text/csv",
                key="dl_ocds_csv",
                use_container_width=True,
            )

    st.divider()

    # ── Tabla de mapeo de campos ─────────────────────────────────────────────
    with st.expander("📋 Mapeo de campos ComprasMX → OCDS"):
        st.markdown("""
| Campo ComprasMX | Ruta OCDS | Notas |
|---|---|---|
| `Número de procedimiento` | `tender.id` / `ocid` (base) | Se sanitiza para uso como identificador |
| `Código del contrato` | `contract.id` | |
| `Institución` | `parties[buyer].identifier.legalName` | |
| `Nombre de la UC` | `parties[buyer].name` / `tender.procuringEntity.name` | |
| `Clave de la UC` | `parties[buyer].identifier.id` (scheme: `MX-UC`) | |
| `rfc` | `parties[supplier].identifier.id` (scheme: `MX-RFC`) | |
| `Proveedor o contratista` | `parties[supplier].name` | |
| `Tipo Simplificado` | `tender.procurementMethod` | `open` / `selective` / `limited` |
| `Artículo de excepción` | `tender.procurementMethodRationale` | Solo AD y excepciones |
| `Descripción excepción` | `tender.procurementMethodDetails` | |
| `Tipo de contratación` | `tender.mainProcurementCategory` | `goods` / `services` / `works` |
| `Importe DRC` | `tender.value.amount` = `award.value.amount` = `contract.value.amount` | |
| `Moneda` | `*.value.currency` | Default `MXN` |
| `Partida específica` | `tender.items[0].classification.id` (scheme: `CUCoP`) | |
| `Fecha de publicación` | `tender.tenderPeriod.startDate` | |
| `Fecha de fallo` | `tender.awardPeriod.endDate` / `award.date` | |
| `Fecha de firma del contrato` | `contract.dateSigned` / `release.date` | |
| `Fecha de inicio del contrato` | `tender.contractPeriod.startDate` / `contract.period.startDate` | |
| `Fecha de fin del contrato` | `tender.contractPeriod.endDate` / `contract.period.endDate` | |
| `Dirección del anuncio` | `contract.documents[0].url` | |
        """)

    st.divider()
    st.caption(f"División de Monitoreo de la Integridad Institucional – IMSS | ComprasMX {_anios_label}")


# NAVEGACIÓN PRINCIPAL (st.navigation)
# ═══════════════════════════════════════════════════════════════
pg = st.navigation(
    {
        "Análisis de Compras": [
            st.Page(pagina_descripcion, title="Descripción de las Compras", icon="🗂️"),
            st.Page(pagina_explorador,  title="Explorador de Gasto",         icon="🔍"),
            st.Page(pagina_historica,   title="Evolución Histórica",          icon="📈"),
        ],
        "Indicadores de Riesgo": [
            st.Page(pagina_riesgo,           title="Indicadores de Riesgo",    icon="🚨"),
            st.Page(pagina_ranking_riesgo,   title="Ranking de Riesgo",        icon="🏆"),
            st.Page(pagina_fragmentacion,    title="Fragmentación",             icon="🧩"),
            st.Page(pagina_colusion,       title="Simulación de Competencia", icon="🕸️"),
            st.Page(pagina_precios,        title="Analítica de Precios",     icon="💊"),
        ],
        "Herramientas": [
            st.Page(pagina_expediente,  title="Expediente de Contrato", icon="🔎"),
            st.Page(pagina_empresa,     title="Ficha de la Empresa",    icon="🏭"),
            st.Page(pagina_mapa_riesgo, title="Perfil UC",              icon="🗺️"),
            st.Page(pagina_ocds,        title="Exportar OCDS",          icon="📤"),
        ],
    }
)
pg.run()

# FOOTER
# ─────────────────────────────────────────────
st.divider()
st.caption(f"División de Monitoreo de la Integridad Institucional – IMSS | ComprasMX {_anios_label}")
